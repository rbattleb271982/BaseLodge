"""
Resort import utility — reads prod_resorts_full.xlsx into the database.

Accepts all dependencies as parameters so it can be called from both
the admin init-db route (inside the app context) and the standalone script.
"""
import os
import openpyxl

PASS_BRAND_MAP = {
    'MountainCollective': 'Mountain Collective',
}


def _normalize_pass_brands(raw):
    """Convert xlsx Pass Brands string to (pass_brands_str, pass_brands_list).

    'None'             -> ('None', ['None'])
    'Ikon'             -> ('Ikon', ['Ikon'])
    'Ikon,MountainCollective' -> ('Ikon,Mountain Collective', ['Ikon', 'Mountain Collective'])
    """
    if not raw or str(raw).strip() == 'None':
        return 'None', ['None']
    parts = [p.strip() for p in str(raw).split(',') if p.strip()]
    normalized = [PASS_BRAND_MAP.get(p, p) for p in parts]
    return ','.join(normalized), normalized


def import_resorts_from_xlsx(xlsx_path, db, Resort, generate_resort_slug, STATE_ABBR_MAP):
    """Upsert all resorts from prod_resorts_full.xlsx into the database.

    Idempotent — matches on (name, state_code, country_code) composite key.
    Handles same-named resorts in different states/countries correctly
    by appending a state/country suffix to the slug when needed.

    Returns a dict: {'added': int, 'updated': int}
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Resort xlsx not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    stats = {'added': 0, 'updated': 0}

    all_existing = Resort.query.all()
    existing_by_key = {}
    for r in all_existing:
        key = (
            (r.name or '').strip().lower(),
            (r.state_code or r.state or '').strip().upper(),
            (r.country_code or r.country or 'US').strip().upper(),
        )
        existing_by_key[key] = r

    used_slugs = {r.slug for r in all_existing}

    for row in ws.iter_rows(min_row=2, values_only=True):
        rid, name, country_code, country_name, state_region, pass_brands_raw, status = row

        if not name:
            continue

        name = name.strip()
        country_code = (country_code or 'US').strip().upper()
        country_name = (country_name or '').strip()
        state_code = (state_region or '').strip()
        is_active = str(status or '').strip().upper() == 'ACTIVE'

        pass_brands_str, pass_brands_list = _normalize_pass_brands(pass_brands_raw)
        brand = pass_brands_list[0] if pass_brands_list else 'Other'

        state_name = STATE_ABBR_MAP.get(state_code.upper()) if state_code else None
        if not state_name and state_code:
            state_name = state_code

        lookup_key = (name.lower(), state_code.upper(), country_code)
        existing = existing_by_key.get(lookup_key)

        if existing:
            existing.name = name
            existing.state = state_code
            existing.state_code = state_code
            existing.state_name = state_name
            existing.country = country_code
            existing.country_code = country_code
            existing.country_name = country_name
            existing.pass_brands = pass_brands_str
            existing.pass_brands_json = pass_brands_list
            existing.brand = brand
            existing.is_active = is_active
            stats['updated'] += 1
        else:
            base_slug = generate_resort_slug(name)
            slug = base_slug
            if slug in used_slugs:
                suffix = state_code.lower() if state_code else country_code.lower()
                slug = f"{base_slug}-{suffix}"
                counter = 2
                while slug in used_slugs:
                    slug = f"{base_slug}-{suffix}-{counter}"
                    counter += 1
            used_slugs.add(slug)

            resort = Resort(
                name=name,
                state=state_code,
                state_code=state_code,
                state_name=state_name,
                country=country_code,
                country_code=country_code,
                country_name=country_name,
                pass_brands=pass_brands_str,
                pass_brands_json=pass_brands_list,
                brand=brand,
                slug=slug,
                is_active=is_active,
            )
            db.session.add(resort)
            stats['added'] += 1

    db.session.commit()

    # Deactivate any existing resort whose (name, state_code, country_code) was not
    # present in the xlsx — these are old dev-seed resorts that production renamed or removed.
    xlsx_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid, name, country_code, country_name, state_region, pass_brands_raw, status = row
        if name:
            xlsx_keys.add((
                name.strip().lower(),
                (state_region or '').strip().upper(),
                (country_code or 'US').strip().upper(),
            ))

    deactivated = 0
    for key, resort in existing_by_key.items():
        if key not in xlsx_keys and resort.is_active:
            resort.is_active = False
            deactivated += 1

    if deactivated:
        db.session.commit()
        stats['deactivated'] = deactivated

    return stats
