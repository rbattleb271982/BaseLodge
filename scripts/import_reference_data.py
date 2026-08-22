#!/usr/bin/env python3
"""Safely import the reviewed resort reference catalog into non-production PostgreSQL.

The importer is deliberately independent of Flask, SQLAlchemy models, app.py,
startup migrations, and notification providers.  Validation is the default
mode; database writes require both the guarded environment and ``--apply``.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
import argparse
import hashlib
from pathlib import Path
import sys
from typing import Any, Callable, Mapping, Sequence

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

from runtime_config import (
    DatabaseConfiguration,
    RuntimeConfigurationError,
    resolve_reference_import_database_config,
)
from utils.countries import COUNTRIES


DEFAULT_WORKBOOK_PATH = (
    ROOT
    / "attached_assets"
    / "BaseLodge_Resort_Master_IMPORT_READY_v3_1778641272852.xlsx"
)
EXPECTED_WORKBOOK_SHA256 = (
    "00e830c1576092e859f2e951ff51d8a2bc570f622964d80040bed636a8f13e09"
)
REQUIRED_SHEETS = ("Resorts_Cleaned", "ResortPassMap_FIXED")
RESORT_COLUMNS = (
    "resort_slug",
    "resort_name",
    "state_code",
    "state_name",
    "country_code",
    "country_name",
    "is_active",
    "is_region",
    "official_resort_name",
    "entity_type",
)
MAPPING_COLUMNS = (
    "resort_slug",
    "resort_name",
    "mvp_pass_name",
    "source_pass_names",
    "source_row_count",
    "match_status",
    "note",
)
ALLOWED_PASS_NAMES = frozenset({"Epic", "Ikon", "Other"})
EXPECTED_RESORT_COUNT = 695
EXPECTED_MAPPING_COUNT = 245
EXPECTED_MAPPED_RESORT_COUNT = 212
EXPECTED_UNMAPPED_RESORT_COUNT = 483
ADVISORY_LOCK_KEY = 6_031_161


class ReferenceImportError(RuntimeError):
    """Raised when validation or a guarded reference import fails."""


@dataclass(frozen=True)
class ReferenceData:
    workbook_path: Path
    workbook_sha256: str
    resorts: tuple[dict[str, Any], ...]
    mappings: tuple[dict[str, str], ...]
    country_codes: tuple[str, ...]


@dataclass(frozen=True)
class ValidationSummary:
    resort_count: int
    mapping_count: int
    mapped_resort_count: int
    unmapped_resort_count: int
    country_codes: tuple[str, ...]
    duplicate_slugs: tuple[str, ...]
    duplicate_mapping_pairs: tuple[tuple[str, str], ...]
    missing_mapping_slugs: tuple[str, ...]
    invalid_pass_names: tuple[str, ...]
    invalid_country_codes: tuple[str, ...]

    @property
    def valid(self) -> bool:
        return not any(
            (
                self.duplicate_slugs,
                self.duplicate_mapping_pairs,
                self.missing_mapping_slugs,
                self.invalid_pass_names,
                self.invalid_country_codes,
            )
        )


def _header(sheet: Any, expected: Sequence[str]) -> None:
    values = tuple(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    if values[: len(expected)] != tuple(expected):
        raise ReferenceImportError(
            f"Worksheet {sheet.title} has unexpected columns; expected the reviewed layout."
        )


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _workbook_sha256(workbook_path: Path) -> str:
    try:
        return hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    except OSError as exc:
        raise ReferenceImportError("Reference workbook could not be read.") from exc


def _load_rows(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        workbook = openpyxl.load_workbook(
            workbook_path, read_only=True, data_only=True
        )
    except (OSError, ValueError) as exc:
        raise ReferenceImportError("Reference workbook could not be read.") from exc

    try:
        missing_sheets = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
        if missing_sheets:
            raise ReferenceImportError(
                "Reference workbook is missing required worksheet(s): "
                + ", ".join(missing_sheets)
            )

        resort_sheet = workbook["Resorts_Cleaned"]
        mapping_sheet = workbook["ResortPassMap_FIXED"]
        _header(resort_sheet, RESORT_COLUMNS)
        _header(mapping_sheet, MAPPING_COLUMNS)

        resorts = []
        for row in resort_sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            resorts.append(
                {
                    "slug": _text(row[0]),
                    "name": _text(row[1]),
                    "state_code": _text(row[2]),
                    "state_name": _text(row[3]),
                    "country_code": _text(row[4]).upper(),
                    "country_name": _text(row[5]),
                    "is_active": row[6],
                    "is_region": row[7],
                }
            )

        mappings = []
        for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            mappings.append(
                {
                    "slug": _text(row[0]),
                    "pass_name": _text(row[2]),
                }
            )
        return resorts, mappings
    finally:
        workbook.close()


def _validate_rows(
    resorts: list[dict[str, Any]],
    mappings: list[dict[str, Any]],
    *,
    workbook_path: Path,
    enforce_expected_counts: bool,
) -> tuple[ReferenceData, ValidationSummary]:
    resort_slugs = [row["slug"] for row in resorts]
    slug_counts = Counter(resort_slugs)
    duplicate_slugs = tuple(sorted(slug for slug, count in slug_counts.items() if count > 1))

    required_fields = (
        "slug",
        "name",
        "state_code",
        "state_name",
        "country_code",
        "country_name",
    )
    missing_fields = sorted(
        {
            field
            for row in resorts
            for field in required_fields
            if not row[field]
        }
    )
    invalid_flags = [
        field
        for row in resorts
        for field in ("is_active", "is_region")
        if not isinstance(row[field], bool)
    ]
    if missing_fields or invalid_flags:
        details = []
        if missing_fields:
            details.append("missing required fields: " + ", ".join(missing_fields))
        if invalid_flags:
            details.append("is_active/is_region must be boolean")
        raise ReferenceImportError("Reference workbook validation failed (" + "; ".join(details) + ").")

    source_slugs = set(resort_slugs)
    normalized_mappings = []
    for row in mappings:
        slug = row["slug"]
        pass_name = row["pass_name"]
        if not slug:
            raise ReferenceImportError("Reference workbook contains a mapping with no resort slug.")
        if not pass_name or pass_name.casefold() == "none":
            continue
        normalized_mappings.append({"slug": slug, "pass_name": pass_name})

    mapping_pairs = [(row["slug"], row["pass_name"]) for row in normalized_mappings]
    pair_counts = Counter(mapping_pairs)
    duplicate_mapping_pairs = tuple(
        sorted(pair for pair, count in pair_counts.items() if count > 1)
    )
    missing_mapping_slugs = tuple(
        sorted({row["slug"] for row in normalized_mappings} - source_slugs)
    )
    invalid_pass_names = tuple(
        sorted({row["pass_name"] for row in normalized_mappings} - ALLOWED_PASS_NAMES)
    )
    country_codes = tuple(sorted({row["country_code"] for row in resorts}))
    invalid_country_codes = tuple(
        sorted(code for code in country_codes if code not in COUNTRIES)
    )

    summary = ValidationSummary(
        resort_count=len(resorts),
        mapping_count=len(normalized_mappings),
        mapped_resort_count=len({row["slug"] for row in normalized_mappings}),
        unmapped_resort_count=len(source_slugs - {row["slug"] for row in normalized_mappings}),
        country_codes=country_codes,
        duplicate_slugs=duplicate_slugs,
        duplicate_mapping_pairs=duplicate_mapping_pairs,
        missing_mapping_slugs=missing_mapping_slugs,
        invalid_pass_names=invalid_pass_names,
        invalid_country_codes=invalid_country_codes,
    )
    if enforce_expected_counts and (
        summary.resort_count != EXPECTED_RESORT_COUNT
        or summary.mapping_count != EXPECTED_MAPPING_COUNT
        or summary.mapped_resort_count != EXPECTED_MAPPED_RESORT_COUNT
        or summary.unmapped_resort_count != EXPECTED_UNMAPPED_RESORT_COUNT
    ):
        raise ReferenceImportError(
            "Reference workbook counts do not match the reviewed canonical source."
        )
    if not summary.valid:
        raise ReferenceImportError(_validation_error(summary))

    data = ReferenceData(
        workbook_path=workbook_path.resolve(),
        workbook_sha256=_workbook_sha256(workbook_path),
        resorts=tuple(resorts),
        mappings=tuple(normalized_mappings),
        country_codes=country_codes,
    )
    return data, summary


def _validation_error(summary: ValidationSummary) -> str:
    findings = []
    if summary.duplicate_slugs:
        findings.append("duplicate slugs: " + ", ".join(summary.duplicate_slugs))
    if summary.duplicate_mapping_pairs:
        findings.append("duplicate mapping pairs")
    if summary.missing_mapping_slugs:
        findings.append("missing mapping slugs: " + ", ".join(summary.missing_mapping_slugs))
    if summary.invalid_pass_names:
        findings.append("invalid pass names: " + ", ".join(summary.invalid_pass_names))
    if summary.invalid_country_codes:
        findings.append("invalid country codes: " + ", ".join(summary.invalid_country_codes))
    return "Reference workbook validation failed: " + "; ".join(findings) + "."


def load_reference_data(
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    *,
    enforce_expected_counts: bool = True,
) -> tuple[ReferenceData, ValidationSummary]:
    """Load and validate the canonical workbook without opening a database."""
    resorts, mappings = _load_rows(workbook_path)
    return _validate_rows(
        resorts,
        mappings,
        workbook_path=workbook_path,
        enforce_expected_counts=enforce_expected_counts,
    )


def validation_report(summary: ValidationSummary) -> str:
    duplicate_slugs = ", ".join(summary.duplicate_slugs) or "none"
    duplicate_pairs = str(len(summary.duplicate_mapping_pairs))
    missing_slugs = ", ".join(summary.missing_mapping_slugs) or "none"
    invalid_passes = ", ".join(summary.invalid_pass_names) or "none"
    invalid_countries = ", ".join(summary.invalid_country_codes) or "none"
    return "\n".join(
        (
            "Workbook validation: PASS",
            f"Planned resorts: {summary.resort_count}",
            f"Planned resort-pass mappings: {summary.mapping_count}",
            f"Mapped resorts: {summary.mapped_resort_count}",
            f"Resorts without a major-pass mapping: {summary.unmapped_resort_count}",
            f"Country-code validation: {'PASS' if not summary.invalid_country_codes else 'FAIL'}",
            f"Duplicate slugs: {duplicate_slugs}",
            f"Duplicate mapping pairs: {duplicate_pairs}",
            f"Missing mapping slugs: {missing_slugs}",
            f"Invalid pass names: {invalid_passes}",
            f"Invalid country codes: {invalid_countries}",
        )
    )


def _connect(configuration: DatabaseConfiguration):
    try:
        connection = psycopg2.connect(configuration.database_url)
    except psycopg2.Error as exc:
        raise ReferenceImportError(
            "Unable to connect to the guarded reference-import target."
        ) from exc
    connection.autocommit = False
    return connection


def _upsert_resorts(connection: Any, resorts: Sequence[dict[str, Any]]) -> dict[str, int]:
    statement = """
        INSERT INTO resort (
            slug, name, state, state_full, country, country_code, country_name,
            state_code, state_name, is_active, is_region
        ) VALUES %s
        ON CONFLICT (slug) DO UPDATE SET
            name = EXCLUDED.name,
            state = EXCLUDED.state,
            state_full = EXCLUDED.state_full,
            country = EXCLUDED.country,
            country_code = EXCLUDED.country_code,
            country_name = EXCLUDED.country_name,
            state_code = EXCLUDED.state_code,
            state_name = EXCLUDED.state_name,
            is_active = EXCLUDED.is_active,
            is_region = EXCLUDED.is_region
    """
    values = [
        (
            row["slug"],
            row["name"],
            row["state_code"],
            row["state_name"],
            row["country_code"],
            row["country_code"],
            row["country_name"],
            row["state_code"],
            row["state_name"],
            row["is_active"],
            row["is_region"],
        )
        for row in resorts
    ]
    with connection.cursor() as cursor:
        execute_values(cursor, statement, values)
        cursor.execute(
            "SELECT id, slug FROM resort WHERE slug = ANY(%s)",
            ([row["slug"] for row in resorts],),
        )
        ids = {slug: resort_id for resort_id, slug in cursor.fetchall()}
    if len(ids) != len(resorts):
        raise ReferenceImportError("Reference import could not resolve every resort slug.")
    return ids


def _replace_resort_passes(
    connection: Any,
    mappings: Sequence[dict[str, str]],
    resort_ids: Mapping[str, int],
) -> None:
    ids = list(resort_ids.values())
    mappings_by_slug: dict[str, list[str]] = {}
    for row in mappings:
        mappings_by_slug.setdefault(row["slug"], []).append(row["pass_name"])

    primary_by_slug = {
        slug: next(
            pass_name
            for pass_name in ("Epic", "Ikon", "Other")
            if pass_name in pass_names
        )
        for slug, pass_names in mappings_by_slug.items()
    }
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM resort_pass WHERE resort_id = ANY(%s)", (ids,))
        values = [
            (
                resort_ids[row["slug"]],
                row["pass_name"],
                row["pass_name"] == primary_by_slug[row["slug"]],
            )
            for row in mappings
        ]
        if values:
            execute_values(
                cursor,
                """
                INSERT INTO resort_pass (resort_id, pass_name, is_primary)
                VALUES %s
                """,
                values,
            )


def apply_reference_data(
    data: ReferenceData,
    configuration: DatabaseConfiguration,
    *,
    connection_factory: Callable[[str], Any] = psycopg2.connect,
    fail_after_resorts: bool = False,
) -> None:
    """Apply validated reference data in one transaction."""
    if (
        data.workbook_path != DEFAULT_WORKBOOK_PATH.resolve()
        or data.workbook_sha256 != EXPECTED_WORKBOOK_SHA256
        or _workbook_sha256(DEFAULT_WORKBOOK_PATH) != EXPECTED_WORKBOOK_SHA256
    ):
        raise ReferenceImportError(
            "Reference import apply only accepts the reviewed canonical workbook."
        )
    try:
        connection = connection_factory(configuration.database_url)
        connection.autocommit = False
    except psycopg2.Error as exc:
        raise ReferenceImportError(
            "Unable to connect to the guarded reference-import target."
        ) from exc

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT pg_advisory_xact_lock(%s)", (ADVISORY_LOCK_KEY,))
        resort_ids = _upsert_resorts(connection, data.resorts)
        if fail_after_resorts:
            raise RuntimeError("forced reference-import failure")
        _replace_resort_passes(connection, data.mappings, resort_ids)
        connection.commit()
    except ReferenceImportError:
        connection.rollback()
        raise
    except Exception as exc:
        connection.rollback()
        raise ReferenceImportError(
            "Reference import failed and was rolled back."
        ) from exc
    finally:
        connection.close()


def run_import(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    apply: bool = False,
    environ: Mapping[str, str] | None = None,
    connection_factory: Callable[[str], Any] = psycopg2.connect,
    fail_after_resorts: bool = False,
) -> ValidationSummary:
    """Validate the source and optionally apply it under explicit authorization."""
    try:
        configuration = resolve_reference_import_database_config(environ)
    except RuntimeConfigurationError as exc:
        raise ReferenceImportError(str(exc)) from exc
    data, summary = load_reference_data(workbook_path)
    if apply:
        apply_reference_data(
            data,
            configuration,
            connection_factory=connection_factory,
            fail_after_resorts=fail_after_resorts,
        )
    return summary


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate or safely import BaseLodge resort reference data."
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Apply the validated workbook; omitted means non-mutating dry-run.",
    )
    args = parser.parse_args(argv)
    try:
        summary = run_import(apply=args.apply)
        print(validation_report(summary))
        if args.apply:
            print("Reference import committed successfully.")
        else:
            print("Dry-run: no database changes were written.")
        return 0
    except ReferenceImportError as exc:
        print(f"Reference import refused or failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())