"""
Resort + ResortPass import script.

Reads:
  - Resorts_Cleaned   tab → upserts Resort rows by slug
  - ResortPassMap_FIXED tab → replaces ResortPass rows per resort

Usage:
    python scripts/import_resorts.py [--dry-run]

Pass --dry-run to preview the import summary without writing to the database.
Idempotent: safe to re-run.
"""

import sys
import os
import argparse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

XLSX_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'attached_assets',
    'BaseLodge_Resort_Master_IMPORT_READY_v3_1778641272852.xlsx',
)

MVP_PASS_NORM = {
    'Epic':               'Epic',
    'Ikon':               'Ikon',
    'Other':              'Other',
    'Mountain Collective':'Other',
    'Indy':               'Other',
    'Freedom':            'Other',
    'Powder Alliance':    'Other',
    'Ski California':     'Other',
    'None':               None,
}


def _load_workbook():
    if not os.path.exists(XLSX_PATH):
        print(f"[ERROR] Workbook not found: {XLSX_PATH}")
        sys.exit(1)
    return openpyxl.load_workbook(XLSX_PATH)


def _build_pass_map(wb):
    """Returns dict: slug → MVP pass name (or None for no pass)."""
    ws = wb['ResortPassMap_FIXED']
    pass_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        slug = row[0]
        mvp_pass_raw = row[2]
        if not slug:
            continue
        slug = str(slug).strip()
        mvp_pass = str(mvp_pass_raw).strip() if mvp_pass_raw else None
        pass_map[slug] = MVP_PASS_NORM.get(mvp_pass) if mvp_pass else None
    return pass_map


def _load_resorts(wb):
    """Returns list of dicts from Resorts_Cleaned."""
    ws = wb['Resorts_Cleaned']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        slug = row[0]
        name = row[1]
        if not slug or not name:
            continue
        state_code    = str(row[2]).strip() if row[2] else ''
        state_name    = str(row[3]).strip() if row[3] else state_code
        country_code  = str(row[4]).strip() if row[4] else 'US'
        country_name  = str(row[5]).strip() if row[5] else ''
        is_active     = bool(row[6]) if row[6] is not None else True
        is_region     = bool(row[7]) if row[7] is not None else False
        rows.append({
            'slug':         str(slug).strip(),
            'name':         str(name).strip(),
            'state_code':   state_code,
            'state_name':   state_name,
            'country_code': country_code,
            'country_name': country_name,
            'is_active':    is_active,
            'is_region':    is_region,
        })
    return rows


def run_import(dry_run=False):
    from app import app, db
    from models import Resort, ResortPass

    wb = _load_workbook()
    pass_map    = _build_pass_map(wb)
    resort_rows = _load_resorts(wb)

    # Warn on duplicate slugs in xlsx
    seen = {}
    for rd in resort_rows:
        s = rd['slug']
        if s in seen:
            print(f"  [WARN] Duplicate slug in workbook: {s}")
        seen[s] = True

    stats = {'added': 0, 'updated': 0, 'pass_deleted': 0, 'pass_added': 0,
             'no_change': 0, 'warnings': 0}

    with app.app_context():
        existing_by_slug = {r.slug: r for r in Resort.query.all()}

        for rd in resort_rows:
            slug = rd['slug']
            existing = existing_by_slug.get(slug)

            if existing:
                changed = (
                    existing.name         != rd['name']         or
                    existing.state_code   != rd['state_code']   or
                    existing.country_code != rd['country_code'] or
                    existing.is_active    != rd['is_active']    or
                    existing.is_region    != rd['is_region']
                )
                if changed:
                    if not dry_run:
                        existing.name         = rd['name']
                        existing.state        = rd['state_code']
                        existing.state_code   = rd['state_code']
                        existing.state_name   = rd['state_name']
                        existing.state_full   = rd['state_name']
                        existing.country      = rd['country_code']
                        existing.country_code = rd['country_code']
                        existing.country_name = rd['country_name']
                        existing.is_active    = rd['is_active']
                        existing.is_region    = rd['is_region']
                    stats['updated'] += 1
                else:
                    stats['no_change'] += 1
            else:
                if not dry_run:
                    resort = Resort(
                        slug         = slug,
                        name         = rd['name'],
                        state        = rd['state_code'],
                        state_code   = rd['state_code'],
                        state_name   = rd['state_name'],
                        state_full   = rd['state_name'],
                        country      = rd['country_code'],
                        country_code = rd['country_code'],
                        country_name = rd['country_name'],
                        is_active    = rd['is_active'],
                        is_region    = rd['is_region'],
                    )
                    db.session.add(resort)
                stats['added'] += 1

        if not dry_run:
            # Advance the sequence past the current max ID to avoid PK conflicts
            db.session.execute(db.text(
                "SELECT setval(pg_get_serial_sequence('resort', 'id'), "
                "GREATEST((SELECT COALESCE(MAX(id), 0) FROM resort), "
                "nextval(pg_get_serial_sequence('resort', 'id')) - 1))"
            ))
            db.session.flush()

        # Re-query so newly added resorts have IDs
        all_by_slug = {r.slug: r for r in Resort.query.all()} if not dry_run else existing_by_slug

        xlsx_slugs = {rd['slug'] for rd in resort_rows}

        for slug in xlsx_slugs:
            resort = all_by_slug.get(slug)

            if not resort:
                # New resort (only exists in dry-run pre-flush state) — count the pass
                mvp_pass = pass_map.get(slug)
                if mvp_pass and mvp_pass != 'None':
                    stats['pass_added'] += 1
                continue

            if not dry_run:
                deleted = ResortPass.query.filter_by(resort_id=resort.id).delete()
                stats['pass_deleted'] += deleted
            else:
                existing_rp = ResortPass.query.filter_by(resort_id=resort.id).count()
                stats['pass_deleted'] += existing_rp

            mvp_pass = pass_map.get(slug)
            if mvp_pass and mvp_pass != 'None':
                if not dry_run:
                    rp = ResortPass(
                        resort_id = resort.id,
                        pass_name = mvp_pass,
                        is_primary= True,
                    )
                    db.session.add(rp)
                stats['pass_added'] += 1

        if not dry_run:
            db.session.commit()
            print(f"[OK] Committed to database.")

    return stats


def main():
    parser = argparse.ArgumentParser(description='Import BaseLodge resort + pass data from Excel.')
    parser.add_argument('--dry-run', action='store_true',
                        help='Preview changes without writing to the database.')
    args = parser.parse_args()

    mode = 'DRY RUN' if args.dry_run else 'LIVE'
    print(f"\n{'='*60}")
    print(f"  BaseLodge Resort Import — {mode}")
    print(f"  Source: {os.path.basename(XLSX_PATH)}")
    print(f"{'='*60}")

    stats = run_import(dry_run=args.dry_run)

    print(f"\n  Resorts")
    print(f"    Added          : {stats['added']}")
    print(f"    Updated        : {stats['updated']}")
    print(f"    No change      : {stats['no_change']}")
    print(f"\n  ResortPass rows")
    print(f"    Deleted        : {stats['pass_deleted']}")
    print(f"    Added          : {stats['pass_added']}")
    if stats.get('warnings'):
        print(f"\n  Warnings       : {stats['warnings']}")
    if args.dry_run:
        print(f"\n  [DRY RUN] No changes written. Re-run without --dry-run to apply.")
    print()


if __name__ == '__main__':
    main()
