"""
========================================
BaseLodge Application
========================================

SYSTEM OF RECORD (as of 2026-01-15):
  Supabase is the single system of record for resorts and core app data.
  
  - Resort data: 693 resorts imported from prod_resorts_full.xlsx
  - Schema: Managed exclusively via Flask-Migrate/Alembic
  - Database: SUPABASE_DATABASE_URL environment variable
  
  DO NOT use db.create_all() or run legacy seed scripts.
  All schema changes must go through migrations.
========================================
"""

import os
import secrets
import time
import json
import threading
import jwt
import httpx
from datetime import datetime, date, timedelta, timezone
from zoneinfo import ZoneInfo
from sqlalchemy.orm import joinedload
from services.pass_utils import (
    normalize_pass, display_pass_label, normalize_passes_string,
    format_passes_for_display, passes_match, is_real_pass,
    normalize_pass_selection, count_real_passes,
    PASS_NORM_MAP, PASS_DISPLAY_MAP, CANONICAL_PASS_ORDER,
)
from constants.equipment import SKI_BRANDS, SNOWBOARD_BRANDS, BOOT_BRANDS, BINDING_TYPES, BINDING_BRANDS_BY_TYPE

# ── Admin timezone helpers — America/Denver display ───────────────────────────
_ADMIN_TZ = ZoneInfo("America/Denver")

# ── Founder login-alert throttle cache ────────────────────────────────────────
# Maps user_id → Denver date string ("YYYY-MM-DD").  Resets on server restart.
# Used by _queue_founder_login_push to suppress duplicate pushes within a day.
_FOP_THROTTLE: dict = {}

def _admin_now():
    """Current datetime in America/Denver (timezone-aware)."""
    return datetime.now(tz=_ADMIN_TZ)

def _admin_today_start_utc():
    """Naive UTC datetime for the start of today in America/Denver."""
    denver_midnight = _admin_now().replace(hour=0, minute=0, second=0, microsecond=0)
    return denver_midnight.astimezone(timezone.utc).replace(tzinfo=None)

def _admin_yesterday_start_utc():
    """Naive UTC datetime for the start of yesterday in America/Denver."""
    denver_midnight = _admin_now().replace(hour=0, minute=0, second=0, microsecond=0)
    return (denver_midnight - timedelta(days=1)).astimezone(timezone.utc).replace(tzinfo=None)

def _fmt_admin_now():
    """Return formatted Denver timestamp, e.g. '2026-05-29 10:45 MDT'."""
    n = _admin_now()
    return n.strftime("%Y-%m-%d %H:%M") + " " + n.strftime("%Z")

# ─────────────────────────────────────────────────────────────────────────────

def _resolve_base_url():
    """Resolve the base URL for invite links and other absolute URLs.

    Priority:
    1. BASE_URL env var — explicit override; always wins (set this in production
       secrets to https://app.baselodgeapp.com).
    2. REPLIT_DEV_DOMAIN — present in Replit IDE/dev environments; used only
       when no explicit BASE_URL is configured.
    3. Hardcoded production fallback.
    """
    explicit = os.getenv("BASE_URL")
    if explicit:
        resolved = explicit.rstrip("/")
        print(f"[BASE_URL] Using BASE_URL env var: {resolved}")
        return resolved
    replit_domain = os.getenv("REPLIT_DEV_DOMAIN")
    if replit_domain:
        resolved = f"https://{replit_domain}"
        print(f"[BASE_URL] Dev mode — using REPLIT_DEV_DOMAIN: {resolved}")
        return resolved
    return "https://app.baselodgeapp.com"

BASE_URL = _resolve_base_url()
import sqlalchemy as sa
from sqlalchemy import func
from urllib.parse import urlparse
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, abort, send_file, current_app, g, make_response
from flask_login import LoginManager, login_required, current_user, login_user, logout_user, user_loaded_from_cookie
from functools import wraps, lru_cache
from flask_migrate import Migrate
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from authlib.integrations.flask_client import OAuth
from models import db, User, SkiTrip, Friend, Invitation, InviteToken, TripInviteToken, Resort, ResortPass, GroupTrip, TripGuest, GuestStatus, check_shared_upcoming_trip, EquipmentSetup, EquipmentSlot, EquipmentDiscipline, AccommodationStatus, TransportationStatus, DismissedNudge, DismissedInsightCard, Event, EmailLog, SkiTripParticipant, ParticipantRole, ParticipantTransportation, ParticipantEquipment, Activity, ActivityType, LessonChoice, CarpoolRole, InviteType, PushDeviceToken, UserAvailability, MessageEventLog, MountainPageView, InviteShareEvent
from services.open_dates import get_open_date_matches, get_available_dates_for_user
from services.ideas_engine import build_overlap_windows, build_wishlist_overlaps
from services.message_events import create_message_event, is_duplicate_event, should_retry
from services.messaging_constants import (
    EventName, Category, DeliveryStatus, SuppressionReason, Channel, Provider,
    MAX_RETRY_COUNT, RETRYABLE_STATUSES,
)
from services.push_providers import send_onesignal_push, send_onesignal_custom_event
from services.message_dispatch import emit_messaging_event
from io import BytesIO
import segno
import random
import click
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
import unicodedata
import re
import analytics as ph_analytics


def generate_resort_slug(name):
    """Generate a URL-safe slug from resort name.
    
    - Lowercase
    - Hyphen-separated
    - ASCII-safe (unicode normalized)
    - Deterministic (same name -> same slug)
    - Never null or empty
    """
    if not name:
        raise ValueError("Resort name is required for slug generation")
    
    # Normalize unicode to ASCII
    slug = unicodedata.normalize('NFKD', str(name))
    slug = slug.encode('ascii', 'ignore').decode('ascii')
    
    # Lowercase
    slug = slug.lower()
    
    # Replace non-alphanumeric with hyphens
    slug = re.sub(r'[^a-z0-9]+', '-', slug)
    
    # Collapse multiple hyphens
    slug = re.sub(r'-+', '-', slug)
    
    # Strip leading/trailing hyphens
    slug = slug.strip('-')
    
    if not slug:
        raise ValueError(f"Could not generate valid slug from name: {name}")
    
    return slug

# ============================================================================
# PROFILE CONSOLIDATION NOTE:
# The app no longer uses /profile or profile.html.
# All profile-related UI lives under /more.
# Do NOT reintroduce profile routes or templates.
# ============================================================================

is_production = os.environ.get("SUPABASE_DATABASE_URL") is not None and "postgresql" in os.environ.get("SUPABASE_DATABASE_URL", "")

# Enforce Supabase connection for resort-related operations
SUPABASE_URL = os.environ.get("SUPABASE_DATABASE_URL")
if not SUPABASE_URL:
    if is_production:
        raise RuntimeError("CRITICAL: SUPABASE_DATABASE_URL is not set in production. App startup aborted.")
    else:
        # Loud warning for development
        print("!" * 70)
        print("⚠️  WARNING: SUPABASE_DATABASE_URL is not set.")
        print("⚠️  Development will fallback to local SQLite but Resort data will be MISSING.")
        print("!" * 70)

app = Flask(__name__)
app.config["PREFERRED_URL_SCHEME"] = "https"

# ── Response compression ───────────────────────────────────────────────────
# Gzip HTML/JSON responses — typically 60-70% size reduction on page HTML.
# Skips already-compressed content (images, pre-gzipped assets).
# Threshold 500 bytes avoids compressing tiny API responses.
from flask_compress import Compress as _Compress
_Compress(app)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://",
)

def _user_or_ip():
    from flask_login import current_user as _cu
    try:
        if _cu.is_authenticated:
            return f"user:{_cu.id}"
    except Exception:
        pass
    return get_remote_address()

# ── CSRF helpers ──────────────────────────────────────────────────────────────
def generate_csrf_token():
    """Return a per-session CSRF token, creating one if absent."""
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)
    return session['_csrf_token']

def validate_csrf_request():
    """Abort 403 if the CSRF token in the request does not match the session."""
    session_token = session.get('_csrf_token')
    request_token = (
        request.form.get('csrf_token')
        or request.headers.get('X-CSRF-Token')
        or request.headers.get('X-CSRFToken')
    )
    if not session_token or not secrets.compare_digest(session_token, request_token or ''):
        abort(403)

app.jinja_env.globals['csrf_token'] = generate_csrf_token

@app.before_request
def redirect_to_canonical_domain():
    parsed_url = urlparse(request.url)
    hostname = parsed_url.hostname.lower() if parsed_url.hostname else ""

    # Allow the Replit dev workspace preview through — do not redirect it.
    # REPLIT_DEV_DOMAIN is only set in the Replit dev environment (not in
    # production or deployed Replit apps), so this exemption is dev-only.
    replit_dev_domain = os.environ.get("REPLIT_DEV_DOMAIN", "").lower()
    if replit_dev_domain and hostname == replit_dev_domain:
        return None

    if hostname.endswith("replit.app") or hostname.endswith("replit.dev"):
        new_url = request.url.replace(
            f"{parsed_url.scheme}://{parsed_url.netloc}",
            "https://app.baselodgeapp.com",
            1
        )
        return redirect(new_url, code=301)

# ============================================================================
# SESSION & SECURITY CONFIGURATION
# ============================================================================
app.config["SECRET_KEY"] = os.environ.get("SESSION_SECRET")
if not app.config["SECRET_KEY"]:
    if not is_production:
        app.config["SECRET_KEY"] = "dev-secret-key-fallback"
    else:
        raise RuntimeError("SESSION_SECRET environment variable is NOT SET in production.")

# Session configuration for Replit iframe environment
app.config.update(
    SESSION_COOKIE_SECURE=is_production,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE=os.environ.get("SESSION_COOKIE_SAMESITE", "Lax"),
    PERMANENT_SESSION_LIFETIME=timedelta(days=7),
    SESSION_REFRESH_EACH_REQUEST=False,
    REMEMBER_COOKIE_SECURE=is_production,
    REMEMBER_COOKIE_HTTPONLY=True,
    REMEMBER_COOKIE_DURATION=timedelta(days=30),
)

# Navigation debug flag — set BL_NAV_DEBUG=1 in environment to enable.
# Server: prints [BL-NAV] timing lines per route to stdout.
# Client: window.__BL_NAV_DEBUG__ controls browser console timing output.
# When disabled: zero overhead — no logs, no JS cost.
app.config["BL_NAV_DEBUG"] = os.environ.get("BL_NAV_DEBUG", "0").strip() == "1"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "auth"
login_manager.login_message = None

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@user_loaded_from_cookie.connect_via(app)
def _on_user_loaded_from_cookie(sender, user, **kwargs):
    """Log whenever Flask-Login silently restores a session from a remember-me cookie.
    Fires once per request where the cookie — not the session — is the auth source.
    Does NOT log passwords, tokens, or cookie contents.
    """
    try:
        app.logger.info(
            "[auth_restore] method=remember_cookie user_id=%s email=%s "
            "ua=%.100s ts=%s ip=%s",
            user.id,
            user.email,
            request.user_agent.string,
            datetime.utcnow().isoformat(),
            request.remote_addr,
        )
    except Exception:
        pass


from utils.countries import COUNTRIES, STATE_ABBR_MAP

@app.context_processor
def inject_countries():
    return {'COUNTRIES': COUNTRIES}


_NOTIF_TYPES = [
    'join_request_received', 'join_request_accepted', 'join_request_declined',
    'connection_accepted', 'trip_invite_received', 'trip_invite_accepted', 'trip_invite_declined',
    'trip_location_changed', 'trip_pass_changed',
]


@app.context_processor
def inject_notif_count():
    """Inject notification unread count into every template."""
    if not current_user.is_authenticated:
        return {'notif_unread_count': 0}
    try:
        q = Activity.query.filter(
            Activity.recipient_user_id == current_user.id,
            Activity.type.in_(_NOTIF_TYPES)
        )
        last_viewed = session.get('notif_last_viewed_at')
        if last_viewed:
            try:
                last_viewed_dt = datetime.fromisoformat(last_viewed)
                q = q.filter(Activity.created_at > last_viewed_dt)
            except (ValueError, TypeError):
                pass
        return {'notif_unread_count': q.count()}
    except Exception:
        return {'notif_unread_count': 0}


@app.context_processor
def inject_pending_friend_count():
    """Inject pending friend-request count into every app-shell template for the nav badge."""
    if not current_user.is_authenticated:
        return {'pending_friend_count': 0}
    try:
        count = Invitation.query.filter(
            Invitation.receiver_id == current_user.id,
            Invitation.status == 'pending',
            Invitation.trip_id.is_(None)
        ).count()
        return {'pending_friend_count': count}
    except Exception:
        return {'pending_friend_count': 0}


# Helper to normalize country code
def normalize_country_code(code):
    if not code:
        return None
    code = code.strip().upper()
    if code in COUNTRIES:
        return code
    return None

@app.route("/health")
def health_check():
    """Health check endpoint for production probes. Returns JSON status and DB connectivity check."""
    try:
        # Lightweight query to verify DB connectivity
        db.session.execute(sa.text("SELECT 1")).fetchone()
        return jsonify({
            "status": "healthy",
            "database": "connected",
            "environment": os.environ.get("FLASK_ENV", "development"),
            "timestamp": datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        app.logger.error(f"Health check failed: {e}")
        return jsonify({
            "status": "unhealthy",
            "database": "disconnected",
            "error": str(e) if not is_production else "Internal Server Error"
        }), 500

def get_upcoming_trip_count(user):
    """
    Returns the count of unique upcoming trips a user is committed to.
    'Committed' means being the owner or an ACCEPTED participant.
    Filters:
    - Deduplicate by trip.id
    - Includes owner or ACCEPTED participant role
    - Filters for end_date >= today (includes in-progress trips, consistent with all route queries)
    - Excludes past, canceled, archived, or pending states
    """
    if not user:
        return 0
    
    today = date.today()
    
    # 1. Trips owned by the user
    owned_trips = SkiTrip.query.filter(
        SkiTrip.user_id == user.id,
        SkiTrip.end_date >= today
    ).all()

    # 2. Trips where the user is an ACCEPTED participant
    participant_trips = (
        db.session.query(SkiTrip)
        .join(SkiTripParticipant, SkiTrip.id == SkiTripParticipant.trip_id)
        .filter(
            SkiTripParticipant.user_id == user.id,
            SkiTripParticipant.status == GuestStatus.ACCEPTED,
            SkiTrip.end_date >= today
        )
        .all()
    )

    # Deduplicate by trip ID
    all_upcoming_trips = {t.id for t in owned_trips}
    for t in participant_trips:
        all_upcoming_trips.add(t.id)

    return len(all_upcoming_trips)


def get_past_trip_count(user):
    """Returns the count of trips owned by user that have already ended."""
    if not user:
        return 0
    return SkiTrip.query.filter(
        SkiTrip.user_id == user.id,
        SkiTrip.end_date < date.today()
    ).count()


def get_trip_status(trip, today=None):
    """
    Canonical trip phase classification.
    Returns one of: 'past', 'in_progress', 'upcoming'.

    Rules:
      past:        end_date < today
      in_progress: start_date <= today <= end_date
      upcoming:    start_date > today

    Trips with missing dates default to 'upcoming'.
    """
    if today is None:
        today = date.today()
    if not trip.start_date or not trip.end_date:
        return 'upcoming'
    if trip.end_date < today:
        return 'past'
    if trip.start_date <= today:
        return 'in_progress'
    return 'upcoming'


@app.template_filter('display_name')
def display_name_filter(user, current_user_id=None):
    """
    Returns full name (First Last) or 'You' if viewing own profile.
    Usage in templates: {{ user|display_name(current_user.id) }}
    """
    try:
        if not user:
            return ""
        user_id = getattr(user, 'id', None)
        if current_user_id and user_id == current_user_id:
            return "You"
        first = getattr(user, 'first_name', '') or ''
        last = getattr(user, 'last_name', '') or ''
        full_name = f"{first} {last}".strip()
        return full_name if full_name else "Friend"
    except Exception:
        return "Friend"


def format_passes_display(pass_type):
    """
    Format a pass_type string for user-facing display.
    Delegates to pass_utils.format_passes_for_display which handles
    normalization and includes no_pass / no_pass_yet states.
    """
    return format_passes_for_display(pass_type)


@app.template_filter('rider_display')
def rider_display_filter(raw):
    """
    Normalize a raw stored rider-type value (or display_rider_type output) to its
    canonical display label.  Safe to pipe any string through — unknown values
    pass through unchanged.

    Mappings:
        "Cross-Country" → "Cross-country"
        All other values pass through unchanged.
    """
    from models import _fmt_rider
    if not raw:
        return raw or ""
    # display_rider_type joins multi-type values with " + "; normalize each part
    parts = str(raw).split(" + ")
    return " + ".join(_fmt_rider(p.strip()) for p in parts)


@app.template_filter('identity_line')
def identity_line_filter(user):
    """
    Formats user identity as: Rider type · Level · Pass(es)
    Examples:
        Skier · Advanced · Epic
        Skier · Advanced · No pass
        Skier · Advanced · No pass yet
    """
    try:
        if not user:
            return ""

        parts = []

        display_rider = getattr(user, 'display_rider_type', None)
        if display_rider:
            parts.append(display_rider)

        skill_level = getattr(user, "skill_level", None)
        if skill_level:
            parts.append(str(skill_level))

        pass_type = getattr(user, "pass_type", None)
        if pass_type:
            formatted_passes = format_passes_for_display(pass_type)
            if formatted_passes:
                for p in formatted_passes.split(' · '):
                    parts.append(p)

        return " · ".join(parts)

    except Exception:
        return ""


@app.template_filter('pass_display')
def pass_display_filter(pass_type):
    """
    Display a pass_type string for UI (e.g. stats card, header).
    Returns 'Epic · Ikon' or 'No pass yet' etc.
    """
    return format_passes_for_display(pass_type)


@app.template_filter('normalize_pass')
def normalize_pass_filter(pass_type):
    """Jinja filter: normalize a raw pass value to snake_case."""
    return normalize_pass(pass_type)


@app.template_filter('country_name')
def country_name_filter(code):
    """
    Returns full country name from ISO code.
    Usage: {{ resort.country_code|country_name }}
    """
    if not code:
        return ""
    return COUNTRIES.get(code.upper(), code)


@app.template_filter('mountain_passes')
def mountain_passes_filter(resort):
    """
    Formats mountain pass brands as 'Epic · Ikon' (no "Pass" suffix).
    Returns empty string if no passes.
    Usage: {{ trip.resort|mountain_passes }}
    """
    if not resort:
        return ""
    passes = []
    if hasattr(resort, 'get_passes'):
        try:
            passes = resort.get_passes() or []
        except Exception:
            passes = []
    if not passes:
        pass_brands = getattr(resort, 'pass_brands_json', None) or getattr(resort, 'pass_brands', None)
        if not pass_brands:
            return ""
        if isinstance(pass_brands, list):
            brands = [str(b).strip() for b in pass_brands if str(b).strip()]
        else:
            brands = [b.strip() for b in str(pass_brands).split(',') if b.strip()]
        passes = [{'pass_name': b, 'is_primary': False} for b in brands]
    brands = [p.get('pass_name') for p in passes if p.get('pass_name') and p.get('pass_name') != 'None']
    if not brands:
        return ""
    return ' · '.join(brands)


@app.template_filter('state_abbrev')
def state_abbrev_filter(resort_or_code):
    """
    Returns state abbreviation (e.g., 'CO').
    Accepts resort object or state_code string.
    Usage: {{ trip.resort|state_abbrev }} or {{ 'Colorado'|state_abbrev }}
    """
    if not resort_or_code:
        return ""
    # If it's a resort object, get state_code
    if hasattr(resort_or_code, 'state_code'):
        return resort_or_code.state_code or ""
    # If it's already a short code (2-3 chars), return as-is
    if isinstance(resort_or_code, str) and len(resort_or_code) <= 3:
        return resort_or_code
    return ""


@app.template_filter('state_fullname')
def state_fullname_filter(resort_or_code):
    """
    Returns full state name (e.g., 'Colorado').
    Accepts resort object or state_name string.
    Falls back to state_code or state field if state_name not available.
    Usage: {{ trip.resort|state_fullname }}
    """
    if not resort_or_code:
        return ""
    # If it's a resort object, get state_name with fallbacks
    if hasattr(resort_or_code, 'state_name'):
        return resort_or_code.state_name or getattr(resort_or_code, 'state', '') or ""
    # Also try state field directly (for legacy objects)
    if hasattr(resort_or_code, 'state'):
        return resort_or_code.state or ""
    # If it's already a string, return as-is
    if isinstance(resort_or_code, str):
        return resort_or_code
    return ""


@app.template_filter('resort_display')
def resort_display_filter(resort):
    """Jinja filter: returns disambiguated resort name when duplicates exist.
    Usage: {{ trip.resort | resort_display }}
    Appends (WA) / (ON) for US/CA; (FR) / (AT) etc. for international.
    Plain name returned when no duplication exists.
    """
    return _resort_display_name(resort, AMBIGUOUS_RESORT_NAMES)


@app.template_filter('relative_time')
def relative_time_filter(dt):
    """
    Convert datetime to relative time string (e.g., "2h ago", "Yesterday").
    """
    if not dt:
        return ''
    
    now = datetime.utcnow()
    diff = now - dt
    
    seconds = diff.total_seconds()
    if seconds < 60:
        return 'just now'
    if seconds < 3600:
        return f'{int(seconds // 60)}m ago'
    if seconds < 86400:
        return f'{int(seconds // 3600)}h ago'
    if seconds < 172800:
        return 'Yesterday'
    return dt.strftime('%b %d')


@app.template_filter('format_name')
def format_name_filter(name):
    from utils.formatting import format_name
    return format_name(name)


def compute_user_state(user):
    """
    Returns the canonical navigation state for the given user.

    States (in priority order):
      ANONYMOUS             — not authenticated
      PENDING_VERIFICATION  — authenticated but is_verified == False
      ONBOARDING            — authenticated, verified, core profile incomplete
      ACTIVE_EMPTY          — onboarded, 0 friends, 0 trips
      ACTIVE_SOCIAL         — onboarded, ≥1 friend, 0 trips
      ACTIVE_FULL           — onboarded, ≥1 trip (friend count irrelevant)

    This is the single source of truth for all navigation decisions.
    Result is cached in flask.g for the duration of the request so that
    multiple callers within the same request pay the DB cost at most once.
    """
    if hasattr(g, "_computed_user_state"):
        return g._computed_user_state

    if not user.is_authenticated:
        state = "ANONYMOUS"
    elif not user.is_verified:
        state = "PENDING_VERIFICATION"
    elif not user.is_core_profile_complete:
        state = "ONBOARDING"
    else:
        trip_count = SkiTrip.query.filter_by(user_id=user.id).count()
        if trip_count > 0:
            state = "ACTIVE_FULL"
        else:
            friend_count = Friend.query.filter_by(user_id=user.id).count()
            state = "ACTIVE_SOCIAL" if friend_count > 0 else "ACTIVE_EMPTY"

    g._computed_user_state = state
    return state


def resolve_navigation(path, user_state, pending_intent=None):
    """
    Returns a redirect path string if the user must be moved, or None to allow through.

    pending_intent is stubbed as None — not yet implemented.
    """
    if user_state == "ANONYMOUS":
        allowed = {"/auth", "/login", "/signup", "/logout", "/auth/logout"}
        if path in allowed or path.startswith("/legal"):
            return None
        return "/auth"

    if user_state == "PENDING_VERIFICATION":
        if path in {"/auth/verify", "/logout", "/auth/logout"}:
            return None
        return "/auth/verify"

    if user_state == "ONBOARDING":
        if path in {"/onboarding", "/logout", "/auth/logout"}:
            return None
        return "/onboarding"

    # ACTIVE_* states — root and auth page redirect to home; everything else allowed
    if path in {"/", "/auth"}:
        return "/home"
    return None


@app.after_request
def bl_nav_timing_log(response):
    """Log per-request server timing when BL_NAV_DEBUG is enabled."""
    if not app.config.get("BL_NAV_DEBUG"):
        return response
    t0 = getattr(g, '_bl_nav_t0', None)
    if t0 is None:
        return response
    total_ms = int((time.monotonic() - t0) * 1000)
    qcount = getattr(g, '_bl_nav_qcount', '?')
    endpoint = request.endpoint or request.path
    print(f"[BL-NAV] {endpoint} queries={qcount} total={total_ms}ms status={response.status_code}")
    response.headers["X-BL-Nav-Ms"] = str(total_ms)
    return response


@app.after_request
def set_security_headers(response):
    """Apply baseline security headers to every response."""
    # Allow same-origin framing only in the Replit dev preview (REPLIT_DEV_DOMAIN
    # is set exclusively in the dev workspace, never in production deployments).
    # Production keeps DENY to block all third-party embedding.
    if os.environ.get("REPLIT_DEV_DOMAIN"):
        response.headers["X-Frame-Options"] = "SAMEORIGIN"
    else:
        response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "no-referrer-when-downgrade"
    # Long-cache versioned static assets — HTML responses are unaffected
    # (they are never served from /static/).
    if request.path.startswith("/static/"):
        response.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return response


@app.before_request
# NOTE: This function is lightweight — zero DB queries for most requests.
# Home-specific DB work lives in the home() route handler, not here.
def before_request_handlers():
    import sys

    # ── Auth session-cookie restore logging ───────────────────────────────────
    # Flask-Login silently re-authenticates from the session cookie on every
    # request with zero logging. We detect the first authenticated request in
    # each Flask session and emit a single [auth_restore] line so the audit
    # trail is not completely blank between explicit logins.
    # remember-me cookie restores are logged separately via the
    # user_loaded_from_cookie signal registered near the user_loader above.
    if (current_user.is_authenticated
            and not request.path.startswith('/static/')
            and not session.get('_auth_session_logged')):
        try:
            app.logger.info(
                "[auth_restore] method=session_cookie user_id=%s email=%s "
                "ua=%.100s ts=%s ip=%s",
                current_user.id,
                current_user.email,
                request.user_agent.string,
                datetime.utcnow().isoformat(),
                request.remote_addr,
            )
            session['_auth_session_logged'] = True
            session.modified = True
        except Exception:
            pass

    # ── Activity heartbeat (throttled to 1 DB write per user per hour) ────────
    # Uses a session timestamp so no extra SELECT is needed; just one UPDATE
    # at most once per hour. Skips static assets. Runs for all authenticated
    # users including admin so metrics reflect real usage accurately.
    if (current_user.is_authenticated
            and not request.path.startswith('/static/')):
        _now_ts = time.time()
        if _now_ts - session.get('_last_active_stamp', 0) > 3600:
            try:
                current_user.last_active_at = datetime.utcnow()
                db.session.commit()
                session['_last_active_stamp'] = _now_ts
                session.modified = True
            except Exception:
                db.session.rollback()

    # ── Founder login-alert push ───────────────────────────────────────────────
    # Moved to login paths (email, Google OAuth, password-reset).
    # _queue_founder_login_push() is called immediately after login_user()
    # in each of those handlers — see routes around /auth and /auth/google/callback.

    # ── Navigation timing (BL_NAV_DEBUG) ─────────────────────────────────────
    # Records wall-clock start and initialises a per-request query counter.
    # The SQLAlchemy event listener is registered lazily on the first debug
    # request so it has no cost whatsoever when the flag is off.
    if app.config.get("BL_NAV_DEBUG"):
        g._bl_nav_t0 = time.monotonic()
        g._bl_nav_qcount = 0
        if not getattr(app, '_bl_qc_registered', False):
            try:
                from sqlalchemy import event as _sa_event
                def _bl_count_q(conn, cursor, statement, parameters, context, executemany):
                    if hasattr(g, '_bl_nav_qcount'):
                        g._bl_nav_qcount += 1
                _sa_event.listen(db.engine, 'before_cursor_execute', _bl_count_q)
                app._bl_qc_registered = True
            except Exception:
                pass

    # Make sessions permanent for Replit iframe compatibility
    session.permanent = True

    # ── Gate skip list ────────────────────────────────────────────────────────
    # These paths bypass the nav gate and are handled by their own logic.
    path = request.path
    if (path.startswith("/static/") or
            path.startswith("/api/") or
            path.startswith("/invite/") or
            path.startswith("/trip-invite/") or   # handles own auth-gate + session storage
            path.startswith("/connect/") or        # handles own auth-gate + post_login_redirect
            path.startswith("/auth/google") or
            path.startswith("/auth/apple") or
            path.startswith("/auth/logout") or
            path.startswith("/admin/") or
            path.startswith("/debug/") or
            path.startswith("/reset-password/") or
            path == "/forgot-password" or
            path == "/logout" or
            path == "/robots.txt" or
            path == "/sitemap.xml" or
            path.startswith("/.well-known/") or
            path == "/privacypolicy" or
            path == "/termsandconditions" or
            path.startswith("/download") or
            request.endpoint in {"health_check"}):
        return None

    # ── Navigation gate ───────────────────────────────────────────────────────
    user_state = compute_user_state(current_user)
    redirect_to = resolve_navigation(path, user_state)
    if redirect_to and redirect_to != path:
        return redirect(redirect_to)
    return None

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        admin_emails_str = os.environ.get("ALLOWED_ADMIN_EMAILS", "")
        admin_emails = [e.strip().lower() for e in admin_emails_str.split(",") if e.strip()]
        if not current_user.is_authenticated or current_user.email.lower() not in admin_emails:
            if request.path.startswith('/api/'):
                return jsonify({'status': 'error', 'message': 'Admin privileges required'}), 403
            return "Admin privileges required.", 403
        return f(*args, **kwargs)
    return wrapper

def emit_event(event_name, user, payload=None):
    """Emit high-signal event, skip if user is seeded (test data)."""
    if user.is_seeded:
        return
    event = Event(
        event_name=event_name,
        user_id=user.id,
        payload=payload or {},
        created_at=datetime.utcnow(),
        environment='prod' if is_production else 'dev'
    )
    db.session.add(event)
    db.session.commit()


def get_friend_ids(user_id):
    """Get all direct friend IDs for a user — only returns IDs where the User row still exists."""
    rows = (
        db.session.query(Friend.friend_id)
        .join(User, User.id == Friend.friend_id)
        .filter(Friend.user_id == user_id)
        .all()
    )
    return [r[0] for r in rows]


def create_activity(actor_user_id, recipient_user_id, activity_type, object_type, object_id, extra_data=None):
    """Create an activity record."""
    if actor_user_id == recipient_user_id:
        return
    activity = Activity(
        actor_user_id=actor_user_id,
        recipient_user_id=recipient_user_id,
        type=activity_type.value if hasattr(activity_type, 'value') else activity_type,
        object_type=object_type,
        object_id=object_id,
        created_at=datetime.utcnow(),
        extra_data=extra_data or None,
    )
    db.session.add(activity)


def emit_trip_created_activities(trip, actor_user_id):
    """Create TRIP_CREATED activities for all friends of the actor."""
    friend_ids = get_friend_ids(actor_user_id)
    for friend_id in friend_ids:
        create_activity(
            actor_user_id=actor_user_id,
            recipient_user_id=friend_id,
            activity_type=ActivityType.TRIP_CREATED,
            object_type='trip',
            object_id=trip.id
        )
    check_and_emit_trip_overlap_activities(trip, actor_user_id)
    emit_availability_overlap_activities_for_trip(trip)


def emit_trip_updated_activities(trip, actor_user_id, dates_changed=False):
    """Create TRIP_UPDATED activities for all friends if dates changed."""
    if not dates_changed:
        return
    friend_ids = get_friend_ids(actor_user_id)
    for friend_id in friend_ids:
        create_activity(
            actor_user_id=actor_user_id,
            recipient_user_id=friend_id,
            activity_type=ActivityType.TRIP_UPDATED,
            object_type='trip',
            object_id=trip.id
        )
    check_and_emit_trip_overlap_activities(trip, actor_user_id)
    emit_availability_overlap_activities_for_trip(trip)


def emit_trip_location_changed_activities(trip, actor_user_id, resort_name):
    """Notify accepted participants when the trip resort is changed."""
    participants = SkiTripParticipant.query.filter_by(
        trip_id=trip.id,
        status=GuestStatus.ACCEPTED
    ).all()
    for participant in participants:
        if participant.user_id == actor_user_id:
            continue
        create_activity(
            actor_user_id=actor_user_id,
            recipient_user_id=participant.user_id,
            activity_type=ActivityType.TRIP_LOCATION_CHANGED,
            object_type='trip',
            object_id=trip.id,
            extra_data={'resort_name': resort_name},
        )
    check_and_emit_trip_overlap_activities(trip, actor_user_id)
    emit_availability_overlap_activities_for_trip(trip)


def emit_trip_pass_changed_activities(trip, actor_user_id, pass_display):
    """Notify accepted participants when the trip pass is changed."""
    participants = SkiTripParticipant.query.filter_by(
        trip_id=trip.id,
        status=GuestStatus.ACCEPTED
    ).all()
    for participant in participants:
        if participant.user_id == actor_user_id:
            continue
        create_activity(
            actor_user_id=actor_user_id,
            recipient_user_id=participant.user_id,
            activity_type=ActivityType.TRIP_PASS_CHANGED,
            object_type='trip',
            object_id=trip.id,
            extra_data={'pass_display': pass_display},
        )


def check_and_emit_trip_overlap_activities(trip, actor_user_id):
    """Check for overlapping trips with friends and emit TRIP_OVERLAP activities."""
    friend_ids = get_friend_ids(actor_user_id)
    if not friend_ids:
        return
    
    overlapping_trips = SkiTrip.query.filter(
        SkiTrip.user_id.in_(friend_ids),
        SkiTrip.start_date <= trip.end_date,
        SkiTrip.end_date >= trip.start_date,
        db.or_(
            SkiTrip.resort_id == trip.resort_id,
            SkiTrip.mountain == trip.mountain
        )
    ).all()
    
    notified_users = set()
    for overlap_trip in overlapping_trips:
        if overlap_trip.user_id not in notified_users:
            create_activity(
                actor_user_id=actor_user_id,
                recipient_user_id=overlap_trip.user_id,
                activity_type=ActivityType.TRIP_OVERLAP,
                object_type='trip',
                object_id=trip.id
            )
            notified_users.add(overlap_trip.user_id)


def emit_connection_accepted_activity(actor_user_id, other_user_id):
    """Create CONNECTION_ACCEPTED activity for both users when a friend request is accepted.
    actor_user_id = acceptor (Richard), other_user_id = original sender (Jonathan).
    Both get a notification so both see it in their activity/notification feed.
    """
    # Notify the original invite sender that their request was accepted
    create_activity(
        actor_user_id=actor_user_id,
        recipient_user_id=other_user_id,
        activity_type=ActivityType.CONNECTION_ACCEPTED,
        object_type='user',
        object_id=actor_user_id
    )
    # Mirror: also notify the acceptor (so both sides see the connection in their feed)
    create_activity(
        actor_user_id=other_user_id,
        recipient_user_id=actor_user_id,
        activity_type=ActivityType.CONNECTION_ACCEPTED,
        object_type='user',
        object_id=other_user_id
    )


def emit_trip_invite_accepted_activity(trip, acceptor_user_id, trip_owner_id):
    """Create TRIP_INVITE_ACCEPTED activity for the trip owner."""
    create_activity(
        actor_user_id=acceptor_user_id,
        recipient_user_id=trip_owner_id,
        activity_type=ActivityType.TRIP_INVITE_ACCEPTED,
        object_type='trip',
        object_id=trip.id
    )


def emit_trip_invite_received_activity(trip, inviter_user_id, invitee_user_id):
    """Create TRIP_INVITE_RECEIVED activity for the invited user."""
    create_activity(
        actor_user_id=inviter_user_id,
        recipient_user_id=invitee_user_id,
        activity_type=ActivityType.TRIP_INVITE_RECEIVED,
        object_type='trip',
        object_id=trip.id
    )


def emit_trip_invite_declined_activity(trip, decliner_user_id, trip_owner_id):
    """Create TRIP_INVITE_DECLINED activity for the trip owner."""
    create_activity(
        actor_user_id=decliner_user_id,
        recipient_user_id=trip_owner_id,
        activity_type=ActivityType.TRIP_INVITE_DECLINED,
        object_type='trip',
        object_id=trip.id
    )


def emit_friend_joined_trip_activities(trip, joiner_user_id):
    """Create FRIEND_JOINED_TRIP activities for other participants on the trip."""
    participants = SkiTripParticipant.query.filter(
        SkiTripParticipant.trip_id == trip.id,
        SkiTripParticipant.user_id != joiner_user_id,
        SkiTripParticipant.status == GuestStatus.ACCEPTED
    ).all()
    
    joiner_friend_ids = set(get_friend_ids(joiner_user_id))
    
    for participant in participants:
        if participant.user_id in joiner_friend_ids:
            create_activity(
                actor_user_id=joiner_user_id,
                recipient_user_id=participant.user_id,
                activity_type=ActivityType.FRIEND_JOINED_TRIP,
                object_type='trip',
                object_id=trip.id
            )


def delete_activities_for_trip(trip_id):
    """Delete all activities related to a trip (application-level cleanup)."""
    Activity.query.filter(
        Activity.object_type == 'trip',
        Activity.object_id == trip_id
    ).delete()


def emit_carpool_activity(user, trip, seats):
    """Emit carpool offered activity to friends with overlapping trip dates.
    
    Only emits to friends with trips that overlap dates and location.
    Group trips do not emit carpool activities (handled at caller).
    """
    if user.is_seeded:
        return
    
    friend_ids = get_friend_ids(user.id)
    if not friend_ids:
        return
    
    # Find friends with overlapping trips at the same location
    overlapping_friends = set()
    for friend_id in friend_ids:
        friend_trips = SkiTrip.query.filter(
            SkiTrip.user_id == friend_id,
            SkiTrip.start_date <= trip.end_date,
            SkiTrip.end_date >= trip.start_date,
            db.or_(
                SkiTrip.mountain == trip.mountain,
                SkiTrip.resort_id == trip.resort_id
            ) if trip.resort_id else SkiTrip.mountain == trip.mountain
        ).first()
        if friend_trips:
            overlapping_friends.add(friend_id)
    
    # Create activity for each overlapping friend
    mountain_name = trip.mountain or (trip.resort.name if trip.resort else 'Unknown')
    for friend_id in overlapping_friends:
        activity = Activity(
            actor_user_id=user.id,
            recipient_user_id=friend_id,
            type=ActivityType.CARPOOL_OFFERED.value,
            object_type='trip',
            object_id=trip.id,
            extra_data={
                'seats': seats,
                'mountain': mountain_name
            }
        )
        db.session.add(activity)
    
    if overlapping_friends:
        db.session.commit()


def coalesce_date_ranges(date_ranges):
    """Merge contiguous or overlapping date ranges into continuous ranges.
    
    Args:
        date_ranges: List of (start_date, end_date) tuples
        
    Returns:
        List of merged (start_date, end_date) tuples
    """
    if not date_ranges:
        return []
    
    sorted_ranges = sorted(date_ranges, key=lambda x: x[0])
    merged = [sorted_ranges[0]]
    
    for start, end in sorted_ranges[1:]:
        last_start, last_end = merged[-1]
        if start <= last_end + timedelta(days=1):
            merged[-1] = (last_start, max(last_end, end))
        else:
            merged.append((start, end))
    
    return merged


def compute_friend_trip_availability_overlaps(user):
    """Compute overlaps between user's open availability and friends' trips.
    
    Returns a list of overlap groups, each containing:
    - overlap_start_date, overlap_end_date
    - list of (friend_id, trip_id, resort_id, resort_name, state, country)
    """
    if not user.open_dates:
        return []
    
    user_open_dates = set()
    for d in user.open_dates:
        try:
            if isinstance(d, str):
                user_open_dates.add(datetime.strptime(d, '%Y-%m-%d').date())
            else:
                user_open_dates.add(d)
        except (ValueError, TypeError):
            continue
    
    if not user_open_dates:
        return []

    friend_ids = get_friend_ids(user.id)
    if not friend_ids:
        return []
    
    friend_trips = SkiTrip.query.filter(
        SkiTrip.user_id.in_(friend_ids),
        SkiTrip.end_date >= date.today(),
        SkiTrip.is_public == True,
    ).all()

    overlap_by_range = {}
    
    for trip in friend_trips:
        trip_dates = set()
        current = trip.start_date
        while current <= trip.end_date:
            trip_dates.add(current)
            current += timedelta(days=1)
        
        overlapping_dates = user_open_dates & trip_dates
        if not overlapping_dates:
            continue

        sorted_dates = sorted(overlapping_dates)
        ranges = []
        range_start = sorted_dates[0]
        range_end = sorted_dates[0]
        
        for d in sorted_dates[1:]:
            if d == range_end + timedelta(days=1):
                range_end = d
            else:
                ranges.append((range_start, range_end))
                range_start = d
                range_end = d
        ranges.append((range_start, range_end))
        
        resort = db.session.get(Resort, trip.resort_id) if trip.resort_id else None
        resort_name = resort.name if resort else (trip.mountain or "Unknown")
        state = resort.state_code if resort else None
        country = resort.country_code if resort else None
        
        # Get friend user data for display
        friend_user = db.session.get(User, trip.user_id)
        friend_first_name = friend_user.first_name if friend_user else "Friend"
        friend_last_name = friend_user.last_name if friend_user else ""
        
        for range_start, range_end in ranges:
            key = (range_start, range_end)
            if key not in overlap_by_range:
                overlap_by_range[key] = []
            overlap_by_range[key].append({
                'friend_id': trip.user_id,
                'trip_id': trip.id,
                'resort_id': trip.resort_id,
                'resort_name': resort_name,
                'state': state,
                'country': country,
                'first_name': friend_first_name,
                'last_name': friend_last_name
            })
    
    coalesced = coalesce_date_ranges(list(overlap_by_range.keys()))
    
    result = []
    for start, end in coalesced:
        friends_data = []
        for (r_start, r_end), data_list in overlap_by_range.items():
            if r_start >= start and r_end <= end:
                friends_data.extend(data_list)
        
        seen_trips = set()
        unique_friends_data = []
        for d in friends_data:
            if d['trip_id'] not in seen_trips:
                unique_friends_data.append(d)
                seen_trips.add(d['trip_id'])
        
        if unique_friends_data:
            result.append({
                'overlap_start_date': start,
                'overlap_end_date': end,
                'friends': unique_friends_data
            })
    
    return result


def delete_availability_overlap_activities_for_trip(trip_id):
    """Delete all FRIEND_TRIP_OVERLAPS_AVAILABILITY activities that reference a specific trip."""
    activities = Activity.query.filter(
        Activity.type == ActivityType.FRIEND_TRIP_OVERLAPS_AVAILABILITY.value
    ).all()
    
    for activity in activities:
        if activity.extra_data and trip_id in activity.extra_data.get('trip_ids', []):
            db.session.delete(activity)


def emit_availability_overlap_activities_for_user(user):
    """Create or update FRIEND_TRIP_OVERLAPS_AVAILABILITY activities for a user.
    
    Called when:
    - User updates their open availability dates
    - A friend creates or edits a trip
    """
    Activity.query.filter(
        Activity.recipient_user_id == user.id,
        Activity.type == ActivityType.FRIEND_TRIP_OVERLAPS_AVAILABILITY.value
    ).delete()
    
    overlaps = compute_friend_trip_availability_overlaps(user)
    
    for overlap in overlaps:
        friend_ids = list(set(f['friend_id'] for f in overlap['friends']))
        trip_ids = list(set(f['trip_id'] for f in overlap['friends']))
        resort_ids = list(set(f['resort_id'] for f in overlap['friends'] if f['resort_id']))
        states = list(set(f['state'] for f in overlap['friends'] if f['state']))
        countries = list(set(f['country'] for f in overlap['friends'] if f['country']))
        
        actor_id = friend_ids[0] if friend_ids else user.id
        
        extra_data = {
            'friend_ids': friend_ids,
            'trip_ids': trip_ids,
            'overlap_start_date': overlap['overlap_start_date'].isoformat(),
            'overlap_end_date': overlap['overlap_end_date'].isoformat(),
            'resort_ids': resort_ids,
            'friends_data': overlap['friends'],
            'state': states,
            'country': countries
        }
        
        activity = Activity(
            actor_user_id=actor_id,
            recipient_user_id=user.id,
            type=ActivityType.FRIEND_TRIP_OVERLAPS_AVAILABILITY.value,
            object_type='availability',
            object_id=user.id,
            created_at=datetime.utcnow(),
            extra_data=extra_data
        )
        db.session.add(activity)


def emit_availability_overlap_activities_for_trip(trip):
    """Recompute availability overlaps for all friends when a trip is created/edited."""
    friend_ids = get_friend_ids(trip.user_id)
    
    for friend_id in friend_ids:
        friend = db.session.get(User, friend_id)
        if friend and friend.open_dates:
            emit_availability_overlap_activities_for_user(friend)


# Database Configuration
supabase_url = os.environ.get("SUPABASE_DATABASE_URL")
if supabase_url:
    supabase_url = supabase_url.strip().strip('"').strip("'")
    if supabase_url.startswith("postgres://"):
        supabase_url = "postgresql://" + supabase_url[len("postgres://"):]
    if supabase_url.startswith("postgresql+psycopg2://"):
        supabase_url = "postgresql://" + supabase_url[len("postgresql+psycopg2://"):]
    for bad_prefix in ["postgresql+asyncpg://", "postgresql+aiopg://"]:
        if supabase_url.startswith(bad_prefix):
            supabase_url = "postgresql://" + supabase_url[len(bad_prefix):]
    if supabase_url.startswith(("postgresql://", "postgres://")) and "@" in supabase_url and "://" in supabase_url:
        app.config["SQLALCHEMY_DATABASE_URI"] = supabase_url
    else:
        app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///baselodge.db"
else:
    if is_production:
        raise RuntimeError("SUPABASE_DATABASE_URL must be set in production.")
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///baselodge.db"

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)
migrate = Migrate(app, db)

oauth = OAuth(app)
oauth.register(
    name="google",
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_id=os.getenv("GOOGLE_CLIENT_ID"),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
    client_kwargs={"scope": "openid email profile"},
)

# Auto-create tables for SQLite (local development only)
if "sqlite" in app.config.get("SQLALCHEMY_DATABASE_URI", ""):
    with app.app_context():
        db.create_all()


def run_equipment_migration():
    """
    Safe one-time schema migration for the multi-setup equipment feature (May 2026).
    Adds is_primary, label, created_at columns if they don't exist, then backfills
    is_primary=TRUE for any row where slot='primary'.
    Uses IF NOT EXISTS so it is a no-op on repeat runs.
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text(
                    "ALTER TABLE equipment_setup ADD COLUMN IF NOT EXISTS is_primary BOOLEAN NOT NULL DEFAULT FALSE"
                ))
                conn.execute(db.text(
                    "ALTER TABLE equipment_setup ADD COLUMN IF NOT EXISTS label VARCHAR(100)"
                ))
                conn.execute(db.text(
                    "ALTER TABLE equipment_setup ADD COLUMN IF NOT EXISTS created_at TIMESTAMP"
                ))
                conn.execute(db.text(
                    "ALTER TABLE equipment_setup ADD COLUMN IF NOT EXISTS binding_brand VARCHAR(100)"
                ))
                conn.execute(db.text(
                    "ALTER TABLE equipment_setup ADD COLUMN IF NOT EXISTS binding_model VARCHAR(100)"
                ))
                # Backfill: existing slot='primary' rows become is_primary=TRUE
                conn.execute(db.text(
                    "UPDATE equipment_setup SET is_primary = TRUE WHERE slot = 'primary' AND is_primary = FALSE"
                ))
                # Backfill created_at for existing rows
                conn.execute(db.text(
                    "UPDATE equipment_setup SET created_at = NOW() WHERE created_at IS NULL"
                ))
                trans.commit()
                print("equipment_migration: schema columns added / backfilled successfully.")
            except Exception as inner_e:
                trans.rollback()
                print(f"equipment_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"equipment_migration: skipped ({e})")


run_equipment_migration()


def run_push_token_migration():
    """Safe one-time schema migration for APNs environment-aware token tracking (May 2026).

    1. Adds apns_environment column if absent (VARCHAR 20, default 'unknown').
    2. Deactivates known sandbox tokens (ids 3, 4) and stamps them 'sandbox'.
       These were registered against api.sandbox.push.apple.com and will always
       return BadEnvironmentKeyInToken when APNS_USE_SANDBOX=false.
    Uses IF NOT EXISTS / safe UPDATE so it is a no-op on repeat runs.
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text(
                    "ALTER TABLE push_device_token "
                    "ADD COLUMN IF NOT EXISTS apns_environment VARCHAR(20) NOT NULL DEFAULT 'unknown'"
                ))
                # Mark the two known sandbox tokens inactive and label them correctly
                conn.execute(db.text(
                    "UPDATE push_device_token "
                    "SET active = FALSE, apns_environment = 'sandbox' "
                    "WHERE id IN (3, 4)"
                ))
                trans.commit()
                print("push_token_migration: apns_environment column ready; sandbox tokens (3,4) deactivated.")
            except Exception as inner_e:
                trans.rollback()
                print(f"push_token_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"push_token_migration: skipped ({e})")


run_push_token_migration()


def run_batch_trip_migration():
    """Add created_in_batch_id column to ski_trip for batch analytics (May 2026).

    Nullable VARCHAR(36) — stores a shared UUID when multiple trips are created
    together via the multi-date flow. Never used for display grouping or series
    logic; analytics and potential undo only.
    Uses IF NOT EXISTS so it is a no-op on repeat runs.
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text(
                    "ALTER TABLE ski_trip ADD COLUMN IF NOT EXISTS created_in_batch_id VARCHAR(36)"
                ))
                trans.commit()
                print("batch_trip_migration: created_in_batch_id column ready.")
            except Exception as inner_e:
                trans.rollback()
                print(f"batch_trip_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"batch_trip_migration: skipped ({e})")


run_batch_trip_migration()


def run_push_notif_pref_migration():
    """Add push_notifications_enabled column to user table (May 2026).

    Boolean, NOT NULL, default TRUE — allows users to opt out of push
    notifications at the app level. Uses IF NOT EXISTS so it is a no-op
    on repeat runs.
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text(
                    "ALTER TABLE \"user\" ADD COLUMN IF NOT EXISTS "
                    "push_notifications_enabled BOOLEAN NOT NULL DEFAULT TRUE"
                ))
                trans.commit()
                print("push_notif_pref_migration: push_notifications_enabled column ready.")
            except Exception as inner_e:
                trans.rollback()
                print(f"push_notif_pref_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"push_notif_pref_migration: skipped ({e})")


run_push_notif_pref_migration()


def run_message_event_log_migration():
    """Ensure message_event_log table exists (safety net for environments where
    flask db upgrade has not been run yet).

    The Alembic migration is the authoritative path. This function only guards
    against a missing table on first startup before the migration is applied.
    Uses IF NOT EXISTS so it is a no-op on every subsequent run.
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS message_event_log (
                        id               SERIAL PRIMARY KEY,
                        event_name       VARCHAR(120) NOT NULL,
                        event_version    INTEGER NOT NULL DEFAULT 1,
                        category         VARCHAR(50) NOT NULL,
                        actor_user_id    INTEGER REFERENCES "user"(id),
                        recipient_user_id INTEGER REFERENCES "user"(id),
                        object_type      VARCHAR(80),
                        object_id        INTEGER,
                        channel          VARCHAR(40),
                        delivery_status  VARCHAR(40) NOT NULL DEFAULT 'pending',
                        suppression_reason VARCHAR(80),
                        provider         VARCHAR(40),
                        provider_message_id VARCHAR(255),
                        payload_json     JSON NOT NULL DEFAULT '{}',
                        message_title    VARCHAR(255),
                        message_body     TEXT,
                        error_message    TEXT,
                        retry_count      INTEGER NOT NULL DEFAULT 0,
                        created_at       TIMESTAMP NOT NULL DEFAULT NOW(),
                        processed_at     TIMESTAMP,
                        sent_at          TIMESTAMP
                    )
                """))
                trans.commit()
                print("message_event_log_migration: table ready.")
            except Exception as inner_e:
                trans.rollback()
                print(f"message_event_log_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"message_event_log_migration: skipped ({e})")


run_message_event_log_migration()


def run_mel_dedupe_index_migration():
    """Add composite partial dedupe index to message_event_log (Phase D-1 Deploy A).

    Index: idx_mel_dedupe
    Columns: (event_name, recipient_user_id, object_type, object_id, created_at)
    Partial:  WHERE delivery_status != 'failed'

    Rationale: is_duplicate_event() filters on all five columns plus the
    delivery_status exclusion. The composite index makes this query an index
    seek rather than a full-table scan as the MEL grows. The partial clause
    excludes FAILED rows — they are never considered duplicates — keeping the
    index smaller and writes cheaper.

    Uses IF NOT EXISTS — idempotent across all restarts. No CONCURRENTLY needed
    at current table size; the standard CREATE INDEX holds a lock for microseconds.

    Rollback: DROP INDEX IF EXISTS idx_mel_dedupe;
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text("""
                    CREATE INDEX IF NOT EXISTS idx_mel_dedupe
                    ON message_event_log (
                        event_name,
                        recipient_user_id,
                        object_type,
                        object_id,
                        created_at
                    )
                    WHERE delivery_status != 'failed'
                """))
                trans.commit()
                print("mel_dedupe_index_migration: idx_mel_dedupe ready.")
            except Exception as inner_e:
                trans.rollback()
                print(f"mel_dedupe_index_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"mel_dedupe_index_migration: skipped ({e})")


run_mel_dedupe_index_migration()


def run_deploy_b_schema_migration():
    """Phase D-1 Deploy B: add parent_mel_id and retry_locked_at to message_event_log.

    parent_mel_id  — nullable FK to self; links retry child rows to the original
                     FAILED row. Flat lineage only: children always reference the
                     original row, never another child.

    retry_locked_at — nullable timestamp; set to NOW() when the retry runner
                      claims a row for processing, reset to NULL after the child
                      MEL row commits. Rows locked for > 15 minutes are treated
                      as stale and become eligible again automatically (no manual
                      unlock needed).

    Both columns are nullable with no DEFAULT — purely additive, no table rewrite,
    no existing row impact. Uses ADD COLUMN IF NOT EXISTS for idempotency.

    Rollback: ALTER TABLE message_event_log
              DROP COLUMN IF EXISTS parent_mel_id,
              DROP COLUMN IF EXISTS retry_locked_at;
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text("""
                    ALTER TABLE message_event_log
                        ADD COLUMN IF NOT EXISTS parent_mel_id   INTEGER
                            REFERENCES message_event_log(id),
                        ADD COLUMN IF NOT EXISTS retry_locked_at TIMESTAMP
                """))
                trans.commit()
                print("deploy_b_schema_migration: parent_mel_id + retry_locked_at ready.")
            except Exception as inner_e:
                trans.rollback()
                print(f"deploy_b_schema_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"deploy_b_schema_migration: skipped ({e})")


run_deploy_b_schema_migration()


def run_ghost_user_cleanup_migration():
    """
    Startup migration: remove Friend and Invitation rows that reference
    a user_id or friend_id that no longer exists in the User table.

    These orphans accumulate when a user row is deleted outside the normal
    delete_account route (e.g. legacy admin tooling, direct DB ops, or a
    partially-failed deletion).  Safe to re-run on every startup — deletes
    nothing if the DB is already clean.
    """
    try:
        with app.app_context():
            # Friend rows where the subject (user_id side) is gone
            d1 = Friend.query.filter(
                ~Friend.user_id.in_(db.session.query(User.id))
            ).delete(synchronize_session=False)

            # Friend rows where the target (friend_id side) is gone
            d2 = Friend.query.filter(
                ~Friend.friend_id.in_(db.session.query(User.id))
            ).delete(synchronize_session=False)

            # Invitation rows where the sender is gone
            d3 = Invitation.query.filter(
                ~Invitation.sender_id.in_(db.session.query(User.id))
            ).delete(synchronize_session=False)

            # Invitation rows where the receiver is gone
            d4 = Invitation.query.filter(
                ~Invitation.receiver_id.in_(db.session.query(User.id))
            ).delete(synchronize_session=False)

            db.session.commit()
            total = d1 + d2 + d3 + d4
            print(f"ghost_user_cleanup_migration: removed {total} orphaned rows "
                  f"(Friend: {d1+d2}, Invitation: {d3+d4}).")
    except Exception as e:
        print(f"ghost_user_cleanup_migration: ERROR — {e}")


run_ghost_user_cleanup_migration()


def run_ski_trip_updated_at_migration():
    """
    Startup migration: add updated_at column to ski_trip.
    Null means the trip was never edited after this feature launched — correct
    fallback; these rows use created_at in Happening label logic.
    """
    try:
        with app.app_context():
            db.session.execute(db.text(
                "ALTER TABLE ski_trip ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP"
            ))
            db.session.commit()
            print("ski_trip_updated_at_migration: updated_at column ready.")
    except Exception as e:
        print(f"ski_trip_updated_at_migration: skipped ({e})")

run_ski_trip_updated_at_migration()


def run_trip_invite_token_migration():
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS trip_invite_token (
                        id              SERIAL PRIMARY KEY,
                        token           VARCHAR(64) UNIQUE NOT NULL,
                        trip_id         INTEGER NOT NULL REFERENCES ski_trip(id) ON DELETE CASCADE,
                        inviter_user_id INTEGER NOT NULL REFERENCES "user"(id) ON DELETE CASCADE,
                        created_at      TIMESTAMP DEFAULT NOW(),
                        used_at         TIMESTAMP,
                        expires_at      TIMESTAMP,
                        is_active       BOOLEAN NOT NULL DEFAULT TRUE
                    )
                """))
                conn.execute(db.text(
                    "CREATE INDEX IF NOT EXISTS idx_trip_invite_token_token ON trip_invite_token(token)"
                ))
                trans.commit()
                print("trip_invite_token_migration: table ready.")
            except Exception as inner_e:
                trans.rollback()
                print(f"trip_invite_token_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"trip_invite_token_migration: skipped ({e})")

run_trip_invite_token_migration()


def run_pass_system_expansion_migration():
    """
    Full ski pass system expansion (May 2026).

    Three idempotent operations:
    1. Widen ski_trip.pass_type column to VARCHAR(100) to safely hold
       multi-pass comma-separated strings (was VARCHAR(50)).
    2. Normalize two known wrong-case user pass values:
         'Epic'     → 'epic'
         'Not Sure' → 'no_pass_yet'
    3. Backfill ResortPass rows for Indy and MountainCollective resorts
       that have those values in Resort.pass_brands but only have an
       'Other' ResortPass row (or none). Inserts are skipped if a row
       for that resort_id / pass_name already exists.
    """
    try:
        with app.app_context():
            conn = db.engine.connect()
            trans = conn.begin()
            try:
                # 1. Widen ski_trip.pass_type
                conn.execute(db.text(
                    "ALTER TABLE ski_trip "
                    "ALTER COLUMN pass_type TYPE VARCHAR(100)"
                ))

                # 2. Normalize bad-case user pass_type values
                conn.execute(db.text(
                    "UPDATE \"user\" SET pass_type = 'epic' "
                    "WHERE pass_type = 'Epic'"
                ))
                conn.execute(db.text(
                    "UPDATE \"user\" SET pass_type = 'no_pass_yet' "
                    "WHERE pass_type = 'Not Sure'"
                ))

                # 3a. Backfill Indy ResortPass rows
                conn.execute(db.text("""
                    INSERT INTO resort_pass (resort_id, pass_name, is_primary, created_at)
                    SELECT r.id, 'Indy', FALSE, NOW()
                    FROM resort r
                    WHERE r.pass_brands LIKE '%Indy%'
                      AND NOT EXISTS (
                          SELECT 1 FROM resort_pass rp
                          WHERE rp.resort_id = r.id AND rp.pass_name = 'Indy'
                      )
                """))

                # 3b. Backfill MountainCollective ResortPass rows
                conn.execute(db.text("""
                    INSERT INTO resort_pass (resort_id, pass_name, is_primary, created_at)
                    SELECT r.id, 'MountainCollective', FALSE, NOW()
                    FROM resort r
                    WHERE r.pass_brands LIKE '%MountainCollective%'
                      AND NOT EXISTS (
                          SELECT 1 FROM resort_pass rp
                          WHERE rp.resort_id = r.id AND rp.pass_name = 'MountainCollective'
                      )
                """))

                trans.commit()
                print("pass_system_expansion_migration: complete.")
            except Exception as inner_e:
                trans.rollback()
                print(f"pass_system_expansion_migration inner error (rolled back): {inner_e}")
            finally:
                conn.close()
    except Exception as e:
        print(f"pass_system_expansion_migration: skipped ({e})")


run_pass_system_expansion_migration()


# ============================================================================
# MOUNTAIN PAGE VIEW MIGRATION — safe CREATE TABLE IF NOT EXISTS
# ============================================================================
def run_mountain_page_view_migration():
    """Create mountain_page_view table + index if they don't exist."""
    try:
        with app.app_context():
            with db.engine.connect() as conn:
                trans = conn.begin()
                try:
                    conn.execute(db.text("""
                        CREATE TABLE IF NOT EXISTS mountain_page_view (
                            id          SERIAL PRIMARY KEY,
                            resort_id   INTEGER NOT NULL,
                            user_id     INTEGER,
                            viewed_at   TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                            session_key VARCHAR(32)
                        )
                    """))
                    conn.execute(db.text("""
                        CREATE INDEX IF NOT EXISTS idx_mpv_resort_time
                        ON mountain_page_view (resort_id, viewed_at)
                    """))
                    trans.commit()
                    print("mountain_page_view_migration: table ready.")
                except Exception as inner_e:
                    trans.rollback()
                    print(f"mountain_page_view_migration inner error (rolled back): {inner_e}")
                finally:
                    conn.close()
    except Exception as e:
        print(f"mountain_page_view_migration: skipped ({e})")


run_mountain_page_view_migration()


# ============================================================================
# RIDER TYPE NORMALIZATION — collapse multi-item arrays to single canonical item
# ============================================================================
def run_rider_type_normalization_migration():
    """
    Startup migration: normalize every user's rider_types value to exactly one
    canonical item stored as a single-element JSON array.

    Handles all legacy forms:
      - Non-list values (strings, None, integers): coerced to list first
      - Comma-packed elements like ['Skier,Snowboarder'] → split before mapping
      - Known lowercase variants  (e.g. 'skier' → 'Skier')
      - Old social labels ('Social / après', 'Social (along for the ride)') → 'Social'
      - Multi-element arrays: drop Social when any real discipline exists and
        keep the first real discipline; Social-only → ['Social']
      - Empty / all-null arrays: left unchanged (user hasn't finished onboarding)
    Idempotent — safe to run on every restart.
    """
    _CANON = {
        'skier': 'Skier',
        'snowboarder': 'Snowboarder',
        'telemark': 'Telemark',
        'cross-country': 'Cross-Country',
        'adaptive': 'Adaptive',
        'interested': 'Interested',
        'social': 'Social',
        'social / après': 'Social',
        'social (along for the ride)': 'Social',
        'both': 'Skier',  # old "Both" value → Skier
    }

    def _to_tokens(raw):
        """Coerce any stored rider_types value into a flat list of string tokens."""
        if raw is None:
            return []
        if isinstance(raw, str):
            # Bare string (not a list) — treat as single value, may be comma-packed
            raw = [raw]
        if not isinstance(raw, list):
            return []
        tokens = []
        for item in raw:
            if not item:
                continue
            # Split comma-packed values stored inside a single array element
            for part in str(item).split(','):
                part = part.strip()
                if part:
                    tokens.append(part)
        return tokens

    def _normalize(raw):
        tokens = _to_tokens(raw)
        if not tokens:
            return raw  # nothing to do — leave unchanged
        cased = [_CANON.get(t.lower(), t) for t in tokens]
        non_social = [t for t in cased if t != 'Social']
        result = [non_social[0]] if non_social else [cased[0]]
        # Return None/empty unchanged; already-canonical single-item arrays untouched
        return result

    try:
        with app.app_context():
            try:
                users = User.query.all()
                changed = 0
                for u in users:
                    raw = u.rider_types
                    # Skip users with no data at all (None, empty list)
                    if raw is None or raw == []:
                        continue
                    normalized = _normalize(raw)
                    if normalized != raw:
                        u.rider_types = normalized
                        changed += 1
                if changed:
                    db.session.commit()
                print(f"rider_type_normalization_migration: normalized {changed} user(s).")
            except Exception as inner_e:
                db.session.rollback()
                print(f"rider_type_normalization_migration inner error (rolled back): {inner_e}")
    except Exception as e:
        print(f"rider_type_normalization_migration: skipped ({e})")

run_rider_type_normalization_migration()


# ============================================================================
# APP STORE METRIC TABLE — safe CREATE TABLE IF NOT EXISTS at startup
# ============================================================================
def _run_app_store_metric_migration():
    """Create app_store_metric table if it does not already exist."""
    try:
        with app.app_context():
            with db.engine.begin() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS app_store_metric (
                        id             SERIAL PRIMARY KEY,
                        platform       VARCHAR(16)  NOT NULL,
                        report_date    DATE         NOT NULL,
                        downloads      INTEGER,
                        page_views     INTEGER,
                        conversion_pct FLOAT,
                        rating         FLOAT,
                        review_count   INTEGER,
                        crashes        FLOAT,
                        anrs           FLOAT,
                        fetched_at     TIMESTAMP    NOT NULL DEFAULT NOW(),
                        CONSTRAINT uq_app_store_metric_platform_date
                            UNIQUE (platform, report_date)
                    )
                """))
                conn.execute(db.text(
                    "CREATE INDEX IF NOT EXISTS ix_app_store_metric_platform "
                    "ON app_store_metric (platform)"
                ))
                conn.execute(db.text(
                    "CREATE INDEX IF NOT EXISTS ix_app_store_metric_report_date "
                    "ON app_store_metric (report_date)"
                ))
        print("app_store_metric_migration: table and indexes ensured.")
    except Exception as _e:
        print(f"app_store_metric_migration: skipped ({_e})")

_run_app_store_metric_migration()


def _run_invite_share_event_migration():
    """Create invite_share_event table and indexes if they do not already exist."""
    try:
        with app.app_context():
            with db.engine.begin() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS invite_share_event (
                        id         SERIAL PRIMARY KEY,
                        user_id    INTEGER      NOT NULL
                                   REFERENCES "user"(id) ON DELETE CASCADE,
                        token_type VARCHAR(16)  NOT NULL,
                        token_id   INTEGER,
                        token      VARCHAR(64),
                        action     VARCHAR(16)  NOT NULL,
                        source     VARCHAR(32)  NOT NULL,
                        user_agent VARCHAR(256),
                        created_at TIMESTAMP    NOT NULL DEFAULT NOW()
                    )
                """))
                conn.execute(db.text(
                    "CREATE INDEX IF NOT EXISTS ix_ise_user_id "
                    "ON invite_share_event (user_id)"
                ))
                conn.execute(db.text(
                    "CREATE INDEX IF NOT EXISTS ix_ise_created_at "
                    "ON invite_share_event (created_at)"
                ))
        print("invite_share_event_migration: table and indexes ensured.")
    except Exception as _e:
        print(f"invite_share_event_migration: skipped ({_e})")

_run_invite_share_event_migration()


# ============================================================================
# RESORT DISAMBIGUATION — compute once at startup, no N+1 in requests
# ============================================================================
from utils.resort_utils import get_ambiguous_resort_names, resort_display_name as _resort_display_name
try:
    with app.app_context():
        AMBIGUOUS_RESORT_NAMES = get_ambiguous_resort_names(db.session, Resort)
    print(f"resort_utils: {len(AMBIGUOUS_RESORT_NAMES)} ambiguous resort name(s) cached.")
except Exception as _rdu_err:
    print(f"resort_utils: could not build AMBIGUOUS_RESORT_NAMES ({_rdu_err}); falling back to empty set.")
    AMBIGUOUS_RESORT_NAMES = frozenset()


# ============================================================================
# STATIC ASSET VERSIONING — computed once at startup from mtime+size
# Injected into Jinja2 globals so templates can append ?v=<hash> to
# fingerprinted URLs, enabling safe long-cache (max-age=31536000).
# ============================================================================
import hashlib as _hashlib

def _asset_version(path):
    """Return 8-char MD5 hex of file mtime+size. Fast — no full file read."""
    try:
        _st = os.stat(path)
        return _hashlib.md5(f"{_st.st_mtime}:{_st.st_size}".encode()).hexdigest()[:8]
    except OSError:
        return "00000000"

STYLES_VERSION    = _asset_version("static/styles.css")
BL_NATIVE_VERSION = _asset_version("static/js/bl-native.js")
ANALYTICS_VERSION = _asset_version("static/analytics.js")
ICONS_VERSION     = _asset_version("static/icons/favicon-32x32.png")


# ============================================================================
# PRODUCTION DIAGNOSTICS - Print on startup
# ============================================================================
def log_startup_diagnostics():
    """Log database and user counts on startup for debugging."""
    try:
        with app.app_context():
            db_url = os.environ.get("SUPABASE_DATABASE_URL")
            
            # Mask credentials
            if db_url and "@" in db_url:
                safe_db = db_url.split("@")[-1]
            elif not is_production:
                safe_db = "DEVELOPMENT FALLBACK: SQLite (baselodge.db)"
            else:
                safe_db = "ERROR: NOT SET"
            
            print("=" * 70)
            print("🔧 BASELODGE STARTUP DIAGNOSTICS")
            print("=" * 70)
            print(f"DATABASE: {safe_db}")
            print(f"PRODUCTION MODE: {is_production}")
            
            # Count records
            user_count = User.query.count()
            friend_count = Friend.query.count()
            trip_count = SkiTrip.query.count()
            
            print("USER COUNT: " + str(user_count))
            print("FRIEND COUNT: " + str(friend_count))
            print("TRIP COUNT: " + str(trip_count))
            print("=" * 70)
            print("✅ BaseLodge started successfully and is ready to serve requests.")
            print("=" * 70)
            
            # Check specific users
            richard = User.query.filter(db.func.lower(User.email) == "richardbattlebaxter@gmail.com").first()
            jonathan = User.query.filter(db.func.lower(User.email) == "jonathanmschmitz@gmail.com").first()
            
            if richard:
                richard_friends = Friend.query.filter_by(user_id=richard.id).count()
                print(f"RICHARD (id={richard.id}): {richard_friends} friends")
            else:
                print("RICHARD: NOT FOUND")
            
            if jonathan:
                jonathan_friends = Friend.query.filter_by(user_id=jonathan.id).count()
                print(f"JONATHAN (id={jonathan.id}): {jonathan_friends} friends")
            else:
                print("JONATHAN: NOT FOUND")
            
            print("=" * 70)
    except Exception as e:
        print(f"⚠️ STARTUP DIAGNOSTICS FAILED: {e}")

# Run diagnostics on import (will show in server logs)
import atexit
@app.before_request
def run_startup_diagnostics_once():
    """Run startup diagnostics once on first request."""
    if not hasattr(app, '_diagnostics_run'):
        app._diagnostics_run = True
        log_startup_diagnostics()

# ============================================================================
# ERROR HANDLERS - Must be registered at module level, not inside functions
# ============================================================================
@app.errorhandler(404)
def not_found_error(error):
    """Handle 404 Not Found errors with user-friendly template."""
    return render_template("404.html"), 404

@app.errorhandler(500)
def internal_error(error):
    """Handle internal server errors with full traceback logging."""
    import traceback
    import sys
    print("=" * 70)
    print("🚨 INTERNAL SERVER ERROR (500)")
    print("=" * 70)
    app.logger.exception(f"Internal server error: {error}")
    print("=" * 70)
    db.session.rollback()
    return render_template("500.html"), 500

@app.errorhandler(Exception)
def handle_exception(e):
    """Handle all exceptions with full traceback logging (except HTTP errors)."""
    from werkzeug.exceptions import HTTPException
    
    # Don't catch HTTP errors like 404 - let them return normally
    if isinstance(e, HTTPException):
        if e.code == 404:
            return render_template("404.html"), 404
        return e
    
    import traceback
    import sys
    print("=" * 70)
    print(f"🚨 UNHANDLED EXCEPTION: {type(e).__name__}")
    print("=" * 70)
    app.logger.exception(f"Unhandled exception: {e}")
    print("=" * 70)
    db.session.rollback()
    return render_template("500.html"), 500

def get_or_create_invite_token(user):
    """Return a valid invite token for user, creating one only when necessary.

    Returns None if user has reached their max invite accepts limit.
    Single-use tokens: each token can only be used once (used_at is set on use).

    Reuses an existing valid (unused + unexpired) token so that a freshly
    shared link is not invalidated the next time the inviter visits /invite or
    /friends.  Token rotation (expiring the current token) happens only at
    unfriend time, not on normal page loads.
    """
    if not can_sender_accept_more_invites(user):
        return None

    now = datetime.utcnow()
    existing = InviteToken.query.filter_by(inviter_id=user.id).all()

    # Reuse the first valid token — avoids rotating a live shareable URL
    for token_obj in existing:
        if not token_obj.is_used() and not token_obj.is_expired():
            return token_obj

    # No valid token exists — create a fresh one
    token = secrets.token_urlsafe(16)
    expires_at = now + timedelta(hours=48)
    invite = InviteToken(token=token, inviter_id=user.id, expires_at=expires_at)
    db.session.add(invite)
    db.session.commit()
    ph_analytics.track(user.id, 'invite_generated', {'source': 'invite_page'})
    return invite


def get_or_create_trip_invite_token(trip_id, inviter_user_id):
    """Get or create a reusable TripInviteToken for this trip+inviter pair."""
    existing = TripInviteToken.query.filter_by(
        trip_id=trip_id,
        inviter_user_id=inviter_user_id,
        is_active=True,
    ).first()
    if existing:
        return existing
    tok = TripInviteToken(
        token=secrets.token_urlsafe(32),
        trip_id=trip_id,
        inviter_user_id=inviter_user_id,
    )
    db.session.add(tok)
    db.session.commit()
    return tok


def can_sender_accept_more_invites(user):
    """Check if sender can accept more invites. Always returns True since invite limits are removed."""
    return True

def count_friends_open_on_same_dates(user):
    """Count UNIQUE friends who have open dates overlapping with the current user.
    
    Returns:
        tuple: (friend_count, user_has_open_dates)
    """
    try:
        today = date.today()
        today_str = today.strftime('%Y-%m-%d')
        
        # Get user's future open dates
        user_open_dates = set(user.open_dates or [])
        user_open_dates = {d for d in user_open_dates if d >= today_str}
        
        # If no open dates, return 0 count but indicate user has no dates
        if not user_open_dates:
            return 0, False
        
        # Get user's friends
        friend_links = Friend.query.filter_by(user_id=user.id).all()
        friend_ids = [f.friend_id for f in friend_links]
        
        if not friend_ids:
            return 0, True
        
        # Get friends' data
        friends = User.query.filter(User.id.in_(friend_ids)).all()
        
        # Count unique friends with overlapping open dates
        matching_friends = set()
        for friend in friends:
            friend_dates = set(friend.open_dates or [])
            # Check if there's any intersection
            if user_open_dates & friend_dates:
                matching_friends.add(friend.id)
        
        return len(matching_friends), True
    except Exception as e:
        app.logger.warning(f"Error counting open date friends: {e}")
        return 0, False

STATE_ABBR = {
    "Alaska": "AK",
    "California": "CA",
    "Colorado": "CO",
    "Idaho": "ID",
    "Maine": "ME",
    "Michigan": "MI",
    "Montana": "MT",
    "New Hampshire": "NH",
    "New Mexico": "NM",
    "New York": "NY",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Utah": "UT",
    "Vermont": "VT",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wyoming": "WY"
}

STATE_NAMES = {v: k for k, v in STATE_ABBR.items()}

COUNTRY_NAMES = {
    "US": "United States",
    "CA": "Canada",
    "JP": "Japan",
    "FR": "France",
    "CH": "Switzerland",
    "AT": "Austria",
    "IT": "Italy",
    "CL": "Chile",
    "ES": "Spain",
    "NO": "Norway",
    "SE": "Sweden",
}

ALL_US_STATES = [
    ("AL", "Alabama"), ("AK", "Alaska"), ("AZ", "Arizona"), ("AR", "Arkansas"),
    ("CA", "California"), ("CO", "Colorado"), ("CT", "Connecticut"), ("DE", "Delaware"),
    ("FL", "Florida"), ("GA", "Georgia"), ("HI", "Hawaii"), ("ID", "Idaho"),
    ("IL", "Illinois"), ("IN", "Indiana"), ("IA", "Iowa"), ("KS", "Kansas"),
    ("KY", "Kentucky"), ("LA", "Louisiana"), ("ME", "Maine"), ("MD", "Maryland"),
    ("MA", "Massachusetts"), ("MI", "Michigan"), ("MN", "Minnesota"), ("MS", "Mississippi"),
    ("MO", "Missouri"), ("MT", "Montana"), ("NE", "Nebraska"), ("NV", "Nevada"),
    ("NH", "New Hampshire"), ("NJ", "New Jersey"), ("NM", "New Mexico"), ("NY", "New York"),
    ("NC", "North Carolina"), ("ND", "North Dakota"), ("OH", "Ohio"), ("OK", "Oklahoma"),
    ("OR", "Oregon"), ("PA", "Pennsylvania"), ("RI", "Rhode Island"), ("SC", "South Carolina"),
    ("SD", "South Dakota"), ("TN", "Tennessee"), ("TX", "Texas"), ("UT", "Utah"),
    ("VT", "Vermont"), ("VA", "Virginia"), ("WA", "Washington"), ("WV", "West Virginia"),
    ("WI", "Wisconsin"), ("WY", "Wyoming")
]

ALL_CANADA_PROVINCES = [
    ("AB", "Alberta"), ("BC", "British Columbia"), ("MB", "Manitoba"),
    ("NB", "New Brunswick"), ("NL", "Newfoundland and Labrador"), ("NS", "Nova Scotia"),
    ("NT", "Northwest Territories"), ("NU", "Nunavut"), ("ON", "Ontario"),
    ("PE", "Prince Edward Island"), ("QC", "Quebec"), ("SK", "Saskatchewan"), ("YT", "Yukon")
]

ALL_JAPAN_REGIONS = [
    ("Hokkaido", "Hokkaido"), ("Nagano", "Nagano"), ("Niigata", "Niigata"),
    ("Gunma", "Gunma"), ("Yamagata", "Yamagata"), ("Iwate", "Iwate")
]

ALL_FRANCE_REGIONS = [
    ("Auvergne-Rhône-Alpes", "Auvergne-Rhône-Alpes"), ("Provence-Alpes-Côte d'Azur", "Provence-Alpes-Côte d'Azur"),
    ("Occitanie", "Occitanie"), ("Nouvelle-Aquitaine", "Nouvelle-Aquitaine")
]

ALL_SWITZERLAND_REGIONS = [
    ("Valais", "Valais"), ("Graubünden", "Graubünden"), ("Bern", "Bern"),
    ("Vaud", "Vaud"), ("Uri", "Uri"), ("Obwalden", "Obwalden")
]

ALL_AUSTRIA_REGIONS = [
    ("Tyrol", "Tyrol"), ("Salzburg", "Salzburg"), ("Vorarlberg", "Vorarlberg"),
    ("Styria", "Styria"), ("Carinthia", "Carinthia")
]

ALL_ITALY_REGIONS = [
    ("Trentino-Alto Adige", "Trentino-Alto Adige"), ("Valle d'Aosta", "Valle d'Aosta"),
    ("Lombardy", "Lombardy"), ("Piedmont", "Piedmont"), ("Veneto", "Veneto")
]

# Canonical states/regions by country code
CANONICAL_STATES_BY_COUNTRY = {
    "US": ALL_US_STATES,
    "CA": ALL_CANADA_PROVINCES,
    "JP": ALL_JAPAN_REGIONS,
    "FR": ALL_FRANCE_REGIONS,
    "CH": ALL_SWITZERLAND_REGIONS,
    "AT": ALL_AUSTRIA_REGIONS,
    "IT": ALL_ITALY_REGIONS,
}

def get_all_countries():
    """Get canonical list of all countries (not derived from resorts).
    Returns list of (country_code, country_name) tuples, sorted with US first.
    """
    countries = list(COUNTRY_NAMES.items())
    # Sort: US first, then alphabetically by name
    def sort_key(item):
        code, name = item
        if code == "US":
            return (0, "")
        return (1, name)
    return sorted(countries, key=sort_key)

def get_all_states_by_country():
    """Get canonical states/regions for all countries (not derived from resorts).
    Returns dict mapping country_code to list of (state_code, state_name) tuples.
    """
    result = {}
    for country_code, states in CANONICAL_STATES_BY_COUNTRY.items():
        result[country_code] = sorted(states, key=lambda x: x[1])
    return result

def get_grouped_locations():
    """Get locations grouped by country for the unified location selector.
    Returns dict with country names as keys and sorted list of (code, name) tuples as values.
    """
    return {
        "United States": sorted(ALL_US_STATES, key=lambda x: x[1]),
        "Canada": sorted(ALL_CANADA_PROVINCES, key=lambda x: x[1])
    }

@lru_cache(maxsize=1)
def get_resorts_for_trip_form():
    """Get all active resorts for the Add Trip form.
    Returns list of dicts with id, name, country_code, state_code, pass_brands.

    ⚠️ CONTRACT: This data is used ONLY for filtering States and Resorts after
    a country is selected. The Country dropdown is populated from COUNTRIES
    in utils/countries.py, NOT from this data. Do not change this contract.

    Excludes region-level entities (is_region=True).

    Cached via lru_cache(maxsize=1) — result is process-scoped and reused across
    requests. Resort data changes rarely; the cache eliminates repeated full-table
    scans on add_trip, edit_trip, and trip_detail (owner) pages.
    Call get_resorts_for_trip_form.cache_clear() after any admin resort mutation.
    """
    resorts = Resort.query.filter_by(is_active=True, is_region=False).order_by(
        Resort.country_code, Resort.state_code, Resort.name
    ).all()
    return tuple(
        {
            "id": r.id,
            "name": r.name,
            "display_name": _resort_display_name(r, AMBIGUOUS_RESORT_NAMES),
            "country_code": r.country_code or r.country,
            "state_code": r.state_code or r.state,
            "state_name": (
                STATE_NAMES.get(r.state_code or r.state or '', '')
                or (r.state_name if r.state_name and r.state_name != (r.state_code or r.state) else '')
                or (r.state_full if r.state_full and r.state_full != (r.state_code or r.state) else '')
                or ''
            ),
            "country_name": (
                r.country_name
                or COUNTRY_NAMES.get(r.country_code or r.country or '', '')
                or ''
            ),
            "pass_brands": r.pass_brands or r.brand or ""
        }
        for r in resorts
    )


@lru_cache(maxsize=1)
def get_all_active_resorts_map():
    """Process-wide {resort_id: SimpleNamespace} for all active non-region resorts.

    Covers every field needed by /mountains (filter UI) and build_destination_feed
    (wishlist resort lookup). Uses SimpleNamespace — not ORM objects — so cached
    values are safe to hold across requests without SQLAlchemy session concerns.

    Cached via lru_cache(maxsize=1) — cleared on app restart (Replit deploy).
    After any admin resort mutation call both:
      get_all_active_resorts_map.cache_clear()
      get_resorts_for_trip_form.cache_clear()
    """
    from types import SimpleNamespace
    from models import ResortPass as _ResortPass

    _PASS_SKIP = frozenset({'no_pass', 'no_pass_yet'})

    # Batch-load all ResortPass rows (1 query)
    _all_rp = _ResortPass.query.all()
    _rp_by_resort = {}
    for _rp in _all_rp:
        _rp_by_resort.setdefault(_rp.resort_id, []).append(_rp.pass_name)

    def _pass_data(resort_id, pass_brands_raw):
        rp_rows = _rp_by_resort.get(resort_id, [])
        labels, keys = [], []
        if rp_rows:
            for pname in rp_rows:
                if not pname or str(pname).lower() in ('none', ''):
                    continue
                norm = normalize_pass(pname)
                if norm and norm not in _PASS_SKIP:
                    label = display_pass_label(norm)
                    if label:
                        labels.append(label)
                        keys.append(norm)
        if keys:
            return ' · '.join(labels), keys
        # Legacy fallback — resort_pass table empty
        raw = pass_brands_raw or ""
        if not raw or str(raw).lower() in ('none', ''):
            return "", []
        for _b in [_b.strip() for _b in str(raw).split(',') if _b.strip()]:
            norm = normalize_pass(_b)
            if norm and norm not in _PASS_SKIP:
                label = display_pass_label(norm)
                if label:
                    labels.append(label)
                    keys.append(norm)
        return ' · '.join(labels), keys

    _resorts = Resort.query.filter_by(is_active=True, is_region=False).order_by(
        Resort.country_code, Resort.state_code, Resort.name
    ).all()

    # Display-only slug → pass label overrides for resorts that resolve to 'other'
    # but belong to a deterministically known pass ecosystem.
    # Do NOT modify normalize_pass(), CANONICAL_PASSES, or any storage logic.
    _RESORT_PASS_OVERRIDES = {
        "palisades-tahoe":              "Ikon",
        "copper-mountain":              "Ikon",
        "steamboat":                    "Ikon",
        "killington":                   "Ikon",
        "mont-tremblant":               "Ikon",
        "deer-valley":                  "Ikon",
        "snowbird":                     "Ikon",
        "big-sky-resort":               "Ikon",
        "jackson-hole-mountain-resort": "Ikon",
        "taos-ski-valley":              "Ikon",
        "park-city-mountain":           "Epic",
        "whiteface":                    "Epic",
    }

    result = {}
    for r in _resorts:
        cc = r.country_code or r.country or ""
        sc = r.state_code or r.state or ""
        sn = r.state_name or r.state_full or ""
        cn = r.country_name or COUNTRY_NAMES.get(cc, cc) or ""
        pl, pk = _pass_data(r.id, r.pass_brands or r.brand or "")
        # Apply display-only override when the resort resolves to generic 'other'
        slug = r.slug or ""
        if pk == ["other"] and slug in _RESORT_PASS_OVERRIDES:
            pl = _RESORT_PASS_OVERRIDES[slug]
        result[r.id] = SimpleNamespace(
            id=r.id,
            name=r.name,
            display_name=_resort_display_name(r, AMBIGUOUS_RESORT_NAMES),
            slug=r.slug or "",
            country_code=cc,
            country_name=cn,
            state_code=sc,
            state_name=sn or sc,
            state=sc,
            pass_labels=pl,
            pass_keys=pk,
            pass_brands=r.pass_brands or r.brand or "",
        )
    return result


RIDER_TYPES = ["Skier", "Snowboarder", "Telemark", "Cross-Country", "Adaptive", "Interested", "Social"]


CANONICAL_PASSES = [
    "no_pass",
    "no_pass_yet",
    "epic",
    "ikon",
    "indy",
    "mountain_collective",
    "powder_alliance",
    "freedom",
    "ski_california",
    "other",
]

def get_sorted_passes():
    """Return passes sorted in canonical display order (snake_case values)."""
    return [
        "epic",
        "ikon",
        "indy",
        "mountain_collective",
        "powder_alliance",
        "freedom",
        "ski_california",
        "other",
        "no_pass",
        "no_pass_yet",
    ]

PASS_OPTIONS = get_sorted_passes()

# Rider-aware copy helpers
def get_gear_term(rider_type):
    """Return rider-aware gear terminology based on stored rider_types[0] value."""
    if rider_type and rider_type.lower() in ['snowboarder', 'snowboarding']:
        return 'board'
    elif rider_type and rider_type.lower() in ['skier', 'skiing', 'telemark']:
        return 'skis'
    return 'gear'

def get_ride_term(rider_type):
    """Return rider-aware action terminology based on stored rider_types[0] value."""
    if rider_type and rider_type.lower() in ['snowboarder', 'snowboarding']:
        return 'ride'
    elif rider_type and rider_type.lower() in ['skier', 'skiing', 'telemark']:
        return 'ski'
    return 'ride'

# Seasonal awareness helper
def get_season_context():
    """Return seasonal copy context based on current date."""
    today = date.today()
    month, day = today.month, today.day
    
    # Pre-season: October 1 – November 30
    if (month == 10) or (month == 11):
        return 'preseason'
    # Mid-season: December 1 – March 15
    elif (month == 12) or (month in [1, 2]) or (month == 3 and day <= 15):
        return 'midseason'
    # Spring: March 16 – April 30
    elif (month == 3 and day > 15) or (month == 4):
        return 'spring'
    # Off-season
    return 'offseason'

def get_seasonal_empty_state(context_type='trip'):
    """Return seasonally appropriate empty state copy."""
    season = get_season_context()
    
    if context_type == 'trip':
        if season == 'preseason':
            return "Plan your first trip of the season"
        elif season == 'midseason':
            return "Who's heading out this week?"
        elif season == 'spring':
            return "Any final turns planned?"
        return "Add a trip to get started"
    
    return ""

# State name mappings for display
STATE_NAMES = {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas', 'CA': 'California',
    'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware', 'FL': 'Florida', 'GA': 'Georgia',
    'HI': 'Hawaii', 'ID': 'Idaho', 'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa',
    'KS': 'Kansas', 'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi', 'MO': 'Missouri',
    'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada', 'NH': 'New Hampshire', 'NJ': 'New Jersey',
    'NM': 'New Mexico', 'NY': 'New York', 'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio',
    'OK': 'Oklahoma', 'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah', 'VT': 'Vermont',
    'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia', 'WI': 'Wisconsin', 'WY': 'Wyoming',
    'DC': 'District of Columbia', 'BC': 'British Columbia', 'AB': 'Alberta', 'ON': 'Ontario', 'QC': 'Quebec'
}

# Note: COUNTRY_NAMES defined above at line ~353

def group_resorts_for_display(resorts):
    """
    Group and sort resorts for display.
    US resorts grouped by state name, non-US resorts grouped by country name.
    Groups sorted alphabetically, resorts within groups sorted alphabetically.
    Returns list of dicts: [{'label': 'Colorado', 'resorts': [resort, ...]}, ...]
    """
    groups = {}
    
    for resort in resorts:
        if resort.country_code == 'US':
            group_key = 'US-' + (resort.state_code or 'Unknown')
            group_label = STATE_NAMES.get(resort.state_code, resort.state_code or 'Unknown')
        else:
            group_key = resort.country_code or 'Unknown'
            group_label = COUNTRY_NAMES.get(resort.country_code, resort.country_code or 'Unknown')
        
        if group_key not in groups:
            groups[group_key] = {'label': group_label, 'resorts': []}
        groups[group_key]['resorts'].append(resort)
    
    # Sort groups alphabetically by label
    sorted_groups = sorted(groups.values(), key=lambda g: g['label'])
    
    # Sort resorts within each group alphabetically by name
    for group in sorted_groups:
        group['resorts'] = sorted(group['resorts'], key=lambda r: r.name.lower())
    
    return sorted_groups

def format_trip_dates(trip):
    """
    Simplified trip date display logic.
    
    Rules:
    1) If start_date == end_date: "Jan 25"
    2) If same month: "Jan 25–27"
    3) If different months: "Jan 25–Feb 1"
    
    No year included.
    """
    try:
        if not trip:
            return ""
        
        start = getattr(trip, 'start_date', None)
        end = getattr(trip, 'end_date', None)
        
        if start and hasattr(start, 'date'):
            start = start.date()
        if end and hasattr(end, 'date'):
            end = end.date()
        
        if not start or not end:
            return ""
        
        if start == end:
            return start.strftime('%b %-d')
        elif start.month == end.month:
            return f"{start.strftime('%b %-d')}–{end.strftime('%-d')}"
        else:
            return f"{start.strftime('%b %-d')}–{end.strftime('%b %-d')}"
    except Exception:
        return ""

# Make functions available to Jinja2 templates
app.jinja_env.globals['get_sorted_passes'] = get_sorted_passes
app.jinja_env.globals['get_gear_term'] = get_gear_term
app.jinja_env.globals['get_ride_term'] = get_ride_term
app.jinja_env.globals['format_trip_dates'] = format_trip_dates
app.jinja_env.globals['get_season_context'] = get_season_context
app.jinja_env.globals['get_seasonal_empty_state'] = get_seasonal_empty_state
app.jinja_env.globals['is_internal_user'] = ph_analytics.is_internal
app.jinja_env.globals['POSTHOG_KEY'] = ph_analytics.POSTHOG_KEY
app.jinja_env.globals['POSTHOG_HOST'] = ph_analytics.POSTHOG_HOST

_PH_NO_PASS_VALS = {'no_pass', 'no_pass_yet', 'none', ''}
def _ph_is_real_pass(pass_str):
    """Return True if pass_str contains at least one real (non-placeholder) pass."""
    return any(s.strip().lower() not in _PH_NO_PASS_VALS
               for s in (pass_str or '').split(',') if s.strip() or not pass_str)
app.jinja_env.globals['ONESIGNAL_APP_ID'] = os.environ.get("ONESIGNAL_APP_ID", "")
app.jinja_env.globals['BL_NAV_DEBUG'] = app.config.get("BL_NAV_DEBUG", False)
app.jinja_env.globals['STYLES_VERSION']    = STYLES_VERSION
app.jinja_env.globals['BL_NATIVE_VERSION'] = BL_NATIVE_VERSION
app.jinja_env.globals['ANALYTICS_VERSION'] = ANALYTICS_VERSION
app.jinja_env.globals['ICONS_VERSION']     = ICONS_VERSION


def group_trips_by_month(trips):
    """
    Group a list of SkiTrip objects (pre-sorted by start_date asc)
    into (month_label, [trips]) tuples based on each trip's start_date month.
    Trips without a start_date are collected under 'TBD'.
    Example output: [("February", [trip1, trip2]), ("March", [trip3])]
    """
    groups = []
    current_label = None
    current_trips = []
    for trip in (trips or []):
        start = getattr(trip, 'start_date', None)
        if start:
            label = start.strftime('%B')
        else:
            label = 'TBD'
        if label != current_label:
            if current_label is not None:
                groups.append((current_label, current_trips))
            current_label = label
            current_trips = [trip]
        else:
            current_trips.append(trip)
    if current_label is not None:
        groups.append((current_label, current_trips))
    return groups


app.jinja_env.globals['group_trips_by_month'] = group_trips_by_month

MOUNTAINS_BY_STATE = {
    "CO": sorted(["Vail", "Breckenridge", "Keystone", "Copper Mountain", "Arapahoe Basin", "Loveland", "Winter Park", "Steamboat", "Aspen Snowmass", "Telluride", "Crested Butte", "Eldora"]),
    "UT": sorted(["Park City", "Deer Valley", "Snowbird", "Alta", "Brighton", "Solitude", "Snowbasin", "Powder Mountain"]),
    "CA": sorted(["Mammoth Mountain", "Palisades Tahoe", "Northstar", "Heavenly", "Kirkwood", "Big Bear", "June Mountain"]),
    "AK": sorted(["Alyeska Resort"]),
    "ID": sorted(["Sun Valley", "Schweitzer", "Bogus Basin", "Brundage Mountain"]),
    "ME": sorted(["Sugarloaf", "Sunday River", "Saddleback"]),
    "MI": sorted(["Boyne Mountain", "Crystal Mountain MI", "Nubs Nob"]),
    "MT": sorted(["Big Sky", "Whitefish Mountain", "Bridger Bowl", "Red Lodge Mountain"]),
    "NH": sorted(["Bretton Woods", "Cannon Mountain", "Loon Mountain", "Wildcat Mountain"]),
    "NM": sorted(["Taos Ski Valley", "Ski Santa Fe", "Angel Fire"]),
    "NY": sorted(["Whiteface", "Gore Mountain", "Hunter Mountain", "Windham Mountain"]),
    "OR": sorted(["Mt. Hood Meadows", "Timberline", "Mt. Bachelor", "Anthony Lakes"]),
    "VT": sorted(["Stowe", "Killington", "Sugarbush", "Jay Peak", "Stratton", "Mount Snow", "Okemo"]),
    "WA": sorted(["Crystal Mountain", "Stevens Pass", "Mt. Baker", "Snoqualmie"]),
    "WY": sorted(["Jackson Hole", "Grand Targhee", "Snow King"])
}

@app.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("5 per hour")
def forgot_password():
    if request.method == "POST":
        _google_account = False
        try:
            email = request.form.get("email", "").lower().strip()
            user = User.query.filter(sa.func.lower(User.email) == email).first()
            
            if user and user.auth_provider == 'email':
                # Only generate a reset token for email-auth accounts.
                # OAuth accounts (Google, etc.) do not have a local password to reset.
                token = user.get_reset_token()
                
                # Send Email via SendGrid
                # Canonical base URL
                reset_url = f"{BASE_URL}/reset-password/{token}"
                
                html_content = f"""
<div style="font-family:Georgia,'Times New Roman',serif;max-width:480px;margin:0 auto;padding:32px 24px;background:#F5F1E8;">
  <p style="font-size:22px;font-weight:500;color:#7A1E1E;margin:0 0 24px;">BaseLodge</p>
  <p style="font-size:16px;color:#1A1A1A;margin:0 0 12px;">Hi {user.first_name or 'there'},</p>
  <p style="font-size:15px;color:#3A3530;line-height:1.6;margin:0 0 28px;">We received a request to reset your BaseLodge password. This link expires in 30 minutes.</p>
  <a href="{reset_url}" style="display:inline-block;background:#7A1E1E;color:#fff;text-decoration:none;padding:14px 28px;border-radius:999px;font-family:sans-serif;font-size:15px;font-weight:500;">Reset my password</a>
  <p style="font-size:13px;color:#9A8F82;margin:24px 0 0;line-height:1.5;">If you didn't request this, you can safely ignore this email — your password won't change.</p>
</div>
"""
                message = Mail(
                    from_email='noreply@baselodgeapp.com',
                    to_emails=user.email,
                    subject='Reset your BaseLodge password',
                    plain_text_content=f'Hi {user.first_name or "there"},\n\nPlease use the following link to reset your password:\n\n{reset_url}\n\nThis link expires in 30 minutes.\n\nIf you didn\'t request this, you can safely ignore this email.',
                    html_content=html_content,
                )
                
                try:
                    key = os.environ.get('SENDGRID_API_KEY', '')
                    sg = SendGridAPIClient(key)
                    app.logger.info(f"SendGrid initialized with key starting with: {key[:4] if key else 'NONE'}")
                    response = sg.send(message)
                    app.logger.info(f"SendGrid response status code: {response.status_code}")
                    app.logger.info(f"Password reset email sent to {user.email}")
                except Exception as e:
                    app.logger.error(f"Error sending password reset email: {e}")

            elif user and user.auth_provider == 'google':
                # Google-auth account — no local password to reset. Flag for clear message.
                _google_account = True
        except Exception as e:
            app.logger.error(f"Error in forgot_password POST handler: {e}")
            db.session.rollback()
        
        if _google_account:
            flash("This account uses a different sign-in method. Please use the method you signed up with.", "info")
        else:
            flash("If an account exists with that email, you'll receive a password reset link.", "info")
        return render_template("forgot_password.html")
        
    return render_template("forgot_password.html")

@app.route("/reset-password", methods=["GET", "POST"])
@app.route("/reset-password/<token>", methods=["GET", "POST"])
@limiter.limit("10 per hour")
def reset_password(token=None):
    # Support both /reset-password?token=... and /reset-password/<token>
    if token is None:
        token = request.args.get("token")

    user = User.verify_reset_token(token)

    if not user:
        flash("This reset link is invalid or has expired.", "error")
        return redirect(url_for("auth"))

    if request.method == "POST":
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")
        
        if not password or len(password) < 8:
            flash("Password must be at least 8 characters.", "error")
            return render_template("reset_password.html", token=token)
            
        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("reset_password.html", token=token)

        user.set_password(password)
        user.password_changed_at = datetime.utcnow()
        user.last_active_at = datetime.utcnow()
        db.session.commit()

        login_user(user)
        # Founder login alert (non-blocking, throttled to 1×/user/day)
        _queue_founder_login_push(user.id, user.email)
        flash("Your password has been reset.", "success")

        # Consume pending invite if user arrived via an invite link
        if "invite_token" in session:
            connected = _connect_pending_inviter(user)
            if connected:
                return redirect(url_for("friends"))

        return redirect("/")

    return render_template("reset_password.html", token=token)

def build_trip_idea(user, idea_type, destination=None, resort_id=None, start_date_str=None, end_date_str=None, social_context=None, friends=None):
    """Canonical Trip Idea builder."""
    has_dates = bool(start_date_str and end_date_str)
    
    # Display Date
    display_date = ""
    if has_dates:
        try:
            d = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            display_date = d.strftime('%b %-d')
        except:
            display_date = start_date_str
            
    # CTA URL
    params = []
    if resort_id: params.append(f"resort_id={resort_id}")
    if has_dates:
        params.append(f"start_date={start_date_str}")
        params.append(f"end_date={end_date_str}")
    
    cta_url = "/add_trip"
    if params:
        cta_url += "?" + "&".join(params)
        
    return {
        "type": idea_type,
        "destination": destination,
        "resort_id": resort_id,
        "social_context": social_context,
        "has_dates": has_dates,
        "start_date_str": start_date_str,
        "end_date_str": end_date_str,
        "display_date": display_date,
        "cta_url": cta_url,
        "friends": friends or []
    }

@app.route("/")
def index():
    return redirect(url_for("home"))

@app.route("/auth", methods=["GET", "POST"])
@limiter.limit("10 per minute")
def auth():
    _ph_reset = session.pop('ph_reset', False)

    # Load inviter from session token for invite preview card
    _invite_inviter = None
    _invite_trips_count = 0
    _invite_token_str = session.get("invite_token")
    if _invite_token_str:
        _invite_obj = InviteToken.query.filter_by(token=_invite_token_str).first()
        if _invite_obj and not _invite_obj.is_used() and not _invite_obj.is_expired():
            _invite_inviter = db.session.get(User, _invite_obj.inviter_id)
            if _invite_inviter:
                _invite_trips_count = get_upcoming_trip_count(_invite_inviter)

    if request.method == "POST":
        form_type = request.form.get("form_type", "login")
        
        if form_type == "signup":
            email = request.form.get("email", "").lower().strip()
            password = request.form.get("password", "")
            first_name = request.form.get("first_name", "").strip()
            last_name = request.form.get("last_name", "").strip()

            if not first_name or not last_name or not email or not password:
                flash("Please fill in all fields.", "error")
                ph_analytics.track(None, 'auth_error', {'error_type': 'missing_fields'})
                return render_template("auth.html", has_invite=("invite_token" in session), posthog_reset=_ph_reset, inviter=_invite_inviter, inviter_trips_count=_invite_trips_count)
            
            if len(password) < 8:
                flash("Password must be at least 8 characters.", "error")
                ph_analytics.track(None, 'auth_error', {'error_type': 'password_too_short'})
                return render_template("auth.html", has_invite=("invite_token" in session), posthog_reset=_ph_reset, inviter=_invite_inviter, inviter_trips_count=_invite_trips_count)
            
            existing_user = User.query.filter_by(email=email).first()
            if existing_user:
                flash("An account with this email already exists.", "error")
                ph_analytics.track(None, 'auth_error', {'error_type': 'email_taken'})
                return render_template("auth.html", has_invite=("invite_token" in session), posthog_reset=_ph_reset, inviter=_invite_inviter, inviter_trips_count=_invite_trips_count)
            
            new_user = User(
                first_name=first_name,
                last_name=last_name,
                email=email,
                auth_provider="email",
                buddy_passes_available=True,
                created_at=datetime.utcnow(),
            )
            new_user.set_password(password)
            
            db.session.add(new_user)
            db.session.commit()
            
            login_user(new_user, remember=True)
            session.modified = True

            # Analytics: alias anon browser id → new user id, then identify
            _ph_anon_id = ph_analytics.get_anon_id(request.cookies)
            ph_analytics.alias(_ph_anon_id, new_user.id)
            ph_analytics.identify(
                new_user.id,
                set_once_props={"is_internal": ph_analytics.is_internal(new_user.email)},
            )
            ph_analytics.track(new_user.id, 'signup_completed', {
                'method': 'email',
                'signup_source': 'invite' if "invite_token" in session else 'organic',
            })

            # Connect with inviter if coming from friend invite link
            if "invite_token" in session:
                # Pre-set post-onboarding redirect to friends before token is consumed
                session["post_onboarding_redirect"] = url_for("friends")
                _connect_pending_inviter(new_user)
            elif "trip_invite_token" in session:
                # Pre-set post-onboarding redirect back to trip invite landing
                session["post_onboarding_redirect"] = url_for(
                    "trip_invite_token_landing", token=session["trip_invite_token"]
                )

            return redirect(url_for("onboarding"))
        
        elif form_type == "login":
            email = request.form.get("email", "").lower().strip()
            password = request.form.get("password", "")

            user = User.query.filter_by(email=email).first()

            # ── Diagnostic logging (safe — no passwords or full hashes) ────────
            _ua = (request.user_agent.string or "")[:160] if request.user_agent else ""
            _hash_prefix = (user.password_hash or "")[:12] if user else ""
            _check = user.check_password(password) if user else False
            app.logger.info(
                "[login_attempt] email_found=%s has_password=%s auth_provider=%s "
                "hash_prefix=%s check_ok=%s ua=%.120s",
                user is not None,
                bool(user.password_hash) if user else None,
                user.auth_provider if user else None,
                _hash_prefix,
                _check,
                _ua,
            )
            # ────────────────────────────────────────────────────────────────────

            if user and _check:
                user.last_active_at = datetime.utcnow()
                login_user(user, remember=True)
                session['_last_active_stamp'] = time.time()
                session.modified = True
                db.session.commit()

                # Analytics: identify on login
                ph_analytics.identify(
                    user.id,
                    set_once_props={"is_internal": ph_analytics.is_internal(user.email)},
                )
                ph_analytics.track(user.id, 'login_completed', {'method': 'email'})

                # Founder login alert (non-blocking, throttled to 1×/user/day)
                _queue_founder_login_push(user.id, user.email)

                # Connect with inviter if coming from friend invite link
                if "invite_token" in session:
                    connected = _connect_pending_inviter(user)
                    if connected:
                        # Redirect to friends page to show the new connection
                        return redirect(url_for("friends"))

                # Return to trip invite landing if coming from trip invite link
                if "trip_invite_token" in session:
                    _ttok = session["trip_invite_token"]
                    return redirect(url_for("trip_invite_token_landing", token=_ttok))

                # Return to any pending post-login destination (e.g. QR connect flow)
                _post_login = session.pop("post_login_redirect", None)
                if _post_login:
                    return redirect(_post_login)

                return redirect(url_for("home"))

            flash("Invalid email or password.", "error")
            ph_analytics.track(None, 'auth_error', {'error_type': 'invalid_credentials'})

    from_invite = "invite_token" in session
    return render_template("auth.html", has_invite=from_invite, from_invite=from_invite, posthog_reset=_ph_reset, inviter=_invite_inviter, inviter_trips_count=_invite_trips_count)


@app.route("/auth/check-email")
@limiter.limit("5 per minute")
def auth_check_email():
    email = request.args.get("email", "").lower().strip()
    if not email:
        return jsonify({"exists": False})
    user = User.query.filter_by(email=email).first()
    return jsonify({"exists": user is not None})


def send_founder_new_user_push(new_user):
    """Send a targeted push to richardbattlebaxter@gmail.com when a new user completes onboarding.

    Never raises — all exceptions are caught and logged so onboarding is never blocked.
    """
    try:
        from services.push_providers import send_onesignal_push as _os_push
        richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
        if not richard:
            app.logger.warning("[FounderAlert] richard account not found — skipping")
            return

        # Build signup line
        first = (new_user.first_name or "").strip()
        last  = (new_user.last_name  or "").strip()
        full  = (first + " " + last).strip() or "Someone"
        state = (new_user.home_state or "").strip()
        signup_line = f"{full} just signed up ({state})" if state else f"{full} just signed up"

        # Build connection line
        friend_rows = Friend.query.filter_by(user_id=new_user.id).all()
        n_friends   = len(friend_rows)
        if n_friends == 0:
            connection_line = "No connection"
        elif n_friends == 1:
            f_user = db.session.get(User, friend_rows[0].friend_id)
            if f_user:
                fn = (f_user.first_name or "").strip()
                ln = (f_user.last_name  or "").strip()
                connection_line = f"Connected to {(fn + ' ' + ln).strip() or 'a friend'}"
            else:
                connection_line = "Connected to 1 friend"
        else:
            connection_line = f"Connected to {n_friends} friends"

        title = "New BaseLodge User 🎿"
        body  = f"{signup_line}\n{connection_line}"

        result = _os_push([richard.id], title, body)
        app.logger.warning(
            "[FounderAlert] push sent to richard (id=%d): success=%s skipped=%s error=%s",
            richard.id, result.get("success"), result.get("skipped"), result.get("error"),
        )
    except Exception as _exc:
        app.logger.exception("[FounderAlert] unexpected error — onboarding not affected: %s", _exc)


def send_founder_app_open_push(user_id):
    """Send a founder-only push to richard when a real user opens BaseLodge.

    Accepts a plain integer user_id so it is safe to call from a background
    thread with its own app context — no SQLAlchemy detached-instance issues.
    Does its own fresh DB lookup inside the function.

    Never raises — all exceptions are caught and logged.
    """
    try:
        from services.push_providers import send_onesignal_push as _os_push

        # Fresh lookup — safe inside a new app context
        user = db.session.get(User, user_id)
        if not user:
            app.logger.warning("[founder_app_open_push] user_id=%d sent=False reason=user_not_found", user_id)
            return

        richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
        if not richard:
            app.logger.warning("[founder_app_open_push] user_id=%d sent=False reason=no_founder_account", user_id)
            return

        first = (user.first_name or "").strip()
        last  = (user.last_name  or "").strip()
        name  = (first + " " + last).strip()
        state = (user.home_state or "").strip()
        lc    = user.login_count or 0  # 0 / None treated as unknown

        def _ordinal(n):
            if 11 <= (n % 100) <= 13:          # 11th, 12th, 13th
                return f"{n}th"
            return f"{n}{['th','st','nd','rd','th','th','th','th','th','th'][n % 10]}"

        login_suffix = f"{_ordinal(lc)} login" if lc > 0 else ""

        parts = [p for p in [state, login_suffix] if p]
        if name and parts:
            body = f"{name} opened the app · {' · '.join(parts)}"
        elif name:
            body = f"{name} opened the app"
        else:
            body = "Someone opened BaseLodge"

        result = _os_push([richard.id], "BaseLodge Opened", body)

        if result.get("success"):
            app.logger.warning(
                "[founder_app_open_push] user_id=%d login_count=%d sent=True reason=sent body=%r",
                user_id, lc, body,
            )
        elif result.get("skipped"):
            app.logger.warning(
                "[founder_app_open_push] user_id=%d sent=False reason=no_push_token skipped_reason=%s",
                user_id, result.get("skipped_reason"),
            )
        else:
            app.logger.warning(
                "[founder_app_open_push] user_id=%d sent=False reason=push_error error=%s",
                user_id, result.get("error"),
            )
    except Exception as _exc:
        app.logger.exception(
            "[founder_app_open_push] user_id=%d sent=False reason=exception error=%s",
            user_id, _exc,
        )


def _queue_founder_login_push(user_id, user_email):
    """Check all gates and fire send_founder_app_open_push in a background thread.

    Called immediately after login_user() at every successful login path.
    Uses module-level _FOP_THROTTLE dict (keyed by user_id, value = Denver date
    string) so the once-per-day limit survives across multiple page loads within
    the same server process.  Resets on server restart.

    Args:
        user_id   (int): plain integer — safe to cross thread boundaries.
        user_email (str): used only for founder exclusion check.

    Never raises — all exceptions are caught and logged.
    """
    try:
        # Gate 1: feature flag
        if os.environ.get("FOUNDER_APP_OPEN_PUSH_ENABLED", "").lower() != "true":
            # Silent — logging every login when feature is off would be too noisy
            return

        # Gate 2: founder / admin exclusion
        _admin_set = {
            e.strip().lower()
            for e in os.environ.get("ALLOWED_ADMIN_EMAILS", "").split(",")
            if e.strip()
        }
        if (user_email or "").lower() in _admin_set:
            app.logger.debug(
                "[founder_app_open_push] user_id=%d sent=False reason=founder_user",
                user_id,
            )
            return

        # Gate 3: once per user per Denver calendar day (in-memory)
        _today = _admin_now().strftime("%Y-%m-%d")
        if _FOP_THROTTLE.get(user_id) == _today:
            app.logger.warning(
                "[founder_app_open_push] user_id=%d sent=False reason=throttled_today date=%s",
                user_id, _today,
            )
            return
        _FOP_THROTTLE[user_id] = _today

        # All gates passed — fire in a background thread (user_id is a plain int).
        # app is the real Flask instance at module level — no _get_current_object() needed.
        def _fire(uid=user_id):
            with app.app_context():
                send_founder_app_open_push(uid)
        threading.Thread(target=_fire, daemon=True).start()
        app.logger.warning(
            "[founder_app_open_push] user_id=%d sent=pending reason=login_success",
            user_id,
        )
    except Exception as _exc:
        app.logger.exception(
            "[founder_app_open_push] user_id=%d gate_error=%s", user_id, _exc,
        )


def _send_founder_invite_share_push(user_id, token_type, action, source):
    """Send a founder-only push notification when an InviteShareEvent is committed.

    Called from a background thread after the DB insert succeeds — never blocks
    the API response. Accepts only plain scalars to avoid DetachedInstanceError.

    token_type — 'friend' | 'trip'
    action     — 'copy' | 'text' | 'share_sheet'
    source     — 'invite_page' | 'friends_empty_state' | 'trip_detail'
    """
    try:
        from services.push_providers import send_onesignal_push as _os_push

        user = db.session.get(User, user_id)
        if not user:
            app.logger.warning("[invite_share_push] user_id=%d sent=False reason=user_not_found", user_id)
            return

        richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
        if not richard:
            app.logger.warning("[invite_share_push] user_id=%d sent=False reason=no_founder_account", user_id)
            return

        first = (user.first_name or "").strip()
        last  = (user.last_name  or "").strip()
        name  = (first + " " + last).strip()
        state = (user.home_state or "").strip()

        type_label   = "friend" if token_type == "friend" else "trip"
        action_label = {"copy": "copy link", "text": "text", "share_sheet": "share sheet"}.get(action, action)
        source_label = {
            "invite_page":          "invite page",
            "friends_empty_state":  "friends page",
            "trip_detail":          "trip detail",
        }.get(source, source)

        parts = [p for p in [state, source_label, action_label] if p]
        if name and parts:
            body = f"{name} sent a {type_label} invite · {' · '.join(parts)}"
        elif name:
            body = f"{name} sent a {type_label} invite"
        else:
            body = f"Someone sent a {type_label} invite · {action_label}"

        result = _os_push([richard.id], "Invite Sent", body)
        if result.get("success"):
            app.logger.warning(
                "[invite_share_push] user_id=%d token_type=%s action=%s sent=True body=%r",
                user_id, token_type, action, body,
            )
        elif result.get("skipped"):
            app.logger.warning(
                "[invite_share_push] user_id=%d sent=False reason=no_push_token",
                user_id,
            )
        else:
            app.logger.warning(
                "[invite_share_push] user_id=%d sent=False reason=push_error error=%s",
                user_id, result.get("error"),
            )
    except Exception as _exc:
        app.logger.exception("[invite_share_push] user_id=%d error=%s", user_id, _exc)


@app.route("/onboarding", methods=["GET", "POST"])
@login_required
def onboarding():
    """Canonical onboarding screen — collects core identity immediately after signup."""
    if request.method == "POST":
        rider_types_raw = request.form.get("rider_types", "")
        rider_types = [r.strip() for r in rider_types_raw.split(",") if r.strip()]
        skill_level = request.form.get("skill_level", "").strip()
        pass_type = request.form.get("pass_type", "").strip()
        home_state = request.form.get("home_state", "").strip()
        backcountry_capable = False
        avi_certified = None

        # Validate required fields
        if not rider_types:
            flash("Please select your rider type.")
            return render_template("identity_setup.html", grouped_locations=get_grouped_locations())

        # Social-only users don't need a skill level (they skip the skill section)
        is_social_only = rider_types == ["Social"]
        if not skill_level and not is_social_only:
            flash("Please select your skill level.")
            return render_template("identity_setup.html", grouped_locations=get_grouped_locations())

        if not pass_type:
            flash("Please select a pass option.")
            return render_template("identity_setup.html", grouped_locations=get_grouped_locations())

        if not home_state:
            flash("Please select your home state or province.")
            return render_template("identity_setup.html", grouped_locations=get_grouped_locations())

        # Save all onboarding data — normalize, dedupe, and canonically order passes
        normalized_pass = normalize_pass_selection(pass_type) or pass_type
        if count_real_passes(normalized_pass) > 3:
            flash("You can select up to 3 passes.")
            return render_template("identity_setup.html", grouped_locations=get_grouped_locations())
        current_user.rider_types = rider_types
        current_user.skill_level = skill_level
        current_user.pass_type = normalized_pass
        current_user.home_state = home_state
        current_user.backcountry_capable = backcountry_capable
        current_user.avi_certified = avi_certified

        db.session.commit()
        ph_analytics.track(current_user.id, 'onboarding_completed', {
            'total_steps': 4,
            'has_pass': _ph_is_real_pass(normalized_pass),
            'has_location': bool(home_state),
            'rider_type_count': len(rider_types) if isinstance(rider_types, list) else (1 if rider_types else 0),
        })
        if _ph_is_real_pass(normalized_pass):
            ph_analytics.track(current_user.id, 'pass_added', {
                'pass_type':    normalized_pass,
                'source':       'onboarding',
                'is_first_pass': True,
            })

        send_founder_new_user_push(current_user)

        # Redirect — invite signups go to friends, others go to home
        next_url = (
            session.pop("post_onboarding_redirect", None)
            or session.pop("next_after_setup", None)
        )
        if next_url:
            return redirect(next_url)
        return redirect(url_for("home"))

    return render_template("identity_setup.html", grouped_locations=get_grouped_locations())


@app.route("/identity-setup")
@login_required
def identity_setup():
    """Legacy URL — redirect to canonical /onboarding."""
    return redirect(url_for("onboarding"))


@app.route("/location-setup", methods=["GET", "POST"])
@login_required
def location_setup():
    """Location setup screen - step 2 of onboarding to collect home state only."""
    user = current_user
    
    # If user already has home_state, redirect to home
    if user.home_state:
        return redirect(url_for("home"))
    
    if request.method == "POST":
        home_state = request.form.get("home_state", "").strip()
        
        # Validate required field
        if not home_state:
            flash("Please select your home state.", "error")
            return redirect(url_for("location_setup"))
        
        # Save home state
        user.home_state = home_state
        db.session.commit()
        
        # Redirect to home - welcome modal will show
        next_url = session.pop("next_after_setup", None)
        if next_url:
            return redirect(next_url)
        return redirect(url_for("home"))
    
    # Get grouped locations for the unified selector
    grouped_locations = get_grouped_locations()
    
    return render_template("location_setup.html", grouped_locations=grouped_locations)


def _apply_invite_token(invite, user):
    """
    Core invite connection logic given a pre-loaded, pre-validated InviteToken and recipient user.
    Caller must have already confirmed: invite is not None, not expired, inviter != user.

    - Creates mutual Friend rows if not already connected (idempotent)
    - Sets user.invited_by_user_id if not already set
    - Marks invite.used_at
    - Commits

    Returns True if a new connection was made, False if users were already friends.
    """
    inviter = db.session.get(User, invite.inviter_id)
    connected = False
    if inviter and inviter.id != user.id:
        existing = Friend.query.filter_by(user_id=user.id, friend_id=inviter.id).first()
        if not existing:
            f1 = Friend(user_id=user.id, friend_id=inviter.id)
            f2 = Friend(user_id=inviter.id, friend_id=user.id)
            db.session.add_all([f1, f2])
            if not user.invited_by_user_id:
                user.invited_by_user_id = inviter.id
            connected = True
            app.logger.info(f"Connected {user.id} with inviter {inviter.id} via token")
        # Mark token used whether or not they were already friends (prevent reuse)
        invite.used_at = datetime.utcnow()
        db.session.commit()
    return connected


def _connect_pending_inviter(user):
    """Helper to connect user with pending inviter from session invite_token."""
    invite_token_str = session.get("invite_token")
    if not invite_token_str:
        return False

    invite = InviteToken.query.filter_by(token=invite_token_str).first()
    # Check if token is invalid, already used, or expired
    if not invite or invite.is_used() or invite.is_expired():
        session.pop("invite_token", None)
        return False

    connected = False
    if invite.inviter_id != user.id:
        connected = _apply_invite_token(invite, user)
        if connected:
            _isc_count = Friend.query.filter_by(user_id=user.id).count()
            ph_analytics.track(user.id, 'friend_connected', {
                'source':          'invite_signup',
                'is_first_friend': _isc_count == 1,
            })

    session.pop("invite_token", None)
    return connected


def _invite_landing_initials(user):
    """Return up to 2 uppercase initials for the avatar, e.g. 'JS' or 'A'."""
    first = (user.first_name or "").strip()
    last  = (user.last_name  or "").strip()
    if first and last:
        return (first[0] + last[0]).upper()
    if first:
        return first[:2].upper()
    return "?"


@app.route("/invite/<token>")
def invite_token_landing(token):
    """Invite landing page — shows a holding page before any acceptance occurs."""
    invite = InviteToken.query.filter_by(token=token).first()

    # 1. Token not found
    if not invite:
        return render_template("invite_invalid.html")

    # 2. Token expired
    if invite.is_expired():
        return render_template("invite_expired.html")

    inviter = db.session.get(User, invite.inviter_id)
    if not inviter:
        return render_template("invite_expired.html")

    # 3. Token already used — explicit single-use enforcement
    if invite.is_used():
        if current_user.is_authenticated:
            # If the current user is already connected to the inviter, redirect cleanly
            existing = Friend.query.filter_by(
                user_id=current_user.id, friend_id=inviter.id
            ).first()
            if existing:
                flash(f"You're already connected with {inviter.first_name}.", "info")
                return redirect(url_for("friends"))
        # Used token, not already connected (or logged-out) — no longer active
        return render_template("invite_expired.html")

    # 4. Inviter is the current user — self-invite guard
    if current_user.is_authenticated and current_user.id == inviter.id:
        flash("That's your own invite link.", "info")
        return redirect(url_for("friends"))

    # 5. Already friends (unused token) — mark used and redirect cleanly
    if current_user.is_authenticated:
        existing = Friend.query.filter_by(
            user_id=current_user.id, friend_id=inviter.id
        ).first()
        if existing:
            invite.used_at = datetime.utcnow()
            db.session.commit()
            flash(f"You're already connected with {inviter.first_name}.", "info")
            return redirect(url_for("friends"))

    # ── Show the holding page — no automatic acceptance ───────────────────────
    # Authenticated users see "Connect with [Name]?" and must confirm explicitly.
    # Unauthenticated users see "Accept Invite" and are routed to auth on confirm.
    initials = _invite_landing_initials(inviter)
    return render_template(
        "invite_landing.html",
        inviter=inviter,
        token=token,
        initials=initials,
    )


@app.route("/invite/<token>/confirm", methods=["POST"])
def invite_token_confirm(token):
    """
    Executes invite acceptance after the user explicitly clicks Accept on the
    landing page. Re-validates the token on every POST (guards against replays,
    expiry races, and double-submits).
    """
    validate_csrf_request()
    invite = InviteToken.query.filter_by(token=token).first()

    if not invite or invite.is_expired() or invite.is_used():
        return render_template("invite_expired.html")

    inviter = db.session.get(User, invite.inviter_id)
    if not inviter:
        return render_template("invite_expired.html")

    # Self-invite guard (user may have signed in on the landing page in another tab)
    if current_user.is_authenticated and current_user.id == inviter.id:
        flash("That's your own invite link.", "info")
        return redirect(url_for("friends"))

    # ── Authenticated: create the friendship now ──────────────────────────────
    if current_user.is_authenticated:
        existing = Friend.query.filter_by(
            user_id=current_user.id, friend_id=inviter.id
        ).first()
        if existing:
            # Already connected — mark token used and show the success screen
            # rather than silently redirecting. Idempotent: safe to call twice.
            invite.used_at = datetime.utcnow()
            db.session.commit()
            return render_template("invite_accepted.html", inviter=inviter)

        # If the user is mid-onboarding, create the friendship now but let
        # before_request gate send them back to /onboarding first.
        # post_onboarding_redirect ensures they land on /friends after setup.
        if not current_user.is_core_profile_complete:
            session["post_onboarding_redirect"] = url_for("friends")

        _apply_invite_token(invite, current_user)
        return render_template("invite_accepted.html", inviter=inviter)

    # ── Unauthenticated: store token and route through auth ───────────────────
    session["invite_token"] = token
    return redirect(url_for("auth"))


@app.route("/setup-profile")
@login_required
def setup_profile():
    """Legacy URL — redirect to canonical /onboarding."""
    return redirect(url_for("onboarding"))


@app.route("/edit_profile", methods=["GET", "POST"])
@login_required
def edit_profile():
    # Explicit permission check: only profile owner can edit
    user = current_user
    if user.id != current_user.id:
        abort(403)
    
    if request.method == "POST":
        new_first = request.form.get("first_name", "").strip()
        new_last = request.form.get("last_name", "").strip()
        if new_first:
            user.first_name = new_first
        if new_last:
            user.last_name = new_last

        user.gender = request.form.get("gender") or None
        birth_year_raw = request.form.get("birth_year")
        user.birth_year = int(birth_year_raw) if birth_year_raw else None
        
        # Handle rider type (single-select, stored as single-item array)
        rider_types_raw = request.form.get("rider_types", "")
        rider_types = [rt.strip() for rt in rider_types_raw.split(",") if rt.strip()]
        user.rider_types = rider_types if rider_types else []
        
        passes_raw = request.form.get("pass_type", "")
        normalized_passes = normalize_pass_selection(passes_raw) or user.pass_type or "no_pass"
        if count_real_passes(normalized_passes) > 3:
            flash("You can select up to 3 passes.", "error")
            return redirect(url_for("edit_profile"))
        _old_pass_ep = user.pass_type  # capture before overwrite for change detection
        user.pass_type = normalized_passes
        user.home_state = request.form.get("home_state") or None
        # Clear skill_level only for Social-only users, otherwise use form value
        is_social_only = rider_types == ["Social"]
        user.skill_level = None if is_social_only else (request.form.get("skill_level") or None)
        user.gear = request.form.get("gear") or None
        home_resort_id_raw = request.form.get("home_resort_id") or None
        if home_resort_id_raw:
            home_resort = db.session.get(Resort, int(home_resort_id_raw))
            if home_resort:
                user.home_resort_id = home_resort.id
            else:
                user.home_resort_id = None
        else:
            user.home_resort_id = None
        
        terrain_raw = request.form.get("terrain_preferences", "")
        terrain_list = [t.strip() for t in terrain_raw.split(",") if t.strip()][:2]
        user.terrain_preferences = terrain_list if terrain_list else []
        user.previous_pass = request.form.get("previous_pass") or None
        user.profile_completed_at = datetime.utcnow()
        user.update_lifecycle_stage()
        
        try:
            db.session.commit()

            # Emit profile_completed event
            emit_event('profile_completed', user)

            if _ph_is_real_pass(normalized_passes):
                ph_analytics.track(user.id, 'pass_added', {
                    'pass_type':    normalized_passes,
                    'source':       'settings',
                    'is_first_pass': not _ph_is_real_pass(_old_pass_ep),
                })

            # ── B3: friend.pass.changed (edit_profile) — centralized dispatch ──
            # One emit per friend → one MEL audit row per recipient.
            if _old_pass_ep != normalized_passes:
                _ep_friend_ids = get_friend_ids(user.id)
                if _ep_friend_ids:
                    _ep_display = format_passes_for_display(normalized_passes).replace(" · ", " + ")
                    current_app.logger.info(
                        "[MESSAGE_DISPATCH] pass_changed (edit_profile): old=%r new=%r friend_count=%d",
                        _old_pass_ep, normalized_passes, len(_ep_friend_ids),
                    )
                    for _friend_id in _ep_friend_ids:
                        emit_messaging_event(
                            event_name=EventName.FRIEND_PASS_CHANGED,
                            actor_user_id=user.id,
                            recipient_user_id=_friend_id,
                            entity_type="user",
                            entity_id=user.id,
                            metadata={
                                "actor_first_name": user.first_name,
                                "new_pass":         normalized_passes,
                                "new_pass_display": _ep_display,
                            },
                            source_route="edit_profile",
                        )

            # Availability nudge: fire once when user just confirmed a real pass
            # but has no open dates set — highest-intent moment to prompt them.
            _should_prompt_avail = (
                _ph_is_real_pass(normalized_passes)
                and not bool(current_user.open_dates)
            )
            if _should_prompt_avail:
                return redirect(url_for("profile", pass_saved="1"))
            return redirect(url_for("profile"))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error saving profile: {e}")
            flash("Something went wrong while saving your profile. Please try again.", "error")
            return redirect(url_for("edit_profile"))
    
    friends_count = Friend.query.filter_by(user_id=user.id).count()
    
    # Build resorts by state from database (exclude region-level entities)
    all_resorts = Resort.query.filter_by(is_active=True, is_region=False).order_by(Resort.state, Resort.name).all()
    resorts_by_state = {}
    for resort in all_resorts:
        if resort.state not in resorts_by_state:
            resorts_by_state[resort.state] = []
        resorts_by_state[resort.state].append({"id": resort.id, "name": resort.name})
    
    return render_template("edit_profile.html", user=user, friends_count=friends_count, state_abbr=STATE_ABBR, pass_options=CANONICAL_PASSES, rider_types=RIDER_TYPES, all_states=ALL_US_STATES, resorts_by_state=resorts_by_state, grouped_locations=get_grouped_locations())

def compute_trip_overlaps(user_trips, friend_trips):
    """
    Canonical trip-overlap detection.

    Compares a user's trips (owned + accepted guest) against friends' public trips
    to find cases where the user and a friend are at the same resort at the same time.

    Resort matching rules (in order):
      1. Both trips have resort_id  → compare resort_id (canonical)
      2. Either trip missing resort_id → compare mountain name strings (legacy fallback)

    Date overlap rule: start1 <= end2 AND start2 <= end1  (inclusive)

    Guards:
      - Skips comparing a trip against itself (my.id == ft.id)
      - Requires both trips to have start_date and end_date

    Returns list of dicts with keys:
        type, my_trip_id, friend_trip_id,
        friend_id, friend_name, friend_first_name,
        mountain, state, brand, resort_id, resort,
        start_date, end_date
    """
    overlaps = []
    for my in user_trips:
        for ft in friend_trips:
            if my.id == ft.id:
                continue
            # Resort match
            if my.resort_id and ft.resort_id:
                same_resort = (my.resort_id == ft.resort_id)
            else:
                my_mtn = my.mountain or (my.resort.name if my.resort else None)
                ft_mtn = ft.mountain or (ft.resort.name if ft.resort else None)
                same_resort = bool(my_mtn and ft_mtn and my_mtn == ft_mtn)
            if not same_resort:
                continue
            if not (my.start_date and my.end_date and ft.start_date and ft.end_date):
                continue
            if not date_ranges_overlap(my.start_date, my.end_date, ft.start_date, ft.end_date):
                continue
            resort = my.resort or ft.resort
            mtn_str = my.mountain or (my.resort.name if my.resort else ft.mountain or (ft.resort.name if ft.resort else None))
            friend_user = ft.user
            friend_first = friend_user.first_name if friend_user else "Friend"
            overlaps.append({
                "type": "trip",
                "my_trip_id": my.id,
                "friend_trip_id": ft.id,
                "friend_id": ft.user_id,
                "friend_name": friend_first,
                "friend_first_name": friend_first,
                "mountain": resort.name if resort else mtn_str,
                "state": resort.state if resort else (getattr(my, 'state', None) or ""),
                "brand": resort.brand if resort else None,
                "resort_id": resort.id if resort else None,
                "resort": resort,
                "start_date": max(my.start_date, ft.start_date),
                "end_date": min(my.end_date, ft.end_date),
            })
    return overlaps


@app.route("/my-trips")
@login_required
def my_trips():
    user = current_user
    today = date.today()
    active_tab = request.args.get("tab", "my_trips")
    _rp_t0 = time.perf_counter()

    # Trip queries (wrapped for production safety)
    _t = time.perf_counter()
    try:
        upcoming_trips = (
            SkiTrip.query
            .options(db.joinedload(SkiTrip.resort))
            .filter(SkiTrip.user_id == current_user.id)
            .filter(SkiTrip.end_date >= today)
            .order_by(SkiTrip.start_date.asc())
            .all()
        ) or []
    except Exception:
        upcoming_trips = []
    if app.debug:
        print(f"[ROUTE_PERF] my_trips.upcoming={time.perf_counter()-_t:.4f}s count={len(upcoming_trips)}")

    _t = time.perf_counter()
    try:
        past_trips = (
            SkiTrip.query
            .options(db.joinedload(SkiTrip.resort))
            .filter(SkiTrip.user_id == current_user.id)
            .filter(SkiTrip.end_date < today)
            .order_by(SkiTrip.start_date.desc())
            .all()
        ) or []
    except Exception:
        past_trips = []
    if app.debug:
        print(f"[ROUTE_PERF] my_trips.past={time.perf_counter()-_t:.4f}s count={len(past_trips)}")

    # Get trips where user is INVITED (pending invites)
    invited_trips = []
    invite_inviters = {}
    try:
        invited_participations = SkiTripParticipant.query.filter(
            SkiTripParticipant.user_id == current_user.id,
            SkiTripParticipant.status == GuestStatus.INVITED
        ).all()
        invited_trip_ids = [p.trip_id for p in invited_participations]
        if invited_trip_ids:
            invited_trips = SkiTrip.query.options(
                db.joinedload(SkiTrip.resort)
            ).filter(
                SkiTrip.id.in_(invited_trip_ids),
                SkiTrip.end_date >= today
            ).order_by(SkiTrip.start_date.asc()).all() or []
            # Batch-load inviters (trip owner = inviter) — one query, no N+1
            inviter_ids = list({t.user_id for t in invited_trips})
            inviter_users = User.query.filter(User.id.in_(inviter_ids)).all() if inviter_ids else []
            _inviter_map = {u.id: u for u in inviter_users}
            invite_inviters = {t.id: _inviter_map.get(t.user_id) for t in invited_trips}
    except Exception as e:
        print(f"  ERROR fetching invited trips: {e}")
        invited_trips = []
        invite_inviters = {}

    # Get trips where user is ACCEPTED as guest (not owner)
    accepted_guest_trips = []
    try:
        accepted_participations = SkiTripParticipant.query.filter(
            SkiTripParticipant.user_id == current_user.id,
            SkiTripParticipant.status == GuestStatus.ACCEPTED
        ).all()
        accepted_trip_ids = [p.trip_id for p in accepted_participations]
        if accepted_trip_ids:
            # Exclude trips the user owns (they're already in upcoming_trips)
            accepted_guest_trips = SkiTrip.query.options(
                db.joinedload(SkiTrip.resort)
            ).filter(
                SkiTrip.id.in_(accepted_trip_ids),
                SkiTrip.user_id != current_user.id,
                SkiTrip.end_date >= today
            ).order_by(SkiTrip.start_date.asc()).all() or []
    except Exception:
        accepted_guest_trips = []

    # Get friends — single join query (replaces get_friend_ids + User.query, saves 1 round trip)
    _t = time.perf_counter()
    try:
        friends = (
            User.query
            .join(Friend, Friend.friend_id == User.id)
            .filter(Friend.user_id == user.id)
            .all()
        )
        friend_ids = [f.id for f in friends]
    except Exception:
        friend_ids = []
        friends = []
    if app.debug:
        print(f"[ROUTE_PERF] my_trips.friends={time.perf_counter()-_t:.4f}s count={len(friends)}")

    # Friends' upcoming trips (wrapped for production safety)
    _t = time.perf_counter()
    friend_trips = []
    try:
        if friend_ids:
            friend_trips = SkiTrip.query.options(
                db.joinedload(SkiTrip.resort)
            ).filter(
                SkiTrip.user_id.in_(friend_ids),
                SkiTrip.end_date >= today,
                SkiTrip.is_public == True
            ).order_by(SkiTrip.start_date.asc()).all() or []
    except Exception:
        friend_trips = []
    if app.debug:
        print(f"[ROUTE_PERF] my_trips.friend_trips={time.perf_counter()-_t:.4f}s count={len(friend_trips)}")

    # Build friends_trips_tab: month + destination grouped rows
    _t = time.perf_counter()
    seven_days_ago_mt = datetime.now() - timedelta(days=7)
    friends_trips_tab = []
    try:
        from collections import OrderedDict as _ODt_mt, defaultdict as _dd_mt
        _friend_map_mt = {f.id: f for f in friends}
        _raw_rows_mt = []
        for _trip in friend_trips:
            _owner = _friend_map_mt.get(_trip.user_id)
            if not _owner:
                continue
            _dest = _trip.resort.name if _trip.resort else (_trip.mountain or 'TBD')
            _status = _trip.trip_status or 'planning'
            _is_new = bool(_trip.created_at and _trip.created_at >= seven_days_ago_mt)
            _fmt_date = format_trip_dates(_trip)
            if _trip.start_date:
                _mkey = _trip.start_date.strftime('%Y-%m')
                _mlabel = _trip.start_date.strftime('%B %Y')
            else:
                _mkey = '9999-99'
                _mlabel = 'Dates TBD'
            _raw_rows_mt.append({
                'destination': _dest,
                'friend_name': f"{_owner.first_name or ''} {_owner.last_name or ''}".strip() or 'Friend',
                'friend_id': _owner.id,
                'status': _status,
                'is_new': _is_new,
                'formatted_date': _fmt_date,
                'month_key': _mkey,
                'month_label': _mlabel,
                'trip_id': _trip.id,
                'trip_start': _trip.start_date,
                'trip_end': _trip.end_date,
            })
        _tab_groups_mt = _dd_mt(list)
        for _row in _raw_rows_mt:
            _tab_groups_mt[(_row['friend_id'], _row['destination'], _row['status'])].append(_row)
        _months_dict_mt = _ODt_mt()
        _seen_gkeys_mt = set()
        for _row in _raw_rows_mt:
            _gkey = (_row['friend_id'], _row['destination'], _row['status'])
            _group = _tab_groups_mt[_gkey]
            _is_grouped = len(_group) >= 3
            if _is_grouped:
                if _gkey in _seen_gkeys_mt:
                    continue
                _seen_gkeys_mt.add(_gkey)
                _sorted_g = sorted([r for r in _group if r['trip_start']], key=lambda r: r['trip_start'])
                if _sorted_g:
                    _first_mo = _sorted_g[0]['trip_start'].strftime('%b')
                    _last_end = _sorted_g[-1]['trip_end']
                    _last_mo = _last_end.strftime('%b') if _last_end else _sorted_g[-1]['trip_start'].strftime('%b')
                    _date_range_lbl = f"{_first_mo}–{_last_mo}" if _first_mo != _last_mo else _first_mo
                else:
                    _date_range_lbl = ''
                _display_row = dict(_row)
                _display_row['grouped'] = True
                _display_row['grouped_count'] = len(_group)
                _display_row['date_range_label'] = _date_range_lbl
                _display_row['grouped_trips'] = [
                    {'trip_id': r['trip_id'], 'formatted_date': r['formatted_date'], 'status': r['status']}
                    for r in sorted(_group, key=lambda r: r['trip_start'] or date.max)
                ]
            else:
                _display_row = dict(_row)
                _display_row['grouped'] = False
            _mk = _display_row['month_key']
            if _mk not in _months_dict_mt:
                _months_dict_mt[_mk] = {'month_label': _display_row['month_label'], 'destinations': _ODt_mt()}
            _dk = _display_row['destination']
            if _dk not in _months_dict_mt[_mk]['destinations']:
                _months_dict_mt[_mk]['destinations'][_dk] = []
            _months_dict_mt[_mk]['destinations'][_dk].append(_display_row)
        friends_trips_tab = [
            {'month_label': _md['month_label'],
             'destinations': [{'name': _dn, 'rows': _dr} for _dn, _dr in _md['destinations'].items()]}
            for _md in _months_dict_mt.values()
        ]
    except Exception:
        friends_trips_tab = []
    if app.debug:
        print(f"[ROUTE_PERF] my_trips.tab_build={time.perf_counter()-_t:.4f}s friend_trip_count={len(friend_trips)}")

    # Build overlaps list — include both owned trips and accepted guest trips
    # so a user who is a guest on a friend's trip at Vail also triggers an overlap
    # with any other friend going to Vail at the same time.
    _t = time.perf_counter()
    try:
        overlaps = compute_trip_overlaps(upcoming_trips + accepted_guest_trips, friend_trips)
    except Exception:
        overlaps = []
    if app.debug:
        print(f"[ROUTE_PERF] my_trips.overlaps={time.perf_counter()-_t:.4f}s count={len(overlaps)}")

    if app.debug:
        print(f"[ROUTE_PERF] route=my_trips total={time.perf_counter()-_rp_t0:.4f}s")
    return render_template(
        "my_trips.html",
        user=user,
        upcoming_trips=upcoming_trips or [],
        past_trips=past_trips or [],
        invited_trips=invited_trips or [],
        invite_inviters=invite_inviters or {},
        accepted_guest_trips=accepted_guest_trips or [],
        active_tab=active_tab,
        friends=friends or [],
        friend_trips=friend_trips or [],
        friends_trips_tab=friends_trips_tab,
        overlaps=overlaps or [],
        today=today
    )

@app.route("/season-snapshot")
@login_required
def season_snapshot():
    today = date.today()
    try:
        upcoming_owned = (
            SkiTrip.query
            .filter(SkiTrip.user_id == current_user.id)
            .filter(SkiTrip.end_date >= today)
            .order_by(SkiTrip.start_date.asc())
            .all()
        ) or []
    except Exception:
        upcoming_owned = []

    try:
        accepted_participations = SkiTripParticipant.query.filter(
            SkiTripParticipant.user_id == current_user.id,
            SkiTripParticipant.status == GuestStatus.ACCEPTED
        ).all()
        accepted_trip_ids = [p.trip_id for p in accepted_participations]
        if accepted_trip_ids:
            accepted_guest_trips = SkiTrip.query.filter(
                SkiTrip.id.in_(accepted_trip_ids),
                SkiTrip.user_id != current_user.id,
                SkiTrip.end_date >= today
            ).order_by(SkiTrip.start_date.asc()).all() or []
        else:
            accepted_guest_trips = []
    except Exception:
        accepted_guest_trips = []

    all_upcoming = sorted(
        upcoming_owned + accepted_guest_trips,
        key=lambda t: t.start_date if t.start_date else date.max
    )

    return render_template(
        "season_snapshot.html",
        user=current_user,
        all_upcoming=all_upcoming,
        today=today,
    )


@app.route("/overlap-detail")
@login_required
def overlap_detail():
    """Overlap detail screen showing friends involved and Start a trip CTA."""
    user = current_user
    
    # Get parameters
    overlap_type = request.args.get('type', 'trip')
    date_str = request.args.get('date')
    mountain = request.args.get('mountain')
    resort_id = request.args.get('resort_id', type=int)
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Parse dates
    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Get friends from the overlap
    friend_ids_str = request.args.get('friends', '')
    friend_ids = [int(fid) for fid in friend_ids_str.split(',') if fid.isdigit()]
    friends = User.query.filter(User.id.in_(friend_ids)).all() if friend_ids else []
    
    # Get resort info if available
    resort = db.session.get(Resort, resort_id) if resort_id else None
    
    # Build display info
    if overlap_type == 'trip':
        title = mountain or (resort.name if resort else 'Trip Overlap')
        subtitle = None
        if start_date and end_date:
            if start_date == end_date:
                subtitle = start_date.strftime('%b %-d')
            else:
                subtitle = f"{start_date.strftime('%b %-d')} – {end_date.strftime('%b %-d')}"
    else:
        # Open date overlap
        if start_date:
            title = f"Open on {start_date.strftime('%b %-d')}"
        else:
            title = "Open Date Overlap"
        subtitle = f"{len(friends)} friend{'s' if len(friends) != 1 else ''} also open"
    
    return render_template(
        "overlap_detail.html",
        user=user,
        overlap_type=overlap_type,
        title=title,
        subtitle=subtitle,
        friends=friends,
        resort=resort,
        resort_id=resort_id,
        mountain=mountain,
        start_date=start_date,
        end_date=end_date,
        date_str=date_str or (start_date_str if start_date_str else '')
    )


@app.route("/mountains")
@login_required
def mountains_tab():
    """Mountains discovery page — shell only. Resort data served via /api/mountains-data
    to keep this HTML response small (~20 KB vs the previous ~210 KB with inline JSON).
    """
    user = current_user
    _rp_t0 = time.perf_counter()

    # ── Warm the resort cache (data itself served via /api/mountains-data) ─────
    _resort_map = get_all_active_resorts_map()
    if app.debug:
        print(f"[ROUTE_PERF] mountains.all_resorts=0.0000s count={len(_resort_map)} (cached)")

    # ── Default filter derived from onboarding state (pure Python, ~0 ms) ──────
    default_state = user.home_state or ""
    default_country = ""
    if default_state:
        for r in _resort_map.values():
            if r.state_code == default_state:
                default_country = r.country_code or ""
                break

    if app.debug:
        print(f"[ROUTE_PERF] route=mountains total={time.perf_counter()-_rp_t0:.4f}s (shell only)")
    return render_template(
        "mountains_tab.html",
        default_country=default_country,
        default_state=default_state,
    )


@app.route("/api/mountains-data")
@login_required
def api_mountains_data():
    """JSON feed for the Mountains tab. Called client-side on DOMContentLoaded.
    Reuses get_all_active_resorts_map() (Phase 1G lru_cache, ~0 ms) plus the
    per-user friend_resort_counts query. Response is authenticated — not cached.
    """
    user = current_user
    _rp_t0 = time.perf_counter()

    _resort_map = get_all_active_resorts_map()

    # ── Friend counts per resort (per-user — must remain dynamic) ─────────────
    _t = time.perf_counter()
    friend_links = Friend.query.filter_by(user_id=user.id).all()
    friend_ids = [f.friend_id for f in friend_links]
    friend_resort_counts = {}
    if friend_ids:
        from sqlalchemy import func as _func
        _counts = db.session.query(
            SkiTrip.resort_id,
            _func.count(_func.distinct(SkiTrip.user_id))
        ).filter(
            SkiTrip.user_id.in_(friend_ids),
            SkiTrip.resort_id.isnot(None)
        ).group_by(SkiTrip.resort_id).all()
        friend_resort_counts = {rid: cnt for rid, cnt in _counts}
    if app.debug:
        print(f"[ROUTE_PERF] api_mountains.friend_resort_counts={time.perf_counter()-_t:.4f}s")

    _t = time.perf_counter()
    resorts_data = [
        {
            "id": r.id,
            "name": r.name,
            "display_name": r.display_name,
            "slug": r.slug,
            "country_code": r.country_code,
            "country_name": r.country_name,
            "state_code": r.state_code,
            "state_name": r.state_name,
            "pass_labels": r.pass_labels,
            "pass_keys": r.pass_keys,
            "friend_count": friend_resort_counts.get(r.id, 0),
        }
        for r in _resort_map.values()
    ]

    # ── Countries (US and CA sorted first) ────────────────────────────────────
    seen_countries = {}
    for rd in resorts_data:
        cc = rd["country_code"]
        if cc and cc not in seen_countries:
            seen_countries[cc] = rd["country_name"] or COUNTRY_NAMES.get(cc, cc)

    def _country_sort(item):
        c = item[0]
        return (0, "") if c == "US" else ((1, "") if c == "CA" else (2, item[1]))

    countries = sorted(seen_countries.items(), key=_country_sort)

    # ── States/regions per country ─────────────────────────────────────────────
    _sbcmap = {}
    for rd in resorts_data:
        cc, sc, sn = rd["country_code"], rd["state_code"], rd["state_name"]
        if cc and sc:
            _sbcmap.setdefault(cc, {})[sc] = sn or sc

    states_by_country = {
        cc: sorted(sd.items(), key=lambda x: x[1])
        for cc, sd in _sbcmap.items()
    }

    # ── Pass filter options (canonical order) ─────────────────────────────────
    _PASS_ORDER = ['epic', 'ikon', 'other']
    seen_pass_keys = set()
    for rd in resorts_data:
        for k in rd['pass_keys']:
            seen_pass_keys.add(k)

    all_passes = []
    seen_in_order = set()
    for k in _PASS_ORDER:
        if k in seen_pass_keys:
            all_passes.append((k, display_pass_label(k)))
            seen_in_order.add(k)
    for k in sorted(seen_pass_keys - seen_in_order):
        all_passes.append((k, display_pass_label(k)))

    if app.debug:
        print(f"[ROUTE_PERF] api_mountains.data_build={time.perf_counter()-_t:.4f}s resort_count={len(resorts_data)}")
        print(f"[ROUTE_PERF] route=api_mountains_data total={time.perf_counter()-_rp_t0:.4f}s resort_count={len(resorts_data)}")

    return jsonify({
        "resorts": resorts_data,
        "states_by_country": states_by_country,
        "all_passes": all_passes,
        "countries": countries,
    })


@app.route("/trip-ideas")
@login_required
def trip_ideas():
    """Deprecated — Ideas feed has moved to Home. Redirect for backwards compatibility."""
    return redirect(url_for("home"))


def _ideas_normalize_pass(pt):
    """
    Short real-pass display for Ideas screens.
    Skips no_pass and no_pass_yet. Returns the display label for the first
    real pass found (including indy, mountain_collective, other, etc.).
    """
    _NON_REAL = frozenset({"no_pass", "no_pass_yet", None, ""})
    if not pt:
        return ""
    for part in str(pt).split(","):
        norm = normalize_pass(part.strip())
        if not norm or norm in _NON_REAL:
            continue
        return display_pass_label(norm)
    return ""


def _ideas_rider_pass_line(user_obj):
    """
    Build 'Skier · Ikon' (or 'Snowboarder · Epic', or just 'Ikon', etc.)
    for a participant row. Returns empty string if neither is set.
    """
    rider = (user_obj.display_rider_type or "").strip()
    norm_pass = _ideas_normalize_pass(user_obj.pass_type)
    if rider and norm_pass:
        return f"{rider} \u00b7 {norm_pass}"
    return rider or norm_pass


def normalize_pass_family(pass_type):
    """
    Map a user's pass_type to its primary real-pass display name for mountain rows.
    Returns 'No pass' when no real pass is found.
    Includes all real passes: epic, ikon, indy, mountain_collective, powder_alliance,
    freedom, ski_california, other.
    """
    _NON_REAL = frozenset({"no_pass", "no_pass_yet", None, ""})
    if not pass_type:
        return "No pass"
    for part in str(pass_type).split(","):
        norm = normalize_pass(part.strip())
        if not norm or norm in _NON_REAL:
            continue
        return display_pass_label(norm)
    return "No pass"


def _mountain_row_identity(rider_type, skill_level, pass_type):
    """
    Build the identity line for a mountain detail person row.
    Format: 'Rider type · Skill level · Pass'  (skill level omitted if blank)
    """
    parts = []
    if rider_type:
        parts.append(rider_type)
    if skill_level:
        parts.append(skill_level)
    parts.append(normalize_pass_family(pass_type))
    return " \u00b7 ".join(parts)


@app.route("/idea/availability")
@login_required
def idea_detail_availability():
    """Detail screen for an availability overlap idea card."""
    from datetime import date as _date
    user = current_user

    # Parse query params
    friend_ids_raw = request.args.get("friend_ids", "")
    try:
        raw_ids = [int(x.strip()) for x in friend_ids_raw.split(",") if x.strip().isdigit()]
    except ValueError:
        raw_ids = []
    resort_id = request.args.get("resort_id", type=int)
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")

    # Security: only expose friends of the current user
    user_friend_ids = {
        f.friend_id for f in Friend.query.filter_by(user_id=user.id).all()
    }
    friend_ids = [fid for fid in raw_ids if fid in user_friend_ids]

    friends = User.query.filter(User.id.in_(friend_ids)).all() if friend_ids else []
    resort = db.session.get(Resort, resort_id) if resort_id else None

    # Format date range for display
    date_range_display = None
    if start_date_str and end_date_str:
        try:
            s = _date.fromisoformat(start_date_str)
            e = _date.fromisoformat(end_date_str)
            if s == e:
                date_range_display = s.strftime("%B %-d")
            elif s.month == e.month:
                date_range_display = f"{s.strftime('%B %-d')} \u2013 {e.strftime('%-d')}"
            else:
                date_range_display = f"{s.strftime('%B %-d')} \u2013 {e.strftime('%B %-d')}"
        except (ValueError, TypeError):
            pass

    participants = []
    for f in friends:
        participants.append({
            "full_name": f"{f.first_name or ''} {f.last_name or ''}".strip(),
            "pass_display": _ideas_rider_pass_line(f),
            "friend_id": f.id,
        })

    user_pass_display = _ideas_rider_pass_line(user)

    _num_words = {1:"One",2:"Two",3:"Three",4:"Four",5:"Five",
                  6:"Six",7:"Seven",8:"Eight",9:"Nine",10:"Ten"}
    n_people = len(friends) + 1
    people_word = _num_words.get(n_people, str(n_people))

    return render_template(
        "idea_detail_availability.html",
        user=user,
        participants=participants,
        resort=resort,
        date_range_display=date_range_display,
        user_pass_display=user_pass_display,
        people_word=people_word,
    )


@app.route("/idea/wishlist")
@login_required
def idea_detail_wishlist():
    """Detail screen for a wishlist overlap idea card."""
    user = current_user

    resort_id = request.args.get("resort_id", type=int)
    friend_ids_raw = request.args.get("friend_ids", "")
    try:
        raw_ids = [int(x.strip()) for x in friend_ids_raw.split(",") if x.strip().isdigit()]
    except ValueError:
        raw_ids = []

    # Security: only expose friends of the current user
    user_friend_ids = {
        f.friend_id for f in Friend.query.filter_by(user_id=user.id).all()
    }
    friend_ids = [fid for fid in raw_ids if fid in user_friend_ids]

    friends = User.query.filter(User.id.in_(friend_ids)).all() if friend_ids else []
    resort = db.session.get(Resort, resort_id) if resort_id else None

    participants = []
    for f in friends:
        participants.append({
            "full_name": f"{f.first_name or ''} {f.last_name or ''}".strip(),
            "pass_display": _ideas_rider_pass_line(f),
            "friend_id": f.id,
        })

    user_pass_display = _ideas_rider_pass_line(user)
    resort_name = _resort_display_name(resort, AMBIGUOUS_RESORT_NAMES) if resort else "this resort"

    return render_template(
        "idea_detail_wishlist.html",
        user=user,
        participants=participants,
        resort=resort,
        resort_name=resort_name,
        user_pass_display=user_pass_display,
    )


@app.route("/idea/trip/<int:trip_id>")
@login_required
def idea_detail_trip(trip_id):
    """Detail screen for a trip overlap idea card — informational only, no CTA."""
    from datetime import date as _date
    user = current_user

    trip = SkiTrip.query.get_or_404(trip_id)

    # Must be a public trip belonging to a friend
    if not trip.is_public:
        abort(404)

    user_friend_ids = {
        f.friend_id for f in Friend.query.filter_by(user_id=user.id).all()
    }
    if trip.user_id not in user_friend_ids and trip.user_id != user.id:
        abort(404)

    trip_owner = db.session.get(User, trip.user_id)
    trip_status = trip.trip_status or "planning"
    anchor_name = trip_owner.first_name if trip_owner else "Your friend"

    # Format date range
    date_range_display = None
    if trip.start_date and trip.end_date:
        s = trip.start_date
        e = trip.end_date
        if s == e:
            date_range_display = s.strftime("%B %-d")
        elif s.month == e.month:
            date_range_display = f"{s.strftime('%B %-d')} \u2013 {e.strftime('%-d')}"
        else:
            date_range_display = f"{s.strftime('%B %-d')} \u2013 {e.strftime('%B %-d')}"

    # Explanatory "why" line varies by trip status
    if trip_status == "going":
        why_line = f"{anchor_name} is going."
    else:
        why_line = f"{anchor_name} is considering this trip."

    # Build participant list: host first, then accepted guests, then "You, if you join"
    participants = []
    if trip_owner:
        participants.append({
            "full_name": f"{trip_owner.first_name or ''} {trip_owner.last_name or ''}".strip(),
            "pass_display": _ideas_rider_pass_line(trip_owner),
            "is_host": True,
            "friend_id": trip_owner.id,
        })

    accepted_rows = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, status=GuestStatus.ACCEPTED
    ).all()
    for row in accepted_rows:
        if row.user_id == trip_owner.id if trip_owner else False:
            continue
        if row.user_id == user.id:
            continue
        guest = db.session.get(User, row.user_id)
        if guest:
            participants.append({
                "full_name": f"{guest.first_name or ''} {guest.last_name or ''}".strip(),
                "pass_display": _ideas_rider_pass_line(guest),
                "is_host": False,
                "friend_id": guest.id,
            })

    user_pass_display = _ideas_rider_pass_line(user)
    resort_name = _resort_display_name(trip.resort, AMBIGUOUS_RESORT_NAMES) if trip.resort else (trip.mountain or "the mountain")

    return render_template(
        "idea_detail_trip.html",
        user=user,
        trip=trip,
        trip_status=trip_status,
        trip_owner=trip_owner,
        anchor_name=anchor_name,
        participants=participants,
        date_range_display=date_range_display,
        why_line=why_line,
        resort_name=resort_name,
        user_pass_display=user_pass_display,
    )

@app.route("/api/mountains/<state>")
def get_mountains(state):
    # DEPRECATED: Not called by any current UI. Kept for external compatibility only.
    # The active Add Trip UI derives states/resorts from the Resort table via get_resorts_for_trip_form().
    state_code = state.upper()
    mountains = MOUNTAINS_BY_STATE.get(state_code, [])
    return jsonify(mountains)

@app.route("/api/trip/create", methods=["POST"])
@login_required
def create_trip():
    user = current_user
    
    data = request.get_json()
    resort_id_raw = data.get("resort_id")
    state = data.get("state")
    mountain = data.get("mountain")
    start_date_str = data.get("start_date")
    end_date_str = data.get("end_date")
    pass_type = data.get("pass_type", user.pass_type or "No Pass")
    is_public = data.get("is_public", True)

    # Resolve resort — canonical source of truth for mountain/state
    resolved_resort = None
    if resort_id_raw:
        try:
            resolved_resort = db.session.get(Resort, int(resort_id_raw))
        except (ValueError, TypeError):
            resolved_resort = None
        if resolved_resort:
            mountain = resolved_resort.name
            state = resolved_resort.state_code or resolved_resort.state

    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else None
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else None
    
    if not start_date or not end_date:
        return jsonify({"success": False, "error": "Please provide both start and end dates."}), 400
    
    if end_date < start_date:
        return jsonify({"success": False, "error": "End date cannot be before start date."}), 400
    
    overlapping = SkiTrip.query.filter(
        SkiTrip.user_id == user.id,
        SkiTrip.start_date <= end_date,
        SkiTrip.end_date >= start_date
    ).first()
    
    if overlapping:
        return jsonify({"success": False, "error": "You already have a trip during these dates."}), 409
    
    # Check if this is a group trip proposal
    friend_id = data.get("friend_id")
    is_group_trip = data.get("is_group", False)
    
    trip = SkiTrip(
        user_id=user.id,
        resort_id=resolved_resort.id if resolved_resort else None,
        state=state,
        mountain=mountain,
        start_date=start_date,
        end_date=end_date,
        pass_type=pass_type,
        is_public=is_public,
        trip_status='planning',
        is_group_trip=is_group_trip or (friend_id is not None),
        created_by_user_id=user.id
    )
    db.session.add(trip)
    db.session.flush()  # Get trip.id before adding participants
    
    # Auto-add owner as participant
    trip.add_owner_as_participant()
    
    # Add participant if friend_id is provided
    if friend_id:
        trip.add_participant(friend_id, GuestStatus.INVITED)
    
    # Track first trip created if not already set
    if not user.first_trip_created_at:
        user.first_trip_created_at = datetime.utcnow()
    
    # Mark planning started (lifecycle signal)
    user.mark_planning_started()
    
    # Update lifecycle stage (planning started)
    user.update_lifecycle_stage()
    
    db.session.commit()

    # ── B5: trip.invite.created (create_trip JSON API) — centralized dispatch ──
    if friend_id:
        emit_messaging_event(
            event_name=EventName.TRIP_INVITE_CREATED,
            actor_user_id=current_user.id,
            recipient_user_id=int(friend_id),
            entity_type="trip",
            entity_id=trip.id,
            metadata={
                "actor_name": current_user.first_name or current_user.username,
                "resort":     mountain or "a trip",
                "trip_id":    trip.id,
            },
            source_route="create_trip",
        )

    # Emit trip_created event
    emit_event('trip_created', user, {
        'trip_id': trip.id,
        'mountain': mountain,
        'state': state
    })
    ph_analytics.track(user.id, 'trip_created', {
        'resort_id':          trip.resort_id,
        'mountain':           mountain,
        'state':              state,
        'is_group':           trip.is_group_trip,
        'has_friend_invited': bool(friend_id),
        'days':               (end_date - start_date).days + 1,
        'source':             'create_trip_api',
    })

    return jsonify({
        "success": True,
        "trip": {
            "id": trip.id,
            "resort_id": trip.resort_id,
            "state": trip.state,
            "state_abbr": STATE_ABBR.get(trip.state, trip.state),
            "mountain": trip.mountain,
            "start_date": trip.start_date.isoformat() if trip.start_date else None,
            "end_date": trip.end_date.isoformat() if trip.end_date else None,
            "pass_type": trip.pass_type,
            "is_public": trip.is_public
        }
    })

@app.route("/api/trip/<int:trip_id>/edit", methods=["POST"])
@login_required
def edit_trip(trip_id):
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    data = request.get_json()
    trip.state = data.get("state")
    trip.mountain = data.get("mountain")
    start_date_str = data.get("start_date")
    end_date_str = data.get("end_date")
    trip.pass_type = data.get("pass_type", trip.pass_type or "No Pass")
    trip.is_public = data.get("is_public", True)
    
    trip.start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else None
    trip.end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else None
    
    if not trip.start_date or not trip.end_date:
        return jsonify({"success": False, "error": "Please provide both start and end dates."}), 400
    
    if trip.end_date < trip.start_date:
        return jsonify({"success": False, "error": "End date cannot be before start date."}), 400
    
    overlapping = SkiTrip.query.filter(
        SkiTrip.user_id == current_user.id,
        SkiTrip.id != trip_id,
        SkiTrip.start_date <= trip.end_date,
        SkiTrip.end_date >= trip.start_date
    ).first()
    
    if overlapping:
        return jsonify({"success": False, "error": "You already have a trip during these dates."}), 409
    
    trip.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({
        "success": True,
        "trip": {
            "id": trip.id,
            "state": trip.state,
            "state_abbr": STATE_ABBR.get(trip.state, trip.state),
            "mountain": trip.mountain,
            "start_date": trip.start_date.isoformat() if trip.start_date else None,
            "end_date": trip.end_date.isoformat() if trip.end_date else None,
            "pass_type": trip.pass_type,
            "is_public": trip.is_public
        }
    })

@app.route("/api/trip/<int:trip_id>/update-dates", methods=["POST"])
@login_required
def update_trip_dates(trip_id):
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    start_date_str = data.get("start_date")
    end_date_str = data.get("end_date")
    if not start_date_str or not end_date_str:
        return jsonify({"success": False, "error": "Both start and end dates are required."}), 400
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"success": False, "error": "Invalid date format."}), 400
    if end_date < start_date:
        return jsonify({"success": False, "error": "End date cannot be before start date."}), 400
    if start_date < date.today():
        return jsonify({"success": False, "error": "Start date cannot be in the past."}), 400
    overlapping = SkiTrip.query.filter(
        SkiTrip.user_id == current_user.id,
        SkiTrip.id != trip_id,
        SkiTrip.start_date <= end_date,
        SkiTrip.end_date >= start_date
    ).first()
    if overlapping:
        return jsonify({"success": False, "error": "You already have a trip during these dates."}), 409
    trip.start_date = start_date
    trip.end_date = end_date
    trip.trip_duration = SkiTrip.calculate_duration(start_date, end_date)
    trip.updated_at = datetime.utcnow()
    try:
        emit_trip_updated_activities(trip, current_user.id, dates_changed=True)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[update_trip_dates] error: {e}")
        return jsonify({"success": False, "error": "Failed to save dates."}), 500
    nights = (end_date - start_date).days
    return jsonify({"success": True, "start_date": start_date.isoformat(), "end_date": end_date.isoformat(), "nights": nights})


@app.route("/api/trip/<int:trip_id>/update-resort", methods=["POST"])
@login_required
def update_trip_resort(trip_id):
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    resort_id = data.get("resort_id")
    if not resort_id:
        return jsonify({"success": False, "error": "A resort is required."}), 400
    resort = Resort.query.filter_by(id=resort_id, is_active=True, is_region=False).first()
    if not resort:
        return jsonify({"success": False, "error": "Invalid resort."}), 400
    resort_actually_changed = trip.resort_id != resort.id
    trip.resort_id = resort.id
    trip.mountain = resort.name
    trip.state = resort.state_code or resort.state
    trip.updated_at = datetime.utcnow()
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[update_trip_resort] error: {e}")
        return jsonify({"success": False, "error": "Failed to save resort."}), 500
    if resort_actually_changed:
        emit_trip_location_changed_activities(trip, current_user.id, resort.name)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"[update_trip_resort] notification error: {e}")
    return jsonify({
        "success": True,
        "resort_id": resort.id,
        "resort_name": resort.name,
        "resort_slug": resort.slug or "",
        "state": trip.state or "",
    })


@app.route("/api/trip/<int:trip_id>/update-pass", methods=["POST"])
@login_required
def update_trip_pass(trip_id):
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    raw_pass = (data.get("pass_type") or "").strip()
    normalized = normalize_pass_selection(raw_pass)
    if not normalized:
        return jsonify({"success": False, "error": "Invalid pass selection."}), 400
    if count_real_passes(normalized) > 3:
        return jsonify({"success": False, "error": "You can select up to 3 passes."}), 400
    pass_actually_changed = trip.pass_type != normalized
    trip.pass_type = normalized
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[update_trip_pass] error: {e}")
        return jsonify({"success": False, "error": "Failed to save pass."}), 500
    display = format_passes_for_display(normalized)
    if pass_actually_changed:
        emit_trip_pass_changed_activities(trip, current_user.id, display)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"[update_trip_pass] notification error: {e}")
    return jsonify({"success": True, "pass_type": normalized, "pass_display": display})


@app.route("/api/trip/<int:trip_id>/delete", methods=["POST"])
@login_required
def delete_trip(trip_id):
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    delete_activities_for_trip(trip_id)
    db.session.delete(trip)
    db.session.commit()
    
    return jsonify({"success": True})


@app.route("/api/trip/<int:trip_id>/participant/settings", methods=["POST"])
@login_required
def update_participant_settings(trip_id):
    """Update current user's lesson and carpool settings for a trip."""
    trip = SkiTrip.query.get_or_404(trip_id)
    
    # Find participant record for current user
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id,
        user_id=current_user.id
    ).first()
    
    if not participant:
        return jsonify({"success": False, "error": "You are not a participant of this trip"}), 403
    
    data = request.get_json() or {}
    
    # Track if carpool changed for activity emission
    old_carpool_role = participant.carpool_role
    old_carpool_seats = participant.carpool_seats
    
    # Update lesson setting
    if 'taking_lesson' in data:
        lesson_val = data['taking_lesson']
        if lesson_val in ['yes', 'no', 'maybe']:
            participant.taking_lesson = LessonChoice(lesson_val)
    
    # Update carpool settings
    if 'carpool_role' in data:
        new_role = data['carpool_role']
        
        if new_role == 'driver':
            participant.carpool_role = CarpoolRole.DRIVER
            participant.needs_ride = None
            seats = data.get('carpool_seats', 1)
            participant.carpool_seats = max(1, min(int(seats), 10)) if seats else 1
        elif new_role == 'rider':
            participant.carpool_role = CarpoolRole.RIDER
            participant.carpool_seats = None
            participant.needs_ride = True
        else:
            # Not participating
            participant.carpool_role = None
            participant.carpool_seats = None
            participant.needs_ride = None
    
    # Handle seat update separately if driver
    if 'carpool_seats' in data and participant.carpool_role == CarpoolRole.DRIVER:
        seats = data['carpool_seats']
        participant.carpool_seats = max(1, min(int(seats), 10)) if seats else 1
    
    db.session.commit()
    
    # Emit activity if user became driver or changed seats
    if participant.carpool_role == CarpoolRole.DRIVER:
        should_emit = (
            old_carpool_role != CarpoolRole.DRIVER or
            old_carpool_seats != participant.carpool_seats
        )
        if should_emit and not trip.is_group_trip:
            emit_carpool_activity(current_user, trip, participant.carpool_seats)
    
    return jsonify({
        "success": True,
        "participant": {
            "taking_lesson": participant.taking_lesson.value if participant.taking_lesson else 'no',
            "carpool_role": participant.carpool_role.value if participant.carpool_role else None,
            "carpool_seats": participant.carpool_seats,
            "needs_ride": participant.needs_ride
        }
    })

@app.route("/api/invite/share", methods=["POST"])
@login_required
@limiter.limit("120 per hour", key_func=_user_or_ip)
def api_invite_share():
    """Record an invite share intent (copy / text / share_sheet) in InviteShareEvent.

    Called from frontend share buttons immediately before or after the copy/share
    action. Non-critical — returns 200 even on internal errors so it never blocks
    the copy/share UX. Fires a founder push in a background thread after commit.

    Body JSON:
        token_type  — 'friend' | 'trip'
        token       — the raw token string (used to resolve token_id)
        action      — 'copy' | 'text' | 'share_sheet'
        source      — 'invite_page' | 'friends_empty_state' | 'trip_detail'
    """
    VALID_ACTIONS     = {"copy", "text", "share_sheet"}
    VALID_SOURCES     = {"invite_page", "friends_empty_state", "trip_detail"}
    VALID_TOKEN_TYPES = {"friend", "trip"}

    try:
        data       = request.get_json(silent=True) or {}
        token_type = (data.get("token_type") or "").strip()
        token_str  = (data.get("token")      or "").strip()
        action     = (data.get("action")     or "").strip()
        source     = (data.get("source")     or "").strip()

        if action     not in VALID_ACTIONS:
            return jsonify({"ok": False, "error": "invalid action"}), 400
        if source     not in VALID_SOURCES:
            return jsonify({"ok": False, "error": "invalid source"}), 400
        if token_type not in VALID_TOKEN_TYPES:
            return jsonify({"ok": False, "error": "invalid token_type"}), 400

        # Resolve token_id from the raw token string when possible
        token_id = None
        if token_str:
            if token_type == "friend":
                _tok = InviteToken.query.filter_by(token=token_str).first()
                token_id = _tok.id if _tok else None
            else:
                _tok = TripInviteToken.query.filter_by(token=token_str).first()
                token_id = _tok.id if _tok else None

        ua = (request.headers.get("User-Agent") or "")[:256]

        evt = InviteShareEvent(
            user_id    = current_user.id,
            token_type = token_type,
            token_id   = token_id,
            token      = token_str or None,
            action     = action,
            source     = source,
            user_agent = ua,
        )
        db.session.add(evt)
        db.session.commit()

        # PostHog supplemental event (supplemental only — admin panel reads DB)
        try:
            ph_analytics.track(current_user.id, "invite_share_intent", {
                "token_type": token_type,
                "action":     action,
                "source":     source,
            })
        except Exception:
            pass

        # Founder push in background thread — plain scalars only, no ORM objects
        _uid = current_user.id
        _tt, _act, _src = token_type, action, source
        def _fire_invite_push():
            with app.app_context():
                _send_founder_invite_share_push(_uid, _tt, _act, _src)
        threading.Thread(target=_fire_invite_push, daemon=True).start()

        app.logger.info(
            "[invite_share] user_id=%d token_type=%s action=%s source=%s",
            current_user.id, token_type, action, source,
        )
        return jsonify({"ok": True})

    except Exception as _exc:
        app.logger.exception("[invite_share] unhandled error: %s", _exc)
        return jsonify({"ok": True})  # always 200 — never block the share UX


@app.route("/api/friends/invite", methods=["POST"])
@login_required
@limiter.limit("20 per hour", key_func=_user_or_ip)
def invite_friend():
    # Authentication guard (already protected by @login_required)
    if not current_user.is_authenticated:
        abort(401)
    
    # Safe form data handling
    data = request.get_json() or {}
    friend_email = data.get("friend_email")
    
    if not friend_email:
        return jsonify({"success": False, "error": "Friend email is required"}), 400
    
    # Validate target user exists
    friend = User.query.filter_by(email=friend_email).first()
    if not friend:
        return jsonify({"success": False, "error": "User not found"}), 404
    
    if friend.id == current_user.id:
        return jsonify({"success": False, "error": "Cannot add yourself as a friend"}), 400
    
    # Prevent duplicate friendships
    existing_friendship = Friend.query.filter_by(user_id=current_user.id, friend_id=friend.id).first()
    if existing_friendship:
        return jsonify({"success": False, "error": "Already friends"}), 409
    
    # Prevent duplicate invites
    existing_invitation = Invitation.query.filter_by(sender_id=current_user.id, receiver_id=friend.id, status='pending').first()
    if existing_invitation:
        return jsonify({"success": False, "error": "Invitation already sent"}), 409
    
    # Database write safety
    try:
        invitation = Invitation(sender_id=current_user.id, receiver_id=friend.id, status='pending')
        db.session.add(invitation)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Invite failed")
        return jsonify({"success": False, "error": "Invite failed"}), 500

    # ── B1: friend.request.created — routed through centralized dispatch ──
    emit_messaging_event(
        event_name=EventName.FRIEND_REQUEST_CREATED,
        actor_user_id=current_user.id,
        recipient_user_id=friend.id,
        entity_type="user",
        entity_id=friend.id,
        metadata={
            "actor_name":    current_user.first_name or current_user.username,
            "invitation_id": invitation.id,
            "user_id":       current_user.id,
        },
        source_route="invite_friend",
    )

    return jsonify({"success": True, "message": "Invitation sent"}), 201

@app.route("/api/friends", methods=["GET"])
@login_required
def get_friends():
    friends = Friend.query.filter_by(user_id=current_user.id).all()
    friends_list = []
    for f in friends:
        if not f.friend:
            continue
        friends_list.append({
            "id": f.friend.id,
            "name": f"{f.friend.first_name} {f.friend.last_name}",
            "email": f.friend.email,
            "pass_type": f.friend.pass_type or "No Pass"
        })
    return jsonify({"success": True, "friends": friends_list}), 200

@app.route("/api/friends/<int:friend_id>", methods=["GET"])
@login_required
def get_friend_profile(friend_id):

    friend = db.session.get(User, friend_id)
    if not friend:
        return jsonify({"success": False, "error": "User not found"}), 404

    if friend.id != current_user.id:
        _auth = Friend.query.filter_by(
            user_id=current_user.id, friend_id=friend.id
        ).first()
        if not _auth:
            return jsonify({"success": False, "error": "Not authorized"}), 403

    return jsonify({
        "success": True,
        "friend": {
            "id": friend.id,
            "name": f"{friend.first_name} {friend.last_name}",
            "pass_type": friend.pass_type or "No Pass",
            "rider_type": friend.display_rider_type or "Not specified"
        }
    }), 200

@app.route("/api/friends/invite/<int:invitation_id>/accept", methods=["POST"])
@login_required
def accept_invitation(invitation_id):
    invitation = db.session.get(Invitation, invitation_id)

    if not invitation:
        return jsonify({"success": False, "error": "Invitation not found"}), 404

    if invitation.receiver_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    # Guard: declined invitations cannot be accepted. Blocks the direct-API edge
    # case where a receiver declines via the UI and then calls accept on the same
    # invitation ID. Only 'pending' invitations may proceed to accept.
    if invitation.status == 'declined':
        return jsonify({"success": False, "error": "This invitation is no longer active"}), 409

    # Idempotency: if users are already friends (double-tap, retry, or QR path raced ahead),
    # mark the invitation accepted and return success without creating duplicate Friend rows
    # or firing duplicate activity/messaging events.
    already_friends = Friend.query.filter_by(
        user_id=current_user.id, friend_id=invitation.sender_id
    ).first()
    if already_friends:
        if invitation.status != 'accepted':
            invitation.status = 'accepted'
            db.session.commit()
        return jsonify({"success": True, "message": "Already friends"}), 200

    invitation.status = 'accepted'

    friend_relationship = Friend(user_id=current_user.id, friend_id=invitation.sender_id)
    reverse_friend = Friend(user_id=invitation.sender_id, friend_id=current_user.id)

    db.session.add(friend_relationship)
    db.session.add(reverse_friend)
    emit_connection_accepted_activity(current_user.id, invitation.sender_id)
    db.session.commit()
    _fc_count = Friend.query.filter_by(user_id=current_user.id).count()
    ph_analytics.track(current_user.id, 'friend_connected', {
        'source':          'invitation_accept',
        'is_first_friend': _fc_count == 1,
    })

    # Store sender name for the one-time home page "connected" moment (acting user only)
    sender = db.session.get(User, invitation.sender_id)
    if sender:
        session['new_connection_name'] = sender.first_name or sender.username or 'your new friend'

    # ── B2: friend.request.accepted — routed through centralized dispatch ──
    emit_messaging_event(
        event_name=EventName.FRIEND_REQUEST_ACCEPTED,
        actor_user_id=current_user.id,
        recipient_user_id=invitation.sender_id,
        entity_type="user",
        entity_id=current_user.id,
        metadata={
            "actor_name": current_user.first_name or current_user.username,
            "user_id":    current_user.id,
        },
        source_route="accept_invitation",
    )

    return jsonify({"success": True, "message": "Friend added"}), 200

@app.route("/api/friends/invite/<int:invitation_id>/decline", methods=["POST"])
@login_required
def decline_invitation(invitation_id):
    invitation = db.session.get(Invitation, invitation_id)
    if not invitation:
        return jsonify({"success": False, "error": "Invitation not found"}), 404
    if invitation.receiver_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    if invitation.status != 'pending':
        return jsonify({"success": True, "message": "Already resolved"}), 200
    invitation.status = 'declined'
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"success": False, "error": "Could not decline"}), 500
    return jsonify({"success": True, "message": "Invitation declined"}), 200


@app.route("/api/friends/<int:friend_id>", methods=["DELETE"])
@login_required
def remove_friend(friend_id):
    friend1 = Friend.query.filter_by(user_id=current_user.id, friend_id=friend_id).first()
    friend2 = Friend.query.filter_by(user_id=friend_id, friend_id=current_user.id).first()

    if not friend1 and not friend2:
        return jsonify({"success": False, "error": "Friendship not found"}), 404

    if friend1:
        db.session.delete(friend1)
    if friend2:
        db.session.delete(friend2)

    # Cancel pending AND accepted friend invitations between the two users in
    # either direction so no invitation row can produce a ghost connected-state
    # after the Friend rows are removed.
    Invitation.query.filter(
        db.or_(
            db.and_(
                Invitation.sender_id == current_user.id,
                Invitation.receiver_id == friend_id,
                Invitation.status.in_(['pending', 'accepted']),
            ),
            db.and_(
                Invitation.sender_id == friend_id,
                Invitation.receiver_id == current_user.id,
                Invitation.status.in_(['pending', 'accepted']),
            ),
        )
    ).update({'status': 'cancelled'}, synchronize_session=False)

    # Expire the initiator's active invite token so a link shared before the
    # unfriend can't be used to immediately reconnect.  A fresh token is issued
    # the next time they visit /invite.
    _now = datetime.utcnow()
    for _tok in InviteToken.query.filter_by(inviter_id=current_user.id).all():
        if not _tok.is_used() and not _tok.is_expired():
            _tok.expires_at = _now

    db.session.commit()

    return jsonify({"success": True, "message": "Friend removed"}), 200


@app.route("/api/friends/<int:friend_id>/set-trip-invites", methods=["POST"])
@login_required
def set_trip_invites(friend_id):
    """Set trip_invites_allowed for a friendship (explicit Yes/No, no toggle)."""
    friendship = Friend.query.filter_by(user_id=current_user.id, friend_id=friend_id).first()
    
    if not friendship:
        return jsonify({"success": False, "error": "Friendship not found"}), 404
    
    data = request.get_json() or {}
    allow = data.get('allow', False)
    
    friendship.trip_invites_allowed = bool(allow)
    db.session.commit()
    
    return jsonify({
        "success": True,
        "trip_invites_allowed": friendship.trip_invites_allowed
    }), 200


@app.route("/api/buddy-pass", methods=["POST"])
@login_required
def update_buddy_pass():
    """Update buddy pass availability for a specific pass."""
    SUPPORTED_PASSES = ['epic', 'ikon', 'other']
    
    data = request.get_json() or {}
    pass_key = data.get('pass_key', '').lower()
    available = data.get('available', False)
    
    if pass_key not in SUPPORTED_PASSES:
        return jsonify({"success": False, "error": "Unsupported pass"}), 400
    
    # Create a fresh dict to ensure SQLAlchemy detects the change
    buddy_passes = dict(current_user.buddy_passes or {})
    buddy_passes[pass_key] = bool(available)
    current_user.buddy_passes = buddy_passes
    
    db.session.commit()
    
    return jsonify({"success": True, "buddy_passes": buddy_passes}), 200


@app.route("/api/push/register-token", methods=["POST"])
@login_required
def push_register_token():
    """Store or refresh an iOS push notification device token for the current user.

    apns_environment resolution order:
      1. Frontend hint via request body ("sandbox" | "production" | "unknown").
         Capacitor.DEBUG=false → production; Capacitor.DEBUG=true → sandbox.
      2. Server inference from APNS_USE_SANDBOX env var (reliable operational default):
         APNS_USE_SANDBOX=false → "production" (TestFlight / App Store)
         APNS_USE_SANDBOX=true  → "sandbox"    (local Xcode installs)
      3. Fall back to "unknown" if neither is available.

    This value is used to select the right token when calling APNs so we never
    send a sandbox token to api.push.apple.com or vice-versa.
    """
    data = request.get_json() or {}
    token = data.get("token", "").strip()
    platform = data.get("platform", "ios").strip() or "ios"

    if not token:
        return jsonify({"success": False, "error": "token is required"}), 400

    token_preview = token[:8] + "…" + token[-6:] if len(token) > 14 else token[:8] + "…"

    # Resolve apns_environment — only meaningful for iOS; Android uses FCM
    if platform == "android":
        apns_env = "n/a"
        env_source = "android"
    else:
        client_hint = (data.get("apns_environment") or "").strip().lower()
        if client_hint in ("sandbox", "production"):
            apns_env = client_hint
            env_source = "client_hint"
        else:
            use_sandbox = os.environ.get("APNS_USE_SANDBOX", "true").lower() not in ("false", "0", "no")
            apns_env = "sandbox" if use_sandbox else "production"
            env_source = "server_inferred"

    try:
        existing = PushDeviceToken.query.filter_by(
            user_id=current_user.id, token=token
        ).first()
        if existing:
            active_before  = existing.active
            env_before     = existing.apns_environment
            existing.active     = True
            existing.updated_at = datetime.utcnow()
            # Environment precedence: only overwrite if the stored value is NULL or
            # "unknown". sandbox/production values are confirmed by APNs behavior
            # (either initial registration or a successful retry correction) and must
            # not be overwritten by a weaker client-side hint.
            if existing.apns_environment in (None, "unknown") or existing.apns_environment == apns_env:
                existing.apns_environment = apns_env
                env_preserved = False
            else:
                env_preserved = True
            env_after = existing.apns_environment
            action    = "refreshed"
            new_row   = None
        else:
            active_before = None
            env_before    = None
            env_preserved = False
            env_after     = apns_env
            new_row = PushDeviceToken(
                user_id=current_user.id,
                token=token,
                platform=platform,
                apns_environment=apns_env,
            )
            db.session.add(new_row)
            action = "inserted"
        db.session.flush()   # populate new_row.id (insert) before commit
        tok_id = existing.id if existing else new_row.id

        # ── Hygiene: deactivate all OTHER active tokens for this user/platform ──
        # This caps the user to 1 active token per platform, preventing stale
        # duplicate rows from causing OneSignal invalid_aliases errors after
        # repeated TestFlight reinstalls.
        stale_others = (
            PushDeviceToken.query
            .filter(
                PushDeviceToken.user_id == current_user.id,
                PushDeviceToken.platform == platform,
                PushDeviceToken.active == True,
                PushDeviceToken.id != tok_id,
            )
            .all()
        )
        for _stale in stale_others:
            _stale.active = False
            current_app.logger.warning(
                "[PushToken] Deactivated stale token id=%s user=%s platform=%s",
                _stale.id, current_user.id, platform,
            )

        db.session.commit()
        current_app.logger.info(
            "[PushToken] action=%s token_id=%s user_id=%s token=%s platform=%s "
            "active_before=%s active_after=True "
            "apns_environment_before=%s apns_environment_after=%s "
            "env_source=%s env_preserved=%s stale_deactivated=%s",
            action, tok_id, current_user.id, token_preview, platform,
            active_before, env_before, env_after,
            env_source, env_preserved, len(stale_others),
        )
    except Exception:
        db.session.rollback()
        current_app.logger.exception(
            "[PushToken] failed for user_id=%s token=%s", current_user.id, token_preview
        )
        return jsonify({"success": False, "error": "Server error"}), 500

    stored_env = existing.apns_environment if existing else apns_env
    return jsonify({
        "success": True,
        "action": action,
        "token_preview": token_preview,
        "apns_environment": stored_env,
        "env_preserved": env_preserved if action == "refreshed" else False,
    }), 200


@app.route("/api/push/beacon", methods=["POST"])
@login_required
def push_debug_beacon():
    """Lightweight step-beacon for push registration diagnostics.

    The TestFlight WKWebView cannot surface console.log to the developer, so
    the push JS POSTs a small beacon at each step — giving server-log visibility
    into exactly how far the script progresses inside the native shell.

    Body: { step: str, data: object }
    Logs: [PushBeacon] user_id=X step=Y data=Z
    """
    body = request.get_json(silent=True) or {}
    step = str(body.get("step", "unknown"))[:64]
    raw_data = body.get("data", {})
    if not isinstance(raw_data, dict):
        raw_data = {"raw": str(raw_data)[:200]}
    safe_data = {str(k)[:32]: str(v)[:200] for k, v in list(raw_data.items())[:20]}
    current_app.logger.warning(
        "[PushBeacon] user_id=%s step=%s data=%s",
        current_user.id, step, safe_data
    )
    return jsonify({"ok": True}), 200


@app.route("/api/push/preferences", methods=["POST"])
@login_required
def push_preferences():
    """Set the current user's push notification preference.

    Body: { "push_enabled": true | false }
    Returns: { "ok": true, "push_enabled": <bool> }

    Writes push_notifications_enabled on the User row. The backend
    send_onesignal_push() helper checks this flag before sending.
    The native client should also call window.blSetPushEnabled() after
    a successful response to opt the OneSignal subscription in/out.
    """
    body = request.get_json(silent=True) or {}
    if "push_enabled" not in body:
        return jsonify({"error": "push_enabled field required"}), 400
    enabled = body["push_enabled"]
    if not isinstance(enabled, bool):
        return jsonify({"error": "push_enabled must be a boolean"}), 400
    current_user.push_notifications_enabled = enabled
    try:
        db.session.commit()
        current_app.logger.warning(
            "[PushPref] user_id=%s push_notifications_enabled=%s",
            current_user.id, enabled,
        )
    except Exception as _e:
        db.session.rollback()
        current_app.logger.error("[PushPref] save failed: %s", _e)
        return jsonify({"error": "save_failed"}), 500
    return jsonify({"ok": True, "push_enabled": enabled}), 200


# ---------------------------------------------------------------------------
# Firebase Cloud Messaging (FCM) — Android push sending
# ---------------------------------------------------------------------------

_firebase_admin_app = None


def _get_firebase_admin():
    """Return an initialized Firebase Admin app, or None if unavailable.

    Reads FIREBASE_SERVICE_ACCOUNT_JSON (full JSON string) from the environment.
    Initializes once and caches. Fails gracefully if the secret is missing.
    """
    global _firebase_admin_app
    if _firebase_admin_app is not None:
        return _firebase_admin_app

    sa_json = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()
    if not sa_json:
        return None

    try:
        import json as _json
        import firebase_admin
        from firebase_admin import credentials as _fb_creds
        sa_dict = _json.loads(sa_json)
        cred = _fb_creds.Certificate(sa_dict)
        _firebase_admin_app = firebase_admin.initialize_app(cred)
        print("[FCM] Firebase Admin SDK initialized.")
        return _firebase_admin_app
    except Exception as _e:
        print(f"[FCM] Firebase Admin init failed: {_e}")
        return None


def send_fcm_push(token, title, body, data=None):
    """Send a push notification via Firebase Cloud Messaging to an Android device.

    Args:
        token  -- FCM registration token
        title  -- notification title string
        body   -- notification body string
        data   -- optional dict of string key/value data payload

    Returns dict with: success (bool), message_id (str|None), error (str|None).
    Never logs the full token value.
    """
    token_preview = token[:8] + "\u2026" + token[-6:] if len(token) > 14 else token[:8] + "\u2026"

    fa = _get_firebase_admin()
    if fa is None:
        current_app.logger.error(
            "[FCM] Cannot send — FIREBASE_SERVICE_ACCOUNT_JSON not configured. "
            "provider=fcm platform=android token=%s", token_preview
        )
        return {"success": False, "error": "Firebase Admin not configured", "message_id": None}

    try:
        import firebase_admin.messaging as _fb_msg
        notification = _fb_msg.Notification(title=title, body=body)
        android_config = _fb_msg.AndroidConfig(
            priority="high",
            notification=_fb_msg.AndroidNotification(
                channel_id="baselodge_default",
            ),
        )
        message = _fb_msg.Message(
            notification=notification,
            android=android_config,
            data={str(k): str(v) for k, v in (data or {}).items()},
            token=token,
        )
        response = _fb_msg.send(message)
        current_app.logger.warning(
            "[FCM] provider=fcm platform=android token=%s message_id=%s status=success",
            token_preview, response,
        )
        return {"success": True, "message_id": response, "error": None}
    except Exception as _fcm_e:
        err_str = str(_fcm_e)[:200]
        current_app.logger.warning(
            "[FCM] provider=fcm platform=android token=%s status=failed error=%s",
            token_preview, err_str,
        )
        return {"success": False, "error": err_str, "message_id": None}


# ---------------------------------------------------------------------------
# APNs push sending
# ---------------------------------------------------------------------------

def _apns_jwt():
    """Create a signed JWT for APNs HTTP/2 bearer auth.

    Reads:
        APNS_KEY_P8   — raw contents of the .p8 private key file
        APNS_KEY_ID   — 10-char Key ID shown in Apple Developer portal
        APNS_TEAM_ID  — 10-char Apple Developer Team ID
    """
    key_p8 = os.environ.get("APNS_KEY_P8", "")
    key_id  = os.environ.get("APNS_KEY_ID", "")
    team_id = os.environ.get("APNS_TEAM_ID", "")
    if not (key_p8 and key_id and team_id):
        raise RuntimeError("APNS_KEY_P8 / APNS_KEY_ID / APNS_TEAM_ID env vars are not set")
    token = jwt.encode(
        {"iss": team_id, "iat": int(time.time())},
        key_p8,
        algorithm="ES256",
        headers={"kid": key_id},
    )
    return token


def send_apns_push(
    device_token: str,
    title: str,
    body: str,
    extra: dict | None = None,
    *,
    prefer_sandbox: bool | None = None,
) -> dict:
    """Send a push notification to one APNs device token.

    On BadEnvironmentKeyInToken, retries against the opposite APNs host once.
    If the retry succeeds, updates the PushDeviceToken row's apns_environment
    to the environment that actually worked and keeps the token active.
    Only marks the token inactive if BOTH environments fail.

    Args:
        device_token:   Raw APNs device token hex string.
        title:          Notification title.
        body:           Notification body.
        extra:          Optional dict merged into the payload root.
        prefer_sandbox: Override the first-attempt APNs environment.
                        True  → try sandbox first.
                        False → try production first.
                        None  → derive from APNS_USE_SANDBOX env var (default).

    Returns a structured dict with keys:
        success, final_success,
        first_attempt_environment, first_attempt_host,
        first_attempt_status_code, first_attempt_error, first_attempt_apns_id,
        retry_attempted, retry_environment, retry_host,
        retry_status_code, retry_error, retry_success, retry_apns_id,
        env_corrected, corrected_token_environment,
        bundle_id

    Environment variables:
        APNS_BUNDLE_ID   — e.g. com.baselodge.app
        APNS_USE_SANDBOX — "true" for local Xcode / simulator.
                           "false" for TestFlight / App Store.
    """
    bundle_id = os.environ.get("APNS_BUNDLE_ID", "com.baselodge.app")
    env_use_sandbox = os.environ.get("APNS_USE_SANDBOX", "true").lower() not in ("false", "0", "no")
    first_sandbox = env_use_sandbox if prefer_sandbox is None else prefer_sandbox

    token_preview = device_token[:8] + "…" + device_token[-6:] if len(device_token) > 14 else device_token[:8] + "…"

    payload = {
        "aps": {"alert": {"title": title, "body": body}, "sound": "default", "badge": 1},
    }
    if extra:
        payload.update(extra)

    current_app.logger.warning("[APNs] payload aps=%s", payload.get("aps"))

    try:
        bearer = _apns_jwt()
    except RuntimeError as exc:
        current_app.logger.error("[APNs] Config error: %s", exc)
        first_env = "sandbox" if first_sandbox else "production"
        first_host = "api.sandbox.push.apple.com" if first_sandbox else "api.push.apple.com"
        return {
            "success": False, "final_success": False,
            "first_attempt_environment": first_env, "first_attempt_host": first_host,
            "first_attempt_status_code": None, "first_attempt_error": str(exc),
            "first_attempt_apns_id": None,
            "retry_attempted": False, "retry_environment": None, "retry_host": None,
            "retry_status_code": None, "retry_error": None, "retry_success": False,
            "retry_apns_id": None, "env_corrected": False,
            "corrected_token_environment": None, "bundle_id": bundle_id,
        }

    def _fire(sandbox: bool) -> tuple:
        """Fire one APNs HTTP/2 request. Returns (status_code, apns_id, reason)."""
        h = "api.sandbox.push.apple.com" if sandbox else "api.push.apple.com"
        hdrs = {
            "authorization": f"bearer {bearer}",
            "apns-topic": bundle_id,
            "apns-push-type": "alert",
            "content-type": "application/json",
        }
        try:
            with httpx.Client(http2=True) as client:
                resp = client.post(
                    f"https://{h}/3/device/{device_token}",
                    content=json.dumps(payload),
                    headers=hdrs,
                    timeout=10,
                )
        except Exception as exc:
            current_app.logger.error("[APNs] HTTP error (sandbox=%s): %s", sandbox, exc)
            return None, None, str(exc)
        aid = resp.headers.get("apns-id")
        current_app.logger.info(
            "[APNs] token=%s host=%s → %d  apns-id=%s  body=%s",
            token_preview, h, resp.status_code, aid, resp.text[:120],
        )
        if resp.status_code == 200:
            return 200, aid, None
        try:
            reason = resp.json().get("reason", resp.text)
        except Exception:
            reason = resp.text
        return resp.status_code, aid, reason

    # ── First attempt ──────────────────────────────────────────────────────────
    first_env_name = "sandbox" if first_sandbox else "production"
    first_host     = "api.sandbox.push.apple.com" if first_sandbox else "api.push.apple.com"
    first_code, first_apns_id, first_error = _fire(first_sandbox)

    current_app.logger.info(
        "[APNs] first attempt: env=%s token=%s status=%s error=%s",
        first_env_name, token_preview, first_code, first_error,
    )

    if first_code == 200:
        return {
            "success": True, "final_success": True,
            "first_attempt_environment": first_env_name, "first_attempt_host": first_host,
            "first_attempt_status_code": 200, "first_attempt_error": None,
            "first_attempt_apns_id": first_apns_id,
            "retry_attempted": False, "retry_environment": None, "retry_host": None,
            "retry_status_code": None, "retry_error": None, "retry_success": False,
            "retry_apns_id": None, "env_corrected": False,
            "corrected_token_environment": None, "bundle_id": bundle_id,
        }

    # ── Decide whether to retry the opposite host ─────────────────────────────
    # BadEnvironmentKeyInToken: Apple explicitly says wrong environment.
    # BadDeviceToken: can mean the stored apns_environment label is wrong and the
    #   token is valid on the other host — always probe opposite before deactivating.
    # Truly permanent errors that no environment change can fix:
    #   Unregistered (app uninstalled / token revoked), DeviceTokenNotForTopic
    #   (wrong bundle), and 410 Gone.
    _RETRY_REASONS    = {"BadEnvironmentKeyInToken", "BadDeviceToken"}
    _PERMANENT_REASONS = {"Unregistered", "DeviceTokenNotForTopic"}
    should_retry = first_error in _RETRY_REASONS
    is_permanent = (
        first_code == 410
        or (first_code == 400 and first_error in _PERMANENT_REASONS)
    )

    if not should_retry and (is_permanent or first_code is None):
        if is_permanent:
            try:
                stale = PushDeviceToken.query.filter_by(token=device_token).first()
                if stale:
                    stale.active = False
                    db.session.commit()
                    current_app.logger.warning(
                        "[APNs] Token marked inactive (%s %s) id=%s token=%s",
                        first_code, first_error, stale.id, token_preview,
                    )
            except Exception:
                db.session.rollback()
                current_app.logger.exception("[APNs] Failed to mark token inactive")
        return {
            "success": False, "final_success": False,
            "first_attempt_environment": first_env_name, "first_attempt_host": first_host,
            "first_attempt_status_code": first_code, "first_attempt_error": first_error,
            "first_attempt_apns_id": first_apns_id,
            "retry_attempted": False, "retry_environment": None, "retry_host": None,
            "retry_status_code": None, "retry_error": None, "retry_success": False,
            "retry_apns_id": None, "env_corrected": False,
            "corrected_token_environment": None, "bundle_id": bundle_id,
        }

    # ── Retry against the opposite host ────────────────────────────────────────
    retry_sandbox   = not first_sandbox
    retry_env_name  = "sandbox" if retry_sandbox else "production"
    retry_host      = "api.sandbox.push.apple.com" if retry_sandbox else "api.push.apple.com"

    current_app.logger.warning(
        "[APNs] %s on %s — retrying opposite env=%s token=%s",
        first_error, first_host, retry_env_name, token_preview,
    )

    retry_code, retry_apns_id, retry_error = _fire(retry_sandbox)

    if retry_code == 200:
        # Retry succeeded — Apple confirmed the token belongs to retry_env_name
        current_app.logger.warning(
            "[APNs] retry succeeded on %s — correcting token env to '%s' token=%s",
            retry_host, retry_env_name, token_preview,
        )
        try:
            tok_row = PushDeviceToken.query.filter_by(token=device_token).first()
            if tok_row:
                tok_row.apns_environment = retry_env_name
                tok_row.active = True
                db.session.commit()
                current_app.logger.warning(
                    "[APNs] Token id=%s apns_environment corrected to '%s', kept active",
                    tok_row.id, retry_env_name,
                )
        except Exception:
            db.session.rollback()
            current_app.logger.exception("[APNs] Failed to correct token environment after retry")
        return {
            "success": True, "final_success": True,
            "first_attempt_environment": first_env_name, "first_attempt_host": first_host,
            "first_attempt_status_code": first_code, "first_attempt_error": first_error,
            "first_attempt_apns_id": first_apns_id,
            "retry_attempted": True, "retry_environment": retry_env_name,
            "retry_host": retry_host, "retry_status_code": 200, "retry_error": None,
            "retry_success": True, "retry_apns_id": retry_apns_id,
            "env_corrected": True, "corrected_token_environment": retry_env_name,
            "bundle_id": bundle_id,
        }

    # Both environments failed — deactivate the token
    current_app.logger.warning(
        "[APNs] retry failed (%s %s) on %s — both envs exhausted, deactivating token=%s",
        retry_code, retry_error, retry_host, token_preview,
    )
    try:
        stale = PushDeviceToken.query.filter_by(token=device_token).first()
        if stale:
            stale.active = False
            # Stamp with the opposite of whichever host was tried first
            stale.apns_environment = retry_env_name
            db.session.commit()
            current_app.logger.warning(
                "[APNs] Token id=%s marked inactive after both environments failed",
                stale.id,
            )
    except Exception:
        db.session.rollback()
        current_app.logger.exception("[APNs] Failed to mark token inactive after both-env failure")

    return {
        "success": False, "final_success": False,
        "first_attempt_environment": first_env_name, "first_attempt_host": first_host,
        "first_attempt_status_code": first_code, "first_attempt_error": first_error,
        "first_attempt_apns_id": first_apns_id,
        "retry_attempted": True, "retry_environment": retry_env_name,
        "retry_host": retry_host, "retry_status_code": retry_code,
        "retry_error": retry_error, "retry_success": False, "retry_apns_id": retry_apns_id,
        "env_corrected": False, "corrected_token_environment": None,
        "bundle_id": bundle_id,
    }


@app.route("/admin/push-diagnostics", methods=["GET"])
@login_required
@admin_required
def admin_push_diagnostics():
    """Full push pipeline diagnostic — all layers in one JSON response.

    Covers: APNs env config, Capacitor/JS setup, all token rows, instructions.
    Check server logs for [PushBeacon] entries after opening the TestFlight app
    to see exactly which step the push script reaches.
    """
    target_user_id = 2

    def _tok_preview(t):
        return t[:8] + "\u2026" + t[-6:] if len(t) > 14 else t[:8] + "\u2026"

    use_sandbox_raw = os.environ.get("APNS_USE_SANDBOX", "true")
    use_sandbox = use_sandbox_raw.lower() not in ("false", "0", "no")
    apns_host = "api.sandbox.push.apple.com" if use_sandbox else "api.push.apple.com"
    bundle_id = os.environ.get("APNS_BUNDLE_ID", "com.baselodge.app")

    all_tokens = (
        PushDeviceToken.query
        .filter_by(user_id=target_user_id, platform="ios")
        .order_by(PushDeviceToken.updated_at.desc())
        .all()
    )
    target_user = db.session.get(User, target_user_id)

    active_tokens = [t for t in all_tokens if t.active]
    target_env = "production" if not use_sandbox else "sandbox"
    # Token selected for the current APNs environment (production or sandbox first,
    # unknown as fallback, mismatched last resort)
    env_matched   = [t for t in active_tokens if t.apns_environment == target_env]
    env_unknown   = [t for t in active_tokens if t.apns_environment == "unknown"]
    env_mismatched = [t for t in active_tokens if t.apns_environment not in (target_env, "unknown")]
    preferred_active = env_matched or env_unknown or env_mismatched
    selected_for_env = preferred_active[0] if preferred_active else None

    def _tok_row(t):
        is_selected = selected_for_env and t.id == selected_for_env.id
        env_match = (
            t.apns_environment == target_env or
            t.apns_environment == "unknown"
        ) if t.active else False
        return {
            "id": t.id,
            "token_preview": _tok_preview(t.token),
            "platform": t.platform,
            "active": t.active,
            "apns_environment": t.apns_environment,
            "selected_for_current_environment": is_selected and bool(t.active),
            "environment_ok": env_match,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "updated_at": t.updated_at.isoformat() if t.updated_at else None,
            "age_days": (datetime.utcnow() - t.updated_at).days if t.updated_at else None,
        }

    return jsonify({
        "server_time": datetime.utcnow().isoformat() + "Z",
        "version_marker": "push-diag-v4-env-aware",
        "auth": {
            "requesting_user_id": current_user.id,
            "requesting_user_email": current_user.email,
        },
        "target_user": {
            "id": target_user_id,
            "email": target_user.email if target_user else None,
            "exists": target_user is not None,
        },
        "apns_env": {
            "APNS_USE_SANDBOX_raw": use_sandbox_raw,
            "use_sandbox": use_sandbox,
            "apns_host": apns_host,
            "target_token_environment": target_env,
            "bundle_id": bundle_id,
            "APNS_KEY_ID_set": bool(os.environ.get("APNS_KEY_ID")),
            "APNS_TEAM_ID_set": bool(os.environ.get("APNS_TEAM_ID")),
            "APNS_KEY_P8_set": bool(os.environ.get("APNS_KEY_P8")),
            "correct_for_testflight": not use_sandbox,
            "note": "TestFlight tokens are production tokens. APNS_USE_SANDBOX=false → api.push.apple.com. Mismatch → BadEnvironmentKeyInToken.",
        },
        "push_js": {
            "template": "templates/components/analytics_head.html",
            "included_in": "templates/base_app.html (all logged-in pages)",
            "dedup_strategy": "window.__pushSetupDone only — cleared on every full page navigation. NO sessionStorage.",
            "beacon_endpoint": "POST /api/push/beacon (@login_required)",
            "beacons_emitted": [
                "script_parsed (sync, before DOMContentLoaded)",
                "dcl_environment (unconditional: ua, has_webkit_handler, has_capacitor)",
                "cap_wait_result (after 2s wait loop)",
                "capacitor_ready / capacitor_missing_in_native / not_native",
                "plugin_check (plugin_found, plugin_via)",
                "listeners_attached",
                "permission_result",
                "register_called",
                "token_received",
                "post_success / post_failed / post_error",
                "registration_error",
            ],
            "note": "Search server logs for [PushBeacon] after opening TestFlight app while logged in.",
        },
        "capacitor": {
            "version": "8.3.1",
            "push_plugin": "@capacitor/push-notifications ^8.0.3",
            "server_url": "https://app.baselodgeapp.com",
            "plugin_access_path": "window.Capacitor.Plugins.PushNotifications (with registerPlugin() fallback)",
            "wait_loop": "2000ms polling every 100ms for window.Capacitor to appear",
            "apns_env_hint": "Capacitor.DEBUG=false → production, Capacitor.DEBUG=true → sandbox; stamped via server inference if DEBUG unavailable",
        },
        "tokens": {
            "count_total": len(all_tokens),
            "count_active": len(active_tokens),
            "count_env_matched": len(env_matched),
            "count_env_unknown": len(env_unknown),
            "count_env_mismatched": len(env_mismatched),
            "selected_token_preview": _tok_preview(selected_for_env.token) if selected_for_env else None,
            "selected_token_env": selected_for_env.apns_environment if selected_for_env else None,
            "all_rows": [_tok_row(t) for t in all_tokens],
        },
        "instructions": {
            "step_1": "Kill the TestFlight app completely (swipe up in app switcher)",
            "step_2": "Reopen and log in — wait 15 seconds",
            "step_3": "Check server logs for [PushBeacon] lines — they show exactly which gate the push script reached",
            "step_4": "If script_parsed appears but dcl_native_user_confirmed does not: DOMContentLoaded or isNativePlatform failed",
            "step_5": "If plugin_check shows plugin_found=False: Capacitor bridge not providing PushNotifications",
            "step_6": "If register_called appears but token_received does not: APNs token delivery failed natively",
            "step_7": "If token_received appears but post_success does not: server-side registration failed",
            "step_8": "Run /admin/test-push after a token is registered to attempt APNs delivery",
        },
    }), 200


# send_onesignal_push and send_onesignal_custom_event have been moved to
# services/push_providers.py (Phase A extraction). They are imported at the
# top of this file via:
#   from services.push_providers import send_onesignal_push, send_onesignal_custom_event
# All callers in app.py continue to work identically — behavior is unchanged.


# send_onesignal_custom_event has been moved to services/push_providers.py
# (Phase A extraction). Imported at the top of this file. Behavior unchanged.


# ── QA push override ─────────────────────────────────────────────────────────
# Temporary pre-launch QA helper: all admin test push routes redirect their
# notification to Richard's device so delivery can be validated before the
# App Store launch.  ONLY affects the four /admin/test-push* routes below.
# Production notification paths (friend requests, trip invites, scheduled
# pushes, automated messaging) are completely unaffected.
_QA_PUSH_OVERRIDE_EMAIL = "richardbattlebaxter@gmail.com"


def _get_qa_push_override_user():
    """Look up the QA override account. Returns None if not found."""
    try:
        return User.query.filter_by(email=_QA_PUSH_OVERRIDE_EMAIL).first()
    except Exception:
        return None


@app.route("/admin/test-push", methods=["GET", "POST"])
@login_required
@admin_required
def admin_test_push():
    """Send a test push notification to a specific target user's most recently updated active token.

    Routing:
      - Reads ?user_id=<id> from the query string; defaults to the current admin user.
      - If the target user's latest active token is platform='android' → FCM (single token).
      - If the target user's latest active token is platform='ios'  → APNs (all matching
        active iOS tokens for that user in the current APNs environment).

    Title: BaseLodge
    Body:  Test push from BaseLodge
    """
    def _tok_preview(t):
        return t[:8] + "\u2026" + t[-6:] if len(t) > 14 else t[:8] + "\u2026"

    # ── Resolve target user (defaults to current admin) ───────────────────────
    try:
        target_user_id = int(request.args.get("user_id", current_user.id))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid user_id parameter."}), 400

    target_user = db.session.get(User, target_user_id)
    if not target_user:
        return jsonify({"error": f"User {target_user_id} not found."}), 404

    # ── QA Override: route all test pushes to Richard's device ────────────────
    _qa_user = _get_qa_push_override_user()
    if _qa_user:
        current_app.logger.warning(
            "[TestPush] QA Override Active — routed test push to Richard "
            "(original target_user_id=%d → qa_user_id=%d email=%s)",
            target_user_id, _qa_user.id, _QA_PUSH_OVERRIDE_EMAIL,
        )
        target_user = _qa_user
        target_user_id = _qa_user.id

    # ── Determine platform from target user's most recently updated active token ──
    latest_token = (
        PushDeviceToken.query
        .filter_by(active=True, user_id=target_user_id)
        .order_by(PushDeviceToken.updated_at.desc())
        .first()
    )

    current_app.logger.warning(
        "[TestPush] target_user_id=%d latest_active_token: id=%s platform=%s",
        target_user_id,
        latest_token.id if latest_token else None,
        latest_token.platform if latest_token else "none",
    )

    # ── Early return when the target user has no active tokens at all ─────────
    if latest_token is None:
        total_token_count = PushDeviceToken.query.filter_by(user_id=target_user_id).count()
        current_app.logger.warning(
            "[TestPush] target_user_id=%d — no active tokens found (total tokens: %d)",
            target_user_id, total_token_count,
        )
        return jsonify({
            "success":       False,
            "final_success": False,
            "reason":        "no_active_tokens",
            "error":         f"No active push token found for user {target_user_id}.",
            "instruction":   "Ask the user to open the BaseLodge app so their device registers a token.",
            "token_counts":  {"total": total_token_count, "active": 0},
            "target_user_id": target_user_id,
            "results_by_token": [],
        }), 200

    # ── Android path ──────────────────────────────────────────────────────────
    if latest_token and latest_token.platform == "android":
        preview = _tok_preview(latest_token.token)
        current_app.logger.warning(
            "[TestPush] provider=fcm platform=android token_id=%d user_id=%d token=%s",
            latest_token.id, latest_token.user_id, preview,
        )
        result = send_fcm_push(
            latest_token.token,
            title="BaseLodge",
            body="Test push from BaseLodge",
        )
        success = result.get("success", False)

        # ── MessageEventLog: FCM test push outcome (android early-return) ──
        try:
            create_message_event(
                event_name=EventName.PUSH_TEST_SENT,
                category=Category.SYSTEM,
                actor_user_id=current_user.id,
                recipient_user_id=latest_token.user_id,
                channel=Channel.PUSH,
                provider=Provider.FCM,
                payload_json={
                    "token_id": latest_token.id,
                    "platform": "android",
                    "source_route": "admin_test_push",
                },
                message_title="BaseLodge",
                message_body="Test push from BaseLodge",
                delivery_status=DeliveryStatus.SENT if success else DeliveryStatus.FAILED,
                error_message=result.get("error") if not success else None,
            )
        except Exception as _mel_err:
            current_app.logger.warning("[MessageEvent] test_push (android) log failed: %s", _mel_err)

        http_status = 200 if success else 502
        return jsonify({
            "provider":              "fcm",
            "platform":              "android",
            "total_tokens_found":    1,
            "total_sent_successfully": 1 if success else 0,
            "total_failed":          0 if success else 1,
            "fcm_secret_set":        bool(os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON")),
            "results_by_token": [{
                "user_id":       latest_token.user_id,
                "token_id":      latest_token.id,
                "token_preview": preview,
                "platform":      "android",
                "success":       success,
                "message_id":    result.get("message_id"),
                "error":         result.get("error"),
            }],
        }), http_status

    # ── iOS path (existing APNs behavior) ─────────────────────────────────────
    use_sandbox_raw = os.environ.get("APNS_USE_SANDBOX", "true")
    use_sandbox = use_sandbox_raw.lower() not in ("false", "0", "no")
    bundle_id   = os.environ.get("APNS_BUNDLE_ID", "com.baselodge.app")
    target_env  = "sandbox" if use_sandbox else "production"

    apns_env_info = {
        "APNS_USE_SANDBOX_raw": use_sandbox_raw,
        "use_sandbox": use_sandbox,
        "target_token_environment": target_env,
        "bundle_id": bundle_id,
        "APNS_KEY_ID_set": bool(os.environ.get("APNS_KEY_ID")),
        "APNS_TEAM_ID_set": bool(os.environ.get("APNS_TEAM_ID")),
        "APNS_KEY_P8_set": bool(os.environ.get("APNS_KEY_P8")),
    }

    # Select active iOS tokens for the target user whose stored env matches the current APNs mode.
    candidate_rows = (
        PushDeviceToken.query
        .filter_by(active=True, platform="ios", apns_environment=target_env,
                   user_id=target_user_id)
        .order_by(PushDeviceToken.updated_at.desc())
        .all()
    )

    current_app.logger.warning(
        "[TestPush] target_user_id=%d provider=apns APNS_USE_SANDBOX=%s target_env=%s active_ios_%s_tokens=%d",
        target_user_id, use_sandbox_raw, target_env, target_env, len(candidate_rows),
    )

    if not candidate_rows:
        current_app.logger.warning(
            "[TestPush] target_user_id=%d no active iOS %s tokens found — nothing to send",
            target_user_id, target_env,
        )
        total_token_count = PushDeviceToken.query.filter_by(user_id=target_user_id).count()
        return jsonify({
            "success":       False,
            "final_success": False,
            "provider":      "apns",
            "platform":      "ios",
            "total_tokens_found":     0,
            "total_sent_successfully": 0,
            "total_failed":           0,
            "apns_env":      apns_env_info,
            "reason":        "no_active_tokens",
            "error":         f"No active iOS {target_env} token found for user {target_user_id}.",
            "instruction":   "Ask the user to open the BaseLodge app so their device registers a token.",
            "token_counts":  {"total": total_token_count, "active": 0},
            "target_user_id": target_user_id,
            "results_by_token": [],
        }), 200

    def _prefer_sandbox(row):
        if row.apns_environment == "sandbox":
            return True
        if row.apns_environment == "production":
            return False
        return None

    results_by_token = []
    total_ok  = 0
    total_bad = 0

    for row in candidate_rows:
        preview = _tok_preview(row.token)
        current_app.logger.warning(
            "[TestPush] sending → provider=apns token_id=%d user_id=%d env=%s token=%s",
            row.id, row.user_id, row.apns_environment, preview,
        )

        result = send_apns_push(
            row.token,
            title="BaseLodge",
            body="Test push from BaseLodge",
            prefer_sandbox=_prefer_sandbox(row),
        )

        final_success = result.get("final_success", result.get("success", False))

        if result.get("retry_attempted"):
            status_code = result.get("retry_status_code")
            error       = result.get("retry_error")
            apns_id     = result.get("retry_apns_id")
            env_used    = result.get("retry_environment")
        else:
            status_code = result.get("first_attempt_status_code")
            error       = result.get("first_attempt_error")
            apns_id     = result.get("first_attempt_apns_id")
            env_used    = result.get("first_attempt_environment")

        if final_success:
            total_ok += 1
            current_app.logger.warning(
                "[APNs TEST] token_id=%d user_id=%d status=success environment=%s response_status=%s",
                row.id, row.user_id, env_used, status_code,
            )
        else:
            total_bad += 1
            current_app.logger.warning(
                "[APNs TEST] token_id=%d user_id=%d status=failed environment=%s response_status=%s reason=%s",
                row.id, row.user_id, env_used, status_code, error or "unknown",
            )

        results_by_token.append({
            "user_id":          row.user_id,
            "token_id":         row.id,
            "token_preview":    preview,
            "platform":         "ios",
            "apns_environment": row.apns_environment,
            "success":          final_success,
            "status_code":      status_code,
            "error":            error,
            "apns_id":          apns_id,
            "env_corrected":    result.get("env_corrected", False),
        })

        # ── MessageEventLog: APNs test push outcome ──
        try:
            create_message_event(
                event_name=EventName.PUSH_TEST_SENT,
                category=Category.SYSTEM,
                actor_user_id=current_user.id,
                recipient_user_id=row.user_id,
                channel=Channel.PUSH,
                provider=Provider.APNS,
                payload_json={
                    "token_id": row.id,
                    "platform": "ios",
                    "source_route": "admin_test_push",
                },
                message_title="BaseLodge",
                message_body="Test push from BaseLodge",
                delivery_status=DeliveryStatus.SENT if final_success else DeliveryStatus.FAILED,
                error_message=error if not final_success else None,
            )
        except Exception as _mel_err:
            current_app.logger.warning("[MessageEvent] test_push (apns) log failed token_id=%d: %s", row.id, _mel_err)

    current_app.logger.warning(
        "[TestPush] done — provider=apns total=%d ok=%d failed=%d",
        len(candidate_rows), total_ok, total_bad,
    )

    overall_http = 200 if total_ok > 0 else 502
    return jsonify({
        "provider":               "apns",
        "platform":               "ios",
        "total_tokens_found":     len(candidate_rows),
        "total_sent_successfully": total_ok,
        "total_failed":           total_bad,
        "apns_env":               apns_env_info,
        "results_by_token":       results_by_token,
    }), overall_http


@app.route("/admin/test-push-all", methods=["GET"])
@login_required
@admin_required
def admin_test_push_all():
    """Send a test push to every active token belonging to the current admin user.

    Loops over all active PushDeviceToken rows for current_user, routing each
    through the correct provider:
      - platform='ios'     → APNs  (send_apns_push)
      - platform='android' → FCM   (send_fcm_push)
      - anything else      → skipped, counted as unsupported

    Title: BaseLodge
    Body:  Test push from BaseLodge

    Never logs full tokens or secrets.
    """
    def _tok_preview(t):
        return t[:8] + "\u2026" + t[-6:] if len(t) > 14 else t[:8] + "\u2026"

    # ── QA Override: send to Richard's tokens instead of current admin's ────────
    _qa_user_all = _get_qa_push_override_user()
    _push_all_user_id = _qa_user_all.id if _qa_user_all else current_user.id

    active_tokens = (
        PushDeviceToken.query
        .filter_by(user_id=_push_all_user_id, active=True)
        .order_by(PushDeviceToken.updated_at.desc())
        .all()
    )

    if _qa_user_all:
        current_app.logger.warning(
            "[TestPushAll] QA Override Active — routed test push to Richard "
            "(original user_id=%d → qa_user_id=%d email=%s active_tokens=%d)",
            current_user.id, _qa_user_all.id, _QA_PUSH_OVERRIDE_EMAIL, len(active_tokens),
        )
    else:
        current_app.logger.warning(
            "[TestPushAll] user_id=%d active_tokens=%d",
            current_user.id, len(active_tokens),
        )

    if not active_tokens:
        return jsonify({
            "route":                  "/admin/test-push-all",
            "user_id":                current_user.id,
            "total_active_tokens":    0,
            "ios_attempted":          0,
            "android_attempted":      0,
            "total_success":          0,
            "total_failed":           0,
            "unsupported_platforms":  0,
            "reason":                 "no_active_tokens",
            "results":                [],
        }), 200

    results        = []
    ios_count      = 0
    android_count  = 0
    success_count  = 0
    failed_count   = 0
    unsupported    = 0

    TEST_TITLE = "BaseLodge"
    TEST_BODY  = "Test push from BaseLodge"
    TEST_DATA  = {"source": "admin_test_push_all"}

    for row in active_tokens:
        preview = _tok_preview(row.token)

        if row.platform == "ios":
            ios_count += 1
            # Derive prefer_sandbox from stored apns_environment
            if row.apns_environment == "sandbox":
                prefer_sandbox = True
            elif row.apns_environment == "production":
                prefer_sandbox = False
            else:
                prefer_sandbox = None  # fall back to APNS_USE_SANDBOX env var

            current_app.logger.warning(
                "[TestPushAll] provider=apns platform=ios token_id=%d user_id=%d token=%s",
                row.id, row.user_id, preview,
            )
            result = send_apns_push(
                row.token,
                title=TEST_TITLE,
                body=TEST_BODY,
                prefer_sandbox=prefer_sandbox,
            )
            final_success = result.get("final_success", result.get("success", False))
            if result.get("retry_attempted"):
                error = result.get("retry_error")
            else:
                error = result.get("first_attempt_error")

            if final_success:
                success_count += 1
            else:
                failed_count += 1

            results.append({
                "token_id":         row.id,
                "platform":         "ios",
                "provider":         "apns",
                "token_preview":    preview,
                "apns_environment": row.apns_environment,
                "success":          final_success,
                "error":            error,
            })

            # ── MessageEventLog: APNs push-all outcome ──
            try:
                create_message_event(
                    event_name=EventName.PUSH_TEST_SENT,
                    category=Category.SYSTEM,
                    actor_user_id=current_user.id,
                    recipient_user_id=row.user_id,
                    channel=Channel.PUSH,
                    provider=Provider.APNS,
                    payload_json={
                        "token_id": row.id,
                        "platform": "ios",
                        "source_route": "admin_test_push_all",
                    },
                    message_title=TEST_TITLE,
                    message_body=TEST_BODY,
                    delivery_status=DeliveryStatus.SENT if final_success else DeliveryStatus.FAILED,
                    error_message=error if not final_success else None,
                )
            except Exception as _mel_err:
                current_app.logger.warning("[MessageEvent] test_push_all (ios) log failed token_id=%d: %s", row.id, _mel_err)

        elif row.platform == "android":
            android_count += 1
            current_app.logger.warning(
                "[TestPushAll] provider=fcm platform=android token_id=%d user_id=%d token=%s",
                row.id, row.user_id, preview,
            )
            result = send_fcm_push(
                row.token,
                title=TEST_TITLE,
                body=TEST_BODY,
                data=TEST_DATA,
            )
            success = result.get("success", False)
            if success:
                success_count += 1
            else:
                failed_count += 1

            results.append({
                "token_id":      row.id,
                "platform":      "android",
                "provider":      "fcm",
                "token_preview": preview,
                "success":       success,
                "message_id":    result.get("message_id"),
                "error":         result.get("error"),
            })

            # ── MessageEventLog: FCM push-all outcome ──
            try:
                create_message_event(
                    event_name=EventName.PUSH_TEST_SENT,
                    category=Category.SYSTEM,
                    actor_user_id=current_user.id,
                    recipient_user_id=row.user_id,
                    channel=Channel.PUSH,
                    provider=Provider.FCM,
                    payload_json={
                        "token_id": row.id,
                        "platform": "android",
                        "source_route": "admin_test_push_all",
                    },
                    message_title=TEST_TITLE,
                    message_body=TEST_BODY,
                    delivery_status=DeliveryStatus.SENT if success else DeliveryStatus.FAILED,
                    error_message=result.get("error") if not success else None,
                )
            except Exception as _mel_err:
                current_app.logger.warning("[MessageEvent] test_push_all (android) log failed token_id=%d: %s", row.id, _mel_err)

        else:
            unsupported += 1
            current_app.logger.warning(
                "[TestPushAll] unsupported platform=%s token_id=%d user_id=%d — skipped",
                row.platform, row.id, row.user_id,
            )
            results.append({
                "token_id":      row.id,
                "platform":      row.platform,
                "provider":      "none",
                "token_preview": preview,
                "success":       False,
                "error":         "unsupported_platform",
            })

    current_app.logger.warning(
        "[TestPushAll] done user_id=%d total=%d ios=%d android=%d success=%d failed=%d unsupported=%d",
        current_user.id, len(active_tokens),
        ios_count, android_count,
        success_count, failed_count, unsupported,
    )

    overall_http = 200 if (success_count > 0 or (ios_count + android_count == 0)) else 502
    return jsonify({
        "route":                 "/admin/test-push-all",
        "user_id":               current_user.id,
        "total_active_tokens":   len(active_tokens),
        "ios_attempted":         ios_count,
        "android_attempted":     android_count,
        "total_success":         success_count,
        "total_failed":          failed_count,
        "unsupported_platforms": unsupported,
        "results":               results,
    }), overall_http


@app.route("/admin/posthog-test", methods=["GET"])
@login_required
@admin_required
def admin_posthog_test():
    """Diagnostic: fire one posthog_server_test event and return capture/flush results."""
    import time as _time
    results = {}

    key = ph_analytics.POSTHOG_KEY
    host = ph_analytics.POSTHOG_HOST
    results["key_set"] = bool(key)
    results["key_prefix"] = (key[:8] + "…") if key else None
    results["host"] = host

    if not key:
        results["outcome"] = "FAIL — POSTHOG_KEY not set"
        return jsonify(results), 200

    try:
        from posthog import Posthog
        _test_client = Posthog(project_api_key=key, host=host)
        results["client_init"] = "OK"
    except Exception as exc:
        results["client_init"] = "FAIL"
        results["client_init_error"] = str(exc)
        results["outcome"] = "FAIL — client init error"
        return jsonify(results), 200

    distinct_id = str(current_user.id)
    event = "posthog_server_test"
    props = {
        "source": "admin_posthog_test",
        "user_id": current_user.id,
        "timestamp": _time.time(),
    }

    try:
        _test_client.capture(event, distinct_id=distinct_id, properties=props)
        results["capture"] = "OK"
    except Exception as exc:
        results["capture"] = "FAIL"
        results["capture_error"] = str(exc)
        results["outcome"] = "FAIL — capture error"
        return jsonify(results), 200

    try:
        _test_client.flush()
        results["flush"] = "OK"
        results["outcome"] = "SUCCESS"
    except Exception as exc:
        results["flush"] = "FAIL"
        results["flush_error"] = str(exc)
        results["outcome"] = "FAIL — flush error"

    app.logger.info("[POSTHOG_DIAG] %s", results)
    return jsonify(results), 200


@app.route("/admin/backfill-posthog", methods=["GET"])
@login_required
@admin_required
def admin_backfill_posthog():
    """Backfill PostHog person properties for all users.

    Uses client.set() only — never capture(). No events, no timeline impact.
    Idempotent: safe to run multiple times.

    Query params:
      ?dry_run=1   Build property dicts and return them without sending anything.
    """
    dry_run = request.args.get("dry_run", "0") == "1"

    # ── 1. Bulk-fetch supporting data (no N+1 inside the user loop) ──────────

    # Sets of user_ids for boolean flags
    trip_owner_ids = {
        r[0] for r in db.session.execute(
            db.text("SELECT DISTINCT user_id FROM ski_trip")
        )
    }
    trip_guest_ids = {
        r[0] for r in db.session.execute(
            db.text("SELECT DISTINCT user_id FROM ski_trip_participant WHERE status = 'accepted'")
        )
    }
    friend_ids = {
        r[0] for r in db.session.execute(
            db.text("SELECT DISTINCT user_id FROM friend")
        )
    }
    generic_invite_ids = {
        r[0] for r in db.session.execute(
            db.text("SELECT DISTINCT inviter_id FROM invite_token")
        )
    }
    trip_invite_ids = {
        r[0] for r in db.session.execute(
            db.text("SELECT DISTINCT inviter_user_id FROM trip_invite_token")
        )
    }

    # Integer counts per user
    friend_counts = {
        r[0]: r[1] for r in db.session.execute(
            db.text("SELECT user_id, COUNT(*) FROM friend GROUP BY user_id")
        )
    }
    trip_counts = {
        r[0]: r[1] for r in db.session.execute(
            db.text("SELECT user_id, COUNT(*) FROM ski_trip GROUP BY user_id")
        )
    }

    all_users = User.query.order_by(User.id).all()

    backfill_date = datetime.utcnow().strftime("%Y-%m-%d")

    # ── 2. Build one property dict per user ──────────────────────────────────

    user_props = []
    for u in all_users:
        uid = u.id
        wl = u.wish_list_resorts or []
        rt = u.rider_types
        if isinstance(rt, list):
            rider_label = ",".join(rt) if rt else "unknown"
        else:
            rider_label = str(rt) if rt else "unknown"

        props = {
            # Activation flags
            "has_completed_signup":     True,
            "has_completed_onboarding": u.lifecycle_stage == "active",
            "has_pass":                 _ph_is_real_pass(u.pass_type),
            "has_availability":         bool(u.open_dates),
            "has_wishlist":             bool(wl),
            "has_trip":                 (uid in trip_owner_ids or uid in trip_guest_ids),
            "has_friend_connection":    uid in friend_ids,
            "has_generated_invite":     (uid in generic_invite_ids or uid in trip_invite_ids),
            # Segmentation helpers (non-PII)
            "lifecycle_stage":          u.lifecycle_stage or "new",
            "pass_type":                (u.pass_type or "none").lower(),
            "rider_type":               rider_label,
            "friend_count":             friend_counts.get(uid, 0),
            "trip_count":               trip_counts.get(uid, 0),
            "wishlist_count":           len(wl),
            "is_internal":              ph_analytics.is_internal(u.email or ""),
            # Backfill sentinel
            "activation_backfilled":    True,
            "activation_backfilled_at": backfill_date,
        }
        user_props.append((uid, props))

    summary = {
        "total_users":  len(user_props),
        "dry_run":      dry_run,
        "sent":         False,
        "flush_ok":     None,
        "flush_error":  None,
        "errors":       [],
        "sample":       [{"user_id": uid, "props": p} for uid, p in user_props[:3]],
    }

    if dry_run:
        app.logger.info("[POSTHOG_BACKFILL] dry_run — %d users, no data sent", len(user_props))
        return jsonify(summary), 200

    # ── 3. Send: one client.set() per user, one flush at the end ─────────────

    client = ph_analytics._get_client()
    if not client:
        summary["errors"].append("PostHog client unavailable — POSTHOG_KEY not set")
        return jsonify(summary), 200

    set_errors = []
    for uid, props in user_props:
        try:
            client.set(distinct_id=str(uid), properties=props)
        except Exception as exc:
            set_errors.append({"user_id": uid, "error": str(exc)})
            app.logger.warning("[POSTHOG_BACKFILL] set failed user_id=%s error=%s", uid, exc)

    try:
        client.flush()
        summary["flush_ok"] = True
        summary["sent"] = True
        app.logger.info(
            "[POSTHOG_BACKFILL] complete — %d users sent, %d errors, flush OK",
            len(user_props), len(set_errors),
        )
    except Exception as exc:
        summary["flush_ok"] = False
        summary["flush_error"] = str(exc)
        app.logger.warning("[POSTHOG_BACKFILL] flush FAILED: %s", exc)

    summary["errors"] = set_errors
    return jsonify(summary), 200


@app.route("/admin/test-push-broadcast", methods=["GET"])
@login_required
@admin_required
def admin_test_push_broadcast():
    """Send a test push to every active token in the database across all users.

    Optional query params:
      ?title=...   override notification title  (default: "BaseLodge")
      ?body=...    override notification body   (default: "Test push from BaseLodge")

    Routes each token through the correct provider:
      platform='ios'     → APNs  (send_apns_push)
      platform='android' → FCM   (send_fcm_push)
      anything else      → skipped, counted as unsupported

    Never logs full tokens or secrets.
    """
    def _tok_preview(t):
        return t[:8] + "\u2026" + t[-6:] if len(t) > 14 else t[:8] + "\u2026"

    title = (request.args.get("title") or "BaseLodge").strip()
    body  = (request.args.get("body")  or "Test push from BaseLodge").strip()

    active_tokens = (
        PushDeviceToken.query
        .filter_by(active=True)
        .order_by(PushDeviceToken.updated_at.desc())
        .all()
    )

    # ── QA Override: narrow broadcast to Richard's tokens only ────────────────
    _qa_user_bc = _get_qa_push_override_user()
    if _qa_user_bc:
        _bc_before = len(active_tokens)
        active_tokens = [t for t in active_tokens if t.user_id == _qa_user_bc.id]
        current_app.logger.warning(
            "[TestPushBroadcast] QA Override Active — routed test push to Richard "
            "(qa_user_id=%d email=%s tokens_before=%d tokens_after=%d)",
            _qa_user_bc.id, _QA_PUSH_OVERRIDE_EMAIL, _bc_before, len(active_tokens),
        )

    unique_users = len({row.user_id for row in active_tokens})

    current_app.logger.warning(
        "[TestPushBroadcast] admin_user_id=%d total_active_tokens=%d unique_users=%d "
        "title=%r body=%r",
        current_user.id, len(active_tokens), unique_users, title[:60], body[:120],
    )

    if not active_tokens:
        return jsonify({
            "route":                 "/admin/test-push-broadcast",
            "admin_user_id":         current_user.id,
            "title_used":            title,
            "body_used":             body,
            "total_active_tokens":   0,
            "unique_users_targeted": 0,
            "ios_attempted":         0,
            "android_attempted":     0,
            "total_success":         0,
            "total_failed":          0,
            "unsupported_platforms": 0,
            "reason":                "no_active_tokens",
            "results":               [],
        }), 200

    results       = []
    ios_count     = 0
    android_count = 0
    success_count = 0
    failed_count  = 0
    unsupported   = 0

    TEST_DATA = {"source": "admin_test_push_broadcast"}

    for row in active_tokens:
        preview = _tok_preview(row.token)

        if row.platform == "ios":
            ios_count += 1
            if row.apns_environment == "sandbox":
                prefer_sandbox = True
            elif row.apns_environment == "production":
                prefer_sandbox = False
            else:
                prefer_sandbox = None

            current_app.logger.warning(
                "[TestPushBroadcast] provider=apns platform=ios "
                "token_id=%d user_id=%d token=%s",
                row.id, row.user_id, preview,
            )
            result       = send_apns_push(row.token, title=title, body=body,
                                          prefer_sandbox=prefer_sandbox)
            final_success = result.get("final_success", result.get("success", False))
            error         = (result.get("retry_error") if result.get("retry_attempted")
                             else result.get("first_attempt_error"))
            if final_success:
                success_count += 1
            else:
                failed_count += 1
            results.append({
                "token_id":         row.id,
                "user_id":          row.user_id,
                "platform":         "ios",
                "provider":         "apns",
                "token_preview":    preview,
                "apns_environment": row.apns_environment,
                "success":          final_success,
                "error":            error,
            })

            # ── MessageEventLog: APNs broadcast outcome ──
            try:
                create_message_event(
                    event_name=EventName.PUSH_BROADCAST_SENT,
                    category=Category.SYSTEM,
                    actor_user_id=current_user.id,
                    recipient_user_id=row.user_id,
                    channel=Channel.PUSH,
                    provider=Provider.APNS,
                    payload_json={
                        "token_id": row.id,
                        "platform": "ios",
                        "source_route": "admin_test_push_broadcast",
                    },
                    message_title=title,
                    message_body=body,
                    delivery_status=DeliveryStatus.SENT if final_success else DeliveryStatus.FAILED,
                    error_message=error if not final_success else None,
                )
            except Exception as _mel_err:
                current_app.logger.warning("[MessageEvent] push_broadcast (ios) log failed token_id=%d: %s", row.id, _mel_err)

        elif row.platform == "android":
            android_count += 1
            current_app.logger.warning(
                "[TestPushBroadcast] provider=fcm platform=android "
                "token_id=%d user_id=%d token=%s",
                row.id, row.user_id, preview,
            )
            result  = send_fcm_push(row.token, title=title, body=body, data=TEST_DATA)
            success = result.get("success", False)
            if success:
                success_count += 1
            else:
                failed_count += 1
            results.append({
                "token_id":      row.id,
                "user_id":       row.user_id,
                "platform":      "android",
                "provider":      "fcm",
                "token_preview": preview,
                "success":       success,
                "message_id":    result.get("message_id"),
                "error":         result.get("error"),
            })

            # ── MessageEventLog: FCM broadcast outcome ──
            try:
                create_message_event(
                    event_name=EventName.PUSH_BROADCAST_SENT,
                    category=Category.SYSTEM,
                    actor_user_id=current_user.id,
                    recipient_user_id=row.user_id,
                    channel=Channel.PUSH,
                    provider=Provider.FCM,
                    payload_json={
                        "token_id": row.id,
                        "platform": "android",
                        "source_route": "admin_test_push_broadcast",
                    },
                    message_title=title,
                    message_body=body,
                    delivery_status=DeliveryStatus.SENT if success else DeliveryStatus.FAILED,
                    error_message=result.get("error") if not success else None,
                )
            except Exception as _mel_err:
                current_app.logger.warning("[MessageEvent] push_broadcast (android) log failed token_id=%d: %s", row.id, _mel_err)

        else:
            unsupported += 1
            current_app.logger.warning(
                "[TestPushBroadcast] unsupported platform=%s token_id=%d user_id=%d — skipped",
                row.platform, row.id, row.user_id,
            )
            results.append({
                "token_id":      row.id,
                "user_id":       row.user_id,
                "platform":      row.platform,
                "provider":      "none",
                "token_preview": preview,
                "success":       False,
                "error":         "unsupported_platform",
            })

    current_app.logger.warning(
        "[TestPushBroadcast] done admin_user_id=%d total=%d ios=%d android=%d "
        "success=%d failed=%d unsupported=%d",
        current_user.id, len(active_tokens),
        ios_count, android_count,
        success_count, failed_count, unsupported,
    )

    overall_http = 200 if (success_count > 0 or (ios_count + android_count == 0)) else 502
    return jsonify({
        "route":                 "/admin/test-push-broadcast",
        "admin_user_id":         current_user.id,
        "title_used":            title,
        "body_used":             body,
        "total_active_tokens":   len(active_tokens),
        "unique_users_targeted": unique_users,
        "ios_attempted":         ios_count,
        "android_attempted":     android_count,
        "total_success":         success_count,
        "total_failed":          failed_count,
        "unsupported_platforms": unsupported,
        "results":               results,
    }), overall_http


@app.route("/admin/list-tokens", methods=["GET"])
@login_required
@admin_required
def admin_list_tokens():
    """Admin read-only diagnostic: list all push device tokens for a user.

    Never sends APNs notifications. Safe to call at any time.

    GET /admin/list-tokens
    GET /admin/list-tokens?user_id=6
    """
    try:
        target_user_id = int(request.args.get("user_id", 2))
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "user_id must be an integer"}), 400

    def _tok_preview(t):
        return t[:8] + "…" + t[-6:] if len(t) > 14 else t[:8] + "…"

    rows = (
        PushDeviceToken.query
        .filter_by(user_id=target_user_id)
        .order_by(PushDeviceToken.updated_at.desc())
        .all()
    )

    active_count   = sum(1 for r in rows if r.active)
    inactive_count = len(rows) - active_count

    return jsonify({
        "success": True,
        "target_user_id": target_user_id,
        "token_counts": {
            "total": len(rows),
            "active": active_count,
            "inactive": inactive_count,
        },
        "tokens": [
            {
                "id": r.id,
                "user_id": r.user_id,
                "platform": r.platform,
                "active": r.active,
                "apns_environment": r.apns_environment,
                "token_preview": _tok_preview(r.token),
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
            for r in rows
        ],
    }), 200


@app.route("/admin/push-token-dedup", methods=["GET"])
@login_required
@admin_required
def admin_push_token_dedup():
    """One-time admin cleanup: deactivate all but the most-recently-updated
    active PushDeviceToken per user/platform pair.

    Safe to run repeatedly — idempotent after first clean run.
    Never deletes rows. Never calls OneSignal or any push provider.

    GET /admin/push-token-dedup
    """
    # Gather all active tokens, grouped by (user_id, platform)
    all_active = (
        PushDeviceToken.query
        .filter_by(active=True)
        .order_by(PushDeviceToken.user_id, PushDeviceToken.platform,
                  PushDeviceToken.updated_at.desc())
        .all()
    )

    # Group by (user_id, platform) — first row in each group is the keeper
    from collections import defaultdict
    groups = defaultdict(list)
    for row in all_active:
        groups[(row.user_id, row.platform)].append(row)

    users_affected    = 0
    tokens_deactivated = 0
    details            = []

    try:
        for (uid, plat), rows in groups.items():
            if len(rows) <= 1:
                continue
            keeper     = rows[0]   # most recently updated active token
            to_deactivate = rows[1:]
            deactivated_ids = []
            for row in to_deactivate:
                row.active = False
                deactivated_ids.append(row.id)
                current_app.logger.warning(
                    "[PushTokenDedup] Deactivated stale token id=%s user=%s platform=%s",
                    row.id, uid, plat,
                )
            users_affected    += 1
            tokens_deactivated += len(deactivated_ids)
            details.append({
                "user_id":             uid,
                "platform":            plat,
                "kept_token_id":       keeper.id,
                "deactivated_token_ids": deactivated_ids,
            })

        db.session.commit()
        current_app.logger.warning(
            "[PushTokenDedup] complete — users_affected=%d tokens_deactivated=%d",
            users_affected, tokens_deactivated,
        )
    except Exception:
        db.session.rollback()
        current_app.logger.exception("[PushTokenDedup] failed — rolled back")
        return jsonify({"success": False, "error": "Server error during dedup"}), 500

    return jsonify({
        "success":           True,
        "users_affected":    users_affected,
        "tokens_deactivated": tokens_deactivated,
        "details":           details,
    }), 200


@app.route("/admin/push-token-audit", methods=["GET"])
@login_required
@admin_required
def admin_push_token_audit():
    """Audit active PushDeviceToken counts per user/platform.

    Flags any user/platform pair with more than 1 active token.
    Read-only — no writes, no push sends.

    GET /admin/push-token-audit
    GET /admin/push-token-audit?user_id=2   (filter to one user)
    """
    try:
        target_user_id = request.args.get("user_id")
        if target_user_id is not None:
            target_user_id = int(target_user_id)
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "user_id must be an integer"}), 400

    query = PushDeviceToken.query.filter_by(active=True)
    if target_user_id is not None:
        query = query.filter_by(user_id=target_user_id)
    active_rows = query.order_by(
        PushDeviceToken.user_id, PushDeviceToken.platform,
        PushDeviceToken.updated_at.desc()
    ).all()

    from collections import defaultdict
    groups = defaultdict(list)
    for row in active_rows:
        groups[(row.user_id, row.platform)].append(row)

    audit_rows  = []
    total_clean = 0
    total_dirty = 0

    def _tok_preview(t):
        return t[:8] + "\u2026" + t[-6:] if len(t) > 14 else t[:8] + "\u2026"

    for (uid, plat), rows in sorted(groups.items()):
        count  = len(rows)
        status = "OK" if count == 1 else "DUPLICATE_ACTIVE_TOKENS"
        if count == 1:
            total_clean += 1
        else:
            total_dirty += 1
        audit_rows.append({
            "user_id":      uid,
            "platform":     plat,
            "active_count": count,
            "status":       status,
            "tokens": [
                {
                    "id":               r.id,
                    "token_preview":    _tok_preview(r.token),
                    "apns_environment": r.apns_environment,
                    "updated_at":       r.updated_at.isoformat() if r.updated_at else None,
                    "created_at":       r.created_at.isoformat() if r.created_at else None,
                }
                for r in rows
            ],
        })

    # Sort: dirty (DUPLICATE) first, then by user_id
    audit_rows.sort(key=lambda x: (0 if x["status"] == "DUPLICATE_ACTIVE_TOKENS" else 1, x["user_id"]))

    return jsonify({
        "success":            True,
        "filter_user_id":     target_user_id,
        "summary": {
            "total_user_platform_pairs": len(audit_rows),
            "clean":     total_clean,
            "duplicate": total_dirty,
        },
        "audit": audit_rows,
    }), 200


def pass_category(pass_type):
    """Categorize pass type into Epic, Ikon, or Other."""
    norm = normalize_pass(pass_type or "")
    if norm == "epic":
        return "Epic"
    if norm == "ikon":
        return "Ikon"
    return "Other"

@app.route("/friends")
@login_required
def friends():
    user = current_user
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    from sqlalchemy.orm import joinedload
    _rp_t0 = time.perf_counter()

    _fp_t0 = time.perf_counter()

    # ── [FRIENDS_PERF] Block 1+2: friend_links + all_friends (single JOIN) ────
    # Replaces two serial queries (Friend→IDs, then User.in_(IDs)) with one JOIN.
    # Saves ~65ms (one Supabase round-trip). Matches the pattern used by /home and
    # /my-trips. Friend objects are still available for friendship_lookup below.
    _t = time.perf_counter()
    _friend_join_rows = (
        db.session.query(User, Friend)
        .join(Friend, Friend.friend_id == User.id)
        .filter(Friend.user_id == user.id)
        .all()
    )
    all_friends = [u for u, _f in _friend_join_rows]
    friend_links = [_f for _u, _f in _friend_join_rows]
    friend_ids = [u.id for u in all_friends]
    if app.debug:
        print(f"[FRIENDS_PERF] friend_links+all_friends={time.perf_counter()-_t:.4f}s count={len(friend_ids)}")

    # ── [FRIENDS_PERF] Block 3: friend_trips ──────────────────────────────────
    _t = time.perf_counter()
    friend_trips = []
    if friend_ids:
        friend_trips = (
            SkiTrip.query
            .options(joinedload(SkiTrip.user), joinedload(SkiTrip.resort))
            .filter(
                SkiTrip.user_id.in_(friend_ids),
                SkiTrip.end_date >= today,
                SkiTrip.is_public == True
            )
            .order_by(SkiTrip.start_date.asc())
            .all()
        )
    if app.debug:
        print(f"[FRIENDS_PERF] friend_trips={time.perf_counter()-_t:.4f}s count={len(friend_trips)}")

    # ── [FRIENDS_PERF] Block 4: user_trips (owned) ────────────────────────────
    _t = time.perf_counter()
    user_trips = (
        SkiTrip.query
        .options(joinedload(SkiTrip.resort))
        .filter(
            SkiTrip.user_id == user.id,
            SkiTrip.end_date >= today
        )
        .all()
    )
    if app.debug:
        print(f"[FRIENDS_PERF] user_trips_owned={time.perf_counter()-_t:.4f}s count={len(user_trips)}")

    # ── [FRIENDS_PERF] Block 5: accepted guest trips ───────────────────────────
    _t = time.perf_counter()
    try:
        _user_accepted_ids = [
            p.trip_id for p in SkiTripParticipant.query.filter(
                SkiTripParticipant.user_id == user.id,
                SkiTripParticipant.status == GuestStatus.ACCEPTED
            ).all()
        ]
        if _user_accepted_ids:
            _user_guest_trips = SkiTrip.query.filter(
                SkiTrip.id.in_(_user_accepted_ids),
                SkiTrip.user_id != user.id,
                SkiTrip.end_date >= today
            ).all()
            user_trips = user_trips + _user_guest_trips
    except Exception:
        pass
    if app.debug:
        print(f"[FRIENDS_PERF] accepted_guest_trips={time.perf_counter()-_t:.4f}s user_trips_total={len(user_trips)}")

    # ── [FRIENDS_PERF] Block 6: compute_trip_overlaps ─────────────────────────
    _t = time.perf_counter()
    trip_overlaps = compute_trip_overlaps(user_trips, friend_trips)
    if app.debug:
        print(f"[FRIENDS_PERF] compute_trip_overlaps={time.perf_counter()-_t:.4f}s overlaps={len(trip_overlaps)}")

    # ── [FRIENDS_PERF] Block 7: open_date_overlaps ────────────────────────────
    _t = time.perf_counter()
    open_date_overlaps = []
    user_open_dates = set(d for d in (user.open_dates or []) if d >= today_str)
    if user_open_dates and friend_ids:
        for friend in all_friends:
            friend_open_dates = set(d for d in (friend.open_dates or []) if d >= today_str)
            shared_dates = user_open_dates & friend_open_dates
            for date_str in shared_dates:
                open_date_overlaps.append({
                    "type": "open",
                    "date_str": date_str,
                    "friend_id": friend.id,
                    "friend_first_name": friend.first_name
                })
    if app.debug:
        print(f"[FRIENDS_PERF] open_date_overlaps={time.perf_counter()-_t:.4f}s overlaps={len(open_date_overlaps)}")

    # ── [FRIENDS_PERF] Block 8: lookup build ──────────────────────────────────
    _t = time.perf_counter()
    friendship_lookup = {f.friend_id: f for f in friend_links}

    trip_overlap_by_friend = {}
    for ov in trip_overlaps:
        fid = ov['friend_id']
        if fid not in trip_overlap_by_friend:
            trip_overlap_by_friend[fid] = []
        trip_overlap_by_friend[fid].append(ov)

    open_overlap_by_friend = {}
    for ov in open_date_overlaps:
        fid = ov['friend_id']
        if fid not in open_overlap_by_friend:
            open_overlap_by_friend[fid] = []
        open_overlap_by_friend[fid].append(ov['date_str'])

    friend_trips_by_id = {}
    for ft in friend_trips:
        fid = ft.user_id
        if fid not in friend_trips_by_id:
            friend_trips_by_id[fid] = []
        friend_trips_by_id[fid].append(ft)
    if app.debug:
        print(f"[FRIENDS_PERF] lookup_build={time.perf_counter()-_t:.4f}s")

    # ── [FRIENDS_PERF] Block 8b: batch owned upcoming trips for all friends ────
    _t = time.perf_counter()
    _batch_owner_trips: list = []
    if friend_ids:
        from sqlalchemy.orm import joinedload
        _batch_owner_trips = (
            SkiTrip.query
            .options(joinedload(SkiTrip.resort))
            .filter(
                SkiTrip.user_id.in_(friend_ids),
                SkiTrip.end_date >= today,
            )
            .all()
        )
    _owner_trips_by_friend: dict = {}
    for _t2 in _batch_owner_trips:
        _owner_trips_by_friend.setdefault(_t2.user_id, []).append(_t2)
    if app.debug:
        print(f"[FRIENDS_PERF] batch_owner_trips={time.perf_counter()-_t:.4f}s rows={len(_batch_owner_trips)}")

    # ── [FRIENDS_PERF] Block 8c: batch accepted participant trips for all friends
    _t = time.perf_counter()
    _batch_part_trips: list = []
    if friend_ids:
        _batch_part_trips = (
            db.session.query(SkiTrip)
            .options(joinedload(SkiTrip.resort))
            .join(SkiTripParticipant, SkiTrip.id == SkiTripParticipant.trip_id)
            .filter(
                SkiTripParticipant.user_id.in_(friend_ids),
                SkiTripParticipant.status == GuestStatus.ACCEPTED,
                SkiTrip.end_date >= today,
            )
            .all()
        )
    # Group by the participant's user_id (stored on the join row), not trip owner
    _part_trips_by_friend: dict = {}
    if friend_ids and _batch_part_trips:
        # Re-query the participant rows to map trip_id → participant user_id
        _part_trip_ids = [t.id for t in _batch_part_trips]
        _part_rows = SkiTripParticipant.query.filter(
            SkiTripParticipant.trip_id.in_(_part_trip_ids),
            SkiTripParticipant.user_id.in_(friend_ids),
            SkiTripParticipant.status == GuestStatus.ACCEPTED,
        ).all()
        _trip_id_to_obj = {t.id: t for t in _batch_part_trips}
        for _pr in _part_rows:
            _trip_obj = _trip_id_to_obj.get(_pr.trip_id)
            if _trip_obj:
                _part_trips_by_friend.setdefault(_pr.user_id, []).append(_trip_obj)
    if app.debug:
        print(f"[FRIENDS_PERF] batch_participant_trips={time.perf_counter()-_t:.4f}s rows={len(_batch_part_trips)}")

    # ── [FRIENDS_PERF] Block 9: per-friend loop ───────────────────────────────
    _loop_t0 = time.perf_counter()
    _acc_trip_count   = 0.0
    _acc_owner_trips  = 0.0
    _acc_part_trips   = 0.0
    _acc_overlap_prep = 0.0

    for friend in all_friends:
        friendship = friendship_lookup.get(friend.id)

        friend._trip_invites_allowed = friendship.trip_invites_allowed if friendship else False
        friend._is_new_friend = bool(friendship and not friendship.has_viewed_profile)

        # [inner] trip count — inline set-union from batched data (no DB call)
        _ti = time.perf_counter()
        _owner_ids = {t.id for t in _owner_trips_by_friend.get(friend.id, [])}
        _part_ids  = {t.id for t in _part_trips_by_friend.get(friend.id, [])}
        friend._upcoming_trip_count = len(_owner_ids | _part_ids)
        friend._has_upcoming_trip = friend._upcoming_trip_count > 0
        _acc_trip_count += time.perf_counter() - _ti

        # [inner] owned upcoming trips — batch lookup (no DB call)
        _ti = time.perf_counter()
        upcoming_owner_trips = _owner_trips_by_friend.get(friend.id, [])
        _acc_owner_trips += time.perf_counter() - _ti

        # [inner] accepted participant trips — batch lookup (no DB call)
        _ti = time.perf_counter()
        upcoming_participant_trips = _part_trips_by_friend.get(friend.id, [])
        _acc_part_trips += time.perf_counter() - _ti

        all_upcoming_trips_dict = {t.id: t for t in upcoming_owner_trips}
        for t in upcoming_participant_trips:
            all_upcoming_trips_dict[t.id] = t

        if all_upcoming_trips_dict:
            latest_created = max(t.created_at for t in all_upcoming_trips_dict.values() if t.created_at)
            friend._latest_upcoming_trip_created_at = latest_created
        else:
            friend._latest_upcoming_trip_created_at = None

        friend._trip_count = friend._upcoming_trip_count
        friend._going_count = sum(
            1 for t in friend_trips_by_id.get(friend.id, [])
            if (t.trip_status or 'planning') == 'going'
        )

        # [inner] overlap label + next trip label
        _ti = time.perf_counter()
        friend._overlap_label = None
        friend._next_trip_label = None

        friend_ov_trips = sorted(trip_overlap_by_friend.get(friend.id, []), key=lambda x: x['start_date'])
        if friend_ov_trips:
            ov = friend_ov_trips[0]
            sd, ed = ov['start_date'], ov['end_date']
            if sd == ed:
                dates_str = sd.strftime('%b %-d')
            elif sd.month == ed.month:
                dates_str = f"{sd.strftime('%b %-d')}–{ed.strftime('%-d')}"
            else:
                dates_str = f"{sd.strftime('%b %-d')}–{ed.strftime('%b %-d')}"
            friend._overlap_label = f"Overlap at {ov['mountain']} · {dates_str}"
        else:
            friend_open_ovs = sorted(open_overlap_by_friend.get(friend.id, []))
            if friend_open_ovs:
                d = datetime.strptime(friend_open_ovs[0], '%Y-%m-%d').date()
                friend._overlap_label = f"Both free {d.strftime('%b %-d')}"

        friend_upcoming_pub = sorted(friend_trips_by_id.get(friend.id, []), key=lambda t: t.start_date)
        if friend_upcoming_pub:
            ft = friend_upcoming_pub[0]
            resort_name = ft.resort.name if ft.resort else (ft.mountain or '')
            state = ft.resort.state if ft.resort else (ft.state or '')
            dates_str = format_trip_dates(ft)
            ft_status = ft.trip_status or 'planning'
            dest = resort_name
            if state:
                dest += f", {state}"
            if dates_str:
                dest += f" · {dates_str}"
            if ft_status == 'going':
                label = f"Going to {dest}"
            else:
                label = f"Planning {dest}"
            friend._next_trip_label = label
        _acc_overlap_prep += time.perf_counter() - _ti

    _loop_total = time.perf_counter() - _loop_t0
    if app.debug:
        print(
            f"[FRIENDS_PERF] per_friend_loop={_loop_total:.4f}s friend_count={len(all_friends)} "
            f"| trip_count_compute={_acc_trip_count:.4f}s "
            f"| owner_trips_lookup={_acc_owner_trips:.4f}s "
            f"| participant_trips_lookup={_acc_part_trips:.4f}s "
            f"| overlap_prep={_acc_overlap_prep:.4f}s"
        )

    # ── [FRIENDS_PERF] Block 10: sort + invite token + alpha_groups ───────────
    _t = time.perf_counter()
    def friend_sort_key(f):
        is_new = 0 if f._is_new_friend else 1
        has_trip = 0 if f._has_upcoming_trip else 1
        latest_ts = 0
        if f._latest_upcoming_trip_created_at:
            latest_ts = -f._latest_upcoming_trip_created_at.timestamp()
        else:
            latest_ts = float('inf')
        first_name = (f.first_name or '').lower()
        return (is_new, has_trip, latest_ts, first_name)

    all_friends_sorted = sorted(all_friends, key=friend_sort_key)
    if app.debug:
        print(f"[FRIENDS_PERF] sort={time.perf_counter()-_t:.4f}s")

    # ── [FRIENDS_PERF] Block 11: invite token ─────────────────────────────────
    _t = time.perf_counter()
    invite_token_obj = get_or_create_invite_token(user)
    invite_url = (
        f"{BASE_URL}{url_for('invite_token_landing', token=invite_token_obj.token)}"
        if invite_token_obj else None
    )
    if app.debug:
        print(f"[FRIENDS_PERF] invite_token={time.perf_counter()-_t:.4f}s")

    # ── friend_count for empty vs populated state switch ──────────────────────
    friend_count = len(all_friends)

    # ── [FRIENDS_PERF] Block 12: alpha_groups ─────────────────────────────────
    _t = time.perf_counter()
    alpha_sorted = sorted(all_friends, key=lambda f: (f.first_name or '').lower())
    alpha_groups = []
    for _f in alpha_sorted:
        _letter = (_f.first_name or '?')[0].upper()
        if not alpha_groups or alpha_groups[-1]['letter'] != _letter:
            alpha_groups.append({'letter': _letter, 'friends': []})
        alpha_groups[-1]['friends'].append(_f)
    if app.debug:
        print(f"[FRIENDS_PERF] alpha_groups={time.perf_counter()-_t:.4f}s groups={len(alpha_groups)}")

    # ── [FRIENDS_PERF] Summary ─────────────────────────────────────────────────
    if app.debug:
        print(f"[FRIENDS_PERF] total={time.perf_counter()-_fp_t0:.4f}s friend_count={friend_count}")
        print(f"[ROUTE_PERF] route=friends total={time.perf_counter()-_rp_t0:.4f}s")

    # ── Pending incoming friend invitations ───────────────────────────────────
    # Separate from the Friend rows (which are confirmed) — these are
    # Invitation rows where the current user is the receiver and status=pending.
    # Ordered newest-first so the most recent request appears at the top.
    pending_incoming = (
        Invitation.query
        .filter_by(receiver_id=current_user.id, status='pending')
        .filter(Invitation.trip_id.is_(None))   # friend invites only, not trip requests
        .order_by(Invitation.created_at.desc())
        .all()
    )
    # Pre-load sender User objects so the template can render names without N+1.
    _sender_ids = [inv.sender_id for inv in pending_incoming]
    _senders_map = {}
    if _sender_ids:
        _senders_map = {u.id: u for u in User.query.filter(User.id.in_(_sender_ids)).all()}
    for inv in pending_incoming:
        inv._sender = _senders_map.get(inv.sender_id)

    return render_template(
        "friends.html",
        user=user,
        friends=all_friends_sorted,
        invite_url=invite_url,
        friend_count=friend_count,
        alpha_groups=alpha_groups,
        pending_incoming=pending_incoming,
    )

@app.route("/friends/<int:friend_id>")
@login_required
def friend_profile(friend_id):
    _rp_t0 = time.perf_counter()
    friend = User.query.get_or_404(friend_id)
    user = current_user

    # Authorization guard: only confirmed friends (or self) may view this profile.
    if friend.id != user.id:
        _auth_friendship = Friend.query.filter_by(
            user_id=user.id, friend_id=friend.id
        ).first()
        if not _auth_friendship:
            abort(403)

    # Mark profile as viewed — clears the NEW badge on the Friends screen.
    # Only touches the current user's side of the relationship; no-op if not found.
    try:
        _friendship = Friend.query.filter_by(
            user_id=user.id, friend_id=friend.id
        ).first()
        if _friendship and not _friendship.has_viewed_profile:
            _friendship.has_viewed_profile = True
            db.session.commit()
    except Exception:
        db.session.rollback()

    # Parse overlap context from URL params (for context banner)
    overlap_context = None
    resort_id = request.args.get('resort_id', type=int)
    overlap_start = request.args.get('overlap_start')
    overlap_end = request.args.get('overlap_end')
    
    if resort_id and overlap_start:
        resort = db.session.get(Resort, resort_id)
        if resort:
            try:
                start_date = datetime.strptime(overlap_start, '%Y-%m-%d').date()
                end_date = datetime.strptime(overlap_end, '%Y-%m-%d').date() if overlap_end else start_date
                overlap_context = {
                    'resort_name': _resort_display_name(resort, AMBIGUOUS_RESORT_NAMES),
                    'start_date': start_date,
                    'end_date': end_date
                }
            except (ValueError, TypeError):
                pass
    
    mountains = friend.mountains_visited or []
    friend_mountains_count = len(mountains)
    friend_mountains_sorted = sorted([m.name if hasattr(m, 'name') else m for m in mountains])
    
    # Get Resort objects for profile card (new unified component)
    friend_visited_resorts = friend.get_visited_resorts()
    friend_wishlist_resorts = friend.get_wishlist_resorts()
    
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    
    # Get friend's primary equipment setup only
    friend_primary_equipment = EquipmentSetup.query.filter_by(
        user_id=friend.id, is_primary=True
    ).first()
    if not friend_primary_equipment:
        friend_primary_equipment = EquipmentSetup.query.filter_by(user_id=friend.id).order_by(
            EquipmentSetup.created_at.asc().nullsfirst(), EquipmentSetup.id.asc()
        ).first()
    friend_secondary_equipment = None  # deprecated — kept for template compat
    
    # Get friend's trips
    _t = time.perf_counter()
    trips = (
        SkiTrip.query
        .filter_by(user_id=friend.id, is_public=True)
        .filter(SkiTrip.end_date >= today)
        .order_by(SkiTrip.start_date.asc())
        .all()
    )

    # Get current user's trips for overlap detection: owned + accepted guest
    user_trips = (
        SkiTrip.query
        .filter_by(user_id=user.id)
        .filter(SkiTrip.end_date >= today)
        .all()
    )
    try:
        _fp_accepted_ids = [
            p.trip_id for p in SkiTripParticipant.query.filter(
                SkiTripParticipant.user_id == user.id,
                SkiTripParticipant.status == GuestStatus.ACCEPTED
            ).all()
        ]
        if _fp_accepted_ids:
            _fp_guest_trips = SkiTrip.query.filter(
                SkiTrip.id.in_(_fp_accepted_ids),
                SkiTrip.user_id != user.id,
                SkiTrip.end_date >= today
            ).all()
            user_trips = user_trips + _fp_guest_trips
    except Exception:
        pass
    if app.debug:
        print(f"[ROUTE_PERF] friend_profile.trips_queries={time.perf_counter()-_t:.4f}s friend_trips={len(trips)} user_trips={len(user_trips)}")

    # Build trip overlaps using the canonical helper
    _t = time.perf_counter()
    _raw_overlaps = compute_trip_overlaps(user_trips, trips)
    if app.debug:
        print(f"[ROUTE_PERF] friend_profile.compute_overlaps={time.perf_counter()-_t:.4f}s overlaps={len(_raw_overlaps)}")

    # Flatten to the shape the template expects + mark each friend trip
    _overlapped_trip_ids = {ov['friend_trip_id'] for ov in _raw_overlaps}
    for trip in trips:
        trip.has_trip_overlap = trip.id in _overlapped_trip_ids

    # Deduplicate by (mountain, start_date, end_date) — one row per overlap window
    _seen = set()
    trip_overlaps = []
    for ov in _raw_overlaps:
        key = (ov['mountain'], ov['start_date'], ov['end_date'])
        if key not in _seen:
            _seen.add(key)
            trip_overlaps.append({
                "mountain": ov['mountain'],
                "state": ov['state'],
                "start_date": ov['start_date'],
                "end_date": ov['end_date'],
            })
    
    # Check pass compatibility - can friend ski at user's upcoming trips?
    friend_passes = set(p.strip() for p in friend.pass_type.split(',')) if friend.pass_type else set()
    can_ski_user_trips = False
    for user_trip in user_trips:
        if user_trip.resort:
            # Use pass_brands if available, fallback to brand
            resort_pass_str = user_trip.resort.pass_brands or user_trip.resort.brand
            if resort_pass_str:
                resort_passes = set(p.strip() for p in resort_pass_str.split(','))
                if friend_passes & resort_passes:
                    can_ski_user_trips = True
                    break
    
    # Get friend's available dates via canonical resolver (UserAvailability first,
    # falls back to legacy open_dates JSON). Returns a set of YYYY-MM-DD strings,
    # already filtered to today and future dates.
    friend_open_dates_set = get_available_dates_for_user(friend)
    friend_open_dates = sorted(friend_open_dates_set)
    friend_open_dates_display = format_open_dates_summary(friend_open_dates) if friend_open_dates else None

    # Get current user's available dates via the same canonical resolver.
    user_open_dates_set = get_available_dates_for_user(user)
    user_open_dates = sorted(user_open_dates_set)

    availability_overlaps = compute_availability_overlaps(user_open_dates, friend_open_dates)
    availability_display, availability_remaining = format_availability_ranges(availability_overlaps)
    
    # Check for shared GroupTrips and existing friendship
    shared_trip_exists = check_shared_upcoming_trip(user.id, friend.id)
    already_friends = Friend.query.filter_by(user_id=user.id, friend_id=friend.id).first() is not None
    show_connect_button = shared_trip_exists and not already_friends
    
    # Get friend's wish list resorts
    friend_wish_list_ids = friend.wish_list_resorts or []
    friend_wish_list = Resort.query.filter(Resort.id.in_(friend_wish_list_ids)).all() if friend_wish_list_ids else []
    
    # Calculate wish list overlap with current user
    user_wish_list_ids = set(user.wish_list_resorts or [])
    wish_list_overlap_ids = [rid for rid in friend_wish_list_ids if rid in user_wish_list_ids]
    wish_list_overlap = Resort.query.filter(Resort.id.in_(wish_list_overlap_ids)).all() if wish_list_overlap_ids else []

    # Calculate "been to" overlap — mountains both users have visited
    try:
        user_visited_resorts = user.get_visited_resorts()
        user_visited_ids = set(r.id for r in user_visited_resorts)
        friend_visited_ids = set(r.id for r in friend_visited_resorts)
        been_to_overlap_ids = user_visited_ids & friend_visited_ids
        been_to_overlap = Resort.query.filter(Resort.id.in_(been_to_overlap_ids)).all() if been_to_overlap_ids else []
    except Exception:
        been_to_overlap = []

    # Whether the current user has upcoming trips they could invite this friend to
    has_user_upcoming_trips = len(user_trips) > 0

    if app.debug:
        print(f"[ROUTE_PERF] route=friend_profile total={time.perf_counter()-_rp_t0:.4f}s")
    return render_template(
        "friend_profile.html",
        friend=friend,
        friend_mountains_count=friend_mountains_count,
        friend_mountains=friend_mountains_sorted,
        trips=trips,
        trip_overlaps=trip_overlaps,
        friend_open_dates=friend_open_dates,
        friend_open_dates_display=friend_open_dates_display,
        has_availability_overlap=len(availability_overlaps) > 0,
        availability_display=availability_display,
        availability_remaining=availability_remaining,
        show_connect_button=show_connect_button,
        friend_primary_equipment=friend_primary_equipment,
        friend_secondary_equipment=friend_secondary_equipment,
        can_ski_user_trips=can_ski_user_trips,
        friend_wish_list=friend_wish_list,
        wish_list_overlap=wish_list_overlap,
        been_to_overlap=been_to_overlap,
        has_user_upcoming_trips=has_user_upcoming_trips,
        visited_resorts=friend_visited_resorts,
        wishlist_resorts=friend_wishlist_resorts,
        overlap_context=overlap_context,
        stat_upcoming=get_upcoming_trip_count(friend),
        stat_mountains=friend.visited_resorts_count,
        stat_past=get_past_trip_count(friend),
        stat_wishlist=len(friend.wish_list_resorts or []),
        stat_trips_total=SkiTrip.query.filter_by(user_id=friend.id).count(),
        is_friend=already_friends,
    )

@app.route("/friends/<int:friend_id>/remove", methods=["POST"])
@login_required
def remove_friend_web(friend_id):
    validate_csrf_request()
    # Only allow removing an actual friend of the current user
    row_a = Friend.query.filter_by(user_id=current_user.id, friend_id=friend_id).first()
    if not row_a:
        flash("Friend not found.", "info")
        return redirect(url_for("friends"))

    # Delete both directions — friendship is always stored as mirrored rows
    row_b = Friend.query.filter_by(user_id=friend_id, friend_id=current_user.id).first()
    db.session.delete(row_a)
    if row_b:
        db.session.delete(row_b)

    # Cancel pending AND accepted friend invitations between the two users in
    # either direction so no invitation row can produce a ghost connected-state
    # after the Friend rows are removed.
    Invitation.query.filter(
        db.or_(
            db.and_(
                Invitation.sender_id == current_user.id,
                Invitation.receiver_id == friend_id,
                Invitation.status.in_(['pending', 'accepted']),
            ),
            db.and_(
                Invitation.sender_id == friend_id,
                Invitation.receiver_id == current_user.id,
                Invitation.status.in_(['pending', 'accepted']),
            ),
        )
    ).update({'status': 'cancelled'}, synchronize_session=False)

    # Expire the initiator's active invite token so a link shared before the
    # unfriend can't be used to immediately reconnect.  A fresh token is issued
    # the next time they visit /invite.
    _now = datetime.utcnow()
    for _tok in InviteToken.query.filter_by(inviter_id=current_user.id).all():
        if not _tok.is_used() and not _tok.is_expired():
            _tok.expires_at = _now

    db.session.commit()

    flash("Friend removed.", "success")
    return redirect(url_for("friends"))

@app.route("/profile/<int:user_id>")
@login_required
def friend_profile_legacy(user_id):
    """Legacy route - redirect to the main friend profile page."""
    # Check if viewing own profile
    if user_id == current_user.id:
        return redirect(url_for("profile"))
    
    # Redirect to the main friend profile route
    return redirect(url_for("friend_profile", friend_id=user_id))

@app.route("/api/profile/update", methods=["POST"])
@login_required
def update_profile():
    user = current_user
    
    data = request.get_json()
    
    if "first_name" in data:
        user.first_name = data.get("first_name", "").strip()
    if "last_name" in data:
        user.last_name = data.get("last_name", "").strip()
    if "rider_type" in data:
        # Update primary_rider_type instead of legacy rider_type
        user.primary_rider_type = data.get("rider_type", "").strip()
    if "primary_rider_type" in data:
        user.primary_rider_type = data.get("primary_rider_type", "").strip()
    if "pass_type" in data:
        _raw_pt = data.get("pass_type", "").strip()
        _norm_pt = normalize_pass_selection(_raw_pt) or _raw_pt
        if count_real_passes(_norm_pt) > 3:
            return jsonify({"success": False, "message": "You can select up to 3 passes."}), 400
        user.pass_type = _norm_pt
    
    db.session.commit()
    
    return jsonify({"success": True, "message": "Profile updated"}), 200

@app.route("/create-trip")
@login_required
def create_trip_page():
    # Legacy route — retired. Redirect to canonical /add_trip, preserving all prefill params.
    params = {k: v for k, v in request.args.items()}
    return redirect(url_for("add_trip", **params), 302)

# ============================================================================
# PLANNING FEATURE
# Shows availability overlap windows between current user and friends
# ============================================================================

def group_dates_into_windows(dates):
    """Group a list of date strings into consecutive windows.
    
    Args:
        dates: List of 'YYYY-MM-DD' date strings (must be sorted)
    
    Returns:
        List of tuples: [(start_date, end_date), ...]
    """
    if not dates:
        return []
    
    windows = []
    sorted_dates = sorted(dates)
    
    window_start = sorted_dates[0]
    window_end = sorted_dates[0]
    
    for i in range(1, len(sorted_dates)):
        current = sorted_dates[i]
        prev = sorted_dates[i - 1]
        
        # Check if consecutive (within 1 day)
        try:
            current_date = datetime.strptime(current, '%Y-%m-%d').date()
            prev_date = datetime.strptime(prev, '%Y-%m-%d').date()
            
            if (current_date - prev_date).days <= 1:
                window_end = current
            else:
                windows.append((window_start, window_end))
                window_start = current
                window_end = current
        except ValueError:
            continue
    
    windows.append((window_start, window_end))
    return windows


def get_planning_windows(user):
    """Get availability overlap windows for the planning feed.
    
    Returns a list of windows, each containing:
    - start_date: 'YYYY-MM-DD'
    - end_date: 'YYYY-MM-DD'
    - friends: list of friend user objects
    - friend_names_display: formatted string for display
    """
    from services.open_dates import get_open_date_matches
    
    matches = get_open_date_matches(user)
    if not matches:
        return []
    
    # Group matches by date
    date_to_friends = {}
    friend_cache = {}
    
    for match in matches:
        d = match['date']
        fid = match['friend_id']
        
        if d not in date_to_friends:
            date_to_friends[d] = set()
        date_to_friends[d].add(fid)
        
        # Cache friend info
        if fid not in friend_cache:
            friend_cache[fid] = {
                'id': fid,
                'name': match['friend_name'],
                'pass': match['friend_pass']
            }
    
    # Filter to future dates only
    today_str = date.today().strftime('%Y-%m-%d')
    future_dates = sorted([d for d in date_to_friends.keys() if d >= today_str])
    
    if not future_dates:
        return []
    
    # Group consecutive dates into windows
    windows = group_dates_into_windows(future_dates)
    
    # For each window, find friends who are available for ALL dates in that window
    result = []
    for start, end in windows:
        # Get all dates in this window
        window_dates = [d for d in future_dates if start <= d <= end]
        
        if not window_dates:
            continue
        
        # Find friends available on ALL dates in window
        friends_per_date = [date_to_friends.get(d, set()) for d in window_dates]
        common_friends = set.intersection(*friends_per_date) if friends_per_date else set()
        
        if not common_friends:
            continue
        
        # Build friend info list
        friends = [friend_cache[fid] for fid in common_friends]
        friends.sort(key=lambda f: f['name'] or '')
        
        # Format display string
        names = [f['name'] for f in friends if f['name']]
        if len(names) == 0:
            continue
        elif len(names) == 1:
            display = f"{names[0]} is free"
        elif len(names) == 2:
            display = f"{names[0]} and {names[1]} are free"
        else:
            display = f"{names[0]} + {len(names) - 1} others are free"
        
        result.append({
            'start_date': start,
            'end_date': end,
            'friends': friends,
            'friend_names_display': display
        })
    
    return result


def format_planning_dates(start_str, end_str):
    """Format date range for planning display (e.g., 'Jan 18–21' or 'Jan 25–Feb 1')."""
    try:
        start = datetime.strptime(start_str, '%Y-%m-%d').date()
        end = datetime.strptime(end_str, '%Y-%m-%d').date()
        
        if start == end:
            return start.strftime('%b %-d')
        elif start.month == end.month:
            return f"{start.strftime('%b %-d')}–{end.day}"
        else:
            return f"{start.strftime('%b %-d')}–{end.strftime('%b %-d')}"
    except ValueError:
        return f"{start_str} – {end_str}"


@app.route("/planning")
@login_required
def planning():
    """Legacy Planning route — redirects to Home."""
    return redirect(url_for('home'))


@app.route("/planning/window/<start_date>/<end_date>")
@login_required
def planning_window(start_date, end_date):
    user = current_user
    
    # Validate date format
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        flash("Invalid date range", "error")
        return redirect(url_for('home'))
    
    # Get friends available in this window
    from services.open_dates import get_open_date_matches
    matches = get_open_date_matches(user)
    
    # Find friends with overlapping dates in this window
    window_dates = set()
    d = start
    while d <= end:
        window_dates.add(d.strftime('%Y-%m-%d'))
        d += timedelta(days=1)
    
    # Group by friend and check if they overlap with this window
    friend_ids_in_window = set()
    for match in matches:
        if match['date'] in window_dates:
            friend_ids_in_window.add(match['friend_id'])
    
    # Load friend details
    friends_data = []
    if friend_ids_in_window:
        friends = User.query.filter(User.id.in_(friend_ids_in_window)).all()
        for friend in friends:
            # Build identity line (same format as elsewhere)
            identity_parts = []
            if friend.display_rider_type:
                identity_parts.append(friend.display_rider_type)
            if friend.pass_type:
                identity_parts.append(friend.pass_type)
            if friend.skill_level:
                identity_parts.append(friend.skill_level)
            
            identity_line = ' · '.join(identity_parts) if identity_parts else ''
            
            friends_data.append({
                'id': friend.id,
                'name': f"{friend.first_name} {friend.last_name}".strip(),
                'identity_line': identity_line
            })
        
        friends_data.sort(key=lambda f: f['name'])
    
    # Format header
    display_dates = format_planning_dates(start_date, end_date)
    
    return render_template(
        "planning_window.html",
        user=user,
        friends=friends_data,
        display_dates=display_dates,
        start_date=start_date,
        end_date=end_date
    )


@app.route("/invite")
@login_required
def invite():
    # Check if user has reached their invite accept limit
    if not can_sender_accept_more_invites(current_user):
        return render_template("invite_limit_reached.html", user=current_user)

    invite_token = get_or_create_invite_token(current_user)
    invite_url = f"{BASE_URL}{url_for('invite_token_landing', token=invite_token.token)}"

    resp = make_response(render_template("invite.html", user=current_user, invite_url=invite_url, remaining_invites=None))
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.route("/my-qr")
@login_required
def my_qr():
    invite_token = get_or_create_invite_token(current_user)
    if not invite_token:
        return render_template("invite_limit_reached.html", user=current_user)
    qr_url = f"{BASE_URL}{url_for('invite_token_landing', token=invite_token.token)}"
    qr = segno.make(qr_url)
    buf = BytesIO()
    qr.save(buf, kind="png", scale=8)
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

@app.route("/connect/<int:user_id>")
def connect_via_qr(user_id):
    inviter = User.query.get_or_404(user_id)
    
    if not current_user.is_authenticated:
        # Store the destination in session so it survives the OAuth redirect cycle.
        # URL params passed to /auth are not consumed after Google OAuth callbacks.
        session["post_login_redirect"] = url_for("connect_via_qr", user_id=user_id)
        return redirect(url_for("auth"))
    
    if current_user.id == inviter.id:
        return render_template("connect_self.html")
    
    existing = Friend.query.filter_by(user_id=current_user.id, friend_id=inviter.id).first()
    if existing:
        return render_template("already_friends.html", friend=inviter)
    
    return render_template("connect_confirm.html", friend=inviter)

@app.route("/connect/<int:user_id>/add", methods=["POST"])
@login_required
@limiter.limit("20 per hour", key_func=_user_or_ip)
def connect_add(user_id):
    inviter = User.query.get_or_404(user_id)

    existing_a_to_b = Friend.query.filter_by(user_id=current_user.id, friend_id=inviter.id).first()
    existing_b_to_a = Friend.query.filter_by(user_id=inviter.id, friend_id=current_user.id).first()

    if not existing_a_to_b:
        new_a_to_b = Friend(user_id=current_user.id, friend_id=inviter.id)
        db.session.add(new_a_to_b)

    if not existing_b_to_a:
        new_b_to_a = Friend(user_id=inviter.id, friend_id=current_user.id)
        db.session.add(new_b_to_a)

    # Mark any pending invitations between these two users as accepted,
    # since the connection was successfully established via QR/direct link.
    Invitation.query.filter(
        db.or_(
            db.and_(
                Invitation.sender_id == current_user.id,
                Invitation.receiver_id == inviter.id,
                Invitation.status == 'pending',
            ),
            db.and_(
                Invitation.sender_id == inviter.id,
                Invitation.receiver_id == current_user.id,
                Invitation.status == 'pending',
            ),
        )
    ).update({'status': 'accepted'}, synchronize_session=False)

    db.session.commit()
    _qr_count = Friend.query.filter_by(user_id=current_user.id).count()
    ph_analytics.track(current_user.id, 'friend_connected', {
        'source':          'qr_scan',
        'is_first_friend': _qr_count == 1,
    })
    return render_template("connect_success.html", friend=inviter)

@app.route("/invite/<int:user_id>")
@login_required
def invite_link(user_id):
    """Legacy integer-based invite URL — retired. Redirect to the invite page."""
    return redirect(url_for("invite"))

def date_ranges_overlap(start1, end1, start2, end2):
    """Check if two date ranges overlap"""
    return start1 <= end2 and start2 <= end1

def format_open_dates_summary(date_strings):
    """Format a list of YYYY-MM-DD strings into human-readable summary grouped by month.
    E.g., ['2024-12-14', '2024-12-18', '2024-12-19', '2025-01-03', '2025-01-04'] 
    -> 'Dec 14,18,19 | Jan 3,4'
    
    Preserves year separation: dates from different years with same month are kept separate.
    """
    if not date_strings:
        return None
    
    from datetime import datetime as dt
    from collections import OrderedDict
    
    # Parse and sort dates
    dates = sorted([dt.strptime(d, '%Y-%m-%d').date() for d in date_strings])
    
    # Group dates by (year, month) to preserve year separation
    months = OrderedDict()
    for d in dates:
        month_key = (d.year, d.month, d.strftime('%b'))  # (2024, 12, 'Dec')
        if month_key not in months:
            months[month_key] = []
        months[month_key].append(d.day)
    
    # Format each month group: "Jan 1,2,3,4"
    # Year is omitted from display but used for correct grouping
    formatted = []
    for (year, month, month_name), days in months.items():
        days_str = ','.join(str(d) for d in days)
        formatted.append(f"{month_name} {days_str}")
    
    return '\n'.join(formatted)

def dates_to_ranges(date_strings):
    """Convert a list of YYYY-MM-DD strings to a list of {start_date, end_date} dicts.
    Groups consecutive dates into ranges.
    E.g., ['2024-12-14', '2024-12-15', '2024-12-16'] -> [{start_date: date(2024-12-14), end_date: date(2024-12-16)}]
    """
    if not date_strings:
        return []
    
    from datetime import datetime as dt
    
    # Parse and sort dates
    dates = sorted([dt.strptime(d, '%Y-%m-%d').date() for d in date_strings])
    
    # Group consecutive dates
    ranges = []
    current_start = dates[0]
    current_end = dates[0]
    
    for i in range(1, len(dates)):
        if (dates[i] - current_end).days == 1:
            current_end = dates[i]
        else:
            ranges.append({"start_date": current_start, "end_date": current_end})
            current_start = dates[i]
            current_end = dates[i]
    
    ranges.append({"start_date": current_start, "end_date": current_end})
    return ranges


def build_home_avail_ranges(date_set, cap=3):
    """
    Converts a set of YYYY-MM-DD availability strings into display-ready range dicts
    for the Home screen availability summary card.

    Returns (capped_ranges, overflow_count) where:
      capped_ranges  — list of up to `cap` dicts, each with a 'display' key
                       e.g. [{'display': 'Dec 14 – Dec 18'}, {'display': 'Jan 3'}]
      overflow_count — int, number of additional ranges beyond the cap
    """
    if not date_set:
        return [], 0

    _SHORT = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    raw = dates_to_ranges(sorted(date_set))
    display_ranges = []
    for r in raw:
        s, e = r['start_date'], r['end_date']
        s_str = f"{_SHORT[s.month - 1]} {s.day}"
        e_str = f"{_SHORT[e.month - 1]} {e.day}"
        display_ranges.append({'display': s_str if s == e else f"{s_str} – {e_str}"})

    overflow = max(0, len(display_ranges) - cap)
    return display_ranges[:cap], overflow


def check_trip_invite_eligibility(user_id, friend_id):
    """
    Check if user can invite friend to a trip.
    Returns True if at least ONE condition is met:
    1. They share a trip overlap (past or upcoming)
    2. They share open availability
    3. The friendship has trip_invites_allowed = True
    """
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    
    # Check 1: Friendship permission
    friendship = Friend.query.filter_by(user_id=user_id, friend_id=friend_id).first()
    if friendship and friendship.trip_invites_allowed:
        return True
    
    # Check 2: Trip overlap (any trip, past or upcoming)
    user_trips = SkiTrip.query.filter_by(user_id=user_id).all()
    friend_trips = SkiTrip.query.filter(
        SkiTrip.user_id == friend_id,
        SkiTrip.is_public == True
    ).all()
    
    for ut in user_trips:
        for ft in friend_trips:
            # Resort match: prefer resort_id (canonical), fall back to mountain string
            if ut.resort_id and ft.resort_id:
                same_resort = (ut.resort_id == ft.resort_id)
            else:
                same_resort = bool(ut.mountain and ft.mountain and ut.mountain == ft.mountain)
            if same_resort and date_ranges_overlap(ut.start_date, ut.end_date, ft.start_date, ft.end_date):
                return True
    
    # Check 3: Shared open availability
    user = db.session.get(User, user_id)
    friend = db.session.get(User, friend_id)
    if user and friend:
        user_open_dates = set(d for d in (user.open_dates or []) if d >= today_str)
        friend_open_dates = set(d for d in (friend.open_dates or []) if d >= today_str)
        if user_open_dates & friend_open_dates:
            return True
    
    return False


def compute_availability_overlaps(user_open_dates, friend_open_dates):
    """Compute overlapping date ranges between user and friend's open dates.
    Returns list of {start_date, end_date} dicts representing overlap ranges.
    Uses exact overlap logic: overlap_start = max(...), overlap_end = min(...).
    Then merges overlapping/adjacent ranges.
    """
    if not user_open_dates or not friend_open_dates:
        return []
    
    # Convert individual dates to ranges
    user_ranges = dates_to_ranges(user_open_dates)
    friend_ranges = dates_to_ranges(friend_open_dates)
    
    # Find all overlaps
    overlaps = []
    for user_r in user_ranges:
        for friend_r in friend_ranges:
            overlap_start = max(user_r["start_date"], friend_r["start_date"])
            overlap_end = min(user_r["end_date"], friend_r["end_date"])
            
            if overlap_start <= overlap_end:
                overlaps.append({"start_date": overlap_start, "end_date": overlap_end})
    
    if not overlaps:
        return []
    
    # Sort by start_date
    overlaps = sorted(overlaps, key=lambda x: x["start_date"])
    
    # Merge overlapping/adjacent ranges
    merged = [overlaps[0]]
    for current in overlaps[1:]:
        last = merged[-1]
        # Check if current overlaps or is adjacent to last (within 1 day)
        if current["start_date"] <= last["end_date"] + timedelta(days=1):
            # Merge
            last["end_date"] = max(last["end_date"], current["end_date"])
        else:
            # No overlap, add as new range
            merged.append(current)
    
    return merged

def format_availability_ranges(ranges):
    """Format a list of {start_date, end_date} dicts for display.
    E.g., [{start: date(2024-12-14), end: date(2024-12-19)}] -> 'Dec 14–19'
    """
    if not ranges:
        return None, 0
    
    formatted = []
    for r in ranges[:3]:  # Only display first 3
        start_str = r["start_date"].strftime('%b %d').replace(' 0', ' ')
        end_str = r["end_date"].strftime('%d').lstrip('0')
        
        # If same month, just show "Dec 14–19"
        if r["start_date"].month == r["end_date"].month:
            formatted.append(f"{start_str}–{end_str}")
        else:
            end_full = r["end_date"].strftime('%b %d').replace(' 0', ' ')
            formatted.append(f"{start_str}–{end_full}")
    
    remaining_count = max(0, len(ranges) - 3)
    return ' · '.join(formatted), remaining_count

def build_trip_overlap_today_card(user, today, friend_ids):
    """
    Returns a card dict if the user has an active trip today AND at least one
    friend has an active trip at the SAME resort today.

    Card dict keys:
      resort_id, resort_name, resort_slug, card_key, friend_count
      friend_id + friend_name (only when friend_count == 1)
    """
    if not friend_ids:
        return None

    # Find user's active trip(s) today
    user_active = SkiTrip.query.filter(
        SkiTrip.user_id == user.id,
        SkiTrip.resort_id.isnot(None),
        SkiTrip.start_date <= today,
        SkiTrip.end_date >= today,
    ).order_by(SkiTrip.start_date.asc()).first()

    if not user_active:
        return None

    resort_id = user_active.resort_id

    # Find friends who are ALSO at the same resort today
    friend_trips_today = SkiTrip.query.filter(
        SkiTrip.user_id.in_(friend_ids),
        SkiTrip.resort_id == resort_id,
        SkiTrip.start_date <= today,
        SkiTrip.end_date >= today,
    ).all()

    if not friend_trips_today:
        return None

    # Distinct qualifying friend IDs
    qualifying_friend_ids = list({t.user_id for t in friend_trips_today})
    friend_count = len(qualifying_friend_ids)

    # Check dismissal scoped to user + mountain + date
    today_str = today.isoformat()
    card_key = f"{resort_id}:{today_str}"
    already_dismissed = DismissedInsightCard.query.filter_by(
        user_id=user.id,
        card_type='trip_overlap_today',
        card_key=card_key,
    ).first()
    if already_dismissed:
        return None

    resort = db.session.get(Resort, resort_id)
    if not resort:
        return None

    card = {
        'resort_id': resort_id,
        'resort_name': _resort_display_name(resort, AMBIGUOUS_RESORT_NAMES),
        'resort_slug': resort.slug,
        'card_key': card_key,
        'friend_count': friend_count,
    }

    if friend_count == 1:
        friend_user = db.session.get(User, qualifying_friend_ids[0])
        if not friend_user:
            return None
        card['friend_id'] = friend_user.id
        card['friend_name'] = f"{friend_user.first_name or ''} {friend_user.last_name or ''}".strip()

    return card


def build_friend_at_mountain_card(user, today, friend_ids):
    """
    Returns a card dict if: user has an upcoming trip whose resort is on their
    wishlist AND at least one friend has a past trip to that same resort.
    Returns None if no qualifying combination exists or it's been dismissed.

    Card dict keys: friend_id, friend_name, resort_id, resort_name, trip_id, card_key
    """
    if not friend_ids:
        return None

    wishlist_set = set(user.wish_list_resorts or [])
    if not wishlist_set:
        return None

    # Find the user's soonest upcoming trip whose resort is on the wishlist
    # AND where at least one friend has a past trip to that resort
    target_trip = None
    target_resort_id = None
    past_friend_trips = []
    for trip in SkiTrip.query.filter(
        SkiTrip.user_id == user.id,
        SkiTrip.resort_id.isnot(None),
        SkiTrip.start_date >= today
    ).order_by(SkiTrip.start_date.asc()).all():
        if trip.resort_id not in wishlist_set:
            continue
        candidate_trips = (
            SkiTrip.query
            .filter(
                SkiTrip.user_id.in_(friend_ids),
                SkiTrip.resort_id == trip.resort_id,
                SkiTrip.end_date < today
            )
            .order_by(SkiTrip.end_date.desc())
            .all()
        )
        if candidate_trips:
            target_trip = trip
            target_resort_id = trip.resort_id
            past_friend_trips = candidate_trips
            break

    if not target_trip or not past_friend_trips:
        return None

    # Build a ranked list: most recent trip → earliest friendship → alphabetical name
    # Collect best trip per friend (already sorted by desc end_date)
    seen_friends = {}
    for pt in past_friend_trips:
        if pt.user_id not in seen_friends:
            seen_friends[pt.user_id] = pt  # most recent trip

    # Load friend records for tie-breaking
    friend_rows = {
        fr.friend_id: fr
        for fr in Friend.query.filter(
            Friend.user_id == user.id,
            Friend.friend_id.in_(list(seen_friends.keys()))
        ).all()
    }

    # Load User objects for name sorting
    friend_users = {
        u.id: u
        for u in User.query.filter(User.id.in_(list(seen_friends.keys()))).all()
    }

    def sort_key(fid):
        trip = seen_friends[fid]
        fu = friend_users.get(fid)
        fr = friend_rows.get(fid)
        # 1. Most recent past trip (negate date for descending)
        recency = -(trip.end_date.toordinal()) if trip.end_date else 0
        # 2. Earliest friendship record
        friendship_age = fr.created_at.toordinal() if fr and fr.created_at else 999999999
        # 3. Alphabetical by full name
        name = f"{fu.first_name or ''} {fu.last_name or ''}".strip().lower() if fu else 'zzz'
        return (recency, friendship_age, name)

    ranked_friend_ids = sorted(seen_friends.keys(), key=sort_key)
    best_friend_id = ranked_friend_ids[0]

    # Check if this exact combination has been dismissed
    card_key = f"{best_friend_id}:{target_resort_id}:{target_trip.id}"
    already_dismissed = DismissedInsightCard.query.filter_by(
        user_id=user.id,
        card_type='friend_at_mountain',
        card_key=card_key
    ).first()
    if already_dismissed:
        return None

    # Build card data
    best_friend = friend_users.get(best_friend_id)
    if not best_friend:
        return None

    resort = db.session.get(Resort, target_resort_id)
    if not resort:
        return None

    friend_full_name = f"{best_friend.first_name or ''} {best_friend.last_name or ''}".strip()
    if not friend_full_name:
        return None

    return {
        'friend_id': best_friend_id,
        'friend_name': friend_full_name,
        'resort_id': target_resort_id,
        'resort_name': _resort_display_name(resort, AMBIGUOUS_RESORT_NAMES),
        'trip_id': target_trip.id,
        'card_key': card_key,
    }


@app.route("/home")
@login_required
def home():
    user = current_user
    today = date.today()
    _rp_t0 = time.perf_counter()

    # One-time connection success message (set by accept_invitation, consumed here)
    new_connection_name = session.pop('new_connection_name', None)

    # Activity-based connection card: surfaces for both the acceptor and the invite sender.
    # Uses DismissedInsightCard so it never reappears once seen.
    # Only shows events from the last 48h to avoid stale product moments.
    sender_connection_card = None
    try:
        cutoff = datetime.utcnow() - timedelta(hours=48)
        _hp_t0 = time.perf_counter()
        recent_connections = Activity.query.filter(
            Activity.recipient_user_id == user.id,
            Activity.type == ActivityType.CONNECTION_ACCEPTED.value,
            Activity.created_at >= cutoff,
        ).order_by(Activity.created_at.desc()).all()
        if app.debug:
            print(f"[HOME_PERF] connection_activity={time.perf_counter()-_hp_t0:.4f}s count={len(recent_connections)}")

        # Batch-fetch dismissed keys and actor users before looping — avoids N+1.
        _conn_card_keys = [
            f"connection:{act.actor_user_id}:{act.recipient_user_id}"
            for act in recent_connections
        ]
        _dismissed_conn_keys = set()
        if _conn_card_keys:
            _dismissed_conn_rows = DismissedInsightCard.query.filter(
                DismissedInsightCard.user_id == user.id,
                DismissedInsightCard.card_type == 'connection_accepted',
                DismissedInsightCard.card_key.in_(_conn_card_keys),
            ).all()
            _dismissed_conn_keys = {d.card_key for d in _dismissed_conn_rows}
        _conn_actor_ids = list({act.actor_user_id for act in recent_connections})
        _conn_actors_map = (
            {u.id: u for u in User.query.filter(User.id.in_(_conn_actor_ids)).all()}
            if _conn_actor_ids else {}
        )
        for act in recent_connections:
            card_key = f"connection:{act.actor_user_id}:{act.recipient_user_id}"
            if card_key not in _dismissed_conn_keys:
                other_user = _conn_actors_map.get(act.actor_user_id)
                if other_user:
                    sender_connection_card = {
                        'name': other_user.first_name or other_user.username or 'your new friend',
                        'card_key': card_key,
                    }
                    break
    except Exception:
        db.session.rollback()
        sender_connection_card = None

    # If the session-based card (acceptor path) is already showing for this same
    # connection, pre-dismiss the Activity card so they don't see two "connected" messages.
    if new_connection_name and sender_connection_card:
        try:
            existing = DismissedInsightCard.query.filter_by(
                user_id=user.id,
                card_type='connection_accepted',
                card_key=sender_connection_card['card_key'],
            ).first()
            if not existing:
                db.session.add(DismissedInsightCard(
                    user_id=user.id,
                    card_type='connection_accepted',
                    card_key=sender_connection_card['card_key'],
                ))
                db.session.commit()
        except Exception:
            db.session.rollback()
        sender_connection_card = None

    # --- Next Trip (created or accepted) ---
    try:
        _hp_t0 = time.perf_counter()
        my_trips = SkiTrip.query.filter(
            SkiTrip.user_id == user.id,
            SkiTrip.end_date >= today
        ).order_by(SkiTrip.start_date.asc()).all()
        if app.debug:
            print(f"[HOME_PERF] my_trips_query={time.perf_counter()-_hp_t0:.4f}s count={len(my_trips)}")
    except Exception:
        db.session.rollback()
        my_trips = []

    try:
        _hp_t0 = time.perf_counter()
        accepted_participations = SkiTripParticipant.query.filter(
            SkiTripParticipant.user_id == user.id,
            SkiTripParticipant.status == GuestStatus.ACCEPTED
        ).all()
        if app.debug:
            print(f"[HOME_PERF] accepted_participations={time.perf_counter() - _hp_t0:.4f}s count={len(accepted_participations)}")
        accepted_trip_ids = [p.trip_id for p in accepted_participations]
        if accepted_trip_ids:
            _hp_t0 = time.perf_counter()
            accepted_guest_trips = SkiTrip.query.filter(
                SkiTrip.id.in_(accepted_trip_ids),
                SkiTrip.user_id != user.id,
                SkiTrip.end_date >= today
            ).order_by(SkiTrip.start_date.asc()).all()
            if app.debug:
                print(f"[HOME_PERF] accepted_guest_trips={time.perf_counter() - _hp_t0:.4f}s count={len(accepted_guest_trips)}")
        else:
            accepted_guest_trips = []
    except Exception:
        db.session.rollback()
        accepted_guest_trips = []

    all_upcoming = sorted(my_trips + accepted_guest_trips, key=lambda t: t.start_date)
    next_trip = all_upcoming[0] if all_upcoming else None
    # --- Friends (single join query: IDs + objects in one round trip) ---
    try:
        _hp_t0 = time.perf_counter()
        all_friends = (
            User.query
            .join(Friend, Friend.friend_id == User.id)
            .filter(Friend.user_id == user.id)
            .all()
        )
        friend_ids = [f.id for f in all_friends]
        if app.debug:
            print(f"[HOME_PERF] friends_load={time.perf_counter() - _hp_t0:.4f}s count={len(friend_ids)}")
    except Exception:
        db.session.rollback()
        friend_ids = []
        all_friends = []

    # --- Trip Invite Banner (soonest active pending trip invite) ---
    banner_invite = None
    banner_invite_count = 0
    trip_invites = []
    try:
        _hp_t0 = time.perf_counter()
        invited_participations = (
            SkiTripParticipant.query
            .options(db.joinedload(SkiTripParticipant.trip).joinedload(SkiTrip.resort))
            .filter(
                SkiTripParticipant.user_id == user.id,
                SkiTripParticipant.status == GuestStatus.INVITED
            )
            .all()
        )
        if app.debug:
            print(f"[HOME_PERF] invited_participations={time.perf_counter() - _hp_t0:.4f}s count={len(invited_participations)}")
        active_invites = sorted(
            [p for p in invited_participations if p.trip and p.trip.end_date >= today],
            key=lambda p: p.trip.start_date
        )
        banner_invite_count = len(active_invites)
        # Batch-load all trip owners in one query instead of one get() per invite
        _inviter_ids = {p.trip.user_id for p in active_invites if p.trip}
        _inviters_map = (
            {u.id: u for u in User.query.filter(User.id.in_(_inviter_ids)).all()}
            if _inviter_ids else {}
        )
        for p in active_invites:
            trip = p.trip
            inviter = _inviters_map.get(trip.user_id)
            resort = trip.resort
            trip_invites.append({
                'trip_id': trip.id,
                'trip': trip,
                'resort': resort,
                'inviter_name': (f"{inviter.first_name or ''} {inviter.last_name or ''}".strip()) if inviter else 'Someone',
            })
        if trip_invites:
            banner_invite = trip_invites[0]
    except Exception:
        db.session.rollback()

    # all_friends already populated above via single join query

    # --- Secondary Card (priority: connect_invite > overlap > friend_trip) ---
    secondary_card = None
    try:
        _hp_t0 = time.perf_counter()
        connect_inv = Invitation.query.filter_by(
            receiver_id=user.id,
            status='pending'
        ).filter(Invitation.trip_id == None).first()
        if app.debug:
            print(f"[HOME_PERF] invitation_query={time.perf_counter()-_hp_t0:.4f}s")
        if connect_inv:
            sender = db.session.get(User, connect_inv.sender_id)
            secondary_card = {
                'type': 'connect_invite',
                'invitation_id': connect_inv.id,
                'sender_name': (f"{sender.first_name or ''} {sender.last_name or ''}".strip()) if sender else 'Someone',
            }
    except Exception:
        db.session.rollback()

    # Diagnostics sentinels — populated inside each try block below
    _diag_opp_engine_count = 0
    _diag_dismissed_count = 0
    _diag_opp_after_dismissal = 0
    _diag_hap_raw_trips = 0
    _diag_hap_candidates = 0
    _diag_hap_opp_suppressed = 0

    # Fetch user availability once here — reused by build_destination_feed (avoids
    # two extra Supabase round-trips: one inside the engine, one at show_add_dates).
    from services.open_dates import get_available_dates_for_user as _get_avail_home
    _hp_t0 = time.perf_counter()
    _user_avail_home = _get_avail_home(user)
    if app.debug:
        print(f"[HOME_PERF] availability_lookup={time.perf_counter() - _hp_t0:.4f}s has_dates={bool(_user_avail_home)}")

    # --- Coordination feed (Home opportunities stream) ---
    dest_feed = []
    _engine_friend_trips = []
    _ideas_engine_diag = {}
    try:
        from services.ideas_engine import build_destination_feed as _build_home_feed
        if all_friends:
            _hp_t0 = time.perf_counter()
            _resort_map = get_all_active_resorts_map()
            _raw_feed, _ideas_engine_diag, _engine_friend_trips = _build_home_feed(
                user, all_friends, user_avail_dates=_user_avail_home,
                user_trips=my_trips, resort_map=_resort_map
            )
            if app.debug:
                print(f"[HOME_PERF] build_destination_feed={time.perf_counter() - _hp_t0:.4f}s raw_count={len(_raw_feed)}")
            _diag_opp_engine_count = len(_raw_feed)
            _dismissed_opp_keys = set()
            try:
                _hp_t0 = time.perf_counter()
                _dismissed_cards = DismissedInsightCard.query.filter_by(
                    user_id=user.id,
                    card_type='opportunity',
                ).all()
                if app.debug:
                    print(f"[HOME_PERF] dismissed_cards_lookup={time.perf_counter() - _hp_t0:.4f}s count={len(_dismissed_cards)}")
                _dismissed_opp_keys = {d.card_key for d in _dismissed_cards}
            except Exception:
                db.session.rollback()
            _diag_dismissed_count = len(_dismissed_opp_keys)
            for _row in _raw_feed:
                # Dismissal key: resort-pinned cards use type:resort_id;
                # no-resort cards use type:friend_ids:start_date (BUG-9 fix)
                if _row.get('resort_id'):
                    _ck = f"{_row['idea_type']}:{_row['resort_id']}"
                else:
                    _fids = "_".join(str(f) for f in sorted(_row.get('friend_ids') or []))
                    _ck = f"{_row['idea_type']}:{_fids}:{_row.get('start_date', 'nodate')}"
                if _ck not in _dismissed_opp_keys:
                    _row['_card_key'] = _ck
                    dest_feed.append(_row)
            _diag_opp_after_dismissal = len(dest_feed)
            dest_feed = dest_feed[:5]
            print(f"[Ideas] dismissed_keys={_diag_dismissed_count} after_dismissal={_diag_opp_after_dismissal} after_cap={len(dest_feed)}")
    except Exception:
        db.session.rollback()
        dest_feed = []

    # Pairs (friend_id, resort_id) already surfaced in Opportunities — used to
    # suppress duplicate signals in Happening.
    _opp_friend_resort_pairs = set()
    for _opp_row in dest_feed:
        _opp_rid = _opp_row.get('resort_id')
        if _opp_rid:
            for _fid in (_opp_row.get('friend_ids') or []):
                _opp_friend_resort_pairs.add((_fid, _opp_rid))

    HOME_HAPPENING_RENDER_CAP = 5
    # --- Happening signals (one row per friend, editorial format, max 5) ---
    # Reuses friend_trips already fetched by build_destination_feed — no second DB query.
    # Re-sorted here by activity_timestamp DESC (most-recent edit/creation first).
    # This sort is Happening-only; Opportunities continues using start_date ASC from the engine.
    #
    # Note: _opp_friend_resort_pairs suppression is intentionally deferred — the two
    # sections serve different purposes (action vs ambient) and the user benefit is low.
    happening_signals = []
    if friend_ids:
        try:
            if app.debug:
                print(f"[HOME_PERF] happening_trips=reused count={len(_engine_friend_trips)}")
            _ft_users_map = {u.id: u for u in all_friends}
            _diag_hap_raw_trips = len(_engine_friend_trips)
            _hap_seen_users = set()
            _now = datetime.utcnow()
            # Re-sort by most-recent activity first (updated_at preferred, else created_at)
            _sorted_friend_trips = sorted(
                _engine_friend_trips,
                key=lambda t: (t.updated_at or t.created_at or datetime.min),
                reverse=True,
            )
            for ft in _sorted_friend_trips:
                ft_user = _ft_users_map.get(ft.user_id)
                if not ft_user:
                    continue
                if ft.user_id in _hap_seen_users:
                    continue
                _hap_seen_users.add(ft.user_id)
                ft_resort = ft.resort
                ft_mountain = ft_resort.name if ft_resort else ft.mountain
                status = ft.trip_status or 'planning'
                full_name = (
                    f"{ft_user.first_name or ''} {ft_user.last_name or ''}".strip()
                ) if ft_user else 'A friend'
                # Line 1: person name only
                person = full_name
                # Line 2: action + mountain (state-based verb, never recency)
                _mtn = ft_mountain or None
                if _mtn:
                    if status == 'going':
                        action_line = f"Going to {_mtn}"
                    elif status in ('confirmed', 'booked'):
                        action_line = f"Heading to {_mtn}"
                    else:
                        action_line = f"Planning {_mtn}"
                else:
                    action_line = "Trip upcoming"
                # Line 3: recency only — no state words
                _activity_ts = ft.updated_at if ft.updated_at else ft.created_at
                _was_updated = ft.updated_at is not None
                _age = (_now - _activity_ts).total_seconds() if _activity_ts else None
                if _age is None:
                    recency_label = "Recently updated"
                elif _age < 86400:
                    recency_label = "Updated today" if _was_updated else "Added today"
                elif _age < 7 * 86400:
                    recency_label = "Updated this week" if _was_updated else "Added this week"
                else:
                    recency_label = "Recently updated"
                happening_signals.append({
                    'person': person,
                    'action_line': action_line,
                    'friend_id': ft.user_id,
                    'recency_label': recency_label,
                    'trip_id': ft.id,
                    '_card_key': f"happening:{ft.id}",
                })
            _diag_hap_candidates = len(happening_signals)
        except Exception:
            db.session.rollback()

    if happening_signals:
        try:
            _dismissed_hap_cards = DismissedInsightCard.query.filter_by(
                user_id=user.id,
                card_type='happening',
            ).all()
            _dismissed_hap_keys = {d.card_key for d in _dismissed_hap_cards}
            if _dismissed_hap_keys:
                happening_signals = [
                    s for s in happening_signals
                    if s['_card_key'] not in _dismissed_hap_keys
                ]
        except Exception:
            db.session.rollback()

    if len(happening_signals) > HOME_HAPPENING_RENDER_CAP:
        happening_signals = happening_signals[:HOME_HAPPENING_RENDER_CAP]


    print(
        f"[HOME_DIAGNOSTICS] happening_friend_ids_count={len(friend_ids)}"
        f" happening_candidate_trips_before_dedupe={_diag_hap_raw_trips}"
        f" happening_group_count={_diag_hap_candidates}"
        f" happening_suppressed_by_opportunities={_diag_hap_opp_suppressed}"
        f" happening_candidate_trips_after_dedupe={max(0, _diag_hap_candidates - _diag_hap_opp_suppressed)}"
        f" happening_render_cap={HOME_HAPPENING_RENDER_CAP}"
        f" happening_after_cap={len(happening_signals)}"
        f" happening_rendered_count={len(happening_signals)}"
    )

    ideas_count = len(dest_feed)
    requests_count = banner_invite_count + (1 if secondary_card else 0)

    # _user_avail_home was fetched once before the coordination feed above.
    has_availability = bool(_user_avail_home)
    show_add_dates = not has_availability
    _home_avail_ranges, _home_avail_overflow = build_home_avail_ranges(_user_avail_home, cap=50)

    # Admin flag for Ideas diagnostic block
    _admin_emails_home = set(
        e.strip().lower()
        for e in os.environ.get("ALLOWED_ADMIN_EMAILS", "").split(",")
        if e.strip()
    )
    is_admin = current_user.is_authenticated and current_user.email.lower() in _admin_emails_home
    show_ideas_diagnostic = (
        is_admin
        and os.environ.get("SHOW_IDEAS_DIAGNOSTIC", "").strip().lower() in {"1", "true", "yes", "on"}
    )

    # Build Ideas diagnostic summary for admin view
    ideas_diag = {
        'raw_friend_trips': _ideas_engine_diag.get('raw_friend_trip', 0),
        'raw_overlap': _ideas_engine_diag.get('raw_overlap', 0),
        'raw_overlap_no_resort': _ideas_engine_diag.get('raw_overlap_no_resort', 0),
        'raw_wishlist': _ideas_engine_diag.get('raw_wishlist', 0),
        'booked_overlap_suppressed': _ideas_engine_diag.get('booked_overlap_suppressed', 0),
        'booked_allowed_existing': _ideas_engine_diag.get('booked_allowed_existing', 0),
        'dismissed': _diag_dismissed_count,
        'after_dismissal': _diag_opp_after_dismissal,
        'final_shown': len(dest_feed),
        'engine_total': _diag_opp_engine_count,
    }

    _hp_t0 = time.perf_counter()
    home_eq = user.get_active_equipment()
    if app.debug:
        print(f"[HOME_PERF] active_equipment={time.perf_counter()-_hp_t0:.4f}s")

    for _row in dest_feed:
        if _row.get('idea_type') == 'availability_overlap' and _row.get('anchor_friend_id'):
            _row['_url'] = url_for('friend_profile', friend_id=_row['anchor_friend_id'])
        elif _row.get('resort') and _row['resort'].slug:
            _row['_url'] = url_for('mountain_detail', slug=_row['resort'].slug)
        else:
            _row['_url'] = url_for('add_trip')

    if app.debug:
        print(f"[ROUTE_PERF] route=home total={time.perf_counter()-_rp_t0:.4f}s")
    _hp_t0 = time.perf_counter()
    _resp = render_template(
        'home.html',
        user=user,
        next_trip=next_trip,
        trip_invites=trip_invites,
        secondary_card=secondary_card,
        happening_signals=happening_signals,
        dest_feed=dest_feed,
        ideas_count=ideas_count,
        requests_count=requests_count,
        show_add_dates=show_add_dates,
        has_availability=has_availability,
        user_avail_ranges=_home_avail_ranges,
        user_avail_overflow=_home_avail_overflow,
        stat_mountains=user.visited_resorts_count,
        stat_trips_total=len(my_trips),
        stat_wishlist=len(user.wish_list_resorts or []),
        stat_trips_url=url_for('my_trips'),
        stat_mountains_url=url_for('mountains_visited'),
        stat_wishlist_url=url_for('settings_wish_list'),
        home_eq=home_eq,
        friend_count=len(friend_ids),
        new_connection_name=new_connection_name,
        sender_connection_card=sender_connection_card,
        is_admin=is_admin,
        show_ideas_diagnostic=show_ideas_diagnostic,
        ideas_diag=ideas_diag,
    )
    if app.debug:
        print(f"[HOME_PERF] render_template={time.perf_counter()-_hp_t0:.4f}s")
    return _resp


@app.route("/onboarding/equipment", methods=["POST"])
@login_required
def save_onboarding_equipment():
    """Save equipment status from progressive completion modal."""
    user = current_user
    
    equipment_status = request.form.get("equipment_status")
    equipment_brand = request.form.get("equipment_brand", "").strip() or None
    equipment_model = request.form.get("equipment_model", "").strip() or None
    boot_brand = request.form.get("boot_brand", "").strip() or None
    boot_model = request.form.get("boot_model", "").strip() or None
    
    if not equipment_status:
        return redirect(url_for("home"))
    
    # Get or create active equipment setup
    equipment = EquipmentSetup.query.filter_by(
        user_id=user.id,
        is_active=True
    ).first()
    
    if not equipment:
        equipment = EquipmentSetup(
            user_id=user.id,
            is_active=True
        )
        db.session.add(equipment)
    
    equipment.equipment_status = equipment_status
    equipment.brand = equipment_brand
    equipment.model = equipment_model
    equipment.boot_brand = boot_brand
    equipment.boot_model = boot_model

    # Mirror status to User so the home header reads it correctly
    if equipment_status in ('needs_rentals', 'have_own_equipment'):
        user.equipment_status = equipment_status

    db.session.commit()
    return redirect(url_for("home"))


@app.route("/onboarding/riding-style", methods=["POST"])
@login_required
def save_onboarding_riding_style():
    """Save terrain preferences from progressive completion modal."""
    user = current_user
    
    # Try hidden input first (populated by JS), then fall back to checkbox values
    terrain_raw = request.form.get("terrain_preferences", "")
    terrain_list = [t.strip() for t in terrain_raw.split(",") if t.strip()][:2]
    
    # Fallback: read directly from checkboxes if hidden input is empty
    if not terrain_list:
        terrain_list = request.form.getlist("terrain_pref")[:2]
    
    if terrain_list:
        user.terrain_preferences = terrain_list
        db.session.commit()
    
    return redirect(url_for("home"))


@app.route("/onboarding/welcome-shown", methods=["POST"])
@login_required
def mark_welcome_shown():
    """Mark the welcome modal as dismissed (once only)."""
    user = current_user
    
    if not user.welcome_modal_seen_at:
        user.welcome_modal_seen_at = datetime.utcnow()
        db.session.commit()
    
    # Support custom redirect destinations from welcome screen
    redirect_to = request.form.get("redirect_to")
    if redirect_to and redirect_to.startswith('/'):
        return redirect(redirect_to)
    
    return redirect(url_for("home"))


@app.route("/dismiss-nudge", methods=["POST"])
@login_required
def dismiss_nudge():
    """Dismiss an availability nudge so it doesn't resurface."""
    date_str = request.form.get("date")
    if date_str:
        try:
            nudge_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            # Store dismissal for this specific date
            existing = DismissedNudge.query.filter_by(
                user_id=current_user.id,
                date_range_start=nudge_date,
                date_range_end=nudge_date
            ).first()
            if not existing:
                dismissal = DismissedNudge(
                    user_id=current_user.id,
                    date_range_start=nudge_date,
                    date_range_end=nudge_date
                )
                db.session.add(dismissal)
                db.session.commit()
        except ValueError:
            pass
    return redirect(url_for("home"))


@app.route("/dismiss-insight-card", methods=["POST"])
@login_required
def dismiss_insight_card():
    """Persist a dismissal for a home insight card so it doesn't resurface."""
    card_type = request.form.get("card_type", "").strip()
    card_key = request.form.get("card_key", "").strip()
    if card_type and card_key:
        try:
            existing = DismissedInsightCard.query.filter_by(
                user_id=current_user.id,
                card_type=card_type,
                card_key=card_key,
            ).first()
            if not existing:
                dismissal = DismissedInsightCard(
                    user_id=current_user.id,
                    card_type=card_type,
                    card_key=card_key,
                )
                db.session.add(dismissal)
                db.session.commit()
        except Exception:
            db.session.rollback()
    return ('', 204)


@app.route("/friend-trip/<int:trip_id>")
@login_required
def friend_trip_details(trip_id):
    """View details of a friend's trip."""
    trip = SkiTrip.query.get_or_404(trip_id)
    friend = db.session.get(User, trip.user_id)

    # Prevent users from viewing trips of non-friends (unless it's their own)
    if trip.user_id != current_user.id:
        is_friend = Friend.query.filter_by(
            user_id=current_user.id,
            friend_id=friend.id
        ).first()

        if not is_friend:
            return render_template('403.html'), 403

    # Calculate overlapping days with current user's trips at the same resort
    your_trips = SkiTrip.query.filter_by(user_id=current_user.id).all()
    
    # Filter for same mountain/resort
    # Note: Resort ID is preferred if both trips have it, otherwise fallback to mountain string
    matching_trips = []
    for yt in your_trips:
        if trip.resort_id and yt.resort_id:
            if trip.resort_id == yt.resort_id:
                matching_trips.append(yt)
        elif yt.mountain == trip.mountain:
            matching_trips.append(yt)

    def overlapping_days(a_start, a_end, b_start, b_end):
        latest_start = max(a_start, b_start)
        earliest_end = min(a_end, b_end)
        delta = (earliest_end - latest_start).days + 1
        return max(0, delta)

    your_overlap_days = 0
    your_overlap_ranges = []

    for yt in matching_trips:
        days = overlapping_days(
            trip.start_date,
            trip.end_date,
            yt.start_date,
            yt.end_date
        )
        if days > 0:
            your_overlap_days += days
            your_overlap_ranges.append(yt)

    has_overlap = your_overlap_days > 0
    
    # PRIVACY: Do NOT query or display other friends' availability.
    # Non-owners should only see their own overlap with this trip.
    # The friends_open_on_trip feature has been removed to prevent
    # showing third-party friend data (e.g., "Richard · Epic") when
    # Jonathan views Charles's trip.

    # Gating for "Request to Join" - Final Production Rules
    # This logic is intentionally simplified for production UX. Do not reintroduce friendship or legacy invite gating.
    today = date.today()
    can_request_join = False
    has_pending_request = False
    is_accepted = False
    pending_request = None
    
    if trip.user_id != current_user.id:
        # 1. Check if viewer is already an accepted participant (CRITICAL: status=ACCEPTED only)
        is_accepted = SkiTripParticipant.query.filter_by(
            trip_id=trip_id, 
            user_id=current_user.id,
            status=GuestStatus.ACCEPTED
        ).first() is not None
        
        # 2. Check for pending join request (status=pending)
        pending_request = Invitation.query.filter_by(
            trip_id=trip_id,
            sender_id=current_user.id,
            invite_type=InviteType.REQUEST,
            status='pending'
        ).first()
        has_pending_request = pending_request is not None
        
        # 3. Trip must be in the future (has not ended)
        is_future = trip.end_date >= today
        
        # Final CTA State Decision:
        # - Already Accepted: Template shows "You're Going"
        # - Future + Not Accepted + No Pending Request: Template shows "Request to Join"
        # - Future + Not Accepted + Has Pending Request: Template shows "Request Sent"
        if is_future and not is_accepted:
            if not has_pending_request:
                can_request_join = True

    pending_request_id = pending_request.id if has_pending_request and pending_request else None

    return render_template(
        "friend_trip_details.html",
        trip=trip,
        friend=friend,
        has_overlap=has_overlap,
        overlap_days=your_overlap_days,
        your_overlap_ranges=your_overlap_ranges,
        friends_open_on_trip=[],  # Privacy protection
        can_request_join=can_request_join,
        has_pending_request=has_pending_request,
        pending_request_id=pending_request_id,
        is_accepted=is_accepted
    )

@app.route("/feedback", methods=["GET", "POST"])
@login_required
def feedback():
    success = False
    error = None
    
    if request.method == "POST":
        feedback_text = request.form.get("feedback_text", "").strip()
        
        if not feedback_text:
            error = "Please enter your feedback before submitting."
        else:
            admin_email = os.environ.get("ADMIN_FEEDBACK_EMAIL")
            sendgrid_api_key = os.environ.get("SENDGRID_API_KEY")
            
            if not admin_email:
                error = "Feedback system is not configured. Please try again later."
            elif not sendgrid_api_key:
                error = "Email service is not configured. Please try again later."
            else:
                try:
                    import sendgrid
                    from sendgrid.helpers.mail import Mail, Email, To, Content
                    
                    sg = sendgrid.SendGridAPIClient(api_key=sendgrid_api_key)
                    
                    from_email = Email("noreply@baselodgeapp.com")
                    to_email = To(admin_email)
                    subject = "New BaseLodge Feedback"
                    
                    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
                    body = f"""New feedback received from BaseLodge:

Feedback:
{feedback_text}

---
User Details:
- Name: {current_user.first_name} {current_user.last_name}
- Email: {current_user.email}
- User ID: {current_user.id}
- Submitted: {timestamp}
"""
                    content = Content("text/plain", body)
                    mail = Mail(from_email, to_email, subject, content)
                    
                    response = sg.client.mail.send.post(request_body=mail.get())
                    
                    if response.status_code in [200, 201, 202]:
                        success = True
                    else:
                        error = "Failed to send feedback. Please try again later."
                        
                except Exception as e:
                    app.logger.error(f"Feedback email error: {e}")
                    error = "Failed to send feedback. Please try again later."
    
    return render_template("feedback.html", success=success, error=error)

@app.route("/more")
@login_required
def more():
    # Legacy URL — skip the intermediate /settings hop and go straight to profile.
    return redirect(url_for('profile'))

@app.route("/profile")
@login_required
def profile():
    _rp_t0 = time.perf_counter()
    mountains_visited_count = current_user.visited_resorts_count

    # Get primary setup in a single query: is_primary rows first, then oldest by created_at.
    primary_equipment = (
        EquipmentSetup.query
        .filter_by(user_id=current_user.id)
        .order_by(
            db.case((EquipmentSetup.is_primary == True, 0), else_=1),
            EquipmentSetup.created_at.asc().nullsfirst(),
            EquipmentSetup.id.asc()
        )
        .first()
    )
    if primary_equipment and not primary_equipment.is_primary:
        primary_equipment.is_primary = True
        db.session.commit()

    has_equipment = primary_equipment is not None
    equipment_summary = ""
    if primary_equipment:
        parts = [primary_equipment.label] if primary_equipment.label else []
        if primary_equipment.brand:
            parts.append(primary_equipment.brand)
        if primary_equipment.model:
            parts.append(primary_equipment.model)
        binding_parts = [x for x in [primary_equipment.binding_brand, primary_equipment.binding_model] if x]
        if binding_parts:
            parts.append("Bindings: " + " ".join(binding_parts))
        equipment_summary = " · ".join(parts) if parts else "Setup saved"

    # Wish list data
    wish_list_ids = current_user.wish_list_resorts or []
    wish_list_count = len(wish_list_ids)
    wish_list_resorts = Resort.query.filter(Resort.id.in_(wish_list_ids)).all() if wish_list_ids else []

    upcoming_trips_count = get_upcoming_trip_count(current_user)

    if app.debug:
        print(f"[ROUTE_PERF] route=profile total={time.perf_counter()-_rp_t0:.4f}s")
    return render_template("profile.html",
                           page_title="Profile",
                           mountains_visited_count=mountains_visited_count,
                           has_equipment=has_equipment,
                           equipment_summary=equipment_summary,
                           wish_list_count=wish_list_count,
                           wish_list_resorts=wish_list_resorts,
                           upcoming_trips_count=upcoming_trips_count,
                           primary_equipment=primary_equipment)

@app.route("/notifications")
@login_required
def notifications():
    """Lightweight in-app notification center using Activity records + pending connect invites."""
    _rp_t0 = time.perf_counter()
    # --- Pending incoming connection requests (not Activity — from Invitation model) ---
    _t = time.perf_counter()
    pending_connects = []
    try:
        connects = Invitation.query.filter_by(
            receiver_id=current_user.id,
            status='pending'
        ).filter(Invitation.trip_id == None).order_by(Invitation.created_at.desc()).all()
        if connects:
            # Bulk-load all sender Users in one query instead of one per invite.
            sender_ids = list({inv.sender_id for inv in connects})
            senders_map = {
                u.id: u
                for u in User.query.filter(User.id.in_(sender_ids)).all()
            }
            for inv in connects:
                sender = senders_map.get(inv.sender_id)
                if sender:
                    sender_name = f"{sender.first_name or ''} {sender.last_name or ''}".strip() or 'Someone'
                    pending_connects.append({
                        'invitation_id': inv.id,
                        'sender_name': sender_name,
                        'sender_id': sender.id,
                    })
    except Exception:
        db.session.rollback()
    if app.debug:
        print(f"[ROUTE_PERF] notifications.pending_connects={time.perf_counter()-_t:.4f}s invite_count={len(pending_connects)}")

    # --- Activity-based notifications ---
    _t = time.perf_counter()
    raw_activities = []
    try:
        # joinedload(Activity.actor) fetches all actor Users in a single JOIN,
        # replacing the previous lazy per-activity SELECT when act.actor is accessed.
        raw_activities = Activity.query.filter(
            Activity.recipient_user_id == current_user.id,
            Activity.type.in_(_NOTIF_TYPES)
        ).options(
            joinedload(Activity.actor)
        ).order_by(Activity.created_at.desc()).limit(50).all()
    except Exception:
        db.session.rollback()
    if app.debug:
        print(f"[ROUTE_PERF] notifications.raw_activities={time.perf_counter()-_t:.4f}s count={len(raw_activities)}")

    _t = time.perf_counter()
    notifs = []
    today = datetime.utcnow()
    for act in raw_activities:
        actor = act.actor
        actor_first = (actor.first_name or 'Someone') if actor else 'Someone'
        actor_name = f"{actor.first_name or ''} {actor.last_name or ''}".strip() if actor else 'Someone'

        trip = act.get_trip() if act.object_type == 'trip' else None
        trip_name = None
        if trip:
            if trip.resort:
                trip_name = trip.resort.name
            else:
                trip_name = trip.mountain or 'a trip'

        action_url = None
        if act.type == 'join_request_received':
            text = f"{actor_name} wants to join your trip{' to ' + trip_name if trip_name else ''}"
            action_url = url_for('trip_detail', trip_id=act.object_id) if act.object_type == 'trip' else None
        elif act.type == 'join_request_accepted':
            text = f"Your request to join {trip_name or 'the trip'} was accepted"
            action_url = url_for('trip_detail', trip_id=act.object_id) if act.object_type == 'trip' else None
        elif act.type == 'join_request_declined':
            text = f"Your request to join {trip_name or 'the trip'} was declined"
            action_url = None
        elif act.type == 'trip_invite_received':
            text = f"{actor_first} invited you to {trip_name or 'a trip'}"
            action_url = url_for('trip_detail', trip_id=act.object_id) if act.object_type == 'trip' else None
        elif act.type == 'trip_invite_accepted':
            text = f"{actor_first} accepted your invite to {trip_name or 'a trip'}"
            action_url = url_for('trip_detail', trip_id=act.object_id) if act.object_type == 'trip' else None
        elif act.type == 'trip_invite_declined':
            text = f"{actor_first} declined your invite to {trip_name or 'a trip'}"
            action_url = None
        elif act.type == 'connection_accepted':
            text = f"{actor_first} accepted your connection request"
            action_url = url_for('friend_profile', friend_id=act.actor_user_id) if actor else None
        elif act.type == 'trip_location_changed':
            extra = act.extra_data or {}
            changed_resort = extra.get('resort_name') or trip_name or 'a new location'
            text = f"Trip location changed to {changed_resort}"
            action_url = url_for('trip_detail', trip_id=act.object_id) if act.object_type == 'trip' else None
        elif act.type == 'trip_pass_changed':
            extra = act.extra_data or {}
            changed_pass = extra.get('pass_display') or 'a new pass'
            text = f"Trip pass updated to {changed_pass}"
            action_url = url_for('trip_detail', trip_id=act.object_id) if act.object_type == 'trip' else None
        else:
            text = f"Update from {actor_first}"
            action_url = None

        # Relative time
        delta = today - act.created_at
        if delta.days == 0:
            if delta.seconds < 60:
                rel_time = "just now"
            elif delta.seconds < 3600:
                rel_time = f"{delta.seconds // 60}m ago"
            else:
                rel_time = f"{delta.seconds // 3600}h ago"
        elif delta.days == 1:
            rel_time = "yesterday"
        elif delta.days < 7:
            rel_time = f"{delta.days}d ago"
        else:
            rel_time = act.created_at.strftime("%-d %b")

        notifs.append({
            'text': text,
            'action_url': action_url,
            'rel_time': rel_time,
            'type': act.type,
        })

    if app.debug:
        print(f"[ROUTE_PERF] notifications.activity_loop={time.perf_counter()-_t:.4f}s activity_count={len(raw_activities)}")
    # Mark all as viewed
    session['notif_last_viewed_at'] = datetime.utcnow().isoformat()

    if app.debug:
        print(f"[ROUTE_PERF] route=notifications total={time.perf_counter()-_rp_t0:.4f}s")
    return render_template('notifications.html',
                           pending_connects=pending_connects,
                           notifs=notifs)


@app.route("/settings")
@login_required
def settings():
    return redirect(url_for("profile"), code=301)


@app.route("/settings/profile", methods=["GET", "POST"])
@login_required
def settings_profile():
    return redirect(url_for('edit_profile'))


@app.route("/settings/equipment")
@login_required
def settings_equipment():
    # Fetch all setups for this user, ordered by creation time
    all_setups = EquipmentSetup.query.filter_by(user_id=current_user.id).order_by(
        EquipmentSetup.created_at.asc().nullsfirst(), EquipmentSetup.id.asc()
    ).all()

    # Ensure at most one is_primary — repair if data is inconsistent
    primary_setups = [s for s in all_setups if s.is_primary]
    if len(primary_setups) > 1:
        # Keep the first, unset the rest
        for extra in primary_setups[1:]:
            extra.is_primary = False
        db.session.commit()
        all_setups = EquipmentSetup.query.filter_by(user_id=current_user.id).order_by(
            EquipmentSetup.created_at.asc().nullsfirst(), EquipmentSetup.id.asc()
        ).all()

    return render_template("settings_equipment.html",
                           all_setups=all_setups,
                           user=current_user,
                           rider_types=current_user.rider_types or [],
                           ski_brands=SKI_BRANDS,
                           board_brands=SNOWBOARD_BRANDS,
                           boot_brands=BOOT_BRANDS,
                           binding_types=BINDING_TYPES,
                           binding_brands_by_type=BINDING_BRANDS_BY_TYPE)


@app.route("/settings/equipment/save", methods=["POST"])
@login_required
def settings_equipment_save():
    """Create or update a single equipment setup."""
    setup_id = request.form.get("setup_id", "")
    discipline_str = request.form.get("discipline", "")
    label = request.form.get("label", "").strip() or None
    brand = request.form.get("brand", "").strip()
    model_val = request.form.get("model", "").strip()
    length_cm = request.form.get("length_cm", "").strip()
    width_mm = request.form.get("width_mm", "").strip()
    binding_type = request.form.get("binding_type", "").strip()
    binding_brand = request.form.get("binding_brand", "").strip()
    binding_model = request.form.get("binding_model", "").strip()
    boot_brand = request.form.get("boot_brand", "").strip()
    boot_model = request.form.get("boot_model", "").strip()
    boot_flex = request.form.get("boot_flex", "").strip()
    purchase_year = request.form.get("purchase_year", "").strip()

    if not discipline_str:
        return jsonify({"error": "Discipline required"}), 400

    discipline = EquipmentDiscipline.SKIER if discipline_str == "Skier" else EquipmentDiscipline.SNOWBOARDER

    existing_setups = EquipmentSetup.query.filter_by(user_id=current_user.id).order_by(
        EquipmentSetup.created_at.asc().nullsfirst(), EquipmentSetup.id.asc()
    ).all()

    is_first = len(existing_setups) == 0

    if setup_id:
        # Update existing
        equipment = EquipmentSetup.query.filter_by(id=int(setup_id), user_id=current_user.id).first()
        if not equipment:
            return jsonify({"error": "Setup not found"}), 404
        old_discipline = equipment.discipline
        equipment.discipline = discipline
        if old_discipline != discipline:
            equipment.binding_type = None
    else:
        # Create new
        equipment = EquipmentSetup(
            user_id=current_user.id,
            discipline=discipline,
            created_at=datetime.utcnow(),
            is_primary=False
        )
        db.session.add(equipment)
        is_first = True  # newly created — will become primary if no primary exists

    equipment.label = label
    equipment.brand = brand if brand else None
    equipment.model = model_val if model_val else None
    equipment.length_cm = int(length_cm) if length_cm else None
    equipment.width_mm = int(width_mm) if width_mm else None
    equipment.binding_type = binding_type if binding_type else None
    equipment.binding_brand = binding_brand if binding_brand else None
    equipment.binding_model = binding_model if binding_model else None
    equipment.boot_brand = boot_brand if boot_brand else None
    equipment.boot_model = boot_model if boot_model else None
    equipment.boot_flex = int(boot_flex) if boot_flex and boot_flex.isdigit() and int(boot_flex) > 0 else None
    equipment.purchase_year = int(purchase_year) if purchase_year and purchase_year.isdigit() else None

    db.session.flush()  # get id if new

    # If no primary exists, make this one primary
    has_primary = any(s.is_primary for s in existing_setups if s.id != equipment.id)
    if not has_primary:
        equipment.is_primary = True

    # Also keep User.equipment_status consistent
    current_user.equipment_status = "have_own_equipment"

    db.session.commit()
    return jsonify({"success": True, "setup_id": equipment.id, "is_primary": equipment.is_primary})


@app.route("/settings/equipment/get/<int:setup_id>")
@login_required
def settings_equipment_get(setup_id):
    """Return JSON data for one equipment setup (used to pre-populate the edit form)."""
    equipment = EquipmentSetup.query.filter_by(id=setup_id, user_id=current_user.id).first()
    if not equipment:
        return jsonify({"error": "Not found"}), 404
    return jsonify({
        "id": equipment.id,
        "label": equipment.label,
        "discipline": equipment.discipline.value if equipment.discipline else "skier",
        "brand": equipment.brand,
        "model": equipment.model,
        "length_cm": equipment.length_cm,
        "width_mm": equipment.width_mm,
        "binding_type": equipment.binding_type,
        "binding_brand": equipment.binding_brand,
        "binding_model": equipment.binding_model,
        "boot_brand": equipment.boot_brand,
        "boot_model": equipment.boot_model,
        "boot_flex": equipment.boot_flex,
        "purchase_year": equipment.purchase_year,
        "is_primary": equipment.is_primary,
    })


@app.route("/settings/equipment/make-primary", methods=["POST"])
@login_required
def settings_equipment_make_primary():
    """Set a specific setup as primary, unsetting all others for this user."""
    setup_id = request.form.get("setup_id", "")
    if not setup_id:
        return jsonify({"error": "setup_id required"}), 400

    equipment = EquipmentSetup.query.filter_by(id=int(setup_id), user_id=current_user.id).first()
    if not equipment:
        return jsonify({"error": "Setup not found"}), 404

    # Unset all primaries for this user
    EquipmentSetup.query.filter_by(user_id=current_user.id, is_primary=True).update({"is_primary": False})
    equipment.is_primary = True
    db.session.commit()
    return jsonify({"success": True})


@app.route("/settings/equipment/delete", methods=["POST"])
@login_required
def settings_equipment_delete():
    """Delete a setup by id. If it was primary and others exist, promote the next one."""
    setup_id = request.form.get("setup_id", "")

    # Legacy fallback: accept slot= param for old callers
    if not setup_id:
        slot_str = request.form.get("slot", "primary")
        slot = EquipmentSlot.PRIMARY if slot_str == "primary" else EquipmentSlot.SECONDARY
        equipment = EquipmentSetup.query.filter_by(user_id=current_user.id, slot=slot).first()
    else:
        equipment = EquipmentSetup.query.filter_by(id=int(setup_id), user_id=current_user.id).first()

    if not equipment:
        return jsonify({"success": True})

    was_primary = equipment.is_primary
    db.session.delete(equipment)
    db.session.flush()

    if was_primary:
        # Promote the next oldest setup
        next_setup = EquipmentSetup.query.filter_by(user_id=current_user.id).order_by(
            EquipmentSetup.created_at.asc().nullsfirst(), EquipmentSetup.id.asc()
        ).first()
        if next_setup:
            next_setup.is_primary = True

    db.session.commit()
    return jsonify({"success": True})


@app.route("/settings/equipment-status", methods=["POST"])
@login_required
def settings_equipment_status():
    """Update user's high-level equipment_status (have_own_equipment or needs_rentals)."""
    status = request.form.get("equipment_status", "have_own_equipment")
    if status not in ["have_own_equipment", "needs_rentals"]:
        return jsonify({"success": False, "error": "Invalid status"}), 400

    current_user.equipment_status = status
    db.session.commit()
    return jsonify({"success": True})


@app.route("/settings/profile/add-rider-type", methods=["POST"])
@login_required
def settings_profile_add_rider_type():
    """Append a single rider type to the current user's profile (idempotent, no duplicates)."""
    rider_type = request.form.get("rider_type", "").strip()
    if not rider_type or rider_type not in RIDER_TYPES:
        return jsonify({"success": False, "error": "Invalid rider type"}), 400
    existing = list(current_user.rider_types or [])
    if rider_type not in existing:
        existing.append(rider_type)
        current_user.rider_types = existing
        db.session.commit()
    return jsonify({"success": True, "rider_types": list(current_user.rider_types)})


# ── Mountain Detail Page ─────────────────────────────────────────────────────

@app.route("/mountain/<slug>")
@login_required
def mountain_detail(slug):
    """
    Focused detail page for a single resort:
    - Hero: state, name, primary pass pill
    - Body: upcoming trips grouped by date window, showing current user + friends
    """
    _rp_t0 = time.perf_counter()
    resort = Resort.query.filter_by(slug=slug, is_active=True).first_or_404()
    today = date.today()

    # Current user's friend IDs (bidirectional friendship model: Friend row = user_id → friend_id)
    friend_ids = set(
        f.friend_id for f in Friend.query.filter_by(user_id=current_user.id).all()
    )
    allowed_ids = friend_ids | {current_user.id}

    # Upcoming + in-progress trips for this resort only (resort_id canonical; no string fallback)
    # is_public=True: private trips must not appear in friend social proof rows.
    raw_trips = (
        SkiTrip.query
        .filter(
            SkiTrip.resort_id == resort.id,
            SkiTrip.end_date >= today,
            SkiTrip.is_public == True,
        )
        .order_by(SkiTrip.start_date.asc())
        .all()
    )

    # Accumulate rows per date window.
    # Key: (start_date, end_date); value: dict of uid -> row_dict (deduped within window)
    _t = time.perf_counter()
    rows_by_window = {}

    # Bulk-load all accepted participants for raw_trips — one query replaces N×2
    # calls to trip.get_accepted_participants() across Pass 1 and Pass 2.
    _raw_trip_ids = [t.id for t in raw_trips]
    _bulk_accepted_uids: dict = {}  # trip_id -> [user_ids]
    if _raw_trip_ids:
        _part_rows = SkiTripParticipant.query.filter(
            SkiTripParticipant.trip_id.in_(_raw_trip_ids),
            SkiTripParticipant.status == GuestStatus.ACCEPTED,
        ).all()
        for _pr in _part_rows:
            _bulk_accepted_uids.setdefault(_pr.trip_id, []).append(_pr.user_id)

    # Pass 1: collect every user ID that will be needed across all trip windows.
    # This lets us do one bulk User query instead of one query per person.
    all_needed_uids = set()
    for trip in raw_trips:
        if get_trip_status(trip, today=today) == 'past':
            continue
        people_ids = {trip.user_id}
        for uid in _bulk_accepted_uids.get(trip.id, []):
            people_ids.add(uid)
        all_needed_uids.update(people_ids & allowed_ids)

    # One bulk query replacing the previous per-uid db.session.get() calls.
    users_by_id = {}
    if all_needed_uids:
        users_by_id = {
            u.id: u
            for u in User.query.filter(User.id.in_(all_needed_uids)).all()
        }

    # Pass 2: build the window rows using the pre-loaded users_by_id dict.
    for trip in raw_trips:
        # Canonical validation: skip any trip get_trip_status classifies as past
        if get_trip_status(trip, today=today) == 'past':
            continue

        # Owner always counts as accepted; also gather explicit accepted participants
        people_ids = {trip.user_id}
        for uid in _bulk_accepted_uids.get(trip.id, []):
            people_ids.add(uid)

        relevant_ids = people_ids & allowed_ids
        if not relevant_ids:
            continue

        window = (trip.start_date, trip.end_date)
        if window not in rows_by_window:
            rows_by_window[window] = {}

        status_label = "Going" if (trip.trip_status == 'going') else "Considering"

        for uid in relevant_ids:
            if uid in rows_by_window[window]:
                # If we see the person again in a different trip for the same window,
                # upgrade status to 'Going' if warranted; otherwise leave as-is.
                if status_label == 'Going':
                    rows_by_window[window][uid]['status_label'] = 'Going'
                continue

            person = users_by_id.get(uid)
            if not person:
                continue

            is_me = (uid == current_user.id)
            rows_by_window[window][uid] = {
                'user_id': uid,
                'is_me': is_me,
                'full_name': f"{(person.first_name or '').strip()} {(person.last_name or '').strip()}".strip(),
                'identity_line': _mountain_row_identity(
                    person.display_rider_type,
                    person.skill_level,
                    person.pass_type,
                ),
                'status_label': status_label,
            }

    if app.debug:
        print(f"[ROUTE_PERF] mountain_detail.window_loop={time.perf_counter()-_t:.4f}s trip_count={len(raw_trips)} users_prefetched={len(users_by_id)}")
    # Build flat sorted rows: chronological by start_date, then alphabetical within same window.
    # "You" sorts as 'You' (Y) — no special priority in the flat layout.
    flat_rows = []
    for (start_date, end_date), rows_map in sorted(rows_by_window.items(), key=lambda x: x[0][0]):
        for row in sorted(rows_map.values(), key=lambda r: 'You' if r['is_me'] else r['full_name']):
            flat_row = dict(row)
            flat_row['start_date'] = start_date
            flat_row['end_date'] = end_date
            flat_rows.append(flat_row)

    primary_pass = resort.get_primary_pass()
    pass_names = resort.get_pass_names()

    # Full state name: STATE_NAMES (abbr→name) first for US states,
    # then fall back to state_name column (may hold full name for non-US resorts)
    state_full = ''
    if resort.state_code:
        state_full = STATE_NAMES.get(resort.state_code, '')
    if not state_full:
        state_full = (resort.state_name or '').strip()
    if not state_full and resort.state_code:
        state_full = resort.state_code

    # Social context counts — friends only (current user excluded), deduped by user_id
    _seen_going = set()
    _seen_considering = set()
    for row in flat_rows:
        if row['is_me']:
            continue
        uid = row['user_id']
        if row['status_label'] == 'Going':
            _seen_going.add(uid)
        else:
            _seen_considering.add(uid)
    social_going = len(_seen_going)
    social_considering = len(_seen_considering)

    # Personal signals: pass coverage + wishlist membership
    import re as _re
    user_passes = [p.strip() for p in (current_user.pass_type or "").split(",") if p.strip()]
    user_pass_covered = False
    user_pass_name = ""
    for _up in user_passes:
        if _up in pass_names:
            user_pass_covered = True
            user_pass_name = _re.sub(r'\s+[Pp]ass$', '', _up).strip()
            break
    is_on_wishlist = resort.id in (current_user.wish_list_resorts or [])

    # Friends who also have this resort on their wishlist
    wishlist_friends = []
    if friend_ids:
        _wl_candidates = User.query.filter(User.id.in_(friend_ids)).all()
        wishlist_friends = [
            u for u in _wl_candidates
            if resort.id in (u.wish_list_resorts or [])
        ]
        wishlist_friends.sort(key=lambda u: ((u.first_name or '').lower(), (u.last_name or '').lower()))

    if app.debug:
        print(f"[ROUTE_PERF] route=mountain_detail total={time.perf_counter()-_rp_t0:.4f}s")
    return render_template(
        'mountain_detail.html',
        resort=resort,
        primary_pass=primary_pass,
        pass_names=pass_names,
        flat_rows=flat_rows,
        state_full=state_full,
        social_going=social_going,
        social_considering=social_considering,
        user_pass_covered=user_pass_covered,
        user_pass_name=user_pass_name,
        is_on_wishlist=is_on_wishlist,
        wishlist_friends=wishlist_friends,
    )


# ── Mountain page-view tracking ───────────────────────────────────────────────
@app.route("/api/mountain/track-view", methods=["POST"])
def track_mountain_view():
    """Fire-and-forget mountain page-view tracker. Never raises to caller."""
    try:
        data      = request.get_json(silent=True) or {}
        resort_id = data.get("resort_id")
        if not resort_id:
            return jsonify({"ok": False, "reason": "missing resort_id"}), 200

        user_id     = current_user.id if current_user.is_authenticated else None
        # session_key is a client-generated UUID sent in the payload (sessionStorage-based)
        raw_sid     = data.get("session_key") or ""
        session_key = str(raw_sid)[:64].strip() or None

        # Dedup: refresh-spam protection only — same session + same resort within 30 seconds
        if session_key:
            spam_cutoff = datetime.utcnow() - timedelta(seconds=30)
            already = db.session.query(MountainPageView.id).filter(
                MountainPageView.resort_id  == resort_id,
                MountainPageView.session_key == session_key,
                MountainPageView.viewed_at  >= spam_cutoff,
            ).first()
            if already:
                return jsonify({"ok": True, "deduped": True})

        db.session.add(MountainPageView(
            resort_id=resort_id,
            user_id=user_id,
            session_key=session_key,
        ))
        db.session.commit()
        return jsonify({"ok": True})
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({"ok": False}), 200  # never fail the page load


@app.route("/settings/mountains-visited")
@login_required
def settings_mountains():
    return redirect(url_for('mountains_visited'))


@app.route("/push-settings")
@login_required
def push_settings():
    return render_template("push_settings.html")


@app.route("/settings/password")
@login_required
def settings_password():
    return redirect(url_for('change_password'))


@app.route("/settings/wish-list")
@login_required
def settings_wish_list():
    # Get all resorts for selection, ordered by country, state, name (exclude regions)
    resorts = Resort.query.filter_by(is_active=True, is_region=False).order_by(
        Resort.country_code, Resort.state_code, Resort.name
    ).all()
    
    # Get current wish list resort IDs
    wish_list_ids = current_user.wish_list_resorts or []
    
    # Get full resort objects for display
    wish_list_resorts = Resort.query.filter(Resort.id.in_(wish_list_ids)).all() if wish_list_ids else []
    
    # Group and sort wish list resorts for display
    grouped_wish_list = group_resorts_for_display(wish_list_resorts)
    
    # Get canonical countries for dropdown from the shared COUNTRIES mapping
    # Safeguard: Dropdown options MUST come from canonical country list, not resort data
    from utils.countries import COUNTRIES
    countries_list = sorted(COUNTRIES.keys(), key=lambda c: (c != 'US', COUNTRIES[c]))
    
    return render_template("settings_wish_list.html",
                           resorts=resorts,
                           wish_list_resorts=wish_list_resorts,
                           grouped_wish_list=grouped_wish_list,
                           wish_list_ids=wish_list_ids,
                           countries=countries_list)


@app.route("/settings/wish-list/save", methods=["POST"])
@login_required
def settings_wish_list_save():
    data = request.get_json()
    resort_ids = data.get("resort_ids", [])
    
    # Enforce max of 15
    if len(resort_ids) > 15:
        return jsonify({"error": "Maximum 15 resorts allowed"}), 400
    
    # Validate resort IDs exist
    valid_ids = []
    for rid in resort_ids:
        resort = db.session.get(Resort, rid)
        if resort:
            valid_ids.append(rid)
    
    _old_wl = list(current_user.wish_list_resorts or [])
    current_user.wish_list_resorts = valid_ids
    db.session.commit()
    if len(valid_ids) > len(_old_wl):
        ph_analytics.track(current_user.id, 'wishlist_added', {
            'added_count': len(valid_ids) - len(_old_wl),
            'total_count': len(valid_ids),
            'source':      'settings',
        })
    return jsonify({"success": True, "count": len(valid_ids)})


# =====================================================================
# INSTANT-SAVE API — Mountains Visited
# =====================================================================

@app.route("/api/mountains-visited/add", methods=["POST"])
@login_required
def api_mountains_visited_add():
    data = request.get_json() or {}
    resort_id = data.get("resort_id")
    if not resort_id:
        return jsonify({"error": "resort_id required"}), 400
    resort = db.session.get(Resort, resort_id)
    if not resort:
        return jsonify({"error": "Resort not found"}), 404
    ids = list(current_user.visited_resort_ids or [])
    if resort_id not in ids:
        ids.append(resort_id)
        names = list(current_user.mountains_visited or [])
        if resort.name not in names:
            names.append(resort.name)
        current_user.visited_resort_ids = ids
        current_user.mountains_visited = names
        db.session.commit()
    return jsonify({"success": True, "count": len(ids)})


@app.route("/api/mountains-visited/remove", methods=["POST"])
@login_required
def api_mountains_visited_remove():
    data = request.get_json() or {}
    resort_id = data.get("resort_id")
    if not resort_id:
        return jsonify({"error": "resort_id required"}), 400
    resort = db.session.get(Resort, resort_id)
    resort_name = resort.name if resort else None
    ids = [i for i in (current_user.visited_resort_ids or []) if i != resort_id]
    names = list(current_user.mountains_visited or [])
    if resort_name and resort_name in names:
        names.remove(resort_name)
    current_user.visited_resort_ids = ids
    current_user.mountains_visited = names
    db.session.commit()
    return jsonify({"success": True, "count": len(ids)})


# =====================================================================
# INSTANT-SAVE API — Wishlist
# =====================================================================

@app.route("/api/wishlist/add", methods=["POST"])
@login_required
def api_wishlist_add():
    data = request.get_json() or {}
    resort_id = data.get("resort_id")
    if not resort_id:
        return jsonify({"error": "resort_id required"}), 400
    resort = db.session.get(Resort, resort_id)
    if not resort:
        return jsonify({"error": "Resort not found"}), 404
    ids = list(current_user.wish_list_resorts or [])
    if len(ids) >= 15:
        return jsonify({"error": "Maximum 15 resorts", "at_limit": True}), 200
    if resort_id not in ids:
        ids.append(resort_id)
        current_user.wish_list_resorts = ids
        db.session.commit()
        ph_analytics.track(current_user.id, 'wishlist_added', {
            'resort_id':   resort_id,
            'total_count': len(ids),
            'source':      'mountain_page',
        })
    return jsonify({"success": True, "count": len(ids), "at_limit": len(ids) >= 15})


@app.route("/api/wishlist/remove", methods=["POST"])
@login_required
def api_wishlist_remove():
    data = request.get_json() or {}
    resort_id = data.get("resort_id")
    if not resort_id:
        return jsonify({"error": "resort_id required"}), 400
    ids = [i for i in (current_user.wish_list_resorts or []) if i != resort_id]
    current_user.wish_list_resorts = ids
    db.session.commit()
    return jsonify({"success": True, "count": len(ids), "at_limit": len(ids) >= 3})


# =====================================================================
# FRIEND READ-ONLY VIEWS — Mountains Visited + Wishlist
# =====================================================================

@app.route("/mountains-visited/<int:user_id>")
@login_required
def friend_mountains_visited(user_id):
    """Read-only view of another user's mountains visited (friends only)."""
    friend = User.query.get_or_404(user_id)
    # Must be a confirmed friend
    is_friend = Friend.query.filter_by(
        user_id=current_user.id, friend_id=user_id
    ).first() is not None
    if not is_friend:
        abort(403)

    # Friend's visited resorts
    visited_ids = friend.visited_resort_ids or []
    visited_resorts = Resort.query.filter(Resort.id.in_(visited_ids)).all() if visited_ids else []
    grouped = group_resorts_for_display(visited_resorts)

    # Current user's wishlist for "On your wishlist" indicator
    user_wishlist_ids = set(current_user.wish_list_resorts or [])

    return render_template(
        "mountains_visited.html",
        read_only=True,
        view_user=friend,
        grouped_selected=grouped,
        mountains_visited_count=len(visited_ids),
        user_wishlist_ids=user_wishlist_ids,
        # own-view fields not needed in read-only — set safe defaults
        resorts=[],
        selected_resort_ids=[],
        countries=[],
        COUNTRIES={},
    )


@app.route("/wishlist/<int:user_id>")
@login_required
def friend_wishlist(user_id):
    """Read-only view of another user's wishlist (friends only)."""
    friend = User.query.get_or_404(user_id)
    is_friend = Friend.query.filter_by(
        user_id=current_user.id, friend_id=user_id
    ).first() is not None
    if not is_friend:
        abort(403)

    # Friend's wishlist
    wish_ids = friend.wish_list_resorts or []
    wish_resorts = Resort.query.filter(Resort.id.in_(wish_ids)).all() if wish_ids else []
    grouped = group_resorts_for_display(wish_resorts)

    # Current user's visited for "You've been here" indicator
    user_visited_ids = set(current_user.visited_resort_ids or [])

    return render_template(
        "settings_wish_list.html",
        read_only=True,
        view_user=friend,
        grouped_wish_list=grouped,
        wish_list_ids=wish_ids,
        user_visited_ids=user_visited_ids,
        # own-view fields not needed in read-only
        resorts=[],
        countries=[],
        COUNTRIES={},
    )


# =====================================================================
# SHARED RESORT FILTERING API (for Wishlist & Mountains Visited only)
# =====================================================================

@app.route("/api/resort-countries")
@login_required
def api_resort_countries():
    """Get all countries that have resorts in the database."""
    countries = db.session.query(Resort.country_code).distinct().filter(
        Resort.country_code.isnot(None),
        Resort.country_code != ''
    ).all()
    
    # Map country codes to display names
    country_names = {
        'US': 'United States',
        'CA': 'Canada',
        'JP': 'Japan',
        'FR': 'France',
        'CH': 'Switzerland',
        'AT': 'Austria',
        'IT': 'Italy'
    }
    
    result = []
    for (code,) in countries:
        if code:
            result.append({
                'code': code,
                'name': country_names.get(code, code)
            })
    
    # Sort by name, with US and CA first
    def sort_key(c):
        if c['code'] == 'US':
            return (0, c['name'])
        elif c['code'] == 'CA':
            return (1, c['name'])
        else:
            return (2, c['name'])
    
    result.sort(key=sort_key)
    return jsonify(result)


@app.route("/api/resort-regions/<country_code>")
@login_required
def api_resort_regions(country_code):
    """Get all regions/states for a specific country."""
    regions = db.session.query(Resort.state_code).distinct().filter(
        Resort.country_code == country_code.upper(),
        Resort.state_code.isnot(None),
        Resort.state_code != ''
    ).all()
    
    result = sorted([r[0] for r in regions if r[0]])
    return jsonify(result)


@app.route("/add-open-dates", methods=["GET", "POST"])
@login_required
def add_open_dates():
    if request.method == "POST":
        # Get selected dates from form (comma-separated YYYY-MM-DD strings)
        selected_dates = request.form.get("selected_dates", "")
        
        if selected_dates:
            dates_list = [d.strip() for d in selected_dates.split(",") if d.strip()]
            # Validate and filter dates
            valid_dates = []
            today_str = date.today().strftime('%Y-%m-%d')
            for d in dates_list:
                try:
                    datetime.strptime(d, '%Y-%m-%d')
                    if d >= today_str:
                        valid_dates.append(d)
                except ValueError:
                    pass
            
            _prev_open_dates = list(current_user.open_dates or [])
            current_user.open_dates = sorted(set(valid_dates))
        else:
            _prev_open_dates = list(current_user.open_dates or [])
            current_user.open_dates = []
        
        db.session.commit()
        if valid_dates:
            ph_analytics.track(current_user.id, 'availability_added', {
                'date_count':    len(valid_dates),
                'is_first_time': not bool(_prev_open_dates),
            })
        
        # Recompute availability overlap activities for this user
        emit_availability_overlap_activities_for_user(current_user)
        db.session.commit()
        
        return redirect(url_for("home"))
    
    # Pre-populate with existing dates
    existing_dates = current_user.open_dates or []

    from services.open_dates import get_available_dates_for_user as _get_avail_od
    _avail_set = _get_avail_od(current_user)
    _avail_ranges, _avail_overflow = build_home_avail_ranges(_avail_set, cap=50)

    return render_template(
        "add_open_dates.html",
        user=current_user,
        existing_dates=existing_dates,
        has_availability=bool(_avail_set),
        user_avail_ranges=_avail_ranges,
        user_avail_overflow=_avail_overflow,
    )

@app.route("/add_trip", methods=["GET", "POST"])
@login_required
def add_trip():
    # Single source of truth: Resort table
    _rp_t0 = time.perf_counter()
    _t = time.perf_counter()
    resorts = get_resorts_for_trip_form()
    _t_resorts = time.perf_counter() - _t
    
    countries_map = COUNTRIES
    states_map = STATE_ABBR_MAP

    user_passes = [p.strip() for p in (current_user.pass_type or "").split(",") if p.strip()]
    
    # Get prefill parameters for "Propose a trip" flow
    def safe_int(val):
        if val in (None, "", "null", "None"):
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None

    raw_friend_id = request.args.get('friend_id')
    raw_resort_id = request.args.get('resort_id')
    
    prefill_friend_id = safe_int(raw_friend_id)
    prefill_start_date = request.args.get('start_date') or None
    prefill_end_date = request.args.get('end_date') or None
    prefill_resort_id = safe_int(raw_resort_id)
    is_group = request.args.get('is_group') == '1'
    
    # Clean up empty strings from query params
    if prefill_start_date == "": prefill_start_date = None
    if prefill_end_date == "": prefill_end_date = None
    
    prefill_friend = None
    if prefill_friend_id is not None:
        prefill_friend = db.session.get(User, prefill_friend_id)

    prefill_resort = None
    if prefill_resort_id is not None:
        prefill_resort = db.session.get(Resort, prefill_resort_id)

    if request.method == "POST":
        resort_id = request.form.get("resort_id")
        start_date_str = request.form.get("start_date")
        end_date_str = request.form.get("end_date")
        is_public = request.form.get("is_public") == "on"
        ride_intent = request.form.get("ride_intent") or None
        trip_equipment_status = request.form.get("trip_equipment_status") or "use_default"
        trip_status_raw = request.form.get("trip_status", "planning")
        trip_status_form = trip_status_raw if trip_status_raw in ("planning", "going") else "planning"
        
        friend_id = request.form.get("friend_id", type=int)
        is_group_trip = request.form.get("is_group") == "1"

        # ── Multi-date batch path ──────────────────────────────────────────
        _date_ranges_json = request.form.get("date_ranges_json")
        if _date_ranges_json and not is_group_trip and not friend_id:
            import uuid as _uuid_mod
            _resort = db.session.get(Resort, resort_id) if resort_id else None
            if not _resort:
                flash("Please select a valid resort.", "error")
                return render_template(
                    "add_trip.html", trip=None, resorts=resorts,
                    countries_map=countries_map, states_map=states_map,
                    user=current_user, form_action=url_for("add_trip"),
                    user_passes=user_passes, prefill_friend=prefill_friend,
                    prefill_start_date=prefill_start_date,
                    prefill_end_date=prefill_end_date,
                    prefill_resort=prefill_resort, is_group=is_group,
                )
            try:
                _ranges = json.loads(_date_ranges_json)
            except (ValueError, TypeError):
                _ranges = []
            if not _ranges:
                flash("Please add at least one date range.", "error")
                return render_template(
                    "add_trip.html", trip=None, resorts=resorts,
                    countries_map=countries_map, states_map=states_map,
                    user=current_user, form_action=url_for("add_trip"),
                    user_passes=user_passes, prefill_friend=prefill_friend,
                    prefill_start_date=prefill_start_date,
                    prefill_end_date=prefill_end_date,
                    prefill_resort=prefill_resort, is_group=is_group,
                )
            _today = date.today()
            _parsed = []
            _batch_errors = []
            for _r in _ranges:
                try:
                    _s = datetime.strptime(_r['start_date'], "%Y-%m-%d").date()
                    _e = datetime.strptime(_r['end_date'], "%Y-%m-%d").date()
                except (KeyError, ValueError, TypeError):
                    _batch_errors.append("Invalid date range — please try again.")
                    continue
                if _e < _s:
                    _batch_errors.append("End date cannot be before start date.")
                    continue
                if _s < _today:
                    _batch_errors.append("A date range cannot be in the past.")
                    continue
                _parsed.append({'start': _s, 'end': _e})
            # Check for mutual overlaps within the batch itself
            _batch_overlap_found = False
            for _i, _pr in enumerate(_parsed):
                if _batch_overlap_found:
                    break
                for _pr2 in _parsed[_i + 1:]:
                    if _pr['start'] <= _pr2['end'] and _pr['end'] >= _pr2['start']:
                        if not _batch_overlap_found:
                            _batch_errors.append("Some of the selected date ranges overlap each other.")
                            _batch_overlap_found = True
                        break
            if _batch_errors:
                for _err in _batch_errors:
                    flash(_err, "error")
                return render_template(
                    "add_trip.html", trip=None, resorts=resorts,
                    countries_map=countries_map, states_map=states_map,
                    user=current_user, form_action=url_for("add_trip"),
                    user_passes=user_passes, prefill_friend=prefill_friend,
                    prefill_start_date=prefill_start_date,
                    prefill_end_date=prefill_end_date,
                    prefill_resort=prefill_resort, is_group=is_group,
                )
            # Check against existing trips in the database
            _final_ranges = []
            _skipped_count = 0
            for _pr in _parsed:
                _existing = SkiTrip.query.filter(
                    SkiTrip.user_id == current_user.id,
                    SkiTrip.end_date >= _today,
                    SkiTrip.start_date <= _pr['end'],
                    SkiTrip.end_date >= _pr['start'],
                ).first()
                if _existing:
                    _skipped_count += 1
                else:
                    _final_ranges.append(_pr)
            if not _final_ranges:
                flash("All selected dates overlap with existing trips.", "error")
                return render_template(
                    "add_trip.html", trip=None, resorts=resorts,
                    countries_map=countries_map, states_map=states_map,
                    user=current_user, form_action=url_for("add_trip"),
                    user_passes=user_passes, prefill_friend=prefill_friend,
                    prefill_start_date=prefill_start_date,
                    prefill_end_date=prefill_end_date,
                    prefill_resort=prefill_resort, is_group=is_group,
                )
            # Create trips atomically
            _batch_id = str(_uuid_mod.uuid4()) if len(_final_ranges) > 1 else None
            _created = []
            try:
                for _pr in _final_ranges:
                    _dur = SkiTrip.calculate_duration(_pr['start'], _pr['end'])
                    _trip = SkiTrip(
                        user_id=current_user.id,
                        resort_id=_resort.id,
                        state=_resort.state_code or _resort.state,
                        mountain=_resort.name,
                        start_date=_pr['start'],
                        end_date=_pr['end'],
                        is_public=is_public,
                        trip_status=trip_status_form,
                        trip_duration=_dur,
                        trip_equipment_status=trip_equipment_status if trip_equipment_status != 'use_default' else None,
                        is_group_trip=False,
                        created_by_user_id=current_user.id,
                        created_in_batch_id=_batch_id,
                    )
                    db.session.add(_trip)
                    _created.append(_trip)
                db.session.flush()
                for _trip in _created:
                    _trip.add_owner_as_participant()
                    emit_trip_created_activities(_trip, current_user.id)
                db.session.commit()
                _n = len(_created)
                if _skipped_count:
                    flash(
                        f"{_skipped_count} date{'s' if _skipped_count > 1 else ''} skipped — already have a trip during those dates.",
                        "info",
                    )
                flash(f"{'Trip added' if _n == 1 else f'{_n} trips added'}.", "trip")
                if _n == 1:
                    return redirect(url_for("trip_detail", trip_id=_created[0].id))
                return redirect(url_for("my_trips"))
            except Exception as _exc:
                db.session.rollback()
                app.logger.error(f"Batch trip creation error: {_exc}")
                flash("Something went wrong saving your trips. Please try again.", "error")
                return render_template(
                    "add_trip.html", trip=None, resorts=resorts,
                    countries_map=countries_map, states_map=states_map,
                    user=current_user, form_action=url_for("add_trip"),
                    user_passes=user_passes, prefill_friend=prefill_friend,
                    prefill_start_date=prefill_start_date,
                    prefill_end_date=prefill_end_date,
                    prefill_resort=prefill_resort, is_group=is_group,
                )
        # ── End multi-date batch path ──────────────────────────────────────

        errors = []

        if not resort_id:
            errors.append("Please select a resort.")
        if not start_date_str:
            errors.append("Please select a start date.")
        if not end_date_str:
            errors.append("Please select an end date.")

        resort = None
        if resort_id:
            resort = db.session.get(Resort, resort_id)
            if not resort:
                errors.append("Invalid resort selected.")

        start_date = None
        end_date = None
        today = date.today()
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                if end_date < start_date:
                    errors.append("End date cannot be before start date.")
                if start_date < today:
                    errors.append("Start date cannot be in the past.")
                if end_date < today:
                    errors.append("End date cannot be in the past.")
            except ValueError:
                errors.append("Invalid date format.")

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template(
                "add_trip.html",
                trip=None,
                resorts=resorts,
                countries_map=countries_map,
                states_map=states_map,
                user=current_user,
                form_action=url_for("add_trip"),
                user_passes=user_passes,
                prefill_friend=prefill_friend,
                prefill_start_date=prefill_start_date,
                prefill_end_date=prefill_end_date,
                prefill_resort=prefill_resort,
                is_group=is_group,
            )
        
        overlapping = SkiTrip.query.filter(
            SkiTrip.user_id == current_user.id,
            SkiTrip.resort_id == resort.id,
            SkiTrip.end_date >= today,
            SkiTrip.start_date <= end_date,
            SkiTrip.end_date >= start_date
        ).first()
        
        if overlapping:
            return render_template(
                "add_trip.html",
                trip=None,
                resorts=resorts,
                countries_map=countries_map,
                states_map=states_map,
                user=current_user,
                form_action=url_for("add_trip"),
                overlap_trip=overlapping,
                overlap_resort_name=resort.name,
                user_passes=user_passes,
                prefill_friend=prefill_friend,
                prefill_start_date=prefill_start_date,
                prefill_end_date=prefill_end_date,
                prefill_resort=prefill_resort,
                is_group=is_group,
            )

        trip_duration = SkiTrip.calculate_duration(start_date, end_date)

        trip = SkiTrip(
            user_id=current_user.id,
            resort_id=resort.id,
            state=resort.state_code or resort.state,
            mountain=resort.name,
            start_date=start_date,
            end_date=end_date,
            is_public=is_public,
            trip_status=trip_status_form,
            ride_intent=ride_intent,
            trip_duration=trip_duration,
            trip_equipment_status=trip_equipment_status if trip_equipment_status != 'use_default' else None,
            is_group_trip=is_group_trip or (friend_id is not None),
            created_by_user_id=current_user.id,
        )
        try:
            db.session.add(trip)
            db.session.flush()
            trip.add_owner_as_participant()
            if friend_id:
                trip.add_participant(friend_id, GuestStatus.INVITED)
            emit_trip_created_activities(trip, current_user.id)
            db.session.commit()

            # ── B5: trip.invite.created (add_trip form) — centralized dispatch ──
            if friend_id:
                emit_messaging_event(
                    event_name=EventName.TRIP_INVITE_CREATED,
                    actor_user_id=current_user.id,
                    recipient_user_id=int(friend_id),
                    entity_type="trip",
                    entity_id=trip.id,
                    metadata={
                        "actor_name": current_user.first_name or current_user.username,
                        "resort":     resort.name if resort else "a trip",
                        "trip_id":    trip.id,
                    },
                    source_route="add_trip",
                )

            flash("Trip added.", "trip")
            return redirect(url_for("trip_detail", trip_id=trip.id))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error adding trip: {e}")
            flash("Something went wrong while saving your trip. Please try again.", "error")
            return render_template(
                "add_trip.html",
                trip=None,
                resorts=resorts,
                countries_map=countries_map,
                states_map=states_map,
                user=current_user,
                form_action=url_for("add_trip"),
                user_passes=user_passes,
                prefill_friend=prefill_friend,
                prefill_start_date=prefill_start_date,
                prefill_end_date=prefill_end_date,
                is_group=is_group,
            )

    # GET - render the add trip form
    # ⚠️ CONTRACT LOCK: countries_map must ALWAYS be passed to template.
    # The country dropdown is server-rendered from COUNTRIES (not derived from resorts).
    if app.debug:
        print(f"[ROUTE_PERF] route=add_trip total={time.perf_counter()-_rp_t0:.4f}s resorts={_t_resorts:.4f}s resort_count={len(resorts)}")
    return render_template(
        "add_trip.html",
        trip=None,
        resorts=resorts,
        countries_map=countries_map,
        states_map=states_map,
        user=current_user,
        form_action=url_for("add_trip"),
        user_passes=user_passes,
        prefill_friend=prefill_friend,
        prefill_start_date=prefill_start_date,
        prefill_end_date=prefill_end_date,
        is_group=is_group,
    )

@app.route("/trips/<int:trip_id>/edit", methods=["GET", "POST"])
@login_required
def edit_trip_form(trip_id):
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id != current_user.id:
        abort(403)
    
    resorts = get_resorts_for_trip_form()
    original_start = trip.start_date
    original_end = trip.end_date
    user_passes = [p.strip() for p in (current_user.pass_type or "").split(",") if p.strip()]
    
    # Get current user's participant record for this trip
    my_participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, user_id=current_user.id
    ).first()
    my_transportation = my_participant.transportation_status.value if my_participant and my_participant.transportation_status else None

    countries_map = COUNTRIES
    states_map = STATE_ABBR_MAP

    if request.method == "POST":
        resort_id = request.form.get("resort_id")
        start_date_str = request.form.get("start_date")
        end_date_str = request.form.get("end_date")
        is_public = request.form.get("is_public") == "on"
        transportation_status = request.form.get("transportation_status") or None
        trip_equipment_status = request.form.get("trip_equipment_status") or "use_default"
        trip_status_raw = request.form.get("trip_status", "planning")
        trip_status = trip_status_raw if trip_status_raw in ("planning", "going") else "planning"

        errors = []

        if not resort_id:
            errors.append("Please select a resort.")
        if not start_date_str:
            errors.append("Please select a start date.")
        if not end_date_str:
            errors.append("Please select an end date.")

        resort = None
        if resort_id:
            resort = db.session.get(Resort, resort_id)
            if not resort:
                errors.append("Invalid resort selected.")

        start_date = None
        end_date = None
        today = date.today()
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                if end_date < start_date:
                    errors.append("End date cannot be before start date.")
                dates_changed = (start_date != original_start or end_date != original_end)
                if dates_changed:
                    if start_date < today:
                        errors.append("Start date cannot be in the past.")
                    if end_date < today:
                        errors.append("End date cannot be in the past.")
            except ValueError:
                errors.append("Invalid date format.")

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template(
                "edit_trip.html",
                trip=trip,
                resorts=resorts,
                countries_map=countries_map,
                states_map=states_map,
                user=current_user,
                form_action=url_for("edit_trip_form", trip_id=trip.id),
                user_passes=user_passes,
                my_transportation=my_transportation,
                trip_status=trip_status,
            )
        
        overlapping = SkiTrip.query.filter(
            SkiTrip.user_id == current_user.id,
            SkiTrip.id != trip.id,
            SkiTrip.resort_id == resort.id,
            SkiTrip.end_date >= today,
            SkiTrip.start_date <= end_date,
            SkiTrip.end_date >= start_date
        ).first()
        
        if overlapping:
            flash("You already have a trip that overlaps these dates.", "error")
            return render_template(
                "edit_trip.html",
                trip=trip,
                resorts=resorts,
                countries_map=countries_map,
                states_map=states_map,
                user=current_user,
                form_action=url_for("edit_trip_form", trip_id=trip.id),
                user_passes=user_passes,
                my_transportation=my_transportation,
                trip_status=trip_status,
            )

        dates_changed = (start_date != original_start or end_date != original_end)
        
        trip.resort_id = resort.id
        trip.state = resort.state_code or resort.state
        trip.mountain = resort.name
        trip.start_date = start_date
        trip.end_date = end_date
        trip.is_public = is_public
        trip.trip_status = trip_status
        trip.trip_equipment_status = trip_equipment_status if trip_equipment_status != 'use_default' else None
        trip.trip_duration = SkiTrip.calculate_duration(start_date, end_date)
        trip.updated_at = datetime.utcnow()
        
        # Update current user's transportation_status on their participant record
        if my_participant and transportation_status:
            try:
                my_participant.transportation_status = ParticipantTransportation(transportation_status)
            except ValueError:
                pass  # Invalid value, ignore
        elif my_participant and not transportation_status:
            my_participant.transportation_status = None
        
        try:
            emit_trip_updated_activities(trip, current_user.id, dates_changed=dates_changed)
            db.session.commit()
            flash("Changes saved.", "trip")
            return redirect(url_for("trip_detail", trip_id=trip.id))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating trip: {e}")
            flash("Something went wrong while saving your trip. Please try again.", "error")
            return render_template(
                "edit_trip.html",
                trip=trip,
                resorts=resorts,
                countries_map=countries_map,
                states_map=states_map,
                user=current_user,
                form_action=url_for("edit_trip_form", trip_id=trip.id),
                user_passes=user_passes,
                my_transportation=my_transportation,
                trip_status=trip_status,
            )

    return render_template(
        "edit_trip.html",
        trip=trip,
        resorts=resorts,
        countries_map=countries_map,
        states_map=states_map,
        user=current_user,
        form_action=url_for("edit_trip_form", trip_id=trip.id),
        user_passes=user_passes,
        my_transportation=my_transportation,
        trip_status=(trip.trip_status or 'planning'),
    )

@app.route("/trips/<int:trip_id>")
@login_required
def trip_detail(trip_id):
    """Trip Detail page - primary hub for viewing and managing a trip."""
    _rp_t0 = time.perf_counter()
    trip = (
        SkiTrip.query
        .options(db.joinedload(SkiTrip.resort))
        .filter_by(id=trip_id)
        .first_or_404()
    )

    # Check if user is owner or a participant
    is_owner = trip.user_id == current_user.id
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, user_id=current_user.id
    ).first()
    is_guest = not is_owner and participant and participant.status == GuestStatus.ACCEPTED
    is_invited = not is_owner and participant and participant.status == GuestStatus.INVITED
    
    # Owner, accepted guests, and invited users can view the trip detail
    if not is_owner and not is_guest and not is_invited:
        abort(404)
    
    # Get participants grouped by status
    invited_participants = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, status=GuestStatus.INVITED
    ).all()
    accepted_participants = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, status=GuestStatus.ACCEPTED
    ).all()
    
    # Get connected friends for invite modal (owner only)
    friends_for_invite = []
    if is_owner:
        friend_links = Friend.query.filter_by(user_id=current_user.id).all()
        friend_ids = [f.friend_id for f in friend_links]
        if friend_ids:
            friends = User.query.filter(User.id.in_(friend_ids)).all()
            
            # Check which friends are already invited or accepted.
            # Use already-loaded lists to avoid a redundant lazy-load of trip.participants.
            existing_participants = {p.user_id: p.status for p in invited_participants + accepted_participants}
            
            for friend in friends:
                status = existing_participants.get(friend.id)
                friends_for_invite.append({
                    'user': friend,
                    'status': status.value if status else None,
                    'disabled': status is not None,
                    'label': 'Already on trip' if status == GuestStatus.ACCEPTED else ('Invite sent' if status == GuestStatus.INVITED else None)
                })
    
    # Generate trip-specific invite URL for external sharing (owner only)
    trip_invite_url = None
    if is_owner:
        try:
            _titok = get_or_create_trip_invite_token(trip.id, current_user.id)
            trip_invite_url = f"{BASE_URL}{url_for('trip_invite_token_landing', token=_titok.token)}" if _titok else None
        except Exception:
            trip_invite_url = None

    # Get trip owner info
    owner = db.session.get(User, trip.user_id)
    
    # Get group signals for aggregated view
    group_signals = trip.get_group_signals()
    
    # Get all participants (owner + accepted guests) for display
    # EXCLUDE organizer from all_participants list for the "Who's Going" section
    all_participants = [p for p in trip.get_all_participants() if p.user_id != trip.user_id]
    
    # Get pending join requests (owner only)
    pending_requests = []
    if is_owner:
        pending_requests = Invitation.query.filter_by(
            trip_id=trip_id,
            invite_type=InviteType.REQUEST,
            status='pending'
        ).all()
    
    # Get current user's participant record for inline editing
    current_user_participant = participant
    if is_owner:
        # Owner needs their own participant record
        current_user_participant = SkiTripParticipant.query.filter_by(
            trip_id=trip_id, user_id=current_user.id, role=ParticipantRole.OWNER
        ).first()
    
    # Calculate participant overlaps (for "You and X overlap for Y days" display)
    _t = time.perf_counter()
    participant_overlaps = []
    today = date.today()
    if trip.start_date and trip.end_date and (is_owner or is_guest):
        trip_dates = set()
        d = trip.start_date
        while d <= trip.end_date:
            trip_dates.add(d)
            d += timedelta(days=1)
        
        # Get other participants' confirmed trips at the same resort during the same period
        other_user_ids = [p.user_id for p in all_participants if p.user_id != current_user.id]
        if other_user_ids:
            # Batch-load all participant users before loop — avoids N+1 (1 query replaces N).
            _participant_users_map = {
                u.id: u for u in User.query.filter(User.id.in_(other_user_ids)).all()
            }
            # Batch-load ALL overlapping trips for all participants in one query —
            # replaces the previous 1-query-per-participant N+1 pattern.
            # Preserves the same resort_id / mountain fallback matching logic.
            if trip.resort_id and trip.mountain:
                _all_other_trips = SkiTrip.query.filter(
                    SkiTrip.user_id.in_(other_user_ids),
                    db.or_(
                        SkiTrip.resort_id == trip.resort_id,
                        db.and_(
                            SkiTrip.resort_id.is_(None),
                            SkiTrip.mountain == trip.mountain
                        )
                    ),
                    SkiTrip.start_date <= trip.end_date,
                    SkiTrip.end_date >= trip.start_date,
                    SkiTrip.end_date >= today
                ).all()
            elif trip.resort_id:
                _all_other_trips = SkiTrip.query.filter(
                    SkiTrip.user_id.in_(other_user_ids),
                    SkiTrip.resort_id == trip.resort_id,
                    SkiTrip.start_date <= trip.end_date,
                    SkiTrip.end_date >= trip.start_date,
                    SkiTrip.end_date >= today
                ).all()
            elif trip.mountain:
                _all_other_trips = SkiTrip.query.filter(
                    SkiTrip.user_id.in_(other_user_ids),
                    SkiTrip.mountain == trip.mountain,
                    SkiTrip.start_date <= trip.end_date,
                    SkiTrip.end_date >= trip.start_date,
                    SkiTrip.end_date >= today
                ).all()
            else:
                _all_other_trips = []

            # Group trips by owner user_id for per-participant overlap calculation
            _other_trips_by_user: dict = {}
            for _ot in _all_other_trips:
                _other_trips_by_user.setdefault(_ot.user_id, []).append(_ot)

            for user_id in other_user_ids:
                other_user = _participant_users_map.get(user_id)
                if not other_user:
                    continue

                other_trips = _other_trips_by_user.get(user_id, [])
                if other_trips:
                    # Calculate overlap days using a SET to avoid double-counting
                    overlap_day_set = set()
                    for ot in other_trips:
                        if ot.start_date and ot.end_date:
                            ot_d = ot.start_date
                            while ot_d <= ot.end_date:
                                if ot_d in trip_dates:
                                    overlap_day_set.add(ot_d)
                                ot_d += timedelta(days=1)

                    if overlap_day_set:
                        participant_overlaps.append({
                            'name': other_user.first_name,
                            'days': len(overlap_day_set)
                        })
    
    if app.debug:
        print(f"[ROUTE_PERF] trip_detail.participant_overlap_loop={time.perf_counter()-_t:.4f}s participant_count={len(all_participants)}")
    # Count how many of the user's friends have this resort on their wishlist
    friends_wishlist_count = 0
    if trip.resort_id:
        _friend_ids = [f.friend_id for f in Friend.query.filter_by(user_id=current_user.id).all()]
        if _friend_ids:
            _wl_friends = User.query.filter(User.id.in_(_friend_ids)).all()
            friends_wishlist_count = sum(
                1 for u in _wl_friends
                if trip.resort_id in (u.wish_list_resorts or [])
            )

    resorts_json = get_resorts_for_trip_form() if is_owner else []

    if app.debug:
        print(f"[ROUTE_PERF] route=trip_detail total={time.perf_counter()-_rp_t0:.4f}s")
    return render_template(
        "trip_detail.html",
        trip=trip,
        owner=owner,
        is_owner=is_owner,
        is_guest=is_guest,
        is_invited=is_invited,
        invited_participants=invited_participants,
        accepted_participants=accepted_participants,
        friends_for_invite=friends_for_invite,
        invite_count=len(invited_participants),
        trip_invite_url=trip_invite_url,
        group_signals=group_signals,
        all_participants=all_participants,
        current_user_participant=current_user_participant,
        participant_overlaps=participant_overlaps,
        pending_requests=pending_requests,
        friends_wishlist_count=friends_wishlist_count,
        today=date.today(),
        resorts_json=resorts_json,
    )


@app.route("/trips/<int:trip_id>/invite")
@login_required
def trip_invite_detail(trip_id):
    """Invite detail page - view trip invitation before accepting or after accepting."""
    trip = SkiTrip.query.get_or_404(trip_id)
    today = date.today()
    
    # Check if invite has expired (trip start date has passed)
    if trip.start_date and trip.start_date < today:
        flash("This invite has expired.", "error")
        return redirect(url_for("my_trips"))
    
    # Check if user has an invite (pending, accepted, or declined) for this trip
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, user_id=current_user.id
    ).first()
    
    if not participant:
        flash("You don't have an invite for this trip.", "error")
        return redirect(url_for("my_trips"))
    
    # Check if user has accepted (unlocked state)
    is_accepted = participant.status == GuestStatus.ACCEPTED
    
    # Check if we should show the nudge (just accepted via query param)
    show_nudge = request.args.get("just_accepted") == "1" and is_accepted
    
    # Get trip owner
    owner = db.session.get(User, trip.user_id)
    
    # Get all participants
    all_participants = SkiTripParticipant.query.filter_by(trip_id=trip_id).all()
    
    # Sort participants by status
    accepted_participants = [p for p in all_participants if p.status == GuestStatus.ACCEPTED]
    other_participants = [p for p in all_participants if p.status != GuestStatus.ACCEPTED]
    
    # Calculate going count: owner (always) + accepted invitees
    going_count = 1 + len(accepted_participants)  # 1 for owner
    
    # Check if user owns equipment (has EquipmentSetup with brand or model)
    user_owns_equipment = EquipmentSetup.query.filter(
        EquipmentSetup.user_id == current_user.id,
        db.or_(
            EquipmentSetup.brand.isnot(None),
            EquipmentSetup.model.isnot(None)
        )
    ).first() is not None
    
    return render_template(
        "trip_invite_detail.html",
        trip=trip,
        owner=owner,
        participant=participant,
        accepted_participants=accepted_participants,
        other_participants=other_participants,
        is_accepted=is_accepted,
        show_nudge=show_nudge,
        going_count=going_count,
        user_owns_equipment=user_owns_equipment,
    )








@app.route("/api/trip/<int:trip_id>/accommodation", methods=["POST"])
@login_required
def update_trip_accommodation(trip_id):
    """Update accommodation status (owner-only)."""
    trip = db.session.get(SkiTrip, trip_id)
    if not trip:
        return jsonify({"status": "error", "message": "Trip not found"}), 404
    
    if trip.user_id != current_user.id:
        return jsonify({"status": "error", "message": "Only the trip organizer can manage accommodations"}), 403

    data = request.json
    status = data.get("status")
    link = data.get("link")
    
    if status == 'none_yet' or not status:
        trip.accommodation_status = None
        trip.accommodation_link = None
    else:
        trip.accommodation_status = status
        trip.accommodation_link = link

    db.session.commit()
    return jsonify({"status": "success"})


@app.route("/api/trip/<int:trip_id>/equipment-override", methods=["POST"])
@login_required
def update_trip_equipment_override(trip_id):
    """Update equipment override status (owner-only)."""
    trip = db.session.get(SkiTrip, trip_id)
    if not trip:
        return jsonify({"status": "error", "message": "Trip not found"}), 404
    
    if trip.user_id != current_user.id:
        return jsonify({"status": "error", "message": "Only the trip organizer can manage equipment overrides"}), 403

    data = request.json
    status = data.get("status") # use_default, have_own_equipment, renting
    
    if status == 'use_default' or not status:
        trip.equipment_override = None
    else:
        trip.equipment_override = status

    db.session.commit()
    
    # Update organizer's participant record to match if they are on the trip
    participant = SkiTripParticipant.query.filter_by(trip_id=trip.id, user_id=current_user.id).first()
    if participant:
        if status == 'have_own_equipment':
            participant.equipment_status = ParticipantEquipment.OWN
        elif status == 'renting':
            participant.equipment_status = ParticipantEquipment.RENTING
        else:
            # Revert to profile logic in the display helper usually, but let's sync the enum if possible
            # ParticipantEquipment doesn't have a 'PROFILE' option, it usually stores the explicit state.
            # For now, we'll let the template/model display property handle the 'None' case.
            participant.equipment_status = None
        db.session.commit()

    return jsonify({"status": "success"})


@app.route("/trip-invite/<token>")
def trip_invite_token_landing(token):
    """External trip invite landing — validates token and shows accept screen."""
    tit = TripInviteToken.query.filter_by(token=token).first()
    if not tit or not tit.is_active:
        return render_template("invite_invalid.html",
                               message="This trip invite link is no longer valid.")
    trip = db.session.get(SkiTrip, tit.trip_id)
    if not trip:
        return render_template("invite_invalid.html",
                               message="This trip invite link is no longer valid.")
    inviter = db.session.get(User, tit.inviter_user_id)
    if not inviter:
        return render_template("invite_invalid.html",
                               message="This trip invite link is no longer valid.")

    # Unauthenticated — store token and send through auth
    if not current_user.is_authenticated:
        session["trip_invite_token"] = token
        return redirect(url_for("auth"))

    # Owner visiting their own trip invite link
    if current_user.id == trip.user_id:
        flash("That's your own trip invite link.", "trip")
        return redirect(url_for("trip_detail", trip_id=trip.id))

    # Already accepted — skip the landing and go straight to the trip
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip.id, user_id=current_user.id
    ).first()
    if participant and participant.status == GuestStatus.ACCEPTED:
        return redirect(url_for("trip_detail", trip_id=trip.id))

    mountain_name = (trip.resort.name if trip.resort else None) or trip.mountain or "the mountain"
    return render_template(
        "trip_invite_token_landing.html",
        trip=trip,
        inviter=inviter,
        token=token,
        participant=participant,
        mountain_name=mountain_name,
    )


@app.route("/trip-invite/<token>/accept", methods=["POST"])
@login_required
def trip_invite_token_accept(token):
    """Accept a trip invite via external token."""
    tit = TripInviteToken.query.filter_by(token=token).first()
    if not tit or not tit.is_active:
        return render_template("invite_invalid.html",
                               message="This trip invite link is no longer valid.")
    trip = db.session.get(SkiTrip, tit.trip_id)
    if not trip:
        return render_template("invite_invalid.html",
                               message="This trip invite link is no longer valid.")

    # Owner guard
    if current_user.id == trip.user_id:
        flash("That's your own trip.", "info")
        return redirect(url_for("trip_detail", trip_id=trip.id))

    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip.id, user_id=current_user.id
    ).first()

    if participant:
        if participant.status == GuestStatus.ACCEPTED:
            flash("You're already on this trip.", "trip")
            return redirect(url_for("trip_detail", trip_id=trip.id))
        # INVITED or DECLINED — upgrade to ACCEPTED
        participant.status = GuestStatus.ACCEPTED
    else:
        participant = SkiTripParticipant(
            trip_id=trip.id,
            user_id=current_user.id,
            status=GuestStatus.ACCEPTED,
            role=ParticipantRole.GUEST,
        )
        db.session.add(participant)

    if not trip.is_group_trip:
        trip.is_group_trip = True

    # Stamp first-use timestamp (informational only — token remains reusable)
    if tit.used_at is None:
        tit.used_at = datetime.utcnow()

    db.session.commit()

    emit_trip_invite_accepted_activity(trip, current_user.id, trip.user_id)
    emit_friend_joined_trip_activities(trip, current_user.id)
    emit_messaging_event(
        event_name=EventName.TRIP_INVITE_ACCEPTED,
        actor_user_id=current_user.id,
        recipient_user_id=trip.user_id,
        entity_type="trip",
        entity_id=trip.id,
        metadata={
            "actor_name": current_user.first_name or current_user.username,
            "resort":     trip.mountain or "your trip",
            "trip_id":    trip.id,
        },
        source_route="trip_invite_token_accept",
    )

    # Clean up session key if present
    session.pop("trip_invite_token", None)

    flash("You're going!", "trip")
    return redirect(url_for("trip_detail", trip_id=trip.id))


@app.route("/trips/<int:trip_id>/invite/cancel", methods=["POST"])
@login_required
def cancel_trip_invite(trip_id):
    """Cancel a pending invite (trip owner only)."""
    trip = SkiTrip.query.get_or_404(trip_id)
    
    # Only trip owner can cancel invites
    if trip.user_id != current_user.id:
        abort(403)
    
    user_id = request.form.get("user_id", type=int)
    if not user_id:
        flash("Invalid request.", "error")
        return redirect(url_for("trip_detail", trip_id=trip_id))
    
    # Find the pending invite
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, user_id=user_id, status=GuestStatus.INVITED
    ).first()
    
    if participant:
        db.session.delete(participant)
        db.session.commit()
        flash("Invite cancelled.", "info")
    
    return redirect(url_for("trip_detail", trip_id=trip_id))


@app.route("/trips/<int:trip_id>/invite", methods=["POST"])
@login_required
def send_trip_invites(trip_id):
    """Send trip invites to selected friends."""
    trip = SkiTrip.query.get_or_404(trip_id)
    
    # Only trip owner can invite
    if trip.user_id != current_user.id:
        abort(403)
    
    friend_ids = request.form.getlist("friend_ids")
    if not friend_ids:
        flash("Please select at least one friend to invite.", "error")
        return redirect(url_for("trip_detail", trip_id=trip_id))
    
    # Validate that all selected users are connected friends
    friend_links = Friend.query.filter_by(user_id=current_user.id).all()
    connected_friend_ids = {f.friend_id for f in friend_links}
    
    newly_invited_user_ids = []
    invites_sent = 0
    for friend_id_str in friend_ids:
        try:
            friend_id = int(friend_id_str)
        except ValueError:
            continue
        
        # Skip if not a connected friend
        if friend_id not in connected_friend_ids:
            continue
        
        # Skip if user is the trip owner
        if friend_id == current_user.id:
            continue
        
        # Check for existing participant record (idempotency)
        existing = SkiTripParticipant.query.filter_by(
            trip_id=trip_id, user_id=friend_id
        ).first()
        
        if not existing:
            participant = SkiTripParticipant(
                trip_id=trip_id,
                user_id=friend_id,
                status=GuestStatus.INVITED
            )
            db.session.add(participant)
            # Emit activity for the invited user
            emit_trip_invite_received_activity(trip, current_user.id, friend_id)
            invites_sent += 1
            newly_invited_user_ids.append(friend_id)
    
    if invites_sent > 0:
        # Mark trip as group trip if not already
        if not trip.is_group_trip:
            trip.is_group_trip = True
        db.session.commit()

        # ── B5: trip.invite.created (trip_detail invite loop) — one emit per recipient ──
        for _invited_uid in newly_invited_user_ids:
            emit_messaging_event(
                event_name=EventName.TRIP_INVITE_CREATED,
                actor_user_id=current_user.id,
                recipient_user_id=_invited_uid,
                entity_type="trip",
                entity_id=trip_id,
                metadata={
                    "actor_name": current_user.first_name or current_user.username,
                    "resort":     trip.mountain or "a trip",
                    "trip_id":    trip_id,
                },
                source_route="trip_detail_invite",
            )

        flash(f"Invite{'s' if invites_sent > 1 else ''} sent to {invites_sent} friend{'s' if invites_sent > 1 else ''}.", "success")
    else:
        flash("No new invites were sent.", "info")
    
    return redirect(url_for("trip_detail", trip_id=trip_id))


@app.route("/trips/<int:trip_id>/request-join", methods=["POST"])
@login_required
@limiter.limit("20 per hour", key_func=_user_or_ip)
def request_to_join_trip(trip_id):
    """Create a join request for a trip."""
    trip = SkiTrip.query.get_or_404(trip_id)
    
    # 1. Trip must not be in the past
    if trip.end_date < date.today():
        return jsonify({"success": False, "error": "This trip has already ended."}), 400
    
    # 2. Requester must not already be an ACCEPTED participant
    is_accepted = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, 
        user_id=current_user.id,
        status=GuestStatus.ACCEPTED
    ).first() is not None
    
    if is_accepted:
        return jsonify({"success": False, "error": "You are already an accepted participant of this trip."}), 400
        
    # 4. Only one pending request per user per trip
    existing_request = Invitation.query.filter_by(
        sender_id=current_user.id,
        receiver_id=trip.user_id,
        trip_id=trip_id,
        invite_type=InviteType.REQUEST,
        status='pending'
    ).first()
    
    if existing_request:
        return jsonify({"success": True, "message": "Request already pending."})
        
    # Create the request
    join_request = Invitation(
        sender_id=current_user.id,
        receiver_id=trip.user_id,
        trip_id=trip_id,
        invite_type=InviteType.REQUEST,
        status='pending'
    )
    db.session.add(join_request)
    # Notify the trip owner
    create_activity(current_user.id, trip.user_id, ActivityType.JOIN_REQUEST_RECEIVED, 'trip', trip.id)
    db.session.commit()
    
    return jsonify({"success": True, "message": "Request sent to owner."})


@app.route("/trips/requests/<int:request_id>/respond", methods=["POST"])
@login_required
def respond_to_join_request(request_id):
    """
    Accept or decline a join request.
    Feature complete as of 2026-01-09. Backend + UI verified.
    """
    invitation = Invitation.query.get_or_404(request_id)
    
    if invitation.invite_type != InviteType.REQUEST:
        return jsonify({"success": False, "error": "Invalid invitation type."}), 400
        
    trip = SkiTrip.query.get_or_404(invitation.trip_id)
    
    # Only the trip owner can respond
    if trip.user_id != current_user.id:
        return jsonify({"success": False, "error": "Only the trip owner can respond to join requests."}), 403
        
    data = request.get_json() or {}
    action = data.get("action")
    
    if action == "accept":
        invitation.status = 'accepted'
        
        # Add requester as participant
        participant = SkiTripParticipant(
            trip_id=trip.id,
            user_id=invitation.sender_id,
            status=GuestStatus.ACCEPTED,
            role=ParticipantRole.GUEST,
            start_date=trip.start_date,
            end_date=trip.end_date
        )
        db.session.add(participant)
        
        # Mark trip as group trip if not already
        if not trip.is_group_trip:
            trip.is_group_trip = True

        # Notify the requester their request was accepted
        create_activity(current_user.id, invitation.sender_id, ActivityType.JOIN_REQUEST_ACCEPTED, 'trip', trip.id)

        db.session.commit()
        return jsonify({"success": True, "message": "Request accepted."})
        
    elif action == "decline":
        invitation.status = 'declined'

        # Notify the requester their request was declined
        create_activity(current_user.id, invitation.sender_id, ActivityType.JOIN_REQUEST_DECLINED, 'trip', trip.id)

        db.session.commit()
        return jsonify({"success": True, "message": "Request declined."})
        
    return jsonify({"success": False, "error": "Invalid action."}), 400


@app.route("/trips/requests/<int:request_id>/cancel", methods=["POST"])
@login_required
def cancel_join_request(request_id):
    """Cancel a pending join request."""
    invitation = Invitation.query.get_or_404(request_id)
    
    if invitation.sender_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    if invitation.invite_type != InviteType.REQUEST:
        return jsonify({"success": False, "error": "Invalid request type"}), 400
        
    if invitation.status != 'pending':
        return jsonify({"success": False, "error": "Only pending requests can be cancelled"}), 400
        
    db.session.delete(invitation)
    db.session.commit()

    if request.is_json or request.accept_mimetypes.best == 'application/json':
        return jsonify({"success": True})

    flash("Join request cancelled.", "info")
    return redirect(url_for("my_trips"))

@app.route("/trips/<int:trip_id>/respond", methods=["POST"])
@login_required
def respond_to_trip_invite(trip_id):
    """Accept or decline a trip invite."""
    trip = SkiTrip.query.get_or_404(trip_id)
    
    # Find the user's participant record
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, user_id=current_user.id
    ).first()
    
    if not participant or participant.status != GuestStatus.INVITED:
        return jsonify({"success": False, "error": "No pending invite found"}), 404
    
    # Support both form data and JSON
    if request.is_json:
        data = request.get_json() or {}
        action = data.get("response") or data.get("action")
    else:
        action = request.form.get("action")
    
    if action == "accept":
        participant.status = GuestStatus.ACCEPTED
        
        # Archive solo trip if it exists
        solo_trip = SkiTrip.query.filter(
            SkiTrip.user_id == current_user.id,
            SkiTrip.id != trip_id,
            SkiTrip.start_date == trip.start_date,
            SkiTrip.end_date == trip.end_date,
            db.or_(
                SkiTrip.resort_id == trip.resort_id,
                SkiTrip.mountain == trip.mountain
            )
        ).first()
        
        if solo_trip:
            # Carry over equipment details from solo trip (SkiTrip doesn't have transportation_status)
            if solo_trip.equipment_override and solo_trip.equipment_override != 'use_default':
                participant.equipment_status = ParticipantEquipment.OWN if solo_trip.equipment_override == 'have_own_equipment' else ParticipantEquipment.RENTING
            
            db.session.delete(solo_trip)
            
        emit_trip_invite_accepted_activity(trip, current_user.id, trip.user_id)
        emit_friend_joined_trip_activities(trip, current_user.id)
        db.session.commit()

        # ── B4: trip.invite.accepted — routed through centralized dispatch ──
        emit_messaging_event(
            event_name=EventName.TRIP_INVITE_ACCEPTED,
            actor_user_id=current_user.id,
            recipient_user_id=trip.user_id,
            entity_type="trip",
            entity_id=trip.id,
            metadata={
                "actor_name": current_user.first_name or current_user.username,
                "resort":     trip.mountain or "your trip",
                "trip_id":    trip_id,
            },
            source_route="respond_to_trip_invite_accept",
        )

        if request.is_json:
            return jsonify({"success": True, "message": "You're going"})
        flash("You're going", "success")
        return redirect(url_for("trip_detail", trip_id=trip_id))
    elif action == "decline":
        participant.status = GuestStatus.DECLINED
        emit_trip_invite_declined_activity(trip, current_user.id, trip.user_id)
        db.session.commit()

        # ── B4: trip.invite.declined — SILENT path, audit row only, no provider call ──
        emit_messaging_event(
            event_name=EventName.TRIP_INVITE_DECLINED,
            actor_user_id=current_user.id,
            recipient_user_id=trip.user_id,
            entity_type="trip",
            entity_id=trip.id,
            metadata={
                "trip_id": trip_id,
                "resort":  trip.mountain or "",
            },
            source_route="respond_to_trip_invite_decline",
        )

        if request.is_json:
            return jsonify({"success": True, "message": "Invite declined"})
        flash("Invite declined.", "info")
        return redirect(url_for("my_trips"))
    else:
        return jsonify({"success": False, "error": "Invalid action"}), 400


@app.route("/api/trips/<int:trip_id>/participant/signals", methods=["POST"])
@login_required
def update_participant_signals(trip_id):
    """Update current user's transportation, equipment, carpool, and lesson signals for a trip."""
    trip = SkiTrip.query.get_or_404(trip_id)
    
    # Find the user's participant record
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, user_id=current_user.id
    ).first()
    
    if not participant:
        return jsonify({"success": False, "error": "You are not a participant of this trip."}), 404
    
    # Only accepted participants or owners can update their signals
    if participant.status != GuestStatus.ACCEPTED and participant.role != ParticipantRole.OWNER:
        return jsonify({"success": False, "error": "You must accept the invite first."}), 403
    
    data = request.get_json() or {}
    
    # Update transportation status
    transportation = data.get("transportation_status")
    if transportation:
        try:
            participant.transportation_status = ParticipantTransportation(transportation)
        except ValueError:
            pass
    
    # Update equipment status
    equipment = data.get("equipment_status")
    if equipment:
        try:
            participant.equipment_status = ParticipantEquipment(equipment)
        except ValueError:
            pass
    elif equipment == "":
        participant.equipment_status = None
    
    # Update carpool role
    carpool = data.get("carpool_role")
    if carpool:
        try:
            participant.carpool_role = CarpoolRole(carpool)
        except ValueError:
            pass
    elif carpool == "":
        participant.carpool_role = None
    
    # Update lesson status
    lesson = data.get("taking_lesson")
    if lesson:
        try:
            participant.taking_lesson = LessonChoice(lesson)
        except ValueError:
            pass
    
    db.session.commit()
    
    # Build display values for response
    carpool_display = None
    if participant.carpool_role:
        carpool_labels = {
            CarpoolRole.DRIVER_WITH_SPACE: "I can drive and have space",
            CarpoolRole.DRIVER_NO_SPACE: "I can drive but have no space",
            CarpoolRole.NEEDS_RIDE: "I need a ride",
            CarpoolRole.NOT_CARPOOLING: "Not carpooling",
            CarpoolRole.OTHER: "Other",
        }
        carpool_display = carpool_labels.get(participant.carpool_role, "Add")
    
    lesson_display = None
    if participant.taking_lesson:
        lesson_labels = {
            LessonChoice.NO: "No",
            LessonChoice.MAYBE: "Considering",
            LessonChoice.YES: "Yes",
        }
        lesson_display = lesson_labels.get(participant.taking_lesson, "Not set")
    
    return jsonify({
        "success": True,
        "transportation_display": participant.get_display_transportation(),
        "equipment_display": participant.get_display_equipment(),
        "carpool_display": carpool_display or "Add",
        "lesson_display": lesson_display or "Not set",
    })


@app.route("/trips/<int:trip_id>/delete", methods=["POST"])
@login_required
def delete_trip_form(trip_id):
    validate_csrf_request()
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id != current_user.id:
        abort(403)

    try:
        # Clean up Invitation rows (bare FK, no cascade)
        Invitation.query.filter(Invitation.trip_id == trip.id).delete()
        # Clean up TripInviteToken rows (nullable=False FK, SQLAlchemy would try to null it)
        TripInviteToken.query.filter_by(trip_id=trip.id).delete()
        # Clean up Activity rows
        delete_activities_for_trip(trip_id)
        db.session.delete(trip)
        db.session.commit()
        app.logger.info(
            "[delete_trip_form] success route=delete_trip_form trip_id=%s user_id=%s",
            trip_id, current_user.id
        )
        flash("Trip deleted.", "trip")
        return redirect(url_for("my_trips"))
    except Exception as e:
        db.session.rollback()
        app.logger.error(
            "[delete_trip_form] error route=delete_trip_form trip_id=%s user_id=%s exc=%s",
            trip_id, current_user.id, e
        )
        flash("Something went wrong while cancelling the trip. Please try again.", "error")
        return redirect(url_for("trip_detail", trip_id=trip_id))

@app.route("/trips/<int:trip_id>/leave", methods=["POST"])
@login_required
def leave_trip(trip_id):
    validate_csrf_request()
    trip = SkiTrip.query.get_or_404(trip_id)
    if trip.user_id == current_user.id:
        abort(403)
    participant = SkiTripParticipant.query.filter_by(
        trip_id=trip_id, user_id=current_user.id, status=GuestStatus.ACCEPTED
    ).first()
    if not participant:
        abort(403)
    try:
        db.session.delete(participant)
        db.session.commit()
        app.logger.info(
            "[leave_trip] success trip_id=%s user_id=%s",
            trip_id, current_user.id
        )
        flash("You've left this trip.", "trip")
        return redirect(url_for("my_trips"))
    except Exception as e:
        db.session.rollback()
        app.logger.error(
            "[leave_trip] error trip_id=%s user_id=%s exc=%s",
            trip_id, current_user.id, e
        )
        flash("Something went wrong. Please try again.", "error")
        return redirect(url_for("trip_detail", trip_id=trip_id))


@app.route("/mountains-visited", methods=["GET", "POST"])
@login_required
def mountains_visited():
    user = current_user
    
    # Get all resorts from database (source of truth, exclude region-level entities)
    all_resorts = Resort.query.filter_by(is_active=True, is_region=False).order_by(
        Resort.country_code, Resort.state_code, Resort.name
    ).all()
    
    if request.method == "POST":
        # Get selected resort IDs from form
        selected_resort_ids = request.form.getlist("resort_ids")
        
        resort_ids = []
        mountain_names = []
        seen_ids = set()
        
        for rid_str in selected_resort_ids:
            try:
                rid = int(rid_str)
                if rid not in seen_ids:
                    resort = db.session.get(Resort, rid)
                    if resort:
                        resort_ids.append(rid)
                        mountain_names.append(resort.name)
                        seen_ids.add(rid)
            except (ValueError, TypeError):
                pass
        
        user.visited_resort_ids = resort_ids
        user.mountains_visited = mountain_names
        
        try:
            db.session.commit()
            return redirect(url_for("settings"))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error saving mountains visited: {e}")
            flash("Something went wrong while saving. Please try again.", "error")
            return redirect(url_for("mountains_visited"))
    
    # Build set of selected resort IDs
    selected_resort_ids = set()
    
    # Include resort IDs from visited_resort_ids
    if user.visited_resort_ids:
        selected_resort_ids.update(user.visited_resort_ids)
    
    # Also try to match legacy mountains_visited names to resort IDs
    if user.mountains_visited:
        for name in user.mountains_visited:
            for resort in all_resorts:
                if resort.name.lower() == name.lower() and resort.id not in selected_resort_ids:
                    selected_resort_ids.add(resort.id)
                    break
    
    mountains_visited_count = len(selected_resort_ids)
    
    # Get selected resort objects for server-side pill rendering
    selected_resorts = []
    if selected_resort_ids:
        selected_resorts = Resort.query.filter(Resort.id.in_(selected_resort_ids)).all()
    
    # Group and sort selected resorts for display
    grouped_selected = group_resorts_for_display(selected_resorts)
    
    # Get canonical countries for dropdown from the shared COUNTRIES mapping
    # Safeguard: Dropdown options MUST come from canonical country list, not resort data
    from utils.countries import COUNTRIES
    countries_list = sorted(COUNTRIES.keys(), key=lambda c: (c != 'US', COUNTRIES[c]))
    
    return render_template(
        "mountains_visited.html",
        resorts=all_resorts,
        selected_resort_ids=list(selected_resort_ids),
        selected_resorts=selected_resorts,
        grouped_selected=grouped_selected,
        mountains_visited_count=mountains_visited_count,
        countries=countries_list,
    )

@app.route("/logout")
@login_required
def logout():
    user_id = current_user.id
    ph_analytics.track(user_id, 'logout')

    # Server-side push token hardening: deactivate all active tokens for this
    # user so no further pushes are dispatched after logout, even if the client
    # blOSLogout() call was bypassed (e.g. direct /logout URL navigation).
    # Tokens are soft-deactivated (not deleted) — the device re-registers on
    # next app open, restoring delivery after the user logs back in.
    # NOTE: This intentionally affects all platforms/devices for this account.
    # A single-device targeted approach is not possible server-side without a
    # per-session device identifier, which is not currently stored in the session.
    try:
        PushDeviceToken.query.filter_by(user_id=user_id, active=True).update(
            {'active': False}, synchronize_session=False
        )
        db.session.commit()
    except Exception as _tok_err:
        db.session.rollback()
        current_app.logger.warning(
            "[logout] push token deactivation failed for user_id=%s: %s",
            user_id, _tok_err,
        )

    logout_user()
    # NOTE: do NOT call session.clear() here — it erases Flask-Login's internal
    # _remember='clear' flag, which prevents the remember_token cookie from being
    # deleted and causes the user to be silently re-authenticated on the next request.
    # logout_user() already removes _user_id, _fresh, and _id from the session.
    session['ph_reset'] = True
    return redirect(url_for("auth"))


@app.route("/login")
def login():
    if current_user.is_authenticated:
        return redirect(url_for("home"))
    return render_template("auth.html", default_tab="login", has_invite=("invite_token" in session))


@app.route("/signup")
def signup():
    if current_user.is_authenticated:
        return redirect(url_for("home"))
    return render_template("auth.html", default_tab="signup", has_invite=("invite_token" in session))


@app.route("/auth/google")
def auth_google():
    try:
        redirect_uri = url_for("auth_google_callback", _external=True)
        return oauth.google.authorize_redirect(redirect_uri)
    except Exception as e:
        app.logger.error(f"Google OAuth redirect error: {e}")
        flash("Google sign-in is not available right now. Please try again.", "error")
        return redirect(url_for("auth"))


@app.route("/auth/google/callback")
def auth_google_callback():
    try:
        token = oauth.google.authorize_access_token()
        userinfo = token.get("userinfo") or oauth.google.parse_id_token(token)
        if not userinfo:
            raise ValueError("No user info returned from Google")

        email = (userinfo.get("email") or "").lower().strip()
        sub = userinfo.get("sub", "")
        given_name = (userinfo.get("given_name") or "").strip()
        family_name = (userinfo.get("family_name") or "").strip() or None

        if not email:
            raise ValueError("Google did not return an email address")

        user = User.query.filter_by(email=email).first()

        _is_new_google_user = False
        if user:
            if user.auth_provider != "google":
                user.auth_provider = "google"
            if not user.provider_id:
                user.provider_id = sub
            db.session.commit()
        else:
            _is_new_google_user = True
            user = User(
                email=email,
                first_name=given_name or email.split("@")[0],
                last_name=family_name,
                auth_provider="google",
                provider_id=sub,
                buddy_passes_available=True,
                created_at=datetime.utcnow(),
            )
            db.session.add(user)
            db.session.commit()
            _has_invite = "invite_token" in session or "trip_invite_token" in session
            ph_analytics.track(user.id, 'signup_completed', {
                'method': 'google',
                'signup_source': 'invite' if _has_invite else 'organic',
                'is_invite_signup': _has_invite,
            })

        user.last_active_at = datetime.utcnow()
        db.session.commit()
        login_user(user, remember=True)
        session['_last_active_stamp'] = time.time()
        session.modified = True

        # Founder login alert (non-blocking, throttled to 1×/user/day)
        _queue_founder_login_push(user.id, user.email)

        if "invite_token" in session:
            session["post_onboarding_redirect"] = url_for("friends")
            _connect_pending_inviter(user)
            # New Google users must complete onboarding before landing on /friends.
            # Existing users (profile complete) go straight to /friends.
            if not user.is_core_profile_complete:
                return redirect(url_for("onboarding"))
            return redirect(url_for("friends"))

        # Preserve trip invite context through the Google OAuth round-trip.
        # The OAuth redirect cycle discards URL params, so the token was stored
        # in the session by trip_invite_token_landing before sending to /auth.
        # Mirror the identical handling in the email-login path (~L2651).
        if "trip_invite_token" in session:
            _ttok = session["trip_invite_token"]
            if not user.is_core_profile_complete:
                # New Google user — finish onboarding first, then return to trip invite.
                session["post_onboarding_redirect"] = url_for(
                    "trip_invite_token_landing", token=_ttok
                )
                return redirect(url_for("onboarding"))
            return redirect(url_for("trip_invite_token_landing", token=_ttok))

        # Return to any pending post-login destination (e.g. QR connect flow).
        _post_login = session.pop("post_login_redirect", None)
        if _post_login:
            return redirect(_post_login)

        if not user.is_core_profile_complete:
            return redirect(url_for("onboarding"))

        return redirect(url_for("home"))

    except Exception as e:
        app.logger.error(f"Google OAuth callback error: {e}")
        flash("Something went wrong signing in with Google. Please try again.", "error")
        return redirect(url_for("auth"))


@app.route("/auth/apple")
def auth_apple():
    return redirect(url_for("auth"))


@app.route("/auth/apple/callback")
def auth_apple_callback():
    return redirect(url_for("auth"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
@limiter.limit("10 per minute", key_func=_user_or_ip)
def change_password():
    # Google (and any future OAuth) accounts have no local password.
    if current_user.auth_provider != 'email':
        return render_template("change_password.html", oauth_user=True)

    if request.method == "POST":
        validate_csrf_request()
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        if not current_user.check_password(current_password):
            flash("Current password is incorrect.", "error")
            return redirect(url_for("change_password"))
        
        if len(new_password) < 8:
            flash("New password must be at least 8 characters.", "error")
            return redirect(url_for("change_password"))
        
        if new_password != confirm_password:
            flash("New passwords do not match.", "error")
            return redirect(url_for("change_password"))
        
        current_user.set_password(new_password)
        current_user.password_changed_at = datetime.utcnow()
        try:
            db.session.commit()
            flash("Password updated successfully.", "success")
            return redirect(url_for("change_password"))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error changing password: {e}")
            flash("Something went wrong while updating your password. Please try again.", "error")
            return redirect(url_for("change_password"))
    
    return render_template("change_password.html", oauth_user=False)


@app.route("/delete-account", methods=["POST"])
@login_required
def delete_account():
    validate_csrf_request()

    # Capture identity first so every log line and error path has context.
    user       = current_user._get_current_object()
    user_id    = user.id
    user_email = user.email

    confirm_email       = request.form.get("confirm_email", "").strip()
    confirmation_matched = confirm_email.lower() == user_email.lower()
    app.logger.info(
        f"[delete_account] attempt user_id={user_id} email={user_email} "
        f"confirmation_matched={confirmation_matched}"
    )

    if not confirm_email:
        flash("Please type your email address to confirm account deletion.", "error")
        return redirect(url_for("profile"))

    if not confirmation_matched:
        flash("Email address did not match. Account was not deleted.", "error")
        return redirect(url_for("profile"))

    try:
        # 1. Activity feed rows (actor or recipient)
        Activity.query.filter(
            db.or_(Activity.actor_user_id == user_id, Activity.recipient_user_id == user_id)
        ).delete(synchronize_session=False)

        # 2. EmailLog rows for this user
        EmailLog.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 3. Null out source_event_id on any remaining EmailLog that references this
        #    user's events (could belong to other users), then delete the events
        user_event_ids = [r[0] for r in db.session.query(Event.id).filter_by(user_id=user_id).all()]
        if user_event_ids:
            db.session.query(EmailLog).filter(
                EmailLog.source_event_id.in_(user_event_ids)
            ).update({EmailLog.source_event_id: None}, synchronize_session=False)
        Event.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 4. Dismissed nudges
        DismissedNudge.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 5. Equipment setups
        EquipmentSetup.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 6. Invite tokens created by this user
        InviteToken.query.filter_by(inviter_id=user_id).delete(synchronize_session=False)

        # 7. Friend invitations (sent or received)
        Invitation.query.filter(
            db.or_(Invitation.sender_id == user_id, Invitation.receiver_id == user_id)
        ).delete(synchronize_session=False)

        # 8. SkiTripParticipant rows where this user is a guest on someone else's trip
        SkiTripParticipant.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 9. TripGuest rows where this user is a guest on someone else's GroupTrip
        TripGuest.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 10. Friend rows (both directions)
        Friend.query.filter(
            db.or_(Friend.user_id == user_id, Friend.friend_id == user_id)
        ).delete(synchronize_session=False)

        # 11. SkiTrips owned by this user — delete other participants first, then trips
        owned_trip_ids = [r[0] for r in db.session.query(SkiTrip.id).filter_by(user_id=user_id).all()]
        if owned_trip_ids:
            SkiTripParticipant.query.filter(
                SkiTripParticipant.trip_id.in_(owned_trip_ids)
            ).delete(synchronize_session=False)
        SkiTrip.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 12. GroupTrips hosted by this user — delete guests first, then trips
        hosted_group_ids = [r[0] for r in db.session.query(GroupTrip.id).filter_by(host_id=user_id).all()]
        if hosted_group_ids:
            TripGuest.query.filter(
                TripGuest.trip_id.in_(hosted_group_ids)
            ).delete(synchronize_session=False)
        GroupTrip.query.filter_by(host_id=user_id).delete(synchronize_session=False)

        # 13. Open availability dates
        UserAvailability.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 14. Dismissed insight cards
        DismissedInsightCard.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 15. Push device tokens
        PushDeviceToken.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 15b. Invite share events — must be deleted explicitly.
        #      InviteShareEvent.user has a backref='invite_share_events' on User
        #      without passive_deletes=True.  When SQLAlchemy processes
        #      db.session.delete(user) it tries to SET user_id=NULL on the related
        #      rows before the user row is removed.  user_id is NOT NULL, so that
        #      UPDATE fails with an IntegrityError even though the DB-level FK has
        #      ON DELETE CASCADE (the ORM intercepts before Postgres can cascade).
        InviteShareEvent.query.filter_by(user_id=user_id).delete(synchronize_session=False)

        # 16a. NULL out invited_by_user_id on any users this person invited.
        #      This is a self-referential FK with no ON DELETE action — leaving it
        #      set would block deletion of the user row with an FK violation.
        User.query.filter_by(invited_by_user_id=user_id).update(
            {User.invited_by_user_id: None}, synchronize_session=False
        )

        # 16b. NULL out created_by_user_id on trips this user organised for
        #      someone else (trip.user_id != user_id).  Those trips survive; we
        #      only clear the organiser reference so the FK no longer blocks deletion.
        SkiTrip.query.filter(
            SkiTrip.created_by_user_id == user_id,
            SkiTrip.user_id != user_id
        ).update({SkiTrip.created_by_user_id: None}, synchronize_session=False)

        # 16c. NULL out MessageEventLog actor/recipient FKs.
        #      These columns are nullable but have a hard FK to user.id with no
        #      ON DELETE action — leaving them set causes a FK violation when the
        #      user row is deleted. We preserve the audit rows and only clear the
        #      user references.
        MessageEventLog.query.filter(
            MessageEventLog.actor_user_id == user_id
        ).update({MessageEventLog.actor_user_id: None}, synchronize_session=False)
        MessageEventLog.query.filter(
            MessageEventLog.recipient_user_id == user_id
        ).update({MessageEventLog.recipient_user_id: None}, synchronize_session=False)

        # 17. Flush to surface any remaining FK issues before the final delete.
        db.session.flush()

        # 18. Delete the user row, commit, THEN log out.
        #     logout_user() must come AFTER a successful commit so that if the
        #     commit raises, the user is still authenticated and can see the error
        #     flash on /profile.  Calling it before the commit (old behaviour) left
        #     the user logged out with their account still intact on rollback.
        db.session.delete(user)
        db.session.commit()

        # NOTE: do NOT call session.clear() — same reason as /logout
        logout_user()

        # Category "message" renders on auth.html (which filters to
        # ['error', 'auth', 'message']).  "success" is silently dropped there.
        flash("Your account has been deleted.", "message")
        return redirect(url_for("auth"))

    except Exception as e:
        db.session.rollback()
        app.logger.error(
            f"[delete_account] user_id={user_id} email={user_email} error={repr(e)}"
        )
        flash("We couldn't delete your account right now. Please try again.", "error")
        # After db.session.delete(user) + rollback, SQLAlchemy expels the user
        # object from its identity map (detached state).  On the next request
        # Flask-Login's user_loader calls db.session.get(User, user_id) on a
        # fresh scoped session — but with pgBouncer/Supabase the connection
        # returned to the pool may still be in an aborted-transaction state,
        # causing the get() to fail and Flask-Login to treat the user as
        # anonymous.  @login_required on /profile then bounces them to /auth
        # where the flash is lost or invisible.
        #
        # Fix: re-query the user on the now-clean session and explicitly
        # re-establish the Flask-Login session so /profile loads correctly.
        try:
            _fresh = db.session.get(User, user_id)
            if _fresh:
                login_user(_fresh, remember=True)
        except Exception:
            pass
        return redirect(url_for("profile"))


@app.route("/skip-pass-prompt")
@login_required
def skip_pass_prompt():
    session["pass_prompt_skipped"] = True
    return redirect(url_for("home"))

@app.route("/select-pass", methods=["GET", "POST"])
@login_required
def select_pass():
    if request.method == "POST":
        chosen = request.form.get("pass_type", "")
        normalized_chosen = normalize_pass_selection(chosen) or chosen
        if count_real_passes(normalized_chosen) > 3:
            flash("You can select up to 3 passes.", "error")
            return redirect(url_for("select_pass"))
        _old_pass_sp = current_user.pass_type  # capture before overwrite for change detection
        current_user.pass_type = normalized_chosen
        try:
            db.session.commit()
            session["pass_prompt_skipped"] = False
            if _ph_is_real_pass(normalized_chosen):
                ph_analytics.track(current_user.id, 'pass_added', {
                    'pass_type':    normalized_chosen,
                    'source':       'select_pass',
                    'is_first_pass': not _ph_is_real_pass(_old_pass_sp),
                })
            # ── B3: friend.pass.changed (select_pass) — centralized dispatch ──
            # One emit per friend → one MEL audit row per recipient.
            if _old_pass_sp != normalized_chosen:
                _sp_friend_ids = get_friend_ids(current_user.id)
                if _sp_friend_ids:
                    _sp_display = format_passes_for_display(normalized_chosen).replace(" · ", " + ")
                    current_app.logger.info(
                        "[MESSAGE_DISPATCH] pass_changed (select_pass): old=%r new=%r friend_count=%d",
                        _old_pass_sp, normalized_chosen, len(_sp_friend_ids),
                    )
                    for _friend_id in _sp_friend_ids:
                        emit_messaging_event(
                            event_name=EventName.FRIEND_PASS_CHANGED,
                            actor_user_id=current_user.id,
                            recipient_user_id=_friend_id,
                            entity_type="user",
                            entity_id=current_user.id,
                            metadata={
                                "actor_first_name": current_user.first_name,
                                "new_pass":         normalized_chosen,
                                "new_pass_display": _sp_display,
                            },
                            source_route="select_pass",
                        )
            return redirect(url_for("profile"))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error saving pass selection: {e}")
            flash("Something went wrong while saving your pass. Please try again.", "error")
            return redirect(url_for("select_pass"))

    return render_template("select_pass.html")

@app.route("/admin/init-db", methods=["POST"])
@login_required
@admin_required
def init_db_http():
    """
    HTTP endpoint for database initialization (backup method for deployment).
    Can be called after deployment to initialize the database.
    
    Usage after deployment:
    POST https://yourapp.replit.dev/admin/init-db
    
    This will:
    - Seed all resorts (idempotent)
    - Be idempotent (safe to call multiple times)
    """
    try:
        with app.app_context():
            # Create all tables
            # Schema creation disabled for Supabase migration.
    # Use 'flask db upgrade' instead.
    # db.create_all()
            
            messages = []
            
            # Seed resorts from xlsx if none exist
            resort_count = Resort.query.count()
            if resort_count == 0:
                from utils.resort_import import import_resorts_from_xlsx
                xlsx_path = os.path.join(os.path.dirname(__file__), 'prod_resorts_full.xlsx')
                if os.path.exists(xlsx_path):
                    stats = import_resorts_from_xlsx(xlsx_path, db, Resort, generate_resort_slug, STATE_ABBR_MAP)
                    total = stats['added'] + stats['updated']
                    messages.append(f"Imported {total} resorts from xlsx ({stats['added']} new, {stats['updated']} updated)")
                else:
                    messages.append("WARNING: prod_resorts_full.xlsx not found — resort seeding skipped")
            else:
                messages.append(f"Resorts already exist ({resort_count})")
            
            return jsonify({
                "status": "success",
                "message": "✅ Database initialized",
                "details": messages
            }), 200
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to initialize database: {str(e)}"
        }), 500


MOUNTAIN_NAME_ALIASES = {
    "Crystal Mountain MI": "Crystal Mountain MI",
    "Whitefish Mountain": "Whitefish",
    "Brundage Mountain": "Brundage",
    "Red Lodge Mountain": "Red Lodge",
    "Wildcat Mountain": "Wildcat",
    "Windham Mountain": "Windham",
}

def find_resort_by_name(mountain_name, state_code=None):
    """
    Find a Resort by legacy mountain name with case-insensitive matching.
    Returns Resort or None if no match found.
    
    Matching strategy:
    1. Check hardcoded alias map first
    2. Exact case-insensitive match
    3. Match with common suffix variations (Resort, Mountain, Ski Area)
    """
    from sqlalchemy import func
    
    if not mountain_name:
        return None
    
    aliased_name = MOUNTAIN_NAME_ALIASES.get(mountain_name, mountain_name)
    
    query = Resort.query.filter(Resort.is_active == True)
    if state_code:
        query = query.filter(Resort.state == state_code)
    
    exact_match = query.filter(func.lower(Resort.name) == aliased_name.lower()).first()
    if exact_match:
        return exact_match
    
    suffix_variations = [
        aliased_name,
        f"{aliased_name} Resort",
        f"{aliased_name} Mountain",
        f"{aliased_name} Mountain Resort",
        f"{aliased_name} Ski Area",
        aliased_name.replace(" Resort", ""),
        aliased_name.replace(" Mountain", ""),
        aliased_name.replace(" Ski Area", ""),
    ]
    
    for variation in suffix_variations:
        match = query.filter(func.lower(Resort.name) == variation.lower()).first()
        if match:
            return match
    
    return None


def build_mountain_to_resort_mapping():
    """
    Build a complete mapping of MOUNTAINS_BY_STATE strings to Resort IDs.
    Returns: (mapping_dict, unmatched_list)
    """
    mapping = {}
    unmatched = []
    
    for state_code, mountains in MOUNTAINS_BY_STATE.items():
        for mountain_name in mountains:
            resort = find_resort_by_name(mountain_name, state_code)
            if resort:
                mapping[mountain_name] = resort.id
            else:
                unmatched.append({"name": mountain_name, "state": state_code})
    
    return mapping, unmatched


@app.route("/admin/backfill-resort-ids", methods=["GET", "POST"])
@login_required
@admin_required
def backfill_resort_ids_endpoint():
    """
    Backfill visited_resort_ids and home_resort_id from legacy string data.
    Idempotent - safe to run multiple times.
    
    GET: Preview mode - shows what would be migrated
    POST: Execute mode - performs the migration
    
    Usage: GET https://yourapp.replit.dev/admin/backfill-resort-ids
    """
    try:
        mapping, unmatched = build_mountain_to_resort_mapping()
        
        is_preview = request.method == "GET"
        
        users_with_visited = User.query.filter(
            db.or_(
                User.mountains_visited.isnot(None),
                User.home_mountain.isnot(None)
            )
        ).all()
        
        results = {
            "mapping_stats": {
                "total_mountains": sum(len(m) for m in MOUNTAINS_BY_STATE.values()),
                "mapped": len(mapping),
                "unmatched": len(unmatched)
            },
            "unmatched_mountains": unmatched,
            "users_processed": 0,
            "visited_mountains_migrated": 0,
            "home_mountains_migrated": 0,
            "unmapped_visited_names": [],
            "unmapped_home_names": []
        }
        
        for user in users_with_visited:
            legacy_visited = user.mountains_visited or []
            legacy_home = user.home_mountain
            
            if legacy_visited and (not user.visited_resort_ids or len(user.visited_resort_ids) == 0):
                new_ids = []
                for mountain_name in legacy_visited:
                    if mountain_name in mapping:
                        new_ids.append(mapping[mountain_name])
                    else:
                        resort = find_resort_by_name(mountain_name)
                        if resort:
                            new_ids.append(resort.id)
                        elif mountain_name not in results["unmapped_visited_names"]:
                            results["unmapped_visited_names"].append(mountain_name)
                
                if new_ids:
                    if not is_preview:
                        user.visited_resort_ids = list(set(new_ids))
                    results["visited_mountains_migrated"] += len(new_ids)
            
            if legacy_home and not user.home_resort_id:
                if legacy_home in mapping:
                    if not is_preview:
                        user.home_resort_id = mapping[legacy_home]
                    results["home_mountains_migrated"] += 1
                else:
                    resort = find_resort_by_name(legacy_home)
                    if resort:
                        if not is_preview:
                            user.home_resort_id = resort.id
                        results["home_mountains_migrated"] += 1
                    elif legacy_home not in results["unmapped_home_names"]:
                        results["unmapped_home_names"].append(legacy_home)
            
            results["users_processed"] += 1
        
        if not is_preview:
            db.session.commit()
        
        return jsonify({
            "status": "success",
            "mode": "preview" if is_preview else "executed",
            "message": f"{'Preview' if is_preview else 'Executed'} backfill for {results['users_processed']} users",
            "results": results
        }), 200
        
    except Exception as e:
        import traceback
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": f"Backfill failed: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route("/admin/seed-test-users", methods=["GET", "POST"])
@login_required
@admin_required
def seed_test_users_endpoint():
    """
    HTTP endpoint to seed test users for demo/testing.
    Creates Richard + 20 friends with complete profiles, trips, and friendships.
    
    Usage: GET https://yourapp.replit.dev/admin/seed-test-users
    
    This is idempotent - safe to call multiple times.
    """
    try:
        from seed_test_users import seed_test_data
        from models import EquipmentSetup, EquipmentSlot, EquipmentDiscipline
        
        results = seed_test_data(
            app, db, User, Friend, SkiTrip, Resort,
            EquipmentSetup, EquipmentSlot, EquipmentDiscipline
        )
        
        return jsonify({
            "status": "success",
            "message": "Test users seeded successfully",
            "details": results
        }), 200
    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "message": f"Failed to seed test users: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route("/admin/seed-narrative-states", methods=["GET", "POST"])
@login_required
@admin_required
def seed_narrative_states_endpoint():
    """
    HTTP endpoint to seed 4 test users for narrative state validation.
    Creates users for State 1, 2, 3, and 4 for testing NBA behavior.
    
    Usage: GET https://yourapp.replit.dev/admin/seed-narrative-states
    
    Test user logins (password: testpass123):
    - state1.test@baselodge.dev (State 1: Early Onboarding)
    - state2.test@baselodge.dev (State 2: Profile Complete, Not Planning)
    - state3.test@baselodge.dev (State 3: Planning Started, Not Fully Active)
    - state4.test@baselodge.dev (State 4: Active User)
    """
    try:
        from seed_test_users import seed_narrative_state_users
        
        results = seed_narrative_state_users(app, db, User, SkiTrip, Resort)
        
        return jsonify({
            "status": "success",
            "message": "Narrative state test users seeded",
            "details": results
        }), 200
    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "message": f"Failed to seed narrative state users: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route("/admin/seed-screenshot-data", methods=["GET", "POST"])
@login_required
@admin_required
def seed_screenshot_data_endpoint():
    """
    Seed App Store / marketing screenshot demo data.

    Creates John Carter (baselodge.screenshots@gmail.com / Ikon) plus 5 confirmed
    friends, 2 pending-request stubs, and realistic SkiTrip rows designed to
    populate every screenshot surface (Home, Friends, Ideas, Mountains, Profile).

    Idempotent — safe to call multiple times. Existing rows are skipped.

    Usage: GET /admin/seed-screenshot-data
    """
    try:
        from seed_screenshots import seed_screenshot_data
        results = seed_screenshot_data(
            app, db,
            User, Friend, SkiTrip, Invitation, SkiTripParticipant,
            Resort, GuestStatus, InviteType,
        )
        return jsonify({
            "status":  "success",
            "message": "Screenshot seed data created successfully",
            "details": results,
        }), 200
    except Exception as e:
        import traceback
        return jsonify({
            "status":    "error",
            "message":   f"Screenshot seed failed: {str(e)}",
            "traceback": traceback.format_exc(),
        }), 500


@app.route("/admin/seed-screenshot-expansion", methods=["GET", "POST"])
@login_required
@admin_required
def seed_screenshot_expansion_endpoint():
    """
    Expand App Store screenshot data — adds 28 new fictional users, 20 accepted
    friendships, 5 pending inbound requests, 3 pending outbound requests, group
    trips, availability windows, and activity notifications for John Carter.

    Idempotent — safe to call multiple times. All writes are get-or-create.

    Usage: GET /admin/seed-screenshot-expansion
    """
    try:
        from seed_screenshots_expansion import seed_screenshot_expansion
        results = seed_screenshot_expansion(
            app, db,
            User, Friend, SkiTrip, Invitation, SkiTripParticipant,
            Resort, UserAvailability, Activity, GuestStatus, InviteType,
        )
        return jsonify({
            "status":  "success",
            "message": "Screenshot expansion seed completed successfully",
            "details": results,
        }), 200
    except Exception as e:
        import traceback
        return jsonify({
            "status":    "error",
            "message":   f"Screenshot expansion seed failed: {str(e)}",
            "traceback": traceback.format_exc(),
        }), 500


@app.route("/admin/backfill-planning-timestamp", methods=["GET", "POST"])
@login_required
@admin_required
def backfill_planning_timestamp_endpoint():
    """
    HTTP endpoint to backfill first_planning_timestamp for existing users.
    
    Usage: GET https://yourapp.replit.dev/admin/backfill-planning-timestamp
    
    This is idempotent and safe to run multiple times.
    Only updates users who have trips but no first_planning_timestamp set.
    """
    try:
        from backfill_first_planning_timestamp import backfill_first_planning_timestamp
        
        try:
            from models import TripGuest
        except ImportError:
            TripGuest = None
        
        results = backfill_first_planning_timestamp(app, db, User, SkiTrip, TripGuest)
        
        return jsonify({
            "status": "success",
            "message": "Backfill completed",
            "details": results
        }), 200
    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "message": f"Backfill failed: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route("/admin/backfill-primary-rider-type", methods=["GET", "POST"])
@login_required
@admin_required
def backfill_primary_rider_type_endpoint():
    """
    HTTP endpoint to backfill primary_rider_type from legacy rider_type for existing users.
    
    Usage: GET https://yourapp.replit.dev/admin/backfill-primary-rider-type
    
    This is idempotent and safe to run multiple times.
    Only updates users who have rider_type but no primary_rider_type set.
    """
    try:
        users_updated = 0
        users_skipped = 0
        
        users = User.query.all()
        for user in users:
            if user.rider_type and not user.primary_rider_type:
                user.primary_rider_type = user.rider_type
                users_updated += 1
            else:
                users_skipped += 1
        
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": "Backfill completed",
            "details": {
                "users_updated": users_updated,
                "users_skipped": users_skipped
            }
        }), 200
    except Exception as e:
        import traceback
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": f"Backfill failed: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route("/admin/backfill-organizers-as-participants", methods=["GET", "POST"])
@login_required
@admin_required
def backfill_organizers_as_participants():
    """
    HTTP endpoint to backfill trip organizers as participants.
    
    Usage: GET https://yourapp.replit.dev/admin/backfill-organizers-as-participants
    
    This is idempotent - safe to run multiple times.
    Only creates participant records for trips where the owner is not already a participant.
    """
    try:
        trips_updated = 0
        trips_skipped = 0
        
        all_trips = SkiTrip.query.all()
        for trip in all_trips:
            # Check if owner already has a participant record
            existing = SkiTripParticipant.query.filter_by(
                trip_id=trip.id,
                user_id=trip.user_id
            ).first()
            
            if existing:
                trips_skipped += 1
            else:
                # Add owner as participant with OWNER role
                trip.add_owner_as_participant()
                trips_updated += 1
        
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": "Backfill completed",
            "details": {
                "trips_updated": trips_updated,
                "trips_skipped": trips_skipped,
                "total_trips": len(all_trips)
            }
        }), 200
    except Exception as e:
        import traceback
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": f"Backfill failed: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500


@app.route("/admin/run-backfill-last-active", methods=["GET", "POST"])
@login_required
@admin_required
def run_backfill_last_active():
    """
    One-time corrective backfill of User.last_active_at from historical activity.

    PERMANENTLY DISABLED after execution — the guard below must never be removed.
    See: .agents/memory/activity-tracking.md for the full audit trail.

    GET  → dry-run preview (no writes)
    POST ?confirm=yes → live write
    """
    # ── Permanent disable guard ───────────────────────────────────────────────
    # This route executed once on 2026-05-28 and is now locked.
    # Remove only if a second targeted backfill is explicitly required.
    _BACKFILL_EXECUTED = True
    if _BACKFILL_EXECUTED:
        return jsonify({
            "status": "disabled",
            "reason": "One-time backfill already executed on 2026-05-28. Route is permanently locked.",
        }), 410

    # ── Build the evidence query ──────────────────────────────────────────────
    # NOTE: LEAST(NOW(), NULL) = NOW() in PostgreSQL (unlike GREATEST).
    # Use CASE WHEN to ensure users with ALL-NULL signals stay NULL.
    _SIGNAL_SQL = text("""
        SELECT u.id, u.first_name, u.is_seeded,
            CASE
              WHEN GREATEST(
                u.profile_completed_at,
                u.password_changed_at,
                (SELECT MAX(COALESCE(st.updated_at, st.created_at))
                 FROM ski_trip st WHERE st.user_id = u.id),
                (SELECT MAX(f.created_at) FROM friend f
                 WHERE f.user_id = u.id),
                (SELECT MAX(i.created_at) FROM invitation i
                 WHERE i.sender_id = u.id),
                (SELECT MAX(es.created_at) FROM equipment_setup es
                 WHERE es.user_id = u.id),
                (SELECT MAX(mpv.viewed_at) FROM mountain_page_view mpv
                 WHERE mpv.user_id = u.id),
                (SELECT MAX(e.created_at) FROM event e
                 WHERE e.user_id = u.id),
                (SELECT MAX(tit.created_at) FROM trip_invite_token tit
                 WHERE tit.inviter_user_id = u.id)
              ) IS NOT NULL
              THEN LEAST(NOW(),
                GREATEST(
                  u.profile_completed_at,
                  u.password_changed_at,
                  (SELECT MAX(COALESCE(st.updated_at, st.created_at))
                   FROM ski_trip st WHERE st.user_id = u.id),
                  (SELECT MAX(f.created_at) FROM friend f
                   WHERE f.user_id = u.id),
                  (SELECT MAX(i.created_at) FROM invitation i
                   WHERE i.sender_id = u.id),
                  (SELECT MAX(es.created_at) FROM equipment_setup es
                   WHERE es.user_id = u.id),
                  (SELECT MAX(mpv.viewed_at) FROM mountain_page_view mpv
                   WHERE mpv.user_id = u.id),
                  (SELECT MAX(e.created_at) FROM event e
                   WHERE e.user_id = u.id),
                  (SELECT MAX(tit.created_at) FROM trip_invite_token tit
                   WHERE tit.inviter_user_id = u.id)
                )
              )
              ELSE NULL
            END AS best_ts
        FROM "user" u
        WHERE u.last_active_at IS NULL
        ORDER BY best_ts DESC NULLS LAST
    """)

    rows = db.session.execute(_SIGNAL_SQL).fetchall()
    will_update = [(r[0], r[1], r[2], r[3]) for r in rows if r[3] is not None]
    will_skip   = [(r[0], r[1], r[2])       for r in rows if r[3] is None]

    # ── Dry-run (GET) ─────────────────────────────────────────────────────────
    if request.method == "GET":
        timestamps = [row[3] for row in will_update]
        return jsonify({
            "mode": "dry_run",
            "total_null_last_active": len(rows),
            "will_update": len(will_update),
            "will_skip":   len(will_skip),
            "newest_ts": str(max(timestamps)) if timestamps else None,
            "oldest_ts": str(min(timestamps)) if timestamps else None,
            "detail": [
                {"id": uid, "name": name, "is_seeded": seeded, "inferred_ts": str(ts)}
                for uid, name, seeded, ts in will_update
            ],
            "skipped_ids": [uid for uid, _, _ in will_skip],
        }), 200

    # ── Live write (POST ?confirm=yes) ────────────────────────────────────────
    if request.args.get("confirm") != "yes":
        return jsonify({
            "status": "error",
            "reason": "POST requires ?confirm=yes query param. Run GET first for preview.",
        }), 400

    updated = []
    skipped_count = len(will_skip)
    try:
        for uid, name, seeded, best_ts in will_update:
            user = db.session.get(User, uid)
            if user and user.last_active_at is None:
                user.last_active_at = best_ts
                updated.append({"id": uid, "name": name, "ts": str(best_ts)})
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "reason": str(exc)}), 500

    # ── Post-write verification ───────────────────────────────────────────────
    now = datetime.utcnow()
    seven_ago  = now - timedelta(days=7)
    thirty_ago = now - timedelta(days=30)
    wau_after = User.query.filter(User.last_active_at >= seven_ago).count()
    mau_after = User.query.filter(User.last_active_at >= thirty_ago).count()
    future_ts = User.query.filter(User.last_active_at > now).count()

    return jsonify({
        "status": "success",
        "updated":  len(updated),
        "skipped":  skipped_count,
        "detail":   updated,
        "verification": {
            "wau_after":          wau_after,
            "mau_after":          mau_after,
            "future_timestamps":  future_ts,
        },
    }), 200


@app.route("/open-data-debug")
@login_required
@admin_required
def open_data_debug():
    """
    Debug endpoint to verify open date matching logic.
    Returns JSON with all open date matches for the current user.
    """
    matches = get_open_date_matches(current_user)
    
    return jsonify({
        "user_id": current_user.id,
        "user_email": current_user.email,
        "user_open_dates": current_user.open_dates or [],
        "user_pass_type": current_user.pass_type,
        "matches_count": len(matches),
        "matches": matches
    })


# ============================================================================
# GroupTrip Social Functionality
# ============================================================================

@app.route("/api/group-trip/create", methods=["POST"])
@login_required
def create_group_trip():
    """Create a new GroupTrip."""
    data = request.get_json()
    
    title = data.get("title", "").strip() or None
    start_date_str = data.get("start_date")
    end_date_str = data.get("end_date")
    
    # Validate dates
    if not start_date_str or not end_date_str:
        return jsonify({"success": False, "error": "Start and end dates are required"}), 400
    
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"success": False, "error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    if end_date < start_date:
        return jsonify({"success": False, "error": "End date cannot be before start date"}), 400
    
    # Create GroupTrip
    trip = GroupTrip(
        host_id=current_user.id,
        title=title,
        start_date=start_date,
        end_date=end_date
    )
    db.session.add(trip)
    db.session.commit()
    ph_analytics.track(current_user.id, 'trip_created', {
        'trip_type': 'group_trip',
        'days':      (end_date - start_date).days + 1,
        'source':    'create_group_trip_api',
    })

    return jsonify({
        "success": True,
        "trip_id": trip.id,
        "redirect_url": url_for("view_group_trip", trip_id=trip.id)
    })


@app.route("/group-trip/<int:trip_id>")
@login_required
def view_group_trip(trip_id):
    """View GroupTrip details with invite form (host only)."""
    trip = GroupTrip.query.get_or_404(trip_id)
    
    # Explicit permission check: only host or guests can view
    is_host = trip.host_id == current_user.id
    is_guest = TripGuest.query.filter_by(trip_id=trip_id, user_id=current_user.id).first() is not None
    
    if not is_host and not is_guest:
        abort(403)
    
    # Get guests with their details
    guests = TripGuest.query.filter_by(trip_id=trip_id).all()
    
    # Get user's friends for invite form (host only)
    user_friends = Friend.query.filter_by(user_id=current_user.id).all()
    friend_ids = [f.friend_id for f in user_friends]
    friends = User.query.filter(User.id.in_(friend_ids)).all() if friend_ids else []
    
    # Filter out already invited/joined friends
    invited_ids = {g.user_id for g in guests}
    available_friends = [f for f in friends if f.id not in invited_ids]
    
    return render_template(
        "group_trip_detail.html",
        trip=trip,
        is_host=is_host,
        guests=guests,
        available_friends=available_friends
    )


@app.route("/group-trip/<int:trip_id>/invite", methods=["POST"])
@login_required
def invite_to_group_trip(trip_id):
    """Host invites a friend to GroupTrip."""
    # Authentication guard (already protected by @login_required)
    if not current_user.is_authenticated:
        abort(401)
    
    # Validate trip exists
    trip = GroupTrip.query.get_or_404(trip_id)
    
    # Permission check: only host can invite
    if trip.host_id != current_user.id:
        abort(403)
    
    # Safe form data handling
    friend_id = request.form.get("friend_id", type=int)
    if not friend_id:
        flash("No friend selected.", "error")
        return redirect(url_for("view_group_trip", trip_id=trip_id))
    
    # Validate target user exists
    friend = db.session.get(User, friend_id)
    if not friend:
        flash("User not found.", "error")
        return redirect(url_for("view_group_trip", trip_id=trip_id))
    
    # Check if user is actually a friend
    is_friend = Friend.query.filter_by(user_id=current_user.id, friend_id=friend_id).first()
    if not is_friend:
        flash("User is not in your friends list.", "error")
        return redirect(url_for("view_group_trip", trip_id=trip_id))
    
    # Prevent duplicate invites
    existing = TripGuest.query.filter_by(trip_id=trip_id, user_id=friend_id).first()
    if existing:
        flash(f"{friend.first_name} is already invited.", "error")
        return redirect(url_for("view_group_trip", trip_id=trip_id))
    
    # Database write safety
    try:
        guest = TripGuest(trip_id=trip_id, user_id=friend_id, status=GuestStatus.INVITED)
        db.session.add(guest)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Invite to group trip failed")
        flash("An error occurred while inviting. Please try again.", "error")
        return redirect(url_for("view_group_trip", trip_id=trip_id))
    
    flash(f"Invited {friend.first_name} to the trip!", "success")
    return redirect(url_for("view_group_trip", trip_id=trip_id))


@app.route("/group-trip/<int:trip_id>/accept", methods=["POST"])
@login_required
def accept_group_trip_invite(trip_id):
    """Accept GroupTrip invite and create global friend connection."""
    # Authentication guard (already protected by @login_required)
    if not current_user.is_authenticated:
        abort(401)
    
    # Validate trip exists
    trip = GroupTrip.query.get_or_404(trip_id)
    
    # Validate guest exists
    guest = TripGuest.query.filter_by(trip_id=trip_id, user_id=current_user.id).first_or_404()
    
    # Only allow if invited (not already accepted)
    if guest.status != GuestStatus.INVITED:
        flash("You've already accepted or this invite is invalid.", "error")
        return redirect(url_for("home"))
    
    # Database write safety
    try:
        # Accept the invite
        guest.status = GuestStatus.ACCEPTED
        
        # Mark planning started (lifecycle signal)
        current_user.mark_planning_started()
        
        db.session.commit()
        
        # Create bidirectional friend connection with trip host if not already friends
        host = trip.host
        existing = Friend.query.filter_by(user_id=current_user.id, friend_id=host.id).first()
        if not existing:
            f1 = Friend(user_id=current_user.id, friend_id=host.id)
            f2 = Friend(user_id=host.id, friend_id=current_user.id)
            db.session.add(f1)
            db.session.add(f2)
            db.session.commit()
            flash(f"Trip accepted! You're now connected with {host.first_name}.", "success")
        else:
            flash("Trip accepted!", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Accept group trip invite failed")
        flash("An error occurred while accepting the invite. Please try again.", "error")
        return redirect(url_for("home"))
    
    return redirect(url_for("home"))


@app.route("/group-trip/<int:trip_id>/leave", methods=["POST"])
@login_required
def leave_group_trip(trip_id):
    """Guest leaves GroupTrip (deletes TripGuest row)."""
    trip = GroupTrip.query.get_or_404(trip_id)
    guest = TripGuest.query.filter_by(trip_id=trip_id, user_id=current_user.id).first_or_404()
    
    # Delete the guest record
    db.session.delete(guest)
    db.session.commit()
    
    flash("You've left the trip.", "success")
    return redirect(url_for("home"))


@app.route("/group-trip/<int:trip_id>/remove-guest/<int:guest_id>", methods=["POST"])
@login_required
def remove_trip_guest(trip_id, guest_id):
    """Host removes a guest from GroupTrip."""
    trip = GroupTrip.query.get_or_404(trip_id)
    
    # Only host can remove
    if trip.host_id != current_user.id:
        return abort(403)
    
    guest = TripGuest.query.filter_by(id=guest_id, trip_id=trip_id).first_or_404()
    guest_user = guest.user
    
    db.session.delete(guest)
    db.session.commit()
    
    flash(f"Removed {guest_user.first_name} from the trip.", "success")
    return redirect(url_for("view_group_trip", trip_id=trip_id))


@app.route("/connect-from-trip/<int:user_id>", methods=["POST"])
@login_required
def connect_from_trip(user_id):
    """Connect with a user via shared GroupTrip (create global friendship)."""
    user_to_connect = User.query.get_or_404(user_id)
    
    # Check eligibility
    if not check_shared_upcoming_trip(current_user.id, user_to_connect.id):
        flash("You don't share an upcoming trip with this user.", "error")
        return redirect(url_for("friend_profile", friend_id=user_id))
    
    # Check if already friends
    existing = Friend.query.filter_by(user_id=current_user.id, friend_id=user_id).first()
    if existing:
        flash("You're already connected with this user.", "info")
        return redirect(url_for("friend_profile", friend_id=user_id))
    
    # Create bidirectional friend connection
    f1 = Friend(user_id=current_user.id, friend_id=user_to_connect.id)
    f2 = Friend(user_id=user_to_connect.id, friend_id=current_user.id)
    db.session.add(f1)
    db.session.add(f2)
    # Same helper + same argument order as accept_invitation: actor=current_user, other=them
    emit_connection_accepted_activity(current_user.id, user_to_connect.id)
    db.session.commit()

    # One-time Home card for the acting user — same session pattern as accept_invitation
    session['new_connection_name'] = (
        user_to_connect.first_name or user_to_connect.username or 'your new friend'
    )

    flash(f"Connected with {user_to_connect.first_name}!", "success")
    return redirect(url_for("friend_profile", friend_id=user_id))


# ============================================================================
# Equipment & GroupTrip Status Management
# ============================================================================

@app.route("/profile/equipment/delete", methods=["POST"])
@login_required
def delete_equipment():
    """Delete equipment setup by slot (Primary/Secondary)."""
    data = request.get_json()
    
    slot_name = data.get("slot", "").upper()  # "PRIMARY" or "SECONDARY"
    
    if slot_name not in ["PRIMARY", "SECONDARY"]:
        return jsonify({"success": False, "error": "Invalid slot"}), 400
    
    try:
        slot = EquipmentSlot[slot_name]
    except KeyError:
        return jsonify({"success": False, "error": "Invalid slot"}), 400
    
    # Find equipment - explicit permission check
    equipment = EquipmentSetup.query.filter_by(user_id=current_user.id, slot=slot).first()
    
    if not equipment:
        return jsonify({"success": False, "error": "Equipment not found"}), 404
    
    if equipment.user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    # Delete
    db.session.delete(equipment)
    db.session.commit()
    
    return jsonify({"success": True, "message": f"{slot.value} equipment deleted"})


@app.route("/profile/equipment", methods=["POST"])
@login_required
def save_equipment():
    """Save or update equipment setup (Primary/Secondary)."""
    data = request.get_json()
    
    slot_name = data.get("slot", "").upper()  # "PRIMARY" or "SECONDARY"
    discipline_name = data.get("discipline", "").upper()  # "SKIER" or "SNOWBOARDER"
    brand = data.get("brand", "").strip()
    length_cm = data.get("length_cm")
    width_mm = data.get("width_mm")
    
    # Validate
    if not brand or slot_name not in ["PRIMARY", "SECONDARY"] or discipline_name not in ["SKIER", "SNOWBOARDER"]:
        return jsonify({"success": False, "error": "Invalid input"}), 400
    
    # Convert to enums
    try:
        slot = EquipmentSlot[slot_name]
        discipline = EquipmentDiscipline[discipline_name]
    except KeyError:
        return jsonify({"success": False, "error": "Invalid slot or discipline"}), 400
    
    # Explicit permission check: only owner can edit
    equipment = EquipmentSetup.query.filter_by(user_id=current_user.id, slot=slot).first()
    if equipment and equipment.user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    # Find or create equipment
    if not equipment:
        equipment = EquipmentSetup(user_id=current_user.id, slot=slot)
    
    equipment.discipline = discipline
    equipment.brand = brand
    equipment.length_cm = int(length_cm) if length_cm else None
    equipment.width_mm = int(width_mm) if width_mm else None
    
    db.session.add(equipment)
    db.session.commit()
    
    return jsonify({"success": True, "message": f"{slot.value} equipment saved"})


@app.route("/group-trip/<int:trip_id>/transportation", methods=["POST"])
@login_required
def update_group_trip_transportation(trip_id):
    """Update transportation status (host-only)."""
    trip = GroupTrip.query.get_or_404(trip_id)
    
    if trip.host_id != current_user.id:
        return jsonify({"success": False, "error": "Only host can update"}), 403
    
    data = request.get_json()
    status_name = data.get("transportation_status", "").upper()
    
    if status_name == "":
        trip.transportation_status = None
    elif status_name in ["HAVE_TRANSPORT", "NEED_TRANSPORT", "NOT_SURE"]:
        try:
            trip.transportation_status = TransportationStatus[status_name]
        except KeyError:
            return jsonify({"success": False, "error": "Invalid status"}), 400
    else:
        return jsonify({"success": False, "error": "Invalid status"}), 400
    
    db.session.commit()
    return jsonify({"success": True})


# ============================================================================
# CANONICAL PASS BRAND MAPPINGS (for backfill)
# ============================================================================

EPIC_RESORT_NAMES = {
    "Heavenly Mountain Resort", "Kirkwood Mountain Resort", "Northstar California Resort",
    "Vail Mountain", "Beaver Creek", "Breckenridge Ski Resort", "Keystone Resort",
    "Crested Butte Mountain Resort", "Mt. Brighton", "Afton Alps", "Hidden Valley Ski Resort",
    "Attitash Mountain Resort", "Wildcat Mountain", "Hunter Mountain", "Boston Mills",
    "Brandywine", "Mad River Mountain", "Jack Frost Big Boulder", "Roundtop Mountain Resort",
    "Whitetail Resort", "Liberty Mountain Resort", "Park City Mountain", "Mount Snow",
    "Okemo Mountain Resort", "Stowe Mountain Resort", "Stevens Pass", "Wilmot Mountain"
}

INDY_RESORT_NAMES = {
    "Bear Valley", "China Peak", "Dodge Ridge", "Sunlight Mountain Resort",
    "Powderhorn Mountain Resort", "Ski Cooper", "Brundage Mountain", "Tamarack Resort",
    "Lookout Pass", "Saddleback Mountain", "Marquette Mountain", "Blacktail Mountain",
    "Lost Trail Powder Mountain", "Red Lodge Mountain", "Cannon Mountain", "Ski Santa Fe",
    "Titus Mountain", "Willamette Pass", "Blue Knob All Seasons Resort", "Beaver Mountain",
    "Eagle Point Resort", "Bolton Valley", "Magic Mountain", "White Pass", "Snow King Mountain"
}

MOUNTAIN_COLLECTIVE_RESORT_NAMES = {
    "Aspen Snowmass", "Alta Ski Area", "Snowbird", "Jackson Hole Mountain Resort",
    "Sun Valley", "Sugarbush Resort", "Taos Ski Valley"
}


@app.cli.command("backfill-pass-brands")
@click.option("--force", is_flag=True, help="Force re-run even if already populated")
def backfill_pass_brands(force):
    """Backfill pass_brands column for all resorts. Idempotent by default."""
    created = 0
    updated = 0
    skipped = 0
    null_count = 0

    with app.app_context():
        resorts = Resort.query.all()

        for resort in resorts:
            # Skip if already populated and not forcing
            if resort.pass_brands and not force:
                skipped += 1
                continue

            original_pass_brands = resort.pass_brands
            new_pass_brands = None

            # Priority 1: Mountain Collective (Ikon overlap)
            if resort.name in MOUNTAIN_COLLECTIVE_RESORT_NAMES:
                new_pass_brands = "Ikon,MountainCollective"
                resort.brand = "Ikon"

            # Priority 2: Epic
            elif resort.name in EPIC_RESORT_NAMES:
                new_pass_brands = "Epic"
                resort.brand = "Epic"

            # Priority 3: Indy
            elif resort.name in INDY_RESORT_NAMES:
                new_pass_brands = "Indy"
                resort.brand = "Indy"

            # Priority 4: Existing Ikon (default)
            elif resort.brand == "Ikon":
                new_pass_brands = "Ikon"

            # Fallback: Use existing brand
            else:
                new_pass_brands = resort.brand or "Other"

            # Update if changed
            if new_pass_brands != original_pass_brands:
                resort.pass_brands = new_pass_brands
                db.session.commit()
                if original_pass_brands is None:
                    created += 1
                    print(f"  ✨ CREATED: {resort.name} ({resort.state}) → {new_pass_brands}")
                else:
                    updated += 1
                    print(f"  ✏️  UPDATED: {resort.name} ({resort.state}) → {new_pass_brands} (was: {original_pass_brands})")
            else:
                skipped += 1

        # Verify no nulls
        null_check = Resort.query.filter(Resort.pass_brands.is_(None)).count()

        print("\n" + "=" * 70)
        print("BACKFILL SUMMARY")
        print("=" * 70)
        print(f"Total resorts: {len(resorts)}")
        print(f"Pass brands created: {created}")
        print(f"Pass brands updated: {updated}")
        print(f"Pass brands skipped: {skipped}")
        print(f"Resorts with NULL pass_brands: {null_check}")
        print()

        # Distribution by pass
        epic_count = Resort.query.filter(Resort.pass_brands.contains("Epic")).count()
        ikon_count = Resort.query.filter(Resort.pass_brands.contains("Ikon")).count()
        indy_count = Resort.query.filter(Resort.pass_brands.contains("Indy")).count()
        mountain_collective_count = Resort.query.filter(Resort.pass_brands.contains("MountainCollective")).count()

        print("Distribution:")
        print(f"  - Epic: {epic_count}")
        print(f"  - Ikon: {ikon_count}")
        print(f"  - Indy: {indy_count}")
        print(f"  - MountainCollective: {mountain_collective_count}")
        print()

        # Sample resorts
        print("Sample Results (before/after):")
        samples = [
            ("Park City Mountain", "Epic"),
            ("Bolton Valley", "Indy"),
            ("Aspen Snowmass", "Ikon,MountainCollective"),
            ("Jackson Hole Mountain Resort", "Ikon,MountainCollective"),
        ]
        for name, expected_brands in samples:
            resort = Resort.query.filter_by(name=name).first()
            if resort:
                status = "✓" if resort.pass_brands == expected_brands else "✗"
                print(f"  {status} {name}: {resort.pass_brands} (expected: {expected_brands})")
            else:
                print(f"  ✗ {name}: NOT FOUND")

        print()
        print("✅ Backfill complete!")
        print("=" * 70)


# ============================================================================
# DEMO DATA SEEDING (FULL WORLD)
# ============================================================================

SKIER_BRANDS = SKI_BRANDS
SNOWBOARDER_BRANDS = SNOWBOARD_BRANDS
PASS_OPTIONS_SEEDING = ["Epic", "Ikon", "MountainCollective", "Indy", "PowderAlliance", "Freedom", "SkiCalifornia", "Other", "None"]

FIRST_NAMES = ["Alex", "Jordan", "Sam", "Casey", "Riley", "Morgan", "Jamie", "Taylor", "Jesse", "Charlie", "Skylar", "Quinn", "Dakota", "Avery", "Blake", "Parker", "Rowan", "Drew", "Phoenix", "River", "Jade", "Connor", "Reese", "Emerson", "Sage", "Justice", "Scout", "Lex", "Hayden", "Aspen", "Storm", "Finley", "Devyn", "Canyon", "Sierra", "Teton", "Range", "Peak", "Boulder", "Summit", "Ridge", "Trail", "Alpine", "Powder", "Mogul", "Gnar", "Shred", "Carve", "Slate", "Blake", "Bailey", "Cameron"]

LAST_NAMES = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin", "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark", "Ramirez", "Lewis", "Robinson"]

TRIP_TITLES = ["powder day mission", "spring corn runs", "lake tahoe adventure", "utah powder week", "colorado peaks", "backcountry tour", "resort lap day", "mogul practice", "tree skiing", "alpine exploration"]


@app.cli.command("seed-full-demo-world")
def seed_full_demo_world():
    """Seed comprehensive demo data for end-to-end testing."""
    print("🌍 SEEDING FULL DEMO WORLD...")
    print("=" * 70)
    
    with app.app_context():
        # ====== FIXED USERS ======
        richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
        if not richard:
            richard = User(
                first_name="Richard", last_name="Battle-Baxter",
                email="richardbattlebaxter@gmail.com",
                primary_rider_type="Skier",
                pass_type="Epic", skill_level="Advanced",
                home_state="Colorado", birth_year=1985
            )
            richard.set_password("12345678")
            db.session.add(richard)
            print("✨ Created: Richard Battle-Baxter")
        else:
            print("⊘ Skipped: Richard (already exists)")
        
        jonathan = User.query.filter_by(email="jonathanmschmitz@gmail.com").first()
        if not jonathan:
            jonathan = User(
                first_name="Jonathan", last_name="Schmitz",
                email="jonathanmschmitz@gmail.com",
                primary_rider_type="Skier",
                pass_type="Ikon,MountainCollective", skill_level="Advanced",
                home_state="Utah", birth_year=1990
            )
            jonathan.set_password("12345678")
            db.session.add(jonathan)
            print("✨ Created: Jonathan Schmitz")
        else:
            print("⊘ Skipped: Jonathan (already exists)")
        
        db.session.commit()
        
        # ====== DUMMY USERS (50) ======
        dummy_users = []
        for i in range(50):
            email = f"user{i+1}@baselodge.local"
            if User.query.filter_by(email=email).first():
                print(f"⊘ Skipped: {email} (already exists)")
                dummy_users.append(User.query.filter_by(email=email).first())
                continue
            
            primary_rt = random.choice(["Skier", "Snowboarder"])
            user = User(
                first_name=random.choice(FIRST_NAMES),
                last_name=random.choice(LAST_NAMES),
                email=email,
                primary_rider_type=primary_rt,
                skill_level=random.choice(["Beginner", "Intermediate", "Advanced", "Expert"]),
                home_state=random.choice(["Colorado", "Utah", "California", "Wyoming", "Montana", "Idaho", "Washington"]),
                birth_year=random.randint(1970, 2005)
            )
            
            # 70% single pass, 30% multi-pass
            if random.random() < 0.7:
                user.pass_type = random.choice(["Epic", "Ikon", "Indy", "Other"])
            else:
                user.pass_type = ",".join(sorted(set(random.sample(PASS_OPTIONS_SEEDING[:-2], 2))))
            
            user.set_password("12345678")
            db.session.add(user)
            dummy_users.append(user)
        
        db.session.commit()
        print(f"✨ Created: {len(dummy_users)} dummy users")
        
        # ====== EQUIPMENT ======
        all_users = [richard, jonathan] + dummy_users
        equipment_count = 0
        for user in all_users:
            if EquipmentSetup.query.filter_by(user_id=user.id, is_primary=True).first():
                continue

            user_rt = user.primary_rider_type or user.rider_type or "Skier"
            discipline = EquipmentDiscipline.SKIER if user_rt == "Skier" else EquipmentDiscipline.SNOWBOARDER
            brands = SKIER_BRANDS if user_rt == "Skier" else SNOWBOARDER_BRANDS

            primary = EquipmentSetup(
                user_id=user.id,
                slot=EquipmentSlot.PRIMARY,
                is_primary=True,
                discipline=discipline,
                brand=random.choice(brands),
                length_cm=random.randint(160, 190) if user_rt == "Skier" else random.randint(150, 165),
                width_mm=random.randint(80, 105) if user_rt == "Skier" else None,
                created_at=datetime.utcnow()
            )
            db.session.add(primary)
            equipment_count += 1

            if random.random() < 0.5:
                secondary = EquipmentSetup(
                    user_id=user.id,
                    slot=EquipmentSlot.SECONDARY,
                    is_primary=False,
                    discipline=discipline,
                    brand=random.choice(brands),
                    length_cm=random.randint(160, 190) if user_rt == "Skier" else random.randint(150, 165),
                    width_mm=random.randint(80, 105) if user_rt == "Skier" else None,
                    created_at=datetime.utcnow()
                )
                db.session.add(secondary)
                equipment_count += 1
        
        db.session.commit()
        print(f"✨ Created: {equipment_count} equipment setups")
        
        # ====== SKI TRIPS ======
        resorts = Resort.query.all()
        trip_count = 0
        today = date.today()
        for user in all_users:
            existing_trips = SkiTrip.query.filter_by(user_id=user.id).count()
            if existing_trips >= 4:
                continue
            
            for _ in range(4 - existing_trips):
                start = today + timedelta(days=random.randint(5, 120))
                end = start + timedelta(days=random.randint(1, 5))
                
                trip = SkiTrip(
                    user_id=user.id,
                    resort_id=random.choice(resorts).id,
                    start_date=start,
                    end_date=end,
                    pass_type=random.choice(user.pass_type.split(",")),
                    is_public=True
                )
                db.session.add(trip)
                trip_count += 1
        
        db.session.commit()
        print(f"✨ Created: {trip_count} ski trips")
        
        # ====== FRIEND CONNECTIONS ======
        friend_count = 0
        for user in dummy_users:
            if Friend.query.filter_by(user_id=user.id, friend_id=richard.id).first():
                continue
            
            f1 = Friend(user_id=user.id, friend_id=richard.id)
            f2 = Friend(user_id=richard.id, friend_id=user.id)
            db.session.add_all([f1, f2])
            friend_count += 2
            
            if not Friend.query.filter_by(user_id=user.id, friend_id=jonathan.id).first():
                f3 = Friend(user_id=user.id, friend_id=jonathan.id)
                f4 = Friend(user_id=jonathan.id, friend_id=user.id)
                db.session.add_all([f3, f4])
                friend_count += 2
        
        db.session.commit()
        print(f"✨ Created: {friend_count} friend connections")
        
        # ====== GROUP TRIPS ======
        grouptrip_count = 0
        tripguest_count = 0
        for i in range(5):
            host = richard if i % 2 == 0 else jonathan
            title = f"{random.choice(['March', 'April', 'May'])} {random.choice(TRIP_TITLES)}"
            start = today + timedelta(days=random.randint(10, 60))
            end = start + timedelta(days=random.randint(2, 5))
            
            trip = GroupTrip(
                host_id=host.id,
                title=title,
                start_date=start,
                end_date=end
            )
            db.session.add(trip)
            db.session.flush()
            
            # Add host as accepted guest
            host_guest = TripGuest(trip_id=trip.id, user_id=host.id, status=GuestStatus.ACCEPTED)
            db.session.add(host_guest)
            tripguest_count += 1
            
            # Add jonathan/richard
            other_host = jonathan if host == richard else richard
            other_guest = TripGuest(trip_id=trip.id, user_id=other_host.id, status=GuestStatus.ACCEPTED)
            db.session.add(other_guest)
            tripguest_count += 1
            
            # Add 5-10 random dummy users
            selected_guests = random.sample(dummy_users, min(random.randint(5, 10), len(dummy_users)))
            for guest_user in selected_guests:
                guest = TripGuest(trip_id=trip.id, user_id=guest_user.id, status=GuestStatus.ACCEPTED)
                db.session.add(guest)
                tripguest_count += 1
            
            grouptrip_count += 1
        
        db.session.commit()
        print(f"✨ Created: {grouptrip_count} group trips, {tripguest_count} trip guests")
        
        # ====== OPEN DATES ======
        open_dates_count = 0
        for user in all_users:
            if user.open_dates:
                continue
            
            num_ranges = random.randint(6, 10) if user in [richard, jonathan] else random.randint(3, 6)
            open_dates = []
            for _ in range(num_ranges):
                start = today + timedelta(days=random.randint(5, 180))
                for j in range(random.randint(1, 4)):
                    date_str = (start + timedelta(days=j)).strftime("%Y-%m-%d")
                    if date_str not in open_dates:
                        open_dates.append(date_str)
            
            user.open_dates = sorted(open_dates)
            open_dates_count += len(open_dates)
        
        db.session.commit()
        print(f"✨ Created: {open_dates_count} open dates")
        
        # ====== VERIFICATION ======
        print("\n" + "=" * 70)
        print("VERIFICATION REPORT")
        print("=" * 70)
        print(f"Total users: {User.query.count()}")
        print(f"  - Fixed: 2 (Richard, Jonathan)")
        print(f"  - Dummy: {len(dummy_users)}")
        print(f"Total SkiTrips: {SkiTrip.query.count()}")
        print(f"Total GroupTrips: {GroupTrip.query.count()}")
        print(f"Total TripGuests: {TripGuest.query.count()}")
        print(f"Total EquipmentSetup: {EquipmentSetup.query.count()}")
        print(f"Total Friend connections: {Friend.query.count()}")
        
        # Sample data
        print(f"\nSample Users:")
        for user in random.sample(all_users, min(5, len(all_users))):
            trips = SkiTrip.query.filter_by(user_id=user.id).count()
            equipment = EquipmentSetup.query.filter_by(user_id=user.id).count()
            open_dates = len(user.open_dates) if user.open_dates else 0
            friends_richard = 1 if Friend.query.filter_by(user_id=user.id, friend_id=richard.id).first() else 0
            friends_jonathan = 1 if Friend.query.filter_by(user_id=user.id, friend_id=jonathan.id).first() else 0
            
            print(f"  {user.email}: passes={user.pass_type}, trips={trips}, equipment={equipment}, open_dates={open_dates}, connected_to_richard={friends_richard}, connected_to_jonathan={friends_jonathan}")
        
        print("\n✅ SEEDING COMPLETE!")
        print("=" * 70)


# ============================================================================
# REPAIR DEMO DATA
# ============================================================================

@app.cli.command("repair-demo-data")
def repair_demo_data():
    """Repair seeded demo data: fix passwords and friend connections."""
    print("🔧 REPAIRING DEMO DATA...")
    print("=" * 70)
    
    with app.app_context():
        # ====== FIX PASSWORDS ======
        password_fixes = 0
        
        richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
        if richard:
            richard.set_password("12345678")
            db.session.add(richard)
            password_fixes += 1
            print("✨ Reset password: Richard Battle-Baxter")
        
        jonathan = User.query.filter_by(email="jonathanmschmitz@gmail.com").first()
        if jonathan:
            jonathan.set_password("12345678")
            db.session.add(jonathan)
            password_fixes += 1
            print("✨ Reset password: Jonathan Schmitz")
        
        db.session.commit()
        
        # ====== FIX FRIEND CONNECTIONS ======
        friend_fixes = 0
        
        # Get all dummy users
        dummy_users = User.query.filter(
            User.email.like("user%@baselodge.local")
        ).all()
        
        print(f"\nProcessing {len(dummy_users)} dummy users...")
        
        for user in dummy_users:
            # Connect to Richard
            if richard:
                existing_1 = Friend.query.filter_by(user_id=user.id, friend_id=richard.id).first()
                existing_2 = Friend.query.filter_by(user_id=richard.id, friend_id=user.id).first()
                
                if not existing_1:
                    f1 = Friend(user_id=user.id, friend_id=richard.id)
                    db.session.add(f1)
                    friend_fixes += 1
                
                if not existing_2:
                    f2 = Friend(user_id=richard.id, friend_id=user.id)
                    db.session.add(f2)
                    friend_fixes += 1
            
            # Connect to Jonathan
            if jonathan:
                existing_3 = Friend.query.filter_by(user_id=user.id, friend_id=jonathan.id).first()
                existing_4 = Friend.query.filter_by(user_id=jonathan.id, friend_id=user.id).first()
                
                if not existing_3:
                    f3 = Friend(user_id=user.id, friend_id=jonathan.id)
                    db.session.add(f3)
                    friend_fixes += 1
                
                if not existing_4:
                    f4 = Friend(user_id=jonathan.id, friend_id=user.id)
                    db.session.add(f4)
                    friend_fixes += 1
        
        db.session.commit()
        print(f"✨ Added/verified: {friend_fixes} friend connections")
        
        # ====== VERIFICATION ======
        print("\n" + "=" * 70)
        print("VERIFICATION REPORT")
        print("=" * 70)
        
        # Check passwords
        test_richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
        test_jonathan = User.query.filter_by(email="jonathanmschmitz@gmail.com").first()
        
        print(f"\nPassword Status:")
        if test_richard and test_richard.check_password("12345678"):
            print(f"  ✓ Richard can log in with 12345678")
        else:
            print(f"  ✗ Richard password FAILED")
        
        if test_jonathan and test_jonathan.check_password("12345678"):
            print(f"  ✓ Jonathan can log in with 12345678")
        else:
            print(f"  ✗ Jonathan password FAILED")
        
        # Check friend connections
        print(f"\nFriend Connection Status:")
        total_dummy = len(dummy_users)
        richard_connections = Friend.query.filter_by(friend_id=richard.id).count() if richard else 0
        jonathan_connections = Friend.query.filter_by(friend_id=jonathan.id).count() if jonathan else 0
        
        print(f"  Dummy users connected to Richard: {richard_connections}/{total_dummy}")
        print(f"  Dummy users connected to Jonathan: {jonathan_connections}/{total_dummy}")
        
        # Sample verification
        print(f"\nSample User Connections:")
        for user in random.sample(dummy_users, min(3, len(dummy_users))):
            friends_richard = 1 if Friend.query.filter_by(user_id=user.id, friend_id=richard.id).first() else 0
            friends_jonathan = 1 if Friend.query.filter_by(user_id=user.id, friend_id=jonathan.id).first() else 0
            
            print(f"  {user.email}: richard_friend={friends_richard}, jonathan_friend={friends_jonathan}")
        
        print("\n✅ REPAIR COMPLETE!")
        print("=" * 70)


@app.cli.command("fix-seeded-users")
def fix_seeded_users():
    """Fix seeded users: reset passwords and ensure friend connections."""
    from werkzeug.security import generate_password_hash
    
    print("🔐 FIXING SEEDED USERS...")
    print("=" * 70)
    
    TARGET_USERS = [
        {
            "email": "richardbattlebaxter@gmail.com",
            "first_name": "Richard",
            "last_name": "Battle-Baxter",
            "password": "12345678"
        },
        {
            "email": "jonathanmschmitz@gmail.com",
            "first_name": "Jonathan",
            "last_name": "Schmitz",
            "password": "12345678"
        }
    ]

    users = {}

    # Step 1: Create or update target users with correct passwords
    print("\n📝 STEP 1: Creating/Updating Target Users")
    for u in TARGET_USERS:
        user = User.query.filter(
            db.func.lower(User.email) == u["email"].lower()
        ).first()

        if not user:
            user = User(
                email=u["email"],
                first_name=u["first_name"],
                last_name=u["last_name"],
                password_hash=generate_password_hash(u["password"])
            )
            db.session.add(user)
            print(f"  ✨ CREATED user {u['email']}")
        else:
            user.password_hash = generate_password_hash(u["password"])
            print(f"  ✏️  RESET password for {u['email']}")

        users[u["email"]] = user

    db.session.commit()

    # Step 2: Fix friend connections
    print("\n🤝 STEP 2: Fixing Friend Connections")
    richard = users["richardbattlebaxter@gmail.com"]
    jonathan = users["jonathanmschmitz@gmail.com"]

    all_users = User.query.all()
    connections_added = 0

    for user in all_users:
        if user.id in (richard.id, jonathan.id):
            continue

        for core in (richard, jonathan):
            exists = Friend.query.filter_by(
                user_id=core.id,
                friend_id=user.id
            ).first()

            if not exists:
                db.session.add(Friend(user_id=core.id, friend_id=user.id))
                db.session.add(Friend(user_id=user.id, friend_id=core.id))
                connections_added += 2

    db.session.commit()
    print(f"  ✨ Added {connections_added} friend connections")

    # Step 3: Fix pass_type cleanup (convert "both" to "epic" for seeded users)
    print("\n🎿 STEP 3: Cleaning up pass_type values")
    seeded_users_with_both = User.query.filter(
        User.is_seeded == True,
        (User.pass_type.ilike('%both%') | (User.pass_type == 'both'))
    ).all()
    
    both_count = 0
    for user in seeded_users_with_both:
        if user.pass_type and 'both' in user.pass_type.lower():
            user.pass_type = user.pass_type.replace('Both', 'Epic').replace('both', 'Epic')
            both_count += 1
    
    if both_count > 0:
        db.session.commit()
        print(f"  ✨ Converted {both_count} seeded users from 'both' to 'epic'")
    else:
        print(f"  ✓ No seeded users with pass_type='both' found")

    # Step 4: Verification
    print("\n" + "=" * 70)
    print("✅ VERIFICATION")
    print("=" * 70)
    
    richard_friends = Friend.query.filter_by(user_id=richard.id).count()
    jonathan_friends = Friend.query.filter_by(user_id=jonathan.id).count()
    
    print(f"Richard friends: {richard_friends}")
    print(f"Jonathan friends: {jonathan_friends}")
    
    # Test password
    richard_pwd_ok = richard.check_password("12345678")
    jonathan_pwd_ok = jonathan.check_password("12345678")
    
    print(f"Richard password check: {richard_pwd_ok}")
    print(f"Jonathan password check: {jonathan_pwd_ok}")
    
    # Check for any remaining "both" values
    users_with_both = User.query.filter(
        User.pass_type.ilike('%both%') | (User.pass_type == 'both')
    ).count()
    print(f"Users with pass_type containing 'both': {users_with_both}")
    
    print("\n✅ FIX COMPLETE!")
    print("=" * 70)


@app.route("/admin/version", methods=["GET"])
@login_required
@admin_required
def admin_version():
    """Simple version check endpoint to verify production deployment."""
    import re
    db_uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")
    db_source = "unknown"
    if db_uri:
        if "supabase" in db_uri.lower():
            db_source = "supabase"
        elif "sqlite" in db_uri.lower():
            db_source = "sqlite"
        else:
            db_source = "external-db"
    return jsonify({
        "app_version": "2026-05-01-ui-sync-check",
        "server_timestamp": datetime.utcnow().isoformat() + "Z",
        "flask_env": os.environ.get("FLASK_ENV", "unknown"),
        "database_url_source": db_source,
        "note": "Used to confirm which version production/TestFlight is loading",
        "status": "ok",
        "endpoints_available": [
            "/admin/version",
            "/admin/backfill-country-codes",
            "/admin/resorts-audit",
            "/admin/init-db",
            "/admin/sync-from-canonical"
        ]
    })


@app.route("/admin/debug-users", methods=["GET"])
@login_required
@admin_required
def debug_users():
    """Inspect production database: user count, first 20 users, and DB URI in use."""
    db_uri = app.config.get("SQLALCHEMY_DATABASE_URI", "not set")
    # Mask password from URI for safe display
    import re
    safe_uri = re.sub(r'(:)[^:@]+(@)', r'\1***\2', db_uri)
    users = User.query.order_by(User.id).limit(20).all()
    total = User.query.count()
    return jsonify({
        "database_uri": safe_uri,
        "total_user_count": total,
        "first_20_users": [
            {
                "id": u.id,
                "email": u.email,
                "first_name": u.first_name,
                "last_name": u.last_name,
            }
            for u in users
        ]
    })


@app.route("/admin/db-status", methods=["GET"])
@login_required
@admin_required
def admin_db_status():
    """
    Read-only diagnostic: confirms which database engine is active, whether
    SQLite fallback is in use, and reports counts for all core tables.
    No writes. Safe to call in production at any time.
    """
    import re

    db_uri = app.config.get("SQLALCHEMY_DATABASE_URI", "not set")
    safe_uri = re.sub(r'(:)[^:@]+(@)', r'\1***\2', db_uri)

    engine_type = (
        "postgresql" if "postgresql" in db_uri or "postgres" in db_uri
        else "sqlite" if "sqlite" in db_uri
        else "unknown"
    )
    is_sqlite_fallback = "sqlite" in db_uri

    raw_env_url = os.environ.get("SUPABASE_DATABASE_URL", "")
    env_url_present = bool(raw_env_url)
    env_url_scheme = raw_env_url.split("://")[0] if "://" in raw_env_url else "not set"
    env_url_host = raw_env_url.split("@")[-1] if "@" in raw_env_url else "no @ found"

    try:
        counts = {
            "users":                  User.query.count(),
            "ski_trips":              SkiTrip.query.count(),
            "ski_trip_participants":  SkiTripParticipant.query.count(),
            "friends":                Friend.query.count(),
            "invitations":            Invitation.query.count(),
            "invite_tokens":          InviteToken.query.count(),
            "group_trips":            GroupTrip.query.count(),
            "trip_guests":            TripGuest.query.count(),
        }
        counts_ok = True
        counts_error = None
    except Exception as e:
        counts = {}
        counts_ok = False
        counts_error = str(e)

    return jsonify({
        "db_engine":             engine_type,
        "active_uri_masked":     safe_uri,
        "is_sqlite_fallback":    is_sqlite_fallback,
        "is_production_flag":    is_production,
        "supabase_env_var": {
            "present":           env_url_present,
            "scheme":            env_url_scheme,
            "host_masked":       env_url_host,
        },
        "table_counts":          counts,
        "table_counts_ok":       counts_ok,
        "table_counts_error":    counts_error,
        "assessed_at":           datetime.utcnow().isoformat() + "Z",
        "note": (
            "SQLITE FALLBACK ACTIVE — users may be writing to a local file, not Supabase"
            if is_sqlite_fallback else
            "Supabase PostgreSQL active — no SQLite fallback"
        ),
    })


@app.route("/admin/export-live-data", methods=["GET"])
@login_required
@admin_required
def export_live_data():
    """Read-only full data rescue export. Returns all critical user-linked table rows as JSON."""
    import enum as _enum

    def _val(v):
        """Serialize a column value to a JSON-safe primitive."""
        if v is None:
            return None
        if isinstance(v, _enum.Enum):
            return v.value
        if hasattr(v, 'isoformat'):
            return v.isoformat()
        return v

    def _row(obj, exclude=None):
        exclude = set(exclude or [])
        return {
            c.name: _val(getattr(obj, c.name))
            for c in obj.__table__.columns
            if c.name not in exclude
        }

    users           = User.query.order_by(User.id).all()
    trips           = SkiTrip.query.order_by(SkiTrip.id).all()
    friends         = Friend.query.order_by(Friend.id).all()
    participants    = SkiTripParticipant.query.order_by(SkiTripParticipant.id).all()
    invitations     = Invitation.query.order_by(Invitation.id).all()
    invite_tokens   = InviteToken.query.order_by(InviteToken.id).all()
    group_trips     = GroupTrip.query.order_by(GroupTrip.id).all()
    trip_guests     = TripGuest.query.order_by(TripGuest.id).all()

    return jsonify({
        "exported_at": datetime.utcnow().isoformat(),
        "database_uri": re.sub(r'(:)[^:@]+(@)', r'\1***\2',
                               app.config.get("SQLALCHEMY_DATABASE_URI", "not set")),
        "counts": {
            "users":            len(users),
            "ski_trips":        len(trips),
            "friends":          len(friends),
            "ski_trip_participants": len(participants),
            "invitations":      len(invitations),
            "invite_tokens":    len(invite_tokens),
            "group_trips":      len(group_trips),
            "trip_guests":      len(trip_guests),
        },
        "users":            [_row(u, exclude=["password_hash"]) for u in users],
        "ski_trips":        [_row(t) for t in trips],
        "friends":          [_row(f) for f in friends],
        "ski_trip_participants": [_row(p) for p in participants],
        "invitations":      [_row(i) for i in invitations],
        "invite_tokens":    [_row(t) for t in invite_tokens],
        "group_trips":      [_row(g) for g in group_trips],
        "trip_guests":      [_row(g) for g in trip_guests],
    })


@app.route("/admin/resorts-audit", methods=["GET"])
@login_required
@admin_required
def resorts_audit():
    """Read-only endpoint to fetch all resorts for audit comparison."""
    resorts = Resort.query.all()
    return jsonify({
        "total": len(resorts),
        "resorts": [
            {
                "name": r.name,
                "state_code": r.state_code or r.state,
                "country_code": r.country_code or r.country,
                "pass_brands": r.pass_brands or r.brand
            }
            for r in resorts
        ]
    })


@app.route("/admin/backfill-country-codes", methods=["GET", "POST"])
@login_required
@admin_required
def backfill_country_codes():
    """
    Backfill country_code and state_code for resorts based on state field.
    v2 - Updated 2025-12-25
    
    Usage: GET https://yourapp.replit.dev/admin/backfill-country-codes
    
    This is idempotent - safe to call multiple times.
    """
    US_STATES = {
        'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA',
        'HI', 'ID', 'IL', 'IN', 'IA', 'KS', 'KY', 'LA', 'ME', 'MD',
        'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ',
        'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC',
        'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY'
    }
    
    CA_PROVINCES = {
        'AB', 'BC', 'MB', 'NB', 'NL', 'NS', 'NT', 'NU', 'ON', 'PE', 'QC', 'SK', 'YT'
    }
    
    try:
        resorts = Resort.query.all()
        updated = []
        skipped = []
        
        for r in resorts:
            state = r.state_code or r.state
            if not state:
                skipped.append(f"{r.name}: no state")
                continue
            
            state_upper = state.upper().strip()
            
            if state_upper in US_STATES:
                if r.country_code != 'US' or r.state_code != state_upper:
                    r.country_code = 'US'
                    r.country = 'US'
                    r.state_code = state_upper
                    updated.append(f"{r.name} -> US/{state_upper}")
            elif state_upper in CA_PROVINCES:
                if r.country_code != 'CA' or r.state_code != state_upper:
                    r.country_code = 'CA'
                    r.country = 'CA'
                    r.state_code = state_upper
                    updated.append(f"{r.name} -> CA/{state_upper}")
            else:
                skipped.append(f"{r.name}: unknown state '{state}'")
        
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": f"Updated {len(updated)} resorts, skipped {len(skipped)}",
            "updated": updated[:50],
            "skipped": skipped[:20],
            "total_resorts": len(resorts)
        }), 200
    except Exception as e:
        import traceback
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": str(e),
            "traceback": traceback.format_exc()
        }), 500



# ============================================================================
# ADMIN COUNTRIES ENDPOINT
# ============================================================================

@app.route("/admin/countries", methods=["POST"])
@login_required
@admin_required
def admin_add_country():
    """Add a new country to the reference table."""
    from models import Country
    
    data = request.get_json()
    name = (data.get("name") or "").strip()
    code = (data.get("code") or "").strip().upper()
    
    if not name or not code:
        return jsonify({"status": "error", "message": "Name and code are required"}), 400
    
    if len(code) != 2:
        return jsonify({"status": "error", "message": "Country code must be 2 letters"}), 400
    
    existing = Country.query.filter(db.func.lower(Country.code) == code.lower()).first()
    if existing:
        return jsonify({"status": "success", "id": existing.id, "code": existing.code, "name": existing.name})
    
    country = Country(code=code, name=name, is_active=True)
    db.session.add(country)
    db.session.commit()
    
    COUNTRY_NAMES[code] = name
    
    return jsonify({"status": "success", "id": country.id, "code": country.code, "name": country.name})


# ============================================================================
# ADMIN TRIPS PAGE
# ============================================================================

@app.route("/admin/trips")
@login_required
@admin_required
def admin_trips():
    """Admin page — trip analytics across all three sections."""
    from collections import Counter as _TrCtr, defaultdict as _TrDD
    from datetime import datetime as _dtt

    now         = datetime.utcnow()
    thirty_ago  = now - timedelta(days=30)
    seven_ago   = now - timedelta(days=7)
    first_of_mo = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_label = now.strftime("%B %Y")
    today       = now.date()

    # ── Core counts ───────────────────────────────────────────────────────────
    total_trips    = SkiTrip.query.count()
    trips_month    = SkiTrip.query.filter(SkiTrip.created_at >= first_of_mo).count()
    trips_last_7   = SkiTrip.query.filter(SkiTrip.created_at >= seven_ago).count()
    trips_last_30  = SkiTrip.query.filter(SkiTrip.created_at >= thirty_ago).count()
    upcoming_trips = SkiTrip.query.filter(SkiTrip.start_date >= today).count()
    past_trips     = SkiTrip.query.filter(SkiTrip.end_date < today).count()
    users_with_trips = db.session.query(
        func.count(func.distinct(SkiTrip.user_id))
    ).scalar() or 0

    # Solo vs. group
    solo_trips_count  = SkiTrip.query.filter(
        db.or_(SkiTrip.is_group_trip == False, SkiTrip.is_group_trip.is_(None))
    ).count()
    group_trips_count = SkiTrip.query.filter(SkiTrip.is_group_trip == True).count()
    group_pct = round(group_trips_count / total_trips * 100) if total_trips else 0

    # Avg trip length (from date pairs)
    _len_rows = db.session.query(SkiTrip.start_date, SkiTrip.end_date).filter(
        SkiTrip.start_date.isnot(None), SkiTrip.end_date.isnot(None)
    ).all()
    if _len_rows:
        _total_days = sum((r.end_date - r.start_date).days for r in _len_rows)
        avg_trip_length = round(_total_days / len(_len_rows), 1)
    else:
        avg_trip_length = 0

    # Distinct resorts in upcoming trips
    resorts_scheduled = db.session.query(
        func.count(func.distinct(SkiTrip.resort_id))
    ).filter(SkiTrip.start_date >= today, SkiTrip.resort_id.isnot(None)).scalar() or 0

    # ── Trips by month — last 12 months (Python bucketing) ───────────────────
    _twelve_ago = now - timedelta(days=365)
    _tm_rows = db.session.query(SkiTrip.created_at).filter(
        SkiTrip.created_at >= _twelve_ago
    ).all()
    _tm_buckets = _TrDD(int)
    for (_ts,) in _tm_rows:
        if _ts:
            _tm_buckets[_ts.strftime("%Y-%m")] += 1
    trips_by_month = []
    for i in range(11, -1, -1):
        _mo = (now.month - i - 1) % 12 + 1
        _yr = now.year if (now.month - i) > 0 else now.year - 1
        _key = f"{_yr}-{_mo:02d}"
        trips_by_month.append({"month": _dtt(_yr, _mo, 1).strftime("%b"),
                                "count": _tm_buckets.get(_key, 0)})
    _tm_max = max((m["count"] for m in trips_by_month), default=1)
    for m in trips_by_month:
        m["pct"] = round(m["count"] / _tm_max * 100) if _tm_max > 0 else 0

    # ── Destination intel ─────────────────────────────────────────────────────
    _tp_rows = (
        db.session.query(Resort.name, func.count(SkiTrip.id).label("cnt"))
        .join(SkiTrip, SkiTrip.resort_id == Resort.id)
        .filter(SkiTrip.resort_id.isnot(None))
        .group_by(Resort.id, Resort.name)
        .order_by(func.count(SkiTrip.id).desc()).limit(5).all()
    )
    top_planned_resorts = [{"name": r.name, "count": r.cnt} for r in _tp_rows]
    _tp_max = max((r["count"] for r in top_planned_resorts), default=1)
    for r in top_planned_resorts:
        r["pct"] = round(r["count"] / _tp_max * 100)

    _ud_rows = (
        db.session.query(Resort.name, func.count(SkiTrip.id).label("cnt"))
        .join(SkiTrip, SkiTrip.resort_id == Resort.id)
        .filter(SkiTrip.resort_id.isnot(None), SkiTrip.start_date >= today)
        .group_by(Resort.id, Resort.name)
        .order_by(func.count(SkiTrip.id).desc()).limit(5).all()
    )
    upcoming_destinations = [{"name": r.name, "count": r.cnt} for r in _ud_rows]
    _ud_max = max((r["count"] for r in upcoming_destinations), default=1)
    for r in upcoming_destinations:
        r["pct"] = round(r["count"] / _ud_max * 100)

    _ts_rows = (
        db.session.query(Resort.state_code, func.count(SkiTrip.id).label("cnt"))
        .join(SkiTrip, SkiTrip.resort_id == Resort.id)
        .filter(SkiTrip.resort_id.isnot(None), Resort.state_code.isnot(None))
        .group_by(Resort.state_code)
        .order_by(func.count(SkiTrip.id).desc()).limit(8).all()
    )
    trips_by_state = [{"name": r.state_code, "count": r.cnt} for r in _ts_rows]
    _ts_max = max((r["count"] for r in trips_by_state), default=1)
    for r in trips_by_state:
        r["pct"] = round(r["count"] / _ts_max * 100)

    _tr_rows = (
        db.session.query(Resort.name, func.count(SkiTrip.id).label("cnt"))
        .join(SkiTrip, SkiTrip.resort_id == Resort.id)
        .filter(SkiTrip.resort_id.isnot(None), SkiTrip.created_at >= first_of_mo)
        .group_by(Resort.id, Resort.name)
        .order_by(func.count(SkiTrip.id).desc()).limit(3).all()
    )
    trending_resorts_month = [{"name": r.name, "count": r.cnt} for r in _tr_rows]
    _tr_max = max((r["count"] for r in trending_resorts_month), default=1)
    for r in trending_resorts_month:
        r["pct"] = round(r["count"] / _tr_max * 100)

    # Upcoming trips table (next 5 with resort name)
    _ut_rows = (
        db.session.query(
            SkiTrip.start_date, SkiTrip.end_date, SkiTrip.trip_status, SkiTrip.mountain,
            Resort.name.label("resort_name")
        )
        .outerjoin(Resort, SkiTrip.resort_id == Resort.id)
        .filter(SkiTrip.start_date >= today)
        .order_by(SkiTrip.start_date.asc()).limit(5).all()
    )
    upcoming_trips_table = [
        {"resort": r.resort_name or r.mountain or "Unknown resort",
         "start_date": r.start_date, "end_date": r.end_date,
         "status": r.trip_status or "planning"}
        for r in _ut_rows
    ]

    # ── Social & Coordination ─────────────────────────────────────────────────
    total_invites    = SkiTripParticipant.query.filter(
        SkiTripParticipant.role == ParticipantRole.GUEST
    ).count()
    accepted_invites = SkiTripParticipant.query.filter(
        SkiTripParticipant.role == ParticipantRole.GUEST,
        SkiTripParticipant.status == GuestStatus.ACCEPTED,
    ).count()
    pending_invites  = SkiTripParticipant.query.filter(
        SkiTripParticipant.role == ParticipantRole.GUEST,
        SkiTripParticipant.status == GuestStatus.INVITED,
    ).count()
    invite_accept_pct = round(accepted_invites / total_invites * 100) if total_invites else None

    trips_with_invites = db.session.query(
        func.count(func.distinct(SkiTripParticipant.trip_id))
    ).filter(SkiTripParticipant.role == ParticipantRole.GUEST).scalar() or 0

    trips_with_accepted = db.session.query(
        func.count(func.distinct(SkiTripParticipant.trip_id))
    ).filter(
        SkiTripParticipant.role == ParticipantRole.GUEST,
        SkiTripParticipant.status == GuestStatus.ACCEPTED,
    ).scalar() or 0

    avg_invites_per_trip = round(total_invites / trips_with_invites, 1) if trips_with_invites else 0

    _org_rows = (
        db.session.query(User.first_name, func.count(SkiTrip.id).label("cnt"))
        .join(SkiTrip, SkiTrip.user_id == User.id)
        .group_by(User.id, User.first_name)
        .order_by(func.count(SkiTrip.id).desc()).limit(5).all()
    )
    most_active_organizers = [{"name": r.first_name or "User", "count": r.cnt}
                               for r in _org_rows]
    _org_max = max((o["count"] for o in most_active_organizers), default=1)
    for o in most_active_organizers:
        o["pct"] = round(o["count"] / _org_max * 100)

    _hit_rows = (
        db.session.query(
            Resort.name.label("resort_name"),
            func.count(SkiTripParticipant.id).label("cnt")
        )
        .join(SkiTrip, SkiTripParticipant.trip_id == SkiTrip.id)
        .outerjoin(Resort, SkiTrip.resort_id == Resort.id)
        .filter(SkiTripParticipant.role == ParticipantRole.GUEST)
        .group_by(SkiTrip.id, Resort.name)
        .order_by(func.count(SkiTripParticipant.id).desc()).limit(5).all()
    )
    high_invite_trips = [{"resort": r.resort_name or "Unknown", "count": r.cnt}
                         for r in _hit_rows]
    _hit_max = max((h["count"] for h in high_invite_trips), default=1)
    for h in high_invite_trips:
        h["pct"] = round(h["count"] / _hit_max * 100)

    return render_template('admin_trips.html',
                         active_tab='trips',
                         month_label=month_label,
                         total_trips=total_trips,
                         trips_month=trips_month,
                         trips_last_7=trips_last_7,
                         trips_last_30=trips_last_30,
                         upcoming_trips=upcoming_trips,
                         past_trips=past_trips,
                         users_with_trips=users_with_trips,
                         solo_trips_count=solo_trips_count,
                         group_trips_count=group_trips_count,
                         group_pct=group_pct,
                         avg_trip_length=avg_trip_length,
                         resorts_scheduled=resorts_scheduled,
                         trips_by_month=trips_by_month,
                         top_planned_resorts=top_planned_resorts,
                         upcoming_destinations=upcoming_destinations,
                         trips_by_state=trips_by_state,
                         trending_resorts_month=trending_resorts_month,
                         upcoming_trips_table=upcoming_trips_table,
                         total_invites=total_invites,
                         accepted_invites=accepted_invites,
                         pending_invites=pending_invites,
                         invite_accept_pct=invite_accept_pct,
                         trips_with_invites=trips_with_invites,
                         trips_with_accepted=trips_with_accepted,
                         avg_invites_per_trip=avg_invites_per_trip,
                         most_active_organizers=most_active_organizers,
                         high_invite_trips=high_invite_trips)


# ============================================================================
# ADMIN RESORTS CURATION PAGE (messaging route moved to intelligence suite)
# ============================================================================

@app.route("/admin/resorts_placeholder_removed")
@login_required
@admin_required
def admin_messaging_old_placeholder():
    """Old messaging route stub — body removed; see /admin/messaging in the intelligence suite."""
    return redirect('/admin/messaging')

# ============================================================================
# ADMIN RESORTS CURATION PAGE
# ============================================================================

@app.route("/admin/resorts")
@login_required
@admin_required
def admin_resorts():
    """
    Admin page for curating resort data.
    DEV is the source of truth. PROD sync via canonical JSON export.
    """
    import os
    import json
    
    # Get all resorts sorted by country, state, name
    resorts = Resort.query.order_by(Resort.country_code, Resort.state_code, Resort.name).all()
    
    # Get unique countries for filters (from existing resorts)
    resort_countries = db.session.query(Resort.country_code).distinct().order_by(Resort.country_code).all()
    resort_countries = set(c[0] for c in resort_countries if c[0])
    
    # Get all countries from Country table for the dropdown
    from models import Country
    db_countries = Country.query.filter_by(is_active=True).order_by(Country.name).all()
    
    # Build unified country list
    dropdown_countries = sorted(COUNTRIES.items(), key=lambda x: x[1])
    
    # Keep countries list for filters (just codes from resorts)
    countries = sorted(resort_countries)
    
    # Check last canonical export timestamp
    canonical_file = os.path.join(os.path.dirname(__file__), 'data', 'canonical_resorts.json')
    last_export_info = None
    if os.path.exists(canonical_file):
        try:
            with open(canonical_file, 'r') as f:
                canonical_data = json.load(f)
            last_export_info = {
                'version': canonical_data.get('version', 'unknown'),
                'exported_at': canonical_data.get('exported_at', 'unknown'),
                'count': len(canonical_data.get('resorts', []))
            }
        except Exception:
            pass
    
    total_count = len(resorts)

    # ── Mountains analytics ─────────────────────────────────────────────────
    from collections import Counter as _MtnCtr
    _today = date.today()

    mtn_active_count   = Resort.query.filter_by(is_active=True).count()
    mtn_inactive_count = total_count - mtn_active_count

    mtn_upcoming_resort_count = db.session.query(
        func.count(func.distinct(SkiTrip.resort_id))
    ).filter(SkiTrip.start_date >= _today, SkiTrip.resort_id.isnot(None)).scalar() or 0

    _mtn_pbj = db.session.query(Resort.pass_brands_json).filter(Resort.is_active == True).all()
    _mtn_pass_ctr = _MtnCtr()
    _mtn_on_pass = 0
    for (pbj,) in _mtn_pbj:
        if pbj and isinstance(pbj, list):
            real = [b for b in pbj if b and b.lower() not in ('none', '')]
            if real:
                _mtn_on_pass += 1
                for b in real:
                    _mtn_pass_ctr[b] += 1

    _mtn_wl_rows = db.session.query(User.wish_list_resorts).filter(User.wish_list_resorts.isnot(None)).all()
    _mtn_wish_ctr = _MtnCtr()
    for (wl,) in _mtn_wl_rows:
        if wl and isinstance(wl, list):
            for rid in wl:
                _mtn_wish_ctr[rid] += 1

    _PASS_ORDER = ['Epic', 'Ikon', 'Mountain Collective', 'Indy', 'Other']
    _pc_max = max(_mtn_pass_ctr.values()) if _mtn_pass_ctr else 1
    mtn_by_pass_brand = [
        {"name": p, "count": _mtn_pass_ctr.get(p, 0),
         "pct": round(_mtn_pass_ctr.get(p, 0) / _pc_max * 100) if _mtn_pass_ctr else 0}
        for p in _PASS_ORDER
    ]

    _rbc = (
        db.session.query(Resort.country_code, Resort.country_name, func.count(Resort.id).label("cnt"))
        .filter(Resort.is_active == True)
        .group_by(Resort.country_code, Resort.country_name)
        .order_by(func.count(Resort.id).desc()).limit(8).all()
    )
    _rbc_max = max((r.cnt for r in _rbc), default=1)
    mtn_by_country = [
        {"name": r.country_name or r.country_code or "Unknown",
         "count": r.cnt, "pct": round(r.cnt / _rbc_max * 100)}
        for r in _rbc
    ]


    mtn_no_pass = sum(
        1 for r in resorts
        if not r.get_pass_brands_list() or all(b.lower() in ('none', '') for b in r.get_pass_brands_list())
    )
    mtn_no_state   = sum(1 for r in resorts if not (r.state_code or r.state))
    mtn_no_country = sum(1 for r in resorts if not r.country_code)

    # ── Destination Traffic (MountainPageView) — graceful fallback ───────────
    try:
        from collections import defaultdict as _mpv_dd
        _six_months_ago  = datetime.utcnow() - timedelta(days=182)
        _first_of_mo_mpv = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        _mpv_ago_7d      = datetime.utcnow() - timedelta(days=7)
        _mpv_ago_30d     = datetime.utcnow() - timedelta(days=30)
        _mpv_ago_60d     = datetime.utcnow() - timedelta(days=60)

        mtn_traffic_total    = db.session.query(func.count(MountainPageView.id)).scalar() or 0
        mtn_traffic_loggedin = db.session.query(func.count(MountainPageView.id)).filter(
            MountainPageView.user_id.isnot(None)
        ).scalar() or 0
        mtn_traffic_anon = mtn_traffic_total - mtn_traffic_loggedin

        # Windowed visit counts
        mtn_traffic_7d = db.session.query(func.count(MountainPageView.id)).filter(
            MountainPageView.viewed_at >= _mpv_ago_7d
        ).scalar() or 0
        mtn_traffic_30d = db.session.query(func.count(MountainPageView.id)).filter(
            MountainPageView.viewed_at >= _mpv_ago_30d
        ).scalar() or 0

        # All viewed resorts ranked (all-time), capped at 20
        _tv_rows = (
            db.session.query(Resort.name, func.count(MountainPageView.id).label("cnt"))
            .join(MountainPageView, MountainPageView.resort_id == Resort.id)
            .group_by(Resort.id, Resort.name)
            .order_by(func.count(MountainPageView.id).desc()).limit(20).all()
        )
        _tv_max = max((r.cnt for r in _tv_rows), default=1)
        mtn_top_viewed = [{"name": r.name, "count": r.cnt,
                           "pct": round(r.cnt / _tv_max * 100)} for r in _tv_rows]

        # Monthly visit volume — last 6 months (Python bucketing)
        _mv_rows = db.session.query(MountainPageView.viewed_at).filter(
            MountainPageView.viewed_at >= _six_months_ago
        ).all()
        _mv_buckets = _mpv_dd(int)
        for (_vt,) in _mv_rows:
            if _vt:
                _mv_buckets[_vt.strftime("%Y-%m")] += 1
        _mpv_now = datetime.utcnow()
        mtn_monthly_views = []
        for i in range(5, -1, -1):
            _mo = (_mpv_now.month - i - 1) % 12 + 1
            _yr = _mpv_now.year if (_mpv_now.month - i) > 0 else _mpv_now.year - 1
            _key = f"{_yr}-{_mo:02d}"
            from datetime import datetime as _mdtt
            mtn_monthly_views.append({
                "month": _mdtt(_yr, _mo, 1).strftime("%b"),
                "count": _mv_buckets.get(_key, 0),
            })
        _mv_max = max((m["count"] for m in mtn_monthly_views), default=1)
        for m in mtn_monthly_views:
            m["pct"] = round(m["count"] / _mv_max * 100) if _mv_max > 0 else 0

        # Resort Demand Overview: impressions + engagement + planning + visitation (top 10)
        _top5_rows = (
            db.session.query(
                Resort.id,
                Resort.name,
                func.count(MountainPageView.id).label("total"),
                func.count(func.distinct(MountainPageView.user_id)).label("uniq_users"),
            )
            .join(MountainPageView, MountainPageView.resort_id == Resort.id)
            .group_by(Resort.id, Resort.name)
            .order_by(func.count(MountainPageView.id).desc()).limit(10).all()
        )
        _today_date = datetime.utcnow().date()
        _trip_planned_map = {
            r.resort_id: r.cnt for r in (
                db.session.query(SkiTrip.resort_id, func.count(SkiTrip.id).label("cnt"))
                .filter(SkiTrip.resort_id.isnot(None))
                .group_by(SkiTrip.resort_id).all()
            )
        }
        _mtn_visited_map = {
            r.resort_id: r.cnt for r in (
                db.session.query(SkiTrip.resort_id, func.count(SkiTrip.id).label("cnt"))
                .filter(SkiTrip.resort_id.isnot(None), SkiTrip.end_date < _today_date)
                .group_by(SkiTrip.resort_id).all()
            )
        }
        mtn_top5_traffic = [
            {
                "name": r.name,
                "total": r.total,
                "uniq_users": r.uniq_users,
                "wishlists": _mtn_wish_ctr.get(r.id, 0),
                "trips_planned": _trip_planned_map.get(r.id, 0),
                "mountains_visited": _mtn_visited_map.get(r.id, 0),
            }
            for r in _top5_rows
        ]

        # Most viewed this month KPI
        _mvm_row = (
            db.session.query(Resort.name, func.count(MountainPageView.id).label("cnt"))
            .join(MountainPageView, MountainPageView.resort_id == Resort.id)
            .filter(MountainPageView.viewed_at >= _first_of_mo_mpv)
            .group_by(Resort.id, Resort.name)
            .order_by(func.count(MountainPageView.id).desc()).first()
        )
        mtn_most_viewed_month     = _mvm_row.name if _mvm_row else None
        mtn_most_viewed_month_cnt = _mvm_row.cnt  if _mvm_row else 0

        # Avg distinct resorts viewed per unique user
        _avg_sub = (
            db.session.query(
                MountainPageView.user_id,
                func.count(func.distinct(MountainPageView.resort_id)).label("rc")
            )
            .filter(MountainPageView.user_id.isnot(None),
                    MountainPageView.resort_id.isnot(None))
            .group_by(MountainPageView.user_id)
            .subquery()
        )
        _avg_val = db.session.query(func.avg(_avg_sub.c.rc)).scalar()
        mtn_avg_resorts_per_user = round(float(_avg_val), 1) if _avg_val else 0

        mtn_traffic_ready = True
    except Exception:
        mtn_top_viewed            = []
        mtn_monthly_views         = []
        mtn_traffic_total         = 0
        mtn_traffic_7d            = 0
        mtn_traffic_30d           = 0
        mtn_top5_traffic          = []
        mtn_most_viewed_month     = None
        mtn_most_viewed_month_cnt = 0
        mtn_avg_resorts_per_user  = 0
        mtn_traffic_ready         = False

    return render_template('admin_resorts.html',
                         resorts=resorts,
                         countries=countries,
                         dropdown_countries=dropdown_countries,
                         last_export_info=last_export_info,
                         total_count=total_count,
                         active_tab='mountains',
                         mtn_active_count=mtn_active_count,
                         mtn_inactive_count=mtn_inactive_count,
                         mtn_upcoming_resort_count=mtn_upcoming_resort_count,
                         mtn_resorts_on_pass=_mtn_on_pass,
                         mtn_by_pass_brand=mtn_by_pass_brand,
                         mtn_by_country=mtn_by_country,
                         mtn_no_pass=mtn_no_pass,
                         mtn_no_state=mtn_no_state,
                         mtn_no_country=mtn_no_country,
                         mtn_traffic_ready=mtn_traffic_ready,
                         mtn_traffic_total=mtn_traffic_total,
                         mtn_traffic_7d=mtn_traffic_7d,
                         mtn_traffic_30d=mtn_traffic_30d,
                         mtn_top_viewed=mtn_top_viewed,
                         mtn_monthly_views=mtn_monthly_views,
                         mtn_top5_traffic=mtn_top5_traffic,
                         mtn_most_viewed_month=mtn_most_viewed_month,
                         mtn_most_viewed_month_cnt=mtn_most_viewed_month_cnt,
                         mtn_avg_resorts_per_user=mtn_avg_resorts_per_user)


@app.route("/admin/resort-operations")
@login_required
@admin_required
def admin_resort_operations():
    """Operational resort management: catalog stats, data quality, full CRUD table."""
    import os
    import json
    from collections import Counter as _MtnCtr

    resorts = Resort.query.order_by(Resort.country_code, Resort.state_code, Resort.name).all()
    resort_countries = db.session.query(Resort.country_code).distinct().order_by(Resort.country_code).all()
    resort_countries = set(c[0] for c in resort_countries if c[0])
    dropdown_countries = sorted(COUNTRIES.items(), key=lambda x: x[1])
    countries = sorted(resort_countries)

    canonical_file = os.path.join(os.path.dirname(__file__), 'data', 'canonical_resorts.json')
    last_export_info = None
    if os.path.exists(canonical_file):
        try:
            with open(canonical_file, 'r') as f:
                canonical_data = json.load(f)
            last_export_info = {
                'version': canonical_data.get('version', 'unknown'),
                'exported_at': canonical_data.get('exported_at', 'unknown'),
                'count': len(canonical_data.get('resorts', []))
            }
        except Exception:
            pass

    total_count        = len(resorts)
    mtn_active_count   = Resort.query.filter_by(is_active=True).count()
    mtn_inactive_count = total_count - mtn_active_count

    _mtn_pbj = db.session.query(Resort.pass_brands_json).filter(Resort.is_active == True).all()
    _mtn_pass_ctr = _MtnCtr()
    _mtn_on_pass = 0
    for (pbj,) in _mtn_pbj:
        if pbj and isinstance(pbj, list):
            real = [b for b in pbj if b and b.lower() not in ('none', '')]
            if real:
                _mtn_on_pass += 1
                for b in real:
                    _mtn_pass_ctr[b] += 1
    mtn_resorts_on_pass = _mtn_on_pass

    _PASS_ORDER = ['Epic', 'Ikon', 'Mountain Collective', 'Indy', 'Other']
    _pc_max = max(_mtn_pass_ctr.values()) if _mtn_pass_ctr else 1
    mtn_by_pass_brand = [
        {"name": p, "count": _mtn_pass_ctr.get(p, 0),
         "pct": round(_mtn_pass_ctr.get(p, 0) / _pc_max * 100) if _mtn_pass_ctr else 0}
        for p in _PASS_ORDER
    ]

    _rbc = (
        db.session.query(Resort.country_code, Resort.country_name, func.count(Resort.id).label("cnt"))
        .filter(Resort.is_active == True)
        .group_by(Resort.country_code, Resort.country_name)
        .order_by(func.count(Resort.id).desc()).limit(8).all()
    )
    _rbc_max = max((r.cnt for r in _rbc), default=1)
    mtn_by_country = [
        {"name": r.country_name or r.country_code or "Unknown",
         "count": r.cnt, "pct": round(r.cnt / _rbc_max * 100)}
        for r in _rbc
    ]

    mtn_no_pass = sum(
        1 for r in resorts
        if not r.get_pass_brands_list() or all(b.lower() in ('none', '') for b in r.get_pass_brands_list())
    )
    mtn_no_state   = sum(1 for r in resorts if not (r.state_code or r.state))
    mtn_no_country = sum(1 for r in resorts if not r.country_code)

    return render_template('admin_resort_ops.html',
                           resorts=resorts,
                           countries=countries,
                           dropdown_countries=dropdown_countries,
                           last_export_info=last_export_info,
                           total_count=total_count,
                           active_tab='resort-ops',
                           mtn_active_count=mtn_active_count,
                           mtn_inactive_count=mtn_inactive_count,
                           mtn_resorts_on_pass=mtn_resorts_on_pass,
                           mtn_by_pass_brand=mtn_by_pass_brand,
                           mtn_by_country=mtn_by_country,
                           mtn_no_pass=mtn_no_pass,
                           mtn_no_state=mtn_no_state,
                           mtn_no_country=mtn_no_country,
                           COUNTRIES=COUNTRIES)


@app.route("/admin/resorts/export-excel")
@login_required
@admin_required
def admin_export_resorts_excel():
    """Export filtered resort list to Excel."""
    from io import BytesIO
    from openpyxl import Workbook
    from datetime import datetime
    
    # Trust query params ONLY for Excel export
    search_query = request.args.get('search', '').lower().strip()
    country_filter = request.args.get('country', '').strip()
    state_filter = request.args.get('state', '').strip()
    status_filter = request.args.get('status', '').lower().strip()
    pass_brand_filter = request.args.get('pass_brand', '').lower().strip()
    
    # Base query
    query = Resort.query
    
    # Apply country filter
    if country_filter:
        query = query.filter(Resort.country_code == country_filter)
        
    # Apply state filter
    if state_filter:
        query = query.filter((Resort.state_code == state_filter) | (Resort.state == state_filter))

    # Apply status filter
    if status_filter == 'active':
        query = query.filter(Resort.is_active == True)
    elif status_filter == 'inactive':
        query = query.filter(Resort.is_active == False)
        
    # Apply pass brand filter
    if pass_brand_filter:
        if pass_brand_filter == 'none':
            query = query.filter((Resort.pass_brands == None) | (Resort.pass_brands == '') | (Resort.pass_brands == 'None'))
        else:
            query = query.filter(Resort.pass_brands.ilike(f'%{pass_brand_filter}%'))

    resorts = query.all()
    
    # Apply search filter in-memory to match JS logic
    if search_query:
        resorts = [r for r in resorts if 
                   (r.name and search_query in r.name.lower()) or 
                   (r.state_code and search_query in r.state_code.lower()) or
                   (r.state and search_query in r.state.lower())]

    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            return jsonify({'status': 'error', 'message': 'Failed to initialize Excel worksheet'}), 500
            
        ws.title = "Resorts Export"
        
        # Headers explicitly defined - use country_code and country_name (exact DB columns)
        headers = [
            "ID",
            "Name",
            "country_code",
            "country_name",
            "State / Region",
            "Pass Brands",
            "Status"
        ]
        
        # Write headers by numeric index ONLY
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            from openpyxl.styles import Font
            cell.font = Font(bold=True)
        
        # Data starts at row 2 - export exactly as stored in DB
        for row_idx, r in enumerate(resorts, start=2):
            ws.cell(row=row_idx, column=1, value=r.id)
            ws.cell(row=row_idx, column=2, value=r.name or '')
            ws.cell(row=row_idx, column=3, value=r.country_code or '')
            ws.cell(row=row_idx, column=4, value=r.country_name or '')
            ws.cell(row=row_idx, column=5, value=r.state_code or r.state or '')
            ws.cell(row=row_idx, column=6, value=r.pass_brands or r.brand or '')
            ws.cell(row=row_idx, column=7, value="ACTIVE" if r.is_active else "INACTIVE")
            
        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        if len(val) > max_length:
                            max_length = len(val)
                    except:
                        pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"resorts_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        # Log the error for admin debugging
        app.logger.error(f"Excel Export Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'Export failed: {str(e)}'}), 500


@app.route("/api/admin/resorts/update-pass-brand", methods=["POST"])
@login_required
@admin_required
def admin_update_pass_brand():
    """Update a single resort's pass brands (supports array)."""
    data = request.get_json()
    resort_id = data.get('resort_id')
    pass_brands = data.get('pass_brands')
    
    # Handle legacy single value format
    if pass_brands is None:
        pass_brands = data.get('pass_brand')
    if isinstance(pass_brands, str):
        pass_brands = [pass_brands] if pass_brands else []
    
    # Valid values
    valid_brands = {'Epic', 'Ikon', 'Mountain Collective', 'Indy', 'Other', 'None'}
    
    # Validation
    for brand in pass_brands:
        if brand not in valid_brands:
            return jsonify({'status': 'error', 'message': f'Invalid pass brand: {brand}'}), 400
    
    # Handle "None" semantics - mutually exclusive with all other passes
    if 'None' in pass_brands:
        pass_brands = ['None']  # Preserve as explicit marker, not empty list
        
    resort = db.session.get(Resort, resort_id)
    if not resort:
        return jsonify({'status': 'error', 'message': 'Resort not found'}), 404
    
    resort.pass_brands_json = pass_brands
    db.session.commit()
    get_resorts_for_trip_form.cache_clear()
    return jsonify({'success': True, 'pass_brands': resort.get_pass_brands_list()})


@app.route("/api/admin/resorts/update-field", methods=["POST"])
@login_required
@admin_required
def admin_update_resort_field():
    """Update a single field on a resort (inline editing)."""
    data = request.get_json()
    resort_id = data.get('resort_id')
    field = data.get('field')
    value = data.get('value', '').strip()
    
    allowed_fields = ['name', 'country_code', 'state_code']
    if field not in allowed_fields:
        return jsonify({'success': False, 'message': f'Field {field} not allowed'}), 400
    
    resort = db.session.get(Resort, resort_id)
    if not resort:
        return jsonify({'success': False, 'message': 'Resort not found'}), 404
    
    if field == 'name':
        resort.name = value
    elif field == 'country_code':
        resort.country_code = value.upper() if value else None
        resort.country = value.upper() if value else None
    elif field == 'state_code':
        resort.state_code = value
        resort.state = value
    
    db.session.commit()
    get_resorts_for_trip_form.cache_clear()
    return jsonify({'success': True})


@app.route("/api/admin/resorts/update-country-name", methods=["POST"])
@login_required
@admin_required
def admin_update_country_name():
    """Update a resort's country name override."""
    data = request.get_json()
    resort_id = data.get('resort_id')
    country_name_override = data.get('country_name_override', '').strip()
    
    resort = db.session.get(Resort, resort_id)
    if not resort:
        return jsonify({'success': False, 'message': 'Resort not found'}), 404
    
    # Set to None if empty (falls back to COUNTRIES lookup)
    resort.country_name_override = country_name_override if country_name_override else None
    db.session.commit()
    get_resorts_for_trip_form.cache_clear()
    return jsonify({'success': True, 'display_country_name': resort.display_country_name})


@app.route("/api/admin/resorts/toggle-active", methods=["POST"])
@login_required
@admin_required
def admin_toggle_resort_active():
    """Toggle a resort's active status."""
    data = request.get_json()
    resort_id = data.get('resort_id')
    is_active = data.get('is_active', True)
    
    resort = db.session.get(Resort, resort_id)
    if not resort:
        return jsonify({'success': False, 'message': 'Resort not found'}), 404
    
    resort.is_active = is_active
    db.session.commit()
    get_resorts_for_trip_form.cache_clear()
    return jsonify({'success': True})


@app.route("/api/admin/resorts/delete", methods=["POST"])
@login_required
@admin_required
def admin_delete_resort_post():
    """Delete a single resort via POST (for frontend compatibility)."""
    data = request.get_json()
    resort_id = data.get('resort_id')
    
    resort = db.session.get(Resort, resort_id)
    if not resort:
        return jsonify({'success': False, 'message': 'Resort not found'}), 404
    
    trip_count = SkiTrip.query.filter_by(resort_id=resort_id).count()
    home_count = User.query.filter_by(home_resort_id=resort_id).count()
    
    visited_count = 0
    wishlist_count = 0
    users = User.query.all()
    for user in users:
        if user.visited_resort_ids and resort_id in user.visited_resort_ids:
            visited_count += 1
        if user.wish_list_resorts and resort_id in user.wish_list_resorts:
            wishlist_count += 1
    
    total_refs = trip_count + home_count + visited_count + wishlist_count
    
    if total_refs > 0:
        return jsonify({
            'success': False,
            'message': f'Cannot delete: resort has {total_refs} references. Deactivate instead or merge first.'
        }), 400
    
    resort_name = resort.name
    db.session.delete(resort)
    db.session.commit()
    get_resorts_for_trip_form.cache_clear()
    return jsonify({'success': True, 'message': f'Deleted resort: {resort_name}'})


@app.route("/api/admin/resorts/bulk-update-pass-brand", methods=["POST"])
@login_required
@admin_required
def admin_bulk_update_pass_brand():
    """Bulk update resorts' pass brands (supports array)."""
    data = request.get_json()
    resort_ids = data.get('resort_ids', [])
    pass_brands = data.get('pass_brands')
    
    # Handle legacy single value format
    if pass_brands is None:
        pass_brands = data.get('pass_brand')
    if isinstance(pass_brands, str):
        pass_brands = [pass_brands] if pass_brands else []
    
    # Valid values
    valid_brands = {'Epic', 'Ikon', 'Mountain Collective', 'Indy', 'Other', 'None'}
    
    # Validation
    for brand in pass_brands:
        if brand not in valid_brands:
            return jsonify({'status': 'error', 'message': f'Invalid pass brand: {brand}'}), 400
        
    if not resort_ids:
        return jsonify({'status': 'error', 'message': 'No resorts selected'}), 400
    
    # Handle "None" semantics - mutually exclusive with all other passes
    if 'None' in pass_brands:
        pass_brands = ['None']  # Preserve as explicit marker, not empty list
    
    # Update each resort (bulk update for JSON column)
    resorts = Resort.query.filter(Resort.id.in_(resort_ids)).all()
    for resort in resorts:
        resort.pass_brands_json = pass_brands
    
    db.session.commit()
    return jsonify({'updated_count': len(resorts)})


@app.route("/admin/resorts/import-excel", methods=["POST"])
@login_required
@admin_required
def admin_import_resorts_excel():
    """Import resort updates from Excel. Supports CREATE (blank ID) and UPDATE (existing ID).
    
    Country field normalization (Section 4 compliant):
    - CASE A: country_code ONLY → validate, derive country_name from mapping
    - CASE B: country_name ONLY → reverse-map to code if unique, else reject
    - CASE C: BOTH provided → validate they match, reject if mismatch
    - CASE D: NEITHER provided → leave unchanged (for updates), reject (for creates)
    """
    from openpyxl import load_workbook
    from io import BytesIO
    from utils.countries import is_valid_country_code, country_name_from_code, country_code_from_name
    
    def normalize_country_fields(code_val, name_val, resort_name):
        """Normalize country fields per Section 4 rules.
        Returns (country_code, country_name, error_message).
        If error_message is not None, the row should be rejected.
        """
        code_str = str(code_val).strip().upper() if code_val and str(code_val).strip() else None
        name_str = str(name_val).strip() if name_val and str(name_val).strip() else None
        
        # CASE D: Neither provided
        if not code_str and not name_str:
            return (None, None, None)  # No error, just leave unchanged
        
        # CASE A: country_code ONLY
        if code_str and not name_str:
            if is_valid_country_code(code_str):
                derived_name = country_name_from_code(code_str)
                return (code_str, derived_name, None)
            else:
                return (None, None, f"Invalid country_code '{code_str}' for resort: {resort_name}")
        
        # CASE B: country_name ONLY
        if name_str and not code_str:
            resolved_code = country_code_from_name(name_str)
            if resolved_code:
                resolved_name = country_name_from_code(resolved_code)
                return (resolved_code, resolved_name, None)
            else:
                return (None, None, f"Could not resolve country_name '{name_str}' to a unique code for resort: {resort_name}")
        
        # CASE C: BOTH provided - validate match
        if code_str and name_str:
            if not is_valid_country_code(code_str):
                return (None, None, f"Invalid country_code '{code_str}' for resort: {resort_name}")
            
            expected_name = country_name_from_code(code_str)
            # Check if names match (case-insensitive, whitespace-normalized)
            if expected_name and expected_name.casefold() == name_str.casefold():
                return (code_str, expected_name, None)
            else:
                return (None, None, f"Country code/name mismatch for resort: {resort_name} (code={code_str}, name={name_str}, expected={expected_name})")
        
        return (None, None, None)
    
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'status': 'error', 'message': 'Invalid file format. Please upload .xlsx'}), 400
        
    try:
        wb = load_workbook(BytesIO(file.read()))
        ws = wb.active
        
        rows_processed = 0
        rows_created = 0
        rows_updated = 0
        rows_skipped = 0
        errors = []
        
        # Detect header row (row 1 now contains headers)
        # Format: ID, Name, country_code, country_name, State/Region, Pass Brands, Status
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] if ws.max_row >= 1 else None
        has_country_name_col = False
        if header_row:
            headers_lower = [str(h).lower().strip() if h else '' for h in header_row]
            has_country_name_col = 'country_name' in headers_lower or 'country name' in headers_lower
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
                
            rows_processed += 1
            resort_id = row[0]
            name = row[1]
            country_code_val = row[2] if len(row) > 2 else None
            
            # Parse based on detected schema
            if has_country_name_col:
                # New 7-column format: ID, Name, country_code, country_name, State/Region, Pass Brands, Status
                country_name_val = row[3] if len(row) > 3 else None
                state_region = row[4] if len(row) > 4 else None
                pass_brands = row[5] if len(row) > 5 else None
                status = str(row[6]).upper() if len(row) > 6 and row[6] else None
            else:
                # Legacy 6-column format (no country_name column)
                country_name_val = None
                state_region = row[3] if len(row) > 3 else None
                pass_brands = row[4] if len(row) > 4 else None
                status = str(row[5]).upper() if len(row) > 5 and row[5] else None
            
            # Validation for status
            if status and status not in ['ACTIVE', 'INACTIVE']:
                rows_skipped += 1
                errors.append({'row': row_idx, 'reason': f'Invalid status: {status}'})
                continue
            
            # Normalize country fields using Section 4 rules
            resort_name_str = str(name).strip() if name else f'Row {row_idx}'
            normalized_code, normalized_name, country_error = normalize_country_fields(
                country_code_val, country_name_val, resort_name_str
            )
            
            # CREATE MODE: ID is blank/null
            if not resort_id:
                # Required fields for creation
                if not name or not str(name).strip():
                    rows_skipped += 1
                    errors.append({'row': row_idx, 'reason': 'Missing required field: name'})
                    continue
                
                # CASE D for CREATE: country is required
                if not normalized_code:
                    if country_error:
                        rows_skipped += 1
                        errors.append({'row': row_idx, 'reason': country_error})
                    else:
                        rows_skipped += 1
                        errors.append({'row': row_idx, 'reason': 'Missing required country_code for new resort'})
                    continue
                
                # Generate slug from name
                resort_name = str(name).strip()
                try:
                    base_slug = generate_resort_slug(resort_name)
                except ValueError as e:
                    rows_skipped += 1
                    errors.append({'row': row_idx, 'reason': str(e)})
                    continue
                
                # Ensure slug is unique by appending suffix if needed
                slug = base_slug
                suffix = 1
                while Resort.query.filter_by(slug=slug).first():
                    slug = f"{base_slug}-{suffix}"
                    suffix += 1
                
                # Create new resort with normalized country fields
                new_resort = Resort(
                    name=resort_name,
                    slug=slug,
                    country_code=normalized_code,
                    country=normalized_code,
                    country_name=normalized_name,
                    state_code=str(state_region).strip() if state_region else None,
                    state=str(state_region).strip() if state_region else None,
                    pass_brands=str(pass_brands).strip() if pass_brands else None,
                    brand=str(pass_brands).strip().split(',')[0] if pass_brands else 'Other',
                    is_active=True if not status else (status == 'ACTIVE')
                )
                db.session.add(new_resort)
                rows_created += 1
                continue
            
            # UPDATE MODE: ID is present
            resort = db.session.get(Resort, resort_id)
            if not resort:
                rows_skipped += 1
                errors.append({'row': row_idx, 'reason': f'Resort ID {resort_id} not found'})
                continue
            
            # Check for country field errors (reject row if mismatch/invalid)
            if country_error:
                rows_skipped += 1
                errors.append({'row': row_idx, 'reason': country_error})
                continue
                
            # Apply updates
            updated = False
            if name and str(name).strip() != resort.name:
                resort.name = str(name).strip()
                updated = True
            
            # Apply normalized country fields (CASE D leaves unchanged)
            if normalized_code and normalized_code != resort.country_code:
                resort.country_code = normalized_code
                resort.country = normalized_code
                updated = True
            if normalized_name and normalized_name != resort.country_name:
                resort.country_name = normalized_name
                updated = True
                
            if state_region is not None:
                new_state = str(state_region).strip()
                if new_state != resort.state_code:
                    resort.state_code = new_state
                    resort.state = new_state
                    updated = True
                    
            if pass_brands is not None:
                new_passes = str(pass_brands).strip()
                if new_passes != resort.pass_brands:
                    resort.pass_brands = new_passes
                    resort.brand = new_passes.split(',')[0] if new_passes else 'Other'
                    updated = True
                    
            if status:
                new_active = (status == 'ACTIVE')
                if new_active != resort.is_active:
                    resort.is_active = new_active
                    updated = True
                    
            if updated:
                rows_updated += 1
                
        db.session.commit()
        
        return jsonify({
            'rows_processed': rows_processed,
            'rows_created': rows_created,
            'rows_updated': rows_updated,
            'rows_skipped': rows_skipped,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route("/api/admin/resorts/<int:resort_id>", methods=["PUT"])
@login_required
@admin_required
def admin_update_resort(resort_id):
    """Update a single resort's editable fields."""
    resort = Resort.query.get_or_404(resort_id)
    data = request.get_json()
    
    # Only allow updating specific fields
    if 'name' in data:
        resort.name = (data.get('name') or '').strip()
    if 'country_code' in data:
        resort.country_code = (data.get('country_code') or '').strip().upper()
        resort.country = resort.country_code
    if 'state_code' in data:
        resort.state_code = (data.get('state_code') or '').strip()
        resort.state = resort.state_code
    if 'pass_brands' in data:
        resort.pass_brands = (data.get('pass_brands') or '').strip() or None
        resort.brand = resort.pass_brands.split(',')[0] if resort.pass_brands else 'Other'
    if 'is_active' in data:
        resort.is_active = bool(data['is_active'])
    
    db.session.commit()
    return jsonify({'status': 'success', 'resort_id': resort_id})


@app.route("/api/admin/resorts/<int:resort_id>", methods=["DELETE"])
@login_required
@admin_required
def admin_delete_resort(resort_id):
    """Hard delete a resort. Checks for FK references first."""
    resort = Resort.query.get_or_404(resort_id)
    
    # Check for existing references
    trip_count = SkiTrip.query.filter_by(resort_id=resort_id).count()
    home_count = User.query.filter_by(home_resort_id=resort_id).count()
    
    # Check JSON arrays for references
    visited_count = 0
    wishlist_count = 0
    users = User.query.all()
    for user in users:
        if user.visited_resort_ids and resort_id in user.visited_resort_ids:
            visited_count += 1
        if user.wish_list_resorts and resort_id in user.wish_list_resorts:
            wishlist_count += 1
    
    total_refs = trip_count + home_count + visited_count + wishlist_count
    
    if total_refs > 0:
        return jsonify({
            'status': 'error',
            'message': f'Cannot delete: resort has {total_refs} references (trips: {trip_count}, home: {home_count}, visited: {visited_count}, wishlist: {wishlist_count}). Deactivate instead or merge first.'
        }), 400
    
    resort_name = resort.name
    db.session.delete(resort)
    db.session.commit()
    
    return jsonify({'status': 'success', 'message': f'Deleted resort: {resort_name}'})


@app.route("/api/admin/resorts/bulk-delete", methods=["POST"])
@login_required
@admin_required
def admin_bulk_delete_resorts():
    """Bulk delete resorts. Returns partial success if some have FK references."""
    data = request.get_json()
    ids = data.get('ids', [])
    
    if not ids:
        return jsonify({'status': 'error', 'message': 'No resorts selected'}), 400
    
    deleted = []
    blocked = []
    
    for resort_id in ids:
        resort = db.session.get(Resort, resort_id)
        if not resort:
            blocked.append({'id': resort_id, 'name': 'Unknown', 'reason': 'Not found'})
            continue
        
        # Check FK references
        trip_count = SkiTrip.query.filter_by(resort_id=resort_id).count()
        home_count = User.query.filter_by(home_resort_id=resort_id).count()
        
        visited_count = 0
        wishlist_count = 0
        users = User.query.all()
        for user in users:
            if user.visited_resort_ids and resort_id in user.visited_resort_ids:
                visited_count += 1
            if user.wish_list_resorts and resort_id in user.wish_list_resorts:
                wishlist_count += 1
        
        total_refs = trip_count + home_count + visited_count + wishlist_count
        
        if total_refs > 0:
            blocked.append({
                'id': resort_id,
                'name': resort.name,
                'reason': f'Has {total_refs} references (trips: {trip_count}, home: {home_count}, visited: {visited_count}, wishlist: {wishlist_count})'
            })
        else:
            deleted.append({'id': resort_id, 'name': resort.name})
            db.session.delete(resort)
    
    db.session.commit()
    
    return jsonify({
        'status': 'success',
        'deleted': deleted,
        'blocked': blocked
    })


@app.route("/api/admin/resorts/bulk-activate", methods=["POST"])
@login_required
@admin_required
def admin_bulk_activate_resorts():
    """Bulk activate resorts."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': 'Invalid request body'}), 400
        
        resort_ids = data.get('resort_ids', [])
        
        if not resort_ids:
            return jsonify({'status': 'error', 'message': 'No resorts selected'}), 400
        
        updated_count = Resort.query.filter(Resort.id.in_(resort_ids)).update(
            {Resort.is_active: True},
            synchronize_session=False
        )
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'updated_count': updated_count
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route("/api/admin/resorts/bulk-deactivate", methods=["POST"])
@login_required
@admin_required
def admin_bulk_deactivate_resorts():
    """Bulk deactivate resorts."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': 'Invalid request body'}), 400
        
        resort_ids = data.get('resort_ids', [])
        
        if not resort_ids:
            return jsonify({'status': 'error', 'message': 'No resorts selected'}), 400
        
        updated_count = Resort.query.filter(Resort.id.in_(resort_ids)).update(
            {Resort.is_active: False},
            synchronize_session=False
        )
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'updated_count': updated_count
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route("/api/admin/resorts/merge", methods=["POST"])
@login_required
@admin_required
def admin_merge_resorts():
    """
    Merge duplicate resorts into a canonical resort.
    Repoints all FK references atomically, then marks non-canonical resorts as inactive.
    """
    data = request.get_json()
    canonical_id = data.get('canonical_id')
    duplicate_ids = data.get('duplicate_ids', [])
    
    if not canonical_id or not duplicate_ids:
        return jsonify({'status': 'error', 'message': 'Missing canonical_id or duplicate_ids'}), 400
    
    if canonical_id in duplicate_ids:
        return jsonify({'status': 'error', 'message': 'Canonical resort cannot be in duplicate list'}), 400
    
    canonical = db.session.get(Resort, canonical_id)
    if not canonical:
        return jsonify({'status': 'error', 'message': 'Canonical resort not found'}), 404
    
    duplicates = Resort.query.filter(Resort.id.in_(duplicate_ids)).all()
    if len(duplicates) != len(duplicate_ids):
        return jsonify({'status': 'error', 'message': 'Some duplicate resorts not found'}), 404
    
    try:
        stats = {
            'trips_updated': 0,
            'home_resorts_updated': 0,
            'visited_lists_updated': 0,
            'wishlist_updated': 0,
            'resorts_deactivated': 0
        }
        
        for dup in duplicates:
            dup_id = dup.id
            
            # 1. Update SkiTrip.resort_id
            trips_updated = SkiTrip.query.filter_by(resort_id=dup_id).update({'resort_id': canonical_id})
            stats['trips_updated'] += trips_updated
            
            # 2. Update User.home_resort_id
            home_updated = User.query.filter_by(home_resort_id=dup_id).update({'home_resort_id': canonical_id})
            stats['home_resorts_updated'] += home_updated
            
            # 3. Update User.visited_resort_ids (JSON array)
            # Must use flag_modified to ensure SQLAlchemy detects JSON changes
            from sqlalchemy.orm.attributes import flag_modified
            users_with_visited = User.query.filter(User.visited_resort_ids.isnot(None)).all()
            for user in users_with_visited:
                if user.visited_resort_ids and dup_id in user.visited_resort_ids:
                    new_list = list(user.visited_resort_ids)  # Copy to new list
                    new_list = [rid for rid in new_list if rid != dup_id]
                    if canonical_id not in new_list:
                        new_list.append(canonical_id)
                    user.visited_resort_ids = new_list
                    flag_modified(user, 'visited_resort_ids')
                    stats['visited_lists_updated'] += 1
            
            # 4. Update User.wish_list_resorts (JSON array)
            users_with_wishlist = User.query.filter(User.wish_list_resorts.isnot(None)).all()
            for user in users_with_wishlist:
                if user.wish_list_resorts and dup_id in user.wish_list_resorts:
                    new_list = list(user.wish_list_resorts)  # Copy to new list
                    new_list = [rid for rid in new_list if rid != dup_id]
                    if canonical_id not in new_list:
                        new_list.append(canonical_id)
                    user.wish_list_resorts = new_list
                    flag_modified(user, 'wish_list_resorts')
                    stats['wishlist_updated'] += 1
            
            # 5. Mark duplicate as inactive
            dup.is_active = False
            stats['resorts_deactivated'] += 1
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'canonical_resort': canonical.name,
            'stats': stats
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route("/api/admin/resorts/add", methods=["POST"])
@login_required
@admin_required
def admin_add_resort():
    """
    Add a new resort via free-form entry.
    No tier logic, no canonical writes, no auto-merge.
    """
    import re
    
    data = request.get_json()
    if not data:
        return jsonify({'status': 'error', 'message': 'No data provided'}), 400
    
    name = (data.get('name') or '').strip()
    country_code = (data.get('country_code') or '').strip().upper()
    state_code = (data.get('state_code') or '').strip() or None
    pass_brands_input = data.get('pass_brands')
    
    # Handle pass_brands as array or comma-separated string
    if isinstance(pass_brands_input, list):
        pass_brands_list = pass_brands_input
    elif isinstance(pass_brands_input, str):
        pass_brands_list = [p.strip() for p in pass_brands_input.split(',') if p.strip()]
    else:
        pass_brands_list = []
    
    # Handle "None" semantics - mutually exclusive with all other passes
    if 'None' in pass_brands_list:
        pass_brands_list = ['None']  # Preserve as explicit marker, not empty list
    
    errors = []
    if not name:
        errors.append('Resort name is required')
    elif len(name) < 2:
        errors.append('Name must be at least 2 characters')
    
    if not country_code:
        errors.append('Please select a country')
    elif len(country_code) != 2:
        errors.append('Invalid country code')
    
    if errors:
        return jsonify({'status': 'error', 'message': '; '.join(errors)}), 400
    
    normalized_name = name.lower()
    existing = Resort.query.filter(
        db.func.lower(Resort.name) == normalized_name,
        db.func.upper(Resort.country_code) == country_code
    ).first()
    
    if existing:
        return jsonify({
            'status': 'error',
            'message': f'Resort "{name}" already exists in {country_code} (ID: {existing.id})'
        }), 409
    
    slug_base = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')
    slug = f"{slug_base}-{country_code.lower()}"
    
    slug_exists = Resort.query.filter_by(slug=slug).first()
    if slug_exists:
        slug = f"{slug}-{int(datetime.utcnow().timestamp())}"
    
    new_resort = Resort(
        name=name,
        country=country_code,
        country_code=country_code,
        country_name=COUNTRIES.get(country_code, country_code),
        state=state_code,
        state_code=state_code,
        state_name=state_code,
        state_full=state_code,
        brand=None,
        pass_brands_json=pass_brands_list,
        slug=slug,
        is_active=True,
        is_region=False
    )
    
    try:
        db.session.add(new_resort)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Resort "{name}" added successfully',
            'resort': {
                'id': new_resort.id,
                'name': new_resort.name,
                'country_code': new_resort.country_code,
                'state_code': new_resort.state_code,
                'pass_brands': new_resort.get_pass_brands_list(),
                'pass_brands_json': new_resort.pass_brands_json,
                'slug': new_resort.slug
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route("/api/admin/resorts/export-canonical", methods=["POST"])
@login_required
@admin_required
def admin_export_canonical():
    """
    Export active resorts to canonical_resorts.json.
    This is the source of truth for PROD sync.
    """
    import os
    import json
    from datetime import datetime
    
    # Get all active resorts
    resorts = Resort.query.filter_by(is_active=True).order_by(Resort.country_code, Resort.state_code, Resort.name).all()
    
    canonical_data = {
        'version': datetime.utcnow().strftime('%Y%m%d_%H%M%S'),
        'exported_at': datetime.utcnow().isoformat() + 'Z',
        'exported_by': current_user.email,
        'total_count': len(resorts),
        'resorts': []
    }
    
    for r in resorts:
        canonical_data['resorts'].append({
            'name': r.name,
            'state_code': r.state_code or r.state,
            'country_code': r.country_code,
            'country_name': r.country_name,
            'pass_brands': r.pass_brands or r.brand or ''
        })
    
    canonical_file = os.path.join(os.path.dirname(__file__), 'data', 'canonical_resorts.json')
    
    try:
        with open(canonical_file, 'w') as f:
            json.dump(canonical_data, f, indent=2)
        
        return jsonify({
            'status': 'success',
            'version': canonical_data['version'],
            'exported_at': canonical_data['exported_at'],
            'count': len(resorts),
            'message': f'Exported {len(resorts)} resorts to canonical_resorts.json'
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route("/admin/sync-from-canonical", methods=["GET", "POST"])
@login_required
@admin_required
def admin_sync_from_canonical():
    """
    Sync resorts from canonical_resorts.json into the database.
    Used in production after deploying updated JSON.
    """
    import os
    import json
    
    canonical_file = os.path.join(os.path.dirname(__file__), 'data', 'canonical_resorts.json')
    
    if not os.path.exists(canonical_file):
        if request.method == 'POST':
            return jsonify({'status': 'error', 'message': 'canonical_resorts.json not found'}), 404
        return "No canonical_resorts.json found", 404
    
    with open(canonical_file, 'r') as f:
        canonical_data = json.load(f)
    
    if request.method == 'GET':
        current_count = Resort.query.filter_by(is_active=True).count()
        return f'''
        <html>
        <head><title>Sync Resorts</title>
        <style>
            body {{ font-family: system-ui; padding: 40px; max-width: 600px; margin: 0 auto; }}
            .info {{ background: #f0f0f0; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
            button {{ background: #8F011B; color: white; padding: 12px 24px; border: none; border-radius: 6px; cursor: pointer; font-size: 16px; }}
            button:hover {{ background: #7a0117; }}
            .warning {{ color: #856404; background: #fff3cd; padding: 15px; border-radius: 6px; margin-bottom: 20px; }}
        </style>
        </head>
        <body>
            <h1>Sync Resorts from Canonical JSON</h1>
            <div class="info">
                <p><strong>JSON Version:</strong> {canonical_data.get('version', 'unknown')}</p>
                <p><strong>Exported:</strong> {canonical_data.get('exported_at', 'unknown')}</p>
                <p><strong>Resorts in JSON:</strong> {canonical_data.get('total_count', len(canonical_data.get('resorts', [])))}</p>
                <p><strong>Current DB Count:</strong> {current_count}</p>
            </div>
            <div class="warning">
                This will update/insert resorts from the canonical JSON. Existing resorts not in the JSON will be deactivated.
            </div>
            <form method="POST">
                <button type="submit">Sync Now</button>
            </form>
        </body>
        </html>
        '''
    
    # POST - do the sync
    try:
        stats = {'added': 0, 'updated': 0, 'deactivated': 0}
        
        canonical_keys = set()
        
        # Build a lookup of existing resorts by normalized key
        all_existing = Resort.query.all()
        existing_by_key = {}
        for r in all_existing:
            key = f"{r.name}|{r.state_code or r.state or ''}|{r.country_code or r.country or 'US'}".lower()
            existing_by_key[key] = r
        
        for r_data in canonical_data.get('resorts', []):
            name = (r_data.get('name') or '').strip()
            state_code = (r_data.get('state_code') or '').strip()
            country_code = (r_data.get('country_code') or 'US').strip()
            country_name = (r_data.get('country_name') or '').strip()
            pass_brands = (r_data.get('pass_brands') or '').strip()
            is_region = r_data.get('is_region', False)

            # Derive state_name from STATE_ABBR_MAP for US/CA codes, else use as-is
            state_name = STATE_ABBR_MAP.get(state_code.upper()) if state_code else None
            if not state_name and state_code:
                state_name = state_code

            # Derive pass_brands_json from comma-separated string
            if pass_brands and pass_brands != 'None':
                pass_brands_json = [p.strip() for p in pass_brands.split(',') if p.strip()]
            else:
                pass_brands_json = ['None']
            brand = pass_brands_json[0] if pass_brands_json else 'Other'

            canonical_key = f"{name}|{state_code}|{country_code}".lower()
            canonical_keys.add(canonical_key)
            
            existing = existing_by_key.get(canonical_key)
            
            if existing:
                existing.name = name
                existing.state_code = state_code
                existing.state = state_code
                existing.state_name = state_name
                existing.country_code = country_code
                existing.country = country_code
                existing.country_name = country_name
                existing.pass_brands = pass_brands
                existing.pass_brands_json = pass_brands_json
                existing.brand = brand
                existing.is_active = True
                existing.is_region = is_region
                stats['updated'] += 1
            else:
                import re
                base_slug = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')
                base_slug = f"{base_slug}-{state_code.lower()}" if state_code else base_slug
                slug = base_slug
                counter = 1
                while Resort.query.filter_by(slug=slug).first():
                    slug = f"{base_slug}-{counter}"
                    counter += 1
                new_resort = Resort(
                    name=name,
                    state=state_code,
                    state_code=state_code,
                    state_name=state_name,
                    country=country_code,
                    country_code=country_code,
                    country_name=country_name,
                    pass_brands=pass_brands,
                    pass_brands_json=pass_brands_json,
                    brand=brand,
                    slug=slug,
                    is_active=True,
                    is_region=is_region
                )
                db.session.add(new_resort)
                stats['added'] += 1
        
        # Deactivate resorts not in canonical
        for key, resort in existing_by_key.items():
            if key not in canonical_keys and resort.is_active:
                resort.is_active = False
                stats['deactivated'] += 1
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'version': canonical_data.get('version'),
            'stats': stats,
            'message': f"Sync complete: {stats['added']} added, {stats['updated']} updated, {stats['deactivated']} deactivated"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# TEMP DEBUG ROUTE — safe to remove after use
@app.route("/admin/debug-resort-duplicates", methods=["GET"])
@login_required
@admin_required
def debug_resort_duplicates():
    """Temporary route to detect duplicate resorts by (name, state_code, country_code)."""
    from sqlalchemy import func

    rows = (
        db.session.query(
            Resort.name,
            Resort.state_code,
            Resort.country_code,
            func.count(Resort.id).label("cnt")
        )
        .group_by(Resort.name, Resort.state_code, Resort.country_code)
        .having(func.count(Resort.id) > 1)
        .order_by(func.count(Resort.id).desc())
        .all()
    )

    if not rows:
        return jsonify({"status": "ok", "duplicates": []})

    return jsonify([
        {
            "name": row.name,
            "state_code": row.state_code,
            "country_code": row.country_code,
            "count": row.cnt
        }
        for row in rows
    ])


@app.route("/admin/resorts/duplicates", methods=["GET"])
@login_required
@admin_required
def admin_resorts_duplicates():
    """Find duplicate resorts grouped by normalized (name, state_code, country_code)."""
    try:
        norm_name = func.lower(func.trim(Resort.name))
        norm_state = func.upper(func.trim(Resort.state_code))
        norm_country = func.upper(func.trim(func.coalesce(Resort.country_code, 'US')))

        groups = (
            db.session.query(
                func.lower(func.trim(Resort.name)).label("norm_name"),
                func.upper(func.trim(Resort.state_code)).label("norm_state"),
                func.upper(func.trim(func.coalesce(Resort.country_code, 'US'))).label("norm_country"),
                func.count(Resort.id).label("cnt")
            )
            .group_by(
                func.lower(func.trim(Resort.name)),
                func.upper(func.trim(Resort.state_code)),
                func.upper(func.trim(func.coalesce(Resort.country_code, 'US')))
            )
            .having(func.count(Resort.id) > 1)
            .order_by(func.count(Resort.id).desc())
            .all()
        )

        result_groups = []
        total_duplicate_rows = 0

        for g in groups:
            matches = Resort.query.filter(
                func.lower(func.trim(Resort.name)) == g.norm_name,
                func.upper(func.trim(Resort.state_code)) == g.norm_state,
                func.upper(func.trim(func.coalesce(Resort.country_code, 'US'))) == g.norm_country
            ).all()

            total_duplicate_rows += len(matches)
            result_groups.append({
                "name": g.norm_name,
                "state_code": g.norm_state,
                "country_code": g.norm_country,
                "count": g.cnt,
                "resorts": [
                    {
                        "id": r.id,
                        "name": r.name,
                        "state_code": r.state_code,
                        "country_code": r.country_code,
                        "slug": r.slug,
                        "is_active": r.is_active
                    }
                    for r in matches
                ]
            })

        return jsonify({
            "status": "success",
            "duplicate_group_count": len(result_groups),
            "total_duplicate_rows": total_duplicate_rows,
            "groups": result_groups
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/download")
def download_page():
    """
    Public download page — no login required.
    iOS/iPadOS users are redirected to the App Store.
    Android users are redirected to Google Play.
    Desktop / unknown browsers see a page with both store buttons.
    """
    _APP_STORE_URL  = "https://apps.apple.com/us/app/baselodge/id6764206581"
    _PLAY_STORE_URL = "https://play.google.com/store/apps/details?id=com.baselodge.app&pli=1"

    ua = (request.user_agent.string or "").lower()
    if any(t in ua for t in ("iphone", "ipad", "ipod")):
        return redirect(_APP_STORE_URL, code=302)
    if "android" in ua:
        return redirect(_PLAY_STORE_URL, code=302)
    return render_template("download.html")


@app.route("/download/qr")
def download_qr():
    """Public QR code preview page — no login required."""
    return render_template("download_qr.html")


@app.route("/privacypolicy")
def privacy_policy():
    return render_template("privacy_policy.html")


@app.route("/termsandconditions")
def terms_and_conditions():
    return render_template("terms_and_conditions.html")


@app.route("/robots.txt")
def robots_txt():
    return send_file("static/robots.txt", mimetype="text/plain", max_age=86400)


# ── Deep-link verification files ─────────────────────────────────────────────
#
# These two routes let iOS (Universal Links) and Android (App Links) verify that
# app.baselodgeapp.com is associated with the BaseLodge app, so that invite URLs
# open directly in the installed app instead of the mobile browser.
#
# CURRENT STATUS: placeholders are served — real values need to be filled in.
#
# HOW TO COMPLETE SETUP:
#
# iOS Universal Links (apple-app-site-association):
#   1. Find your Apple Team ID at https://developer.apple.com → Membership
#   2. Replace "XXXXXXXXXX" below with your 10-character Team ID
#   3. The bundle ID is already correct: com.baselodge.app
#   4. In Xcode: target → Signing & Capabilities → + Capability → Associated Domains
#      Add: applinks:app.baselodgeapp.com
#   5. In Capacitor iOS (AppDelegate.swift or capacitor.config.json), handle
#      application(_:continue:restorationHandler:) to parse the invite URL
#
# Android App Links (assetlinks.json):
#   1. In Android Studio: App Signing → upload certificate SHA-256 fingerprint
#      (from Google Play Console → App Signing → App signing key certificate)
#   2. Replace the placeholder SHA-256 below with your real certificate fingerprint
#   3. In android/app/src/main/AndroidManifest.xml, add an intent-filter with
#      android:autoVerify="true" for https://app.baselodgeapp.com/invite/*
#   4. In Capacitor Android, handle appUrlOpen via @capacitor/app's App.addListener
#
# Server-side token preservation (both platforms):
#   The Capacitor App plugin fires appUrlOpen when the app is opened via a deep link.
#   Read the URL, extract the token, and navigate to /invite/<token> inside the webview.
#   The session cookie is already present (same domain), so the accept flow works normally.
#
@app.route("/.well-known/apple-app-site-association")
def apple_app_site_association():
    """iOS Universal Links verification file.

    Tells iOS that app.baselodgeapp.com/invite/* links should open in BaseLodge.
    Replace XXXXXXXXXX with the real Apple Team ID before submitting to the App Store.
    """
    import json as _json
    aasa = {
        "applinks": {
            "apps": [],
            "details": [
                {
                    "appID": "XXXXXXXXXX.com.baselodge.app",  # ← replace Team ID
                    "paths": [
                        "/invite/*",
                        "/trip-invite/*"
                    ]
                }
            ]
        }
    }
    return app.response_class(
        response=_json.dumps(aasa),
        status=200,
        mimetype="application/json"
    )


@app.route("/.well-known/assetlinks.json")
def assetlinks_json():
    """Android App Links verification file.

    Tells Android that app.baselodgeapp.com/invite/* links should open in BaseLodge.
    Replace the placeholder SHA-256 fingerprint with the real signing certificate
    fingerprint from Google Play Console → App Signing before releasing to Play Store.
    """
    import json as _json
    assetlinks = [
        {
            "relation": ["delegate_permission/common.handle_all_urls"],
            "target": {
                "namespace": "android_app",
                "package_name": "com.baselodge.app",
                "sha256_cert_fingerprints": [
                    # ← replace with real SHA-256 from Google Play Console → App Signing
                    "00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00:00"
                ]
            }
        }
    ]
    return app.response_class(
        response=_json.dumps(assetlinks),
        status=200,
        mimetype="application/json"
    )


@app.route("/sitemap.xml")
def sitemap_xml():
    return send_file("static/sitemap.xml", mimetype="application/xml", max_age=86400)


@app.route("/admin/test-onesignal-push", methods=["GET", "POST"])
@login_required
@admin_required
def admin_test_onesignal_push():
    """Send a test push via OneSignal REST API to the current admin user.

    Targets the logged-in admin's BaseLodge user ID as the OneSignal external ID.
    Returns JSON — HTTP 200 on success or silent opt-out skip, HTTP 502 on delivery failure.

    Verifies without exposing:
      - ONESIGNAL_APP_ID is set (logged and reflected in response)
      - ONESIGNAL_REST_API_KEY is set (only presence reported, value never logged)
    """
    current_app.logger.warning(
        "[OneSignal-Test] triggered by admin user_id=%d email=%s",
        current_user.id, current_user.email,
    )

    # ── QA Override: route test push to Richard's device ──────────────────────
    _qa_user_os = _get_qa_push_override_user()
    _os_target_id = current_user.id
    if _qa_user_os:
        _os_target_id = _qa_user_os.id
        current_app.logger.warning(
            "[OneSignal-Test] QA Override Active — routed test push to Richard "
            "(original user_id=%d → qa_user_id=%d email=%s)",
            current_user.id, _qa_user_os.id, _QA_PUSH_OVERRIDE_EMAIL,
        )

    result = send_onesignal_push(
        user_ids=[_os_target_id],
        title="BaseLodge",
        body="Test push from BaseLodge (OneSignal)",
        data={"source": "admin_test"},
    )

    # ── MessageEventLog: OneSignal test push outcome ──
    _os_success = result.get("success") and not result.get("skipped")
    _os_skipped = result.get("skipped", False)
    try:
        create_message_event(
            event_name=EventName.PUSH_TEST_SENT,
            category=Category.SYSTEM,
            actor_user_id=current_user.id,
            recipient_user_id=current_user.id,
            channel=Channel.PUSH,
            provider=Provider.ONESIGNAL,
            payload_json={
                "notification_id": result.get("provider_message_id"),
                "source_route": "admin_test_onesignal_push",
            },
            message_title="BaseLodge",
            message_body="Test push from BaseLodge (OneSignal)",
            delivery_status=(
                DeliveryStatus.SENT    if _os_success else
                DeliveryStatus.SKIPPED if _os_skipped else
                DeliveryStatus.FAILED
            ),
            suppression_reason=SuppressionReason.USER_OPTED_OUT if _os_skipped else None,
            error_message=result.get("error") if not _os_success and not _os_skipped else None,
        )
    except Exception as _mel_err:
        current_app.logger.warning("[MessageEvent] test_onesignal log failed: %s", _mel_err)

    http_status = 200 if result.get("success") else 502
    return jsonify({
        "provider":               "onesignal",
        "target_user_id":         current_user.id,
        "target_external_id":     str(current_user.id),
        "onesignal_app_id_set":   bool(os.environ.get("ONESIGNAL_APP_ID")),
        "onesignal_rest_key_set": bool(os.environ.get("ONESIGNAL_REST_API_KEY")),
        "result":                 result,
    }), http_status


@app.route("/admin/message-events", methods=["GET"])
@login_required
@admin_required
def admin_message_events():
    """Admin debug view — shows the last 200 MessageEventLog rows, newest first.

    Human-readable table for operational visibility and debugging.
    No writes, no notifications, no side effects.
    """
    rows = (
        MessageEventLog.query
        .options(db.joinedload(MessageEventLog.recipient))
        .order_by(MessageEventLog.created_at.desc())
        .limit(200)
        .all()
    )

    status_counts = (
        db.session.query(
            MessageEventLog.delivery_status,
            db.func.count(MessageEventLog.id),
        )
        .group_by(MessageEventLog.delivery_status)
        .all()
    )
    count_map = {status: count for status, count in status_counts}
    stats = {
        "total":   sum(count_map.values()),
        "sent":    count_map.get(DeliveryStatus.SENT, 0),
        "skipped": count_map.get(DeliveryStatus.SKIPPED, 0),
        "failed":  count_map.get(DeliveryStatus.FAILED, 0),
        "pending": count_map.get(DeliveryStatus.PENDING, 0),
    }

    return render_template("admin_message_events.html", rows=rows, stats=stats)


@app.route("/admin/test-message-event", methods=["GET"])
@login_required
@admin_required
def admin_test_message_event():
    """Admin-only route that creates three sample MessageEventLog rows.

    Purpose: verify helper behavior, dedupe logic, and admin page visibility.
    Sends NO real notifications. All rows are labeled source=admin_test.

    Rows created:
      1. sent      — push.test.sent (bypasses dedupe)
      2. skipped   — overlap.detected / digest_only suppression
      3. skipped   — overlap.detected / duplicate_event (dedupe fires on row 2)
    """
    current_app.logger.warning(
        "[MessageEvent-Test] triggered by admin user_id=%d email=%s",
        current_user.id, current_user.email,
    )

    create_message_event(
        event_name=EventName.PUSH_TEST_SENT,
        category=Category.SYSTEM,
        actor_user_id=current_user.id,
        recipient_user_id=current_user.id,
        channel=Channel.PUSH,
        provider=Provider.INTERNAL,
        payload_json={"source": "admin_test", "label": "TEST_SENT"},
        message_title="BaseLodge Test",
        message_body="Test sent row created by admin test route.",
        delivery_status=DeliveryStatus.SENT,
    )

    create_message_event(
        event_name=EventName.OVERLAP_DETECTED,
        category=Category.OVERLAP,
        recipient_user_id=current_user.id,
        channel=Channel.DIGEST,
        delivery_status=DeliveryStatus.SKIPPED,
        suppression_reason=SuppressionReason.DIGEST_ONLY,
        payload_json={"source": "admin_test", "label": "TEST_SKIPPED"},
    )

    is_dup = is_duplicate_event(
        EventName.OVERLAP_DETECTED,
        current_user.id,
    )
    current_app.logger.info(
        "[MessageEvent-Test] dedupe check for overlap.detected → is_duplicate=%s",
        is_dup,
    )

    create_message_event(
        event_name=EventName.OVERLAP_DETECTED,
        category=Category.OVERLAP,
        recipient_user_id=current_user.id,
        channel=Channel.DIGEST,
        delivery_status=DeliveryStatus.SKIPPED,
        suppression_reason=SuppressionReason.DUPLICATE_EVENT,
        payload_json={"source": "admin_test", "label": "TEST_DEDUPE", "dedupe_fired": is_dup},
    )

    return redirect(url_for("admin_message_events"))


# ─── Phase D-1 Deploy B: Retry Runner ────────────────────────────────────────
#
# RETRY_EXECUTION_ENABLED gates the POST execute route.
#
# Set True ONLY after the Deploy B monitoring period confirms:
#   - provider_message_id populated on all new SENT rows
#   - sent_at and processed_at populated on all new rows
#   - Dry-run GET returns a stable, correct eligible set
#   - No unexpected retry_locked_at values anywhere in the table
#
# Changing this to True and restarting Flask is the ONLY code change needed
# to enable retry execution.
# ─────────────────────────────────────────────────────────────────────────────
RETRY_EXECUTION_ENABLED = False

_RETRY_LOOKBACK_HOURS  = 72
_RETRY_STALE_LOCK_MINS = 15
_RETRY_BATCH_LIMIT     = 50   # hard limit — non-configurable per D-1 spec

# Suppression reasons that permanently exclude a row from retry eligibility.
# Mirrors RETRYABLE_STATUSES frozenset in messaging_constants.py.
_NON_RETRYABLE_SUPPRESSION = (
    SuppressionReason.USER_OPTED_OUT,
    SuppressionReason.DUPLICATE_EVENT,
    SuppressionReason.CHANNEL_UNAVAILABLE,
    SuppressionReason.MISSING_REQUIRED_PAYLOAD,
    SuppressionReason.RECIPIENT_INELIGIBLE,
)

# Shared eligibility SQL — used by both GET (dry-run) and POST (execute).
# Non-retryable suppression values are hardcoded string literals (avoids
# IN-clause parameter binding complexity with db.text()).
# Flat lineage: parent_mel_id IS NULL ensures only original FAILED rows are
# eligible; retry child rows are never themselves retried.
_RETRY_ELIGIBILITY_SQL = """
    SELECT mel.id,
           mel.event_name,
           mel.recipient_user_id,
           mel.retry_count,
           mel.created_at,
           mel.payload_json,
           mel.message_title,
           mel.message_body,
           mel.object_type,
           mel.object_id,
           mel.actor_user_id,
           mel.category,
           mel.channel
    FROM message_event_log mel
    WHERE mel.delivery_status   = 'failed'
      AND mel.parent_mel_id     IS NULL
      AND mel.retry_count       < :max_retry_count
      AND (
          mel.suppression_reason IS NULL
          OR mel.suppression_reason NOT IN (
              'user_opted_out', 'duplicate_event', 'channel_unavailable',
              'missing_required_payload', 'recipient_ineligible'
          )
      )
      AND (
          mel.retry_locked_at IS NULL
          OR mel.retry_locked_at < :stale_cutoff
      )
      AND NOT EXISTS (
          SELECT 1 FROM message_event_log child
          WHERE child.parent_mel_id  = mel.id
            AND child.delivery_status = 'sent'
      )
      AND mel.created_at > :lookback_cutoff
    ORDER BY mel.created_at ASC
    LIMIT :batch_limit
"""


@app.route("/admin/retry-failed-events", methods=["GET", "POST"])
@login_required
@admin_required
def admin_retry_failed_events():
    """Admin retry runner — dry-run GET, execute POST (execute ships disabled).

    GET (always enabled — no sends, no mutations):
        Returns a JSON inspection report of retryable rows.
        Response fields:
            execute_enabled     current value of RETRY_EXECUTION_ENABLED
            eligible_count      rows that would be processed on next execute
            locked_count        rows currently claimed (retry_locked_at IS NOT NULL)
            stale_locked_count  locked rows overdue > 15 min (auto-eligible next run)
            eligible_rows       list of {id, event_name, recipient_user_id,
                                         retry_count, created_at}

    POST (execute — RETRY_EXECUTION_ENABLED must be True):
        When disabled: HTTP 200, {"status": "disabled", "message": "..."}.
        When enabled:  processes up to 50 eligible rows using strict ordering.

    Retry execution ordering:
        STEP 1  Atomic claim: UPDATE retry_locked_at=NOW() WHERE NULL.
                If 0 rows updated, a concurrent runner already claimed it — skip.
        STEP 2  Increment retry_count + COMMIT before any external call.
        STEP 3  Re-check push_notifications_enabled live from DB.
                If opted out: write SKIPPED child, proceed to unlock.
        STEP 4  Execute send_onesignal_push().
        STEP 5  Write retry child MEL row with parent_mel_id, provider_message_id,
                sent_at, processed_at (internal), source_route='admin_retry_runner'.
        STEP 6  Unlock (retry_locked_at = NULL) ONLY after child MEL commit.
                If child write fails, do NOT unlock — stale-lock timeout recovers.

    Flat lineage: only rows with parent_mel_id IS NULL are eligible.
    All child rows point to the original FAILED row — never to another child.
    """
    _now             = datetime.utcnow()
    _stale_cutoff    = _now - timedelta(minutes=_RETRY_STALE_LOCK_MINS)
    _lookback_cutoff = _now - timedelta(hours=_RETRY_LOOKBACK_HOURS)
    _params = {
        "max_retry_count": MAX_RETRY_COUNT,
        "stale_cutoff":    _stale_cutoff,
        "lookback_cutoff": _lookback_cutoff,
        "batch_limit":     _RETRY_BATCH_LIMIT,
    }

    # ── GET — dry-run (no sends, no mutations) ─────────────────────────────
    if request.method == "GET":
        try:
            eligible_rows = db.session.execute(
                db.text(_RETRY_ELIGIBILITY_SQL), _params
            ).fetchall()

            locked_count = db.session.execute(db.text(
                "SELECT COUNT(*) FROM message_event_log "
                "WHERE retry_locked_at IS NOT NULL"
            )).scalar() or 0

            stale_locked_count = db.session.execute(db.text(
                "SELECT COUNT(*) FROM message_event_log "
                "WHERE retry_locked_at IS NOT NULL AND retry_locked_at < :cutoff"
            ), {"cutoff": _stale_cutoff}).scalar() or 0

        except Exception as _qe:
            current_app.logger.exception("[RetryRunner] dry-run query failed: %s", _qe)
            return jsonify({"error": f"query failed: {_qe}"}), 500

        return jsonify({
            "execute_enabled":    RETRY_EXECUTION_ENABLED,
            "eligible_count":     len(eligible_rows),
            "locked_count":       int(locked_count),
            "stale_locked_count": int(stale_locked_count),
            "eligible_rows": [
                {
                    "id":                row[0],
                    "event_name":        row[1],
                    "recipient_user_id": row[2],
                    "retry_count":       row[3],
                    "created_at":        row[4].isoformat() if row[4] else None,
                }
                for row in eligible_rows
            ],
        })

    # ── POST — execute ─────────────────────────────────────────────────────
    if not RETRY_EXECUTION_ENABLED:
        return jsonify({
            "status":  "disabled",
            "message": "retry execution not yet enabled — use dry-run GET",
        }), 200

    current_app.logger.warning(
        "[RetryRunner] execute triggered by admin user_id=%d email=%s",
        current_user.id, current_user.email,
    )

    try:
        eligible_rows = db.session.execute(
            db.text(_RETRY_ELIGIBILITY_SQL), _params
        ).fetchall()
    except Exception as _qe:
        current_app.logger.exception("[RetryRunner] execute eligibility query failed: %s", _qe)
        return jsonify({"error": f"eligibility query failed: {_qe}"}), 500

    results = {
        "attempted":          0,
        "sent":               0,
        "skipped_opted_out":  0,
        "failed":             0,
        "skipped_concurrent": 0,
    }

    for row in eligible_rows:
        row_id       = row[0]
        _recipient_id = row[2]

        # STEP 1 — Atomic claim: SET retry_locked_at=NOW() WHERE NULL.
        # rowcount=0 means a concurrent runner claimed this row — skip it.
        try:
            _claim = db.session.execute(db.text(
                "UPDATE message_event_log "
                "SET retry_locked_at = NOW() "
                "WHERE id = :row_id AND retry_locked_at IS NULL"
            ), {"row_id": row_id})
            db.session.commit()
        except Exception as _ce:
            db.session.rollback()
            current_app.logger.warning(
                "[RetryRunner] claim failed row_id=%d: %s", row_id, _ce
            )
            results["skipped_concurrent"] += 1
            continue

        if _claim.rowcount == 0:
            current_app.logger.warning(
                "[RetryRunner] row_id=%d already claimed — skipping", row_id
            )
            results["skipped_concurrent"] += 1
            continue

        # STEP 2 — Increment retry_count + COMMIT before any external call.
        # Burning the attempt slot first prevents silent double-sends if the
        # process is killed mid-flight.
        try:
            db.session.execute(db.text(
                "UPDATE message_event_log "
                "SET retry_count = retry_count + 1 "
                "WHERE id = :row_id"
            ), {"row_id": row_id})
            db.session.commit()
        except Exception as _ie:
            db.session.rollback()
            current_app.logger.warning(
                "[RetryRunner] retry_count increment failed row_id=%d: %s", row_id, _ie
            )
            results["failed"] += 1
            continue

        results["attempted"] += 1

        # STEP 3 — Re-check push_notifications_enabled live from DB.
        try:
            _opt_row  = db.session.execute(db.text(
                'SELECT push_notifications_enabled FROM "user" WHERE id = :uid'
            ), {"uid": _recipient_id}).fetchone()
            _opted_in = bool(_opt_row[0]) if _opt_row else False
        except Exception as _oe:
            current_app.logger.warning(
                "[RetryRunner] opt-out recheck failed row_id=%d: %s", row_id, _oe
            )
            _opted_in = False

        _child_status      = None
        _child_suppression = None
        _child_error       = None
        _child_prov_msg_id = None
        _child_sent_at     = None

        if not _opted_in:
            # Recipient opted out since the original failure — write SKIPPED child.
            _child_status      = DeliveryStatus.SKIPPED
            _child_suppression = SuppressionReason.USER_OPTED_OUT
            results["skipped_opted_out"] += 1
        else:
            # STEP 4 — Execute send_onesignal_push().
            _title   = row[6]   # message_title
            _body    = row[7]   # message_body
            _payload = dict(row[5] or {})
            _payload["source_route"] = "admin_retry_runner"

            try:
                _send_result = send_onesignal_push(
                    user_ids=[_recipient_id],
                    title=_title or "",
                    body=_body or "",
                    data=_payload or None,
                )
            except Exception as _se:
                current_app.logger.warning(
                    "[RetryRunner] send raised row_id=%d: %s", row_id, _se
                )
                _send_result = {
                    "success": False, "provider_message_id": None,
                    "skipped": False, "error": f"send_raised: {_se}",
                }

            _send_skipped = _send_result.get("skipped", False)
            _send_success = bool(_send_result.get("success")) and not _send_skipped

            if _send_success:
                _child_status      = DeliveryStatus.SENT
                _child_prov_msg_id = _send_result.get("provider_message_id")
                _child_sent_at     = datetime.utcnow()
                results["sent"] += 1
            elif _send_skipped:
                _child_status      = DeliveryStatus.SKIPPED
                _child_suppression = SuppressionReason.USER_OPTED_OUT
                results["skipped_opted_out"] += 1
            else:
                _child_status = DeliveryStatus.FAILED
                _child_error  = _send_result.get("error") or "unknown_onesignal_error"
                results["failed"] += 1

        # STEP 5 — Write retry child MEL row.
        # parent_mel_id always points to the original FAILED row (flat lineage).
        # processed_at is set internally by create_message_event().
        _child_committed = False
        try:
            create_message_event(
                event_name=row[1],
                category=row[11],
                actor_user_id=row[10],
                recipient_user_id=_recipient_id,
                object_type=row[8],
                object_id=row[9],
                channel=row[12] or Channel.PUSH,
                provider=Provider.ONESIGNAL,
                payload_json={**(row[5] or {}), "source_route": "admin_retry_runner"},
                message_title=row[6],
                message_body=row[7],
                delivery_status=_child_status,
                suppression_reason=_child_suppression,
                error_message=_child_error,
                provider_message_id=_child_prov_msg_id,
                sent_at=_child_sent_at,
                parent_mel_id=row_id,
            )
            _child_committed = True
        except Exception as _mel_err:
            current_app.logger.warning(
                "[RetryRunner] child MEL write failed row_id=%d: %s", row_id, _mel_err
            )

        # STEP 6 — Unlock ONLY after successful child MEL commit.
        # If the child write failed, leave retry_locked_at set — the 15-minute
        # stale-lock timeout will make the row eligible again automatically.
        if _child_committed:
            try:
                db.session.execute(db.text(
                    "UPDATE message_event_log "
                    "SET retry_locked_at = NULL "
                    "WHERE id = :row_id"
                ), {"row_id": row_id})
                db.session.commit()
            except Exception as _ue:
                db.session.rollback()
                current_app.logger.warning(
                    "[RetryRunner] unlock failed row_id=%d "
                    "(stale-lock timeout will recover): %s",
                    row_id, _ue,
                )

    current_app.logger.warning("[RetryRunner] execute complete: %s", results)
    return jsonify({"status": "ok", **results})


# ============================================================================
# ADMIN CONSOLE — V1
# ============================================================================

@app.route("/admin")
@login_required
@admin_required
def admin_console():
    """Entry point — redirect to dashboard tab."""
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/dashboard")
@login_required
@admin_required
def admin_dashboard():
    """V1 Admin Dashboard — operational KPIs from existing models, no new tables."""
    from datetime import timedelta

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _trend(current, prior, unit="pct", invert=False, label="prior period",
               min_prior=1):
        """Return (display_text, css_class) or None if insufficient data.
        unit: 'pct' for % change, 'pts' for percentage-point difference.
        invert: True when higher is worse (e.g. failure counts).
        """
        if prior is None or current is None:
            return None
        if unit == "pct":
            if prior < min_prior:
                return None
            pct = round((current - prior) / prior * 100)
            if pct == 0:
                return f"flat vs {label}", "ad-trend--flat"
            text = f"{'+' if pct > 0 else ''}{pct}% vs {label}"
            going_up = pct > 0
        else:  # pts
            pts = round(current - prior)
            if pts == 0:
                return f"flat vs {label}", "ad-trend--flat"
            text = f"{'+' if pts > 0 else ''}{pts} pts vs {label}"
            going_up = pts > 0
        good = (going_up and not invert) or (not going_up and invert)
        return text, ("ad-trend--pos" if good else "ad-trend--neg")

    # ── Base date ranges ──────────────────────────────────────────────────────
    now          = datetime.utcnow()
    thirty_ago   = now - timedelta(days=30)
    sixty_ago    = now - timedelta(days=60)
    first_of_mo  = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_label  = now.strftime("%B %Y")

    # Prior month — same calendar-day window ("MTD" parity)
    days_elapsed   = now.day               # days into current month (1-indexed)
    prior_mo       = (now.month - 2) % 12 + 1
    prior_yr       = now.year if now.month > 1 else now.year - 1
    first_of_prior = now.replace(
        year=prior_yr, month=prior_mo, day=1,
        hour=0, minute=0, second=0, microsecond=0,
    )
    prior_mtd_end  = first_of_prior + timedelta(days=days_elapsed)

    # ── User counts ──────────────────────────────────────────────────────────
    seven_ago       = now - timedelta(days=7)
    total_users     = User.query.count()
    wau             = User.query.filter(User.last_active_at >= seven_ago).count()
    mau             = User.query.filter(User.last_active_at >= thirty_ago).count()
    new_users_month = User.query.filter(User.created_at >= first_of_mo).count()

    # ── Trip counts (core) ───────────────────────────────────────────────────
    today       = now.date()
    total_trips = SkiTrip.query.count()
    trips_week  = SkiTrip.query.filter(SkiTrip.created_at >= seven_ago).count()
    trips_month = SkiTrip.query.filter(SkiTrip.created_at >= first_of_mo).count()
    past_trips  = SkiTrip.query.filter(SkiTrip.end_date   <  today).count()

    # ── Trips — rolling 30d (distinct from MTD) ──────────────────────────────
    trips_30d = SkiTrip.query.filter(SkiTrip.created_at >= thirty_ago).count()

    # ── Mountains ─────────────────────────────────────────────────────────────
    total_active_resorts = Resort.query.filter_by(is_active=True).count()
    mtn_page_views = db.session.query(MountainPageView.id).count()

    # ── Social — avg accepted friend connections per user ─────────────────────
    avg_friends_per_user = round(Friend.query.count() / total_users, 1) if total_users else 0.0

    # Pass coverage — active resorts with at least one real (non-None) pass brand
    _pbj_rows = db.session.query(Resort.pass_brands_json).filter(Resort.is_active == True).all()
    _resorts_with_pass = 0
    for (pbj,) in _pbj_rows:
        if pbj and isinstance(pbj, list):
            real_brands = [b for b in pbj if b and b.lower() not in ('none', '')]
            if real_brands:
                _resorts_with_pass += 1
    mtn_pass_coverage_pct = round(_resorts_with_pass / total_active_resorts * 100) if total_active_resorts else 0

    # ── Opt-in rates ─────────────────────────────────────────────────────────
    push_opt_in  = User.query.filter_by(push_notifications_enabled=True).count()
    email_opt_in = User.query.filter_by(email_opt_in=True).count()
    push_rate    = round(push_opt_in  / total_users * 100) if total_users else 0
    email_rate   = round(email_opt_in / total_users * 100) if total_users else 0

    # ── Platform mix (push-token proxy) ──────────────────────────────────────
    # PushDeviceToken.platform is 'ios' or 'android'; users without any active
    # token are likely web-only or haven't granted push permission.
    push_ios     = PushDeviceToken.query.filter_by(active=True, platform='ios').count()
    push_android = PushDeviceToken.query.filter_by(active=True, platform='android').count()
    _users_with_push = db.session.query(
        PushDeviceToken.user_id.distinct()
    ).filter_by(active=True).count()
    push_web_proxy = max(total_users - _users_with_push, 0)

    # ── Pass on file — single headline KPI ────────────────────────────────────
    _no_pass_slugs = frozenset({"no_pass", "no_pass_yet"})
    pass_on_file_count = 0
    for (pt,) in db.session.query(User.pass_type).all():
        if not pt:
            continue
        for raw in str(pt).split(","):
            raw = raw.strip()
            if not raw:
                continue
            norm = normalize_pass(raw)
            if norm and norm not in _no_pass_slugs:
                pass_on_file_count += 1
                break
    pass_on_file_pct = round(pass_on_file_count / total_users * 100) if total_users else 0


    # ── New L30 (rolling, not MTD) ───────────────────────────────────────────
    new_users_l30 = User.query.filter(User.created_at >= thirty_ago).count()

    # ── Connected user rate ───────────────────────────────────────────────────
    _connected_n        = db.session.query(Friend.user_id.distinct()).count()
    connected_user_rate = round(_connected_n / total_users * 100) if total_users else 0

    # ── Reachable users (alias of push-token distinct already computed above) ─
    reachable_users = _users_with_push

    # ── Friend connections (unique pairs = bidirectional rows / 2) ────────────
    friend_connections = Friend.query.count() // 2

    # ── Push failure rate from MEL ────────────────────────────────────────────
    _mel_push = db.session.execute(db.text(
        "SELECT "
        "  COUNT(*) FILTER (WHERE delivery_status='sent'),  "
        "  COUNT(*) FILTER (WHERE delivery_status='failed') "
        "FROM message_event_log WHERE channel='push'"
    )).fetchone()
    _mel_sent        = _mel_push[0] or 0
    _mel_failed      = _mel_push[1] or 0
    push_fail_rate   = round(_mel_failed / (_mel_sent + _mel_failed) * 100) \
                       if (_mel_sent + _mel_failed) else 0
    push_fail_attempted = _mel_sent + _mel_failed
    push_fail_failed    = _mel_failed

    # ── Pass without availability (for Insight 2) ─────────────────────────────
    import json as _json_dash
    _no_pass_d = {'no_pass', 'no_pass_yet', 'none', ''}

    def _dash_has_pass(pt):
        for s in (pt or '').split(','):
            if s.strip().lower() not in _no_pass_d:
                return True
        return False

    def _dash_has_avail(od):
        if od is None:
            return False
        if isinstance(od, list):
            return len(od) > 0
        if isinstance(od, str):
            try:
                p = _json_dash.loads(od)
                return isinstance(p, list) and len(p) > 0
            except Exception:
                return False
        return False

    pass_no_avail_count = sum(
        1 for (pt, od) in db.session.execute(
            db.text('SELECT pass_type, open_dates FROM "user"')
        ).fetchall()
        if _dash_has_pass(pt) and not _dash_has_avail(od)
    )

    # ── Active users reachable by push (for Insight 3) ────────────────────────
    _active_ids_dash = {r[0] for r in db.session.execute(
        db.text("SELECT id FROM \"user\" WHERE lifecycle_stage = 'active'")
    ).fetchall()}
    _push_ids_dash = {r[0] for r in db.session.execute(
        db.text("SELECT DISTINCT user_id FROM push_device_token WHERE active = true")
    ).fetchall()}
    active_users_reachable = len(_active_ids_dash & _push_ids_dash)
    active_users_total     = len(_active_ids_dash)

    # ── 14-day daily series (for sparklines) ─────────────────────────────────
    from sqlalchemy import func as _func

    _fourteen_ago = now - timedelta(days=14)

    # ── Prior-period deltas for dashboard KPI tiles ───────────────────────────
    prior_wau = User.query.filter(
        User.last_active_at >= _fourteen_ago,
        User.last_active_at <  seven_ago,
    ).count()
    trend_wau = _trend(wau, prior_wau, label="prior L7")

    prior_trips_week = SkiTrip.query.filter(
        SkiTrip.created_at >= _fourteen_ago,
        SkiTrip.created_at <  seven_ago,
    ).count()
    trend_trips_week = _trend(trips_week, prior_trips_week, label="prior L7")

    prior_trips_30d_val = SkiTrip.query.filter(
        SkiTrip.created_at >= sixty_ago,
        SkiTrip.created_at <  thirty_ago,
    ).count()
    trend_trips_30d = _trend(trips_30d, prior_trips_30d_val, label="prior L30")

    mtn_views_30d = db.session.query(MountainPageView.id).filter(
        MountainPageView.viewed_at >= thirty_ago,
    ).count()
    prior_mtn_views_30d = db.session.query(MountainPageView.id).filter(
        MountainPageView.viewed_at >= sixty_ago,
        MountainPageView.viewed_at <  thirty_ago,
    ).count()
    trend_mtn_views = _trend(mtn_views_30d, prior_mtn_views_30d, label="prior L30")
    # Build an ordered date spine: oldest first → today
    _date_spine = [
        (now - timedelta(days=i)).strftime('%Y-%m-%d')
        for i in range(13, -1, -1)
    ]

    # Cumulative total users: for each day, count users created on or before that day
    _total_users_series = []
    for _d in _date_spine:
        _cutoff = _d + " 23:59:59"
        _count  = User.query.filter(User.created_at <= _cutoff).count()
        _total_users_series.append(_count)
    total_users_series = _total_users_series

    # Active users: users whose last_active_at falls on each day
    _mau_rows = (
        db.session.query(
            _func.date(User.last_active_at).label('d'),
            _func.count(User.id).label('n'),
        )
        .filter(User.last_active_at >= _fourteen_ago)
        .group_by(_func.date(User.last_active_at))
        .all()
    )
    _mau_map   = {str(r.d): r.n for r in _mau_rows}
    mau_series = [_mau_map.get(d, 0) for d in _date_spine]
    # 7-day slice — last 7 elements of the 14-day series, free
    wau_series = mau_series[-7:]

    # Trips created: trips where created_at falls on each day
    _trips_rows = (
        db.session.query(
            _func.date(SkiTrip.created_at).label('d'),
            _func.count(SkiTrip.id).label('n'),
        )
        .filter(SkiTrip.created_at >= _fourteen_ago)
        .group_by(_func.date(SkiTrip.created_at))
        .all()
    )
    _trips_map  = {str(r.d): r.n for r in _trips_rows}
    trips_series = [_trips_map.get(d, 0) for d in _date_spine]

    # New users: users where created_at falls on each day
    _new_rows = (
        db.session.query(
            _func.date(User.created_at).label('d'),
            _func.count(User.id).label('n'),
        )
        .filter(User.created_at >= _fourteen_ago)
        .group_by(_func.date(User.created_at))
        .all()
    )
    _new_map  = {str(r.d): r.n for r in _new_rows}
    new_users_series = [_new_map.get(d, 0) for d in _date_spine]

    # ── Trend calculations ────────────────────────────────────────────────────

    # 1. New Users This Month — MTD vs prior MTD
    prior_new_users = User.query.filter(
        User.created_at >= first_of_prior,
        User.created_at <  prior_mtd_end,
    ).count()

    # 2. MAU — rolling 30d vs prior 30d
    prior_mau = User.query.filter(
        User.last_active_at >= sixty_ago,
        User.last_active_at <  thirty_ago,
    ).count()
    trend_mau = _trend(mau, prior_mau, unit="pct", label="prior 30d")

    # 3. New Users This Month — MTD vs prior MTD (same window as #1)
    trend_new_users = _trend(
        new_users_month, prior_new_users,
        unit="pct", label="prior MTD",
    )

    # 4. Trips This Month — MTD vs prior MTD
    prior_trips = SkiTrip.query.filter(
        SkiTrip.created_at >= first_of_prior,
        SkiTrip.created_at <  prior_mtd_end,
    ).count()
    trend_trips_month = _trend(
        trips_month, prior_trips,
        unit="pct", label="prior MTD",
    )

    # ── Hub section computations (command center additions) ───────────────
    from collections import Counter as _Counter
    _ns_users = User.query.filter(User.is_seeded == False).all()
    _ns_ids   = [u.id for u in _ns_users]
    _ns_total = len(_ns_ids) or 1

    # activation_pct — non-seeded users who have activated
    _act_n         = sum(1 for u in _ns_users if u.is_active_user)
    activation_pct = round(_act_n / _ns_total * 100)

    # invite_redemption_pct — InviteToken redeemed / generated (non-seeded)
    _it_gen = db.session.query(InviteToken.id).filter(
        InviteToken.inviter_id.in_(_ns_ids)).count()
    _it_red = db.session.query(InviteToken.id).filter(
        InviteToken.inviter_id.in_(_ns_ids),
        InviteToken.used_at != None).count()
    invite_redemption_pct = round(_it_red / _it_gen * 100) if _it_gen else 0

    # top_state / top_state_pct — most common home state among non-seeded
    _state_ctr   = _Counter(u.home_state for u in _ns_users if u.home_state)
    _state_total = sum(_state_ctr.values())
    top_state    = _state_ctr.most_common(1)[0][0] if _state_ctr else None
    top_state_pct = round(_state_ctr[top_state] / _state_total * 100) \
                    if (top_state and _state_total) else 0

    # avg_product_score — identical formula to Product Dashboard
    _no_pass_s = {'no_pass', 'no_pass_yet', 'none', ''}
    def _hub_has_pass(u):
        pt = u.pass_type or ''
        return any(r.strip().lower() not in _no_pass_s
                   for r in str(pt).split(',') if r.strip())
    def _hub_has_avail(u): return isinstance(u.open_dates, list) and len(u.open_dates) > 0
    def _hub_has_wl(u):    return isinstance(u.wish_list_resorts, list) and len(u.wish_list_resorts) > 0
    _hub_equip = set(r[0] for r in db.session.query(EquipmentSetup.user_id.distinct())
                     .filter(EquipmentSetup.user_id.in_(_ns_ids)).all())
    _hub_fr    = set(r[0] for r in db.session.query(Friend.user_id.distinct())
                     .filter(Friend.user_id.in_(_ns_ids)).all())
    _hub_tc    = {r[0]: r[1] for r in db.session.query(
                     SkiTrip.user_id, db.func.count(SkiTrip.id))
                  .filter(SkiTrip.user_id.in_(_ns_ids))
                  .group_by(SkiTrip.user_id).all()}
    _hub_gt    = set(r[0] for r in db.session.query(SkiTrip.user_id.distinct())
                     .filter(SkiTrip.user_id.in_(_ns_ids),
                             SkiTrip.is_group_trip == True).all())
    def _hub_score(u):
        s, tc = 0, _hub_tc.get(u.id, 0)
        if _hub_has_pass(u):   s += 1
        if _hub_has_avail(u):  s += 1
        if _hub_has_wl(u):     s += 1
        if u.id in _hub_equip: s += 1
        if u.id in _hub_fr:    s += 1
        s += tc * 2
        if u.id in _hub_gt:    s += 3
        if tc >= 2:            s += 1
        return s
    _hub_scores       = [_hub_score(u) for u in _ns_users]
    avg_product_score = round(sum(_hub_scores) / len(_hub_scores), 1) if _hub_scores else 0

    # ── Founder Pulse — today vs yesterday ───────────────────────────────────
    _pulse_today_start = _admin_today_start_utc()
    _pulse_yest_start  = _admin_yesterday_start_utc()
    _pulse_yest_end    = _pulse_today_start

    def _pulse_delta(today_val, yest_val):
        """Return (text, css_class) for an absolute today-vs-yesterday delta."""
        if yest_val is None:
            return "—", "fp-trend--flat"
        diff = today_val - yest_val
        if diff == 0:
            return "Flat vs yesterday", "fp-trend--flat"
        sign = "+" if diff > 0 else ""
        cls  = "fp-trend--pos" if diff > 0 else "fp-trend--flat"
        return f"{sign}{diff} vs yesterday", cls

    # 1. Active Today — users who touched the app today
    p_active_today = User.query.filter(User.last_active_at >= _pulse_today_start).count()
    p_active_yest  = User.query.filter(
        User.last_active_at >= _pulse_yest_start,
        User.last_active_at <  _pulse_yest_end,
    ).count()

    # 2. Sessions Today — Activity rows as proxy for total in-app events
    p_sessions_today = Activity.query.filter(Activity.created_at >= _pulse_today_start).count()
    p_sessions_yest  = Activity.query.filter(
        Activity.created_at >= _pulse_yest_start,
        Activity.created_at <  _pulse_yest_end,
    ).count()

    # 3. New Users Today
    p_new_users_today = User.query.filter(User.created_at >= _pulse_today_start).count()
    p_new_users_yest  = User.query.filter(
        User.created_at >= _pulse_yest_start,
        User.created_at <  _pulse_yest_end,
    ).count()

    # 4. Trips Created Today
    p_trips_today = SkiTrip.query.filter(SkiTrip.created_at >= _pulse_today_start).count()
    p_trips_yest  = SkiTrip.query.filter(
        SkiTrip.created_at >= _pulse_yest_start,
        SkiTrip.created_at <  _pulse_yest_end,
    ).count()

    # 5. Mountain Views Today
    p_mtn_views_today = MountainPageView.query.filter(
        MountainPageView.viewed_at >= _pulse_today_start
    ).count()
    p_mtn_views_yest  = MountainPageView.query.filter(
        MountainPageView.viewed_at >= _pulse_yest_start,
        MountainPageView.viewed_at <  _pulse_yest_end,
    ).count()

    # 6. Friend Connections Today
    # Friend table is bidirectional (A→B + B→A rows per pair); divide by 2.
    p_friends_today = Friend.query.filter(Friend.created_at >= _pulse_today_start).count() // 2
    p_friends_yest  = Friend.query.filter(
        Friend.created_at >= _pulse_yest_start,
        Friend.created_at <  _pulse_yest_end,
    ).count() // 2

    # 7a. In-App Invites Today — friend Invitations + SkiTripParticipant INVITED rows
    #     (structured invites between users already in the system)
    p_finv_today = Invitation.query.filter(
        Invitation.created_at >= _pulse_today_start
    ).count()
    p_finv_yest  = Invitation.query.filter(
        Invitation.created_at >= _pulse_yest_start,
        Invitation.created_at <  _pulse_yest_end,
    ).count()
    p_tinv_today = SkiTripParticipant.query.filter(
        SkiTripParticipant.created_at >= _pulse_today_start,
        SkiTripParticipant.status == GuestStatus.INVITED,
    ).count()
    p_tinv_yest  = SkiTripParticipant.query.filter(
        SkiTripParticipant.created_at >= _pulse_yest_start,
        SkiTripParticipant.created_at <  _pulse_yest_end,
        SkiTripParticipant.status == GuestStatus.INVITED,
    ).count()
    p_in_app_invites_today = p_finv_today + p_tinv_today
    p_in_app_invites_yest  = p_finv_yest  + p_tinv_yest

    # 7b. Invite Shares Today — InviteShareEvent rows (copy/text/share_sheet intent)
    #     Counts outbound share actions to non-users; independent of 7a.
    p_invite_share_today = InviteShareEvent.query.filter(
        InviteShareEvent.created_at >= _pulse_today_start,
    ).count()
    p_invite_share_yest  = InviteShareEvent.query.filter(
        InviteShareEvent.created_at >= _pulse_yest_start,
        InviteShareEvent.created_at <  _pulse_yest_end,
    ).count()

    # 8. Pushes Delivered Today — channel='push', delivery_status='sent'
    # sent_at is nullable; fall back to created_at for rows that lack it.
    p_pushes_today = MessageEventLog.query.filter(
        MessageEventLog.channel == 'push',
        MessageEventLog.delivery_status == 'sent',
        db.or_(
            MessageEventLog.sent_at >= _pulse_today_start,
            db.and_(
                MessageEventLog.sent_at == None,
                MessageEventLog.created_at >= _pulse_today_start,
            ),
        ),
    ).count()
    p_pushes_yest = MessageEventLog.query.filter(
        MessageEventLog.channel == 'push',
        MessageEventLog.delivery_status == 'sent',
        db.or_(
            db.and_(MessageEventLog.sent_at >= _pulse_yest_start,
                    MessageEventLog.sent_at <  _pulse_yest_end),
            db.and_(MessageEventLog.sent_at == None,
                    MessageEventLog.created_at >= _pulse_yest_start,
                    MessageEventLog.created_at <  _pulse_yest_end),
        ),
    ).count()

    # Assemble pulse dict — each key: (today_value, (delta_text, delta_css_class))
    pulse = dict(
        active    = (p_active_today,    _pulse_delta(p_active_today,    p_active_yest)),
        sessions  = (p_sessions_today,  _pulse_delta(p_sessions_today,  p_sessions_yest)),
        new_users = (p_new_users_today, _pulse_delta(p_new_users_today, p_new_users_yest)),
        trips     = (p_trips_today,     _pulse_delta(p_trips_today,     p_trips_yest)),
        mtn_views = (p_mtn_views_today, _pulse_delta(p_mtn_views_today, p_mtn_views_yest)),
        friends   = (p_friends_today,   _pulse_delta(p_friends_today,   p_friends_yest)),
        in_app_invites = (p_in_app_invites_today, _pulse_delta(p_in_app_invites_today, p_in_app_invites_yest)),
        invite_share   = (p_invite_share_today,   _pulse_delta(p_invite_share_today,   p_invite_share_yest)),
        pushes         = (p_pushes_today,          _pulse_delta(p_pushes_today,         p_pushes_yest)),
    )
    _dn = _admin_now()
    pulse_time = _dn.strftime("%H:%M") + " " + _dn.strftime("%Z")

    # ── App Store summary tile (reads from DB only — no live API calls) ──────
    try:
        from models import AppStoreMetric
        from datetime import date as _date, timedelta as _td
        _today     = _date.today()
        _yesterday = _today - _td(days=1)
        _l30_start = _today - _td(days=30)

        def _asm_dl_yesterday(platform):
            row = AppStoreMetric.query.filter_by(
                platform=platform, report_date=_yesterday
            ).first()
            return row.downloads if row and row.downloads is not None else None

        def _asm_dl_l30(platform):
            rows = AppStoreMetric.query.filter(
                AppStoreMetric.platform    == platform,
                AppStoreMetric.report_date >= _l30_start,
                AppStoreMetric.downloads   != None,
            ).all()
            return sum(r.downloads for r in rows) if rows else None

        def _asm_rating(platform):
            row = AppStoreMetric.query.filter(
                AppStoreMetric.platform == platform,
                AppStoreMetric.rating   != None,
            ).order_by(AppStoreMetric.report_date.desc()).first()
            return row.rating if row else None

        store_summary = dict(
            ios_dl_yesterday  = _asm_dl_yesterday("ios"),
            android_dl_yesterday = _asm_dl_yesterday("android"),
            ios_dl_l30        = _asm_dl_l30("ios"),
            android_dl_l30    = _asm_dl_l30("android"),
            ios_rating        = _asm_rating("ios"),
            android_rating    = _asm_rating("android"),
            has_data          = AppStoreMetric.query.count() > 0,
        )
    except Exception:
        store_summary = dict(
            ios_dl_yesterday=None, android_dl_yesterday=None,
            ios_dl_l30=None, android_dl_l30=None,
            ios_rating=None, android_rating=None,
            has_data=False,
        )

    return render_template(
        "admin_dashboard.html",
        active_tab         = "dashboard",
        now                = _fmt_admin_now(),
        month_label        = month_label,
        total_users        = total_users,
        mau                = mau,
        new_users_month    = new_users_month,
        new_users_l30      = new_users_l30,
        total_trips        = total_trips,
        trips_week         = trips_week,
        trips_month        = trips_month,
        past_trips         = past_trips,
        push_opt_in        = push_opt_in,
        email_opt_in       = email_opt_in,
        push_rate          = push_rate,
        email_rate         = email_rate,
        pass_on_file_count = pass_on_file_count,
        pass_on_file_pct   = pass_on_file_pct,
        total_active_resorts  = total_active_resorts,
        mtn_pass_coverage_pct = mtn_pass_coverage_pct,
        trend_new_users    = trend_new_users,
        trend_mau          = trend_mau,
        trend_trips_month  = trend_trips_month,
        wau                = wau,
        wau_series         = wau_series,
        total_users_series = total_users_series,
        mau_series         = mau_series,
        trips_series       = trips_series,
        new_users_series   = new_users_series,
        push_ios             = push_ios,
        push_android         = push_android,
        push_web_proxy       = push_web_proxy,
        trips_30d            = trips_30d,
        mtn_page_views       = mtn_page_views,
        avg_friends_per_user = avg_friends_per_user,
        trend_wau            = trend_wau,
        trend_trips_week     = trend_trips_week,
        trend_trips_30d      = trend_trips_30d,
        trend_mtn_views      = trend_mtn_views,
        mtn_views_30d        = mtn_views_30d,
        activation_pct       = activation_pct,
        invite_redemption_pct = invite_redemption_pct,
        top_state            = top_state,
        top_state_pct        = top_state_pct,
        avg_product_score    = avg_product_score,
        # ── Executive Dashboard additions ──────────────────────────────
        connected_user_rate    = connected_user_rate,
        reachable_users        = reachable_users,
        friend_connections     = friend_connections,
        push_fail_rate         = push_fail_rate,
        push_fail_attempted    = push_fail_attempted,
        push_fail_failed       = push_fail_failed,
        pass_no_avail_count    = pass_no_avail_count,
        active_users_reachable = active_users_reachable,
        active_users_total     = active_users_total,
        pulse                  = pulse,
        pulse_time             = pulse_time,
        store_summary          = store_summary,
    )


@app.route("/admin/user-insights")
@login_required
@admin_required
def admin_user_insights():
    """Admin User Insights — read-only analytics on user composition and equipment."""
    from collections import Counter
    from sqlalchemy import func
    from services.pass_utils import normalize_pass, PASS_DISPLAY_MAP, CANONICAL_PASS_ORDER

    all_users = User.query.all()
    total = len(all_users)

    def pct(n):
        return round(n / total * 100) if total else 0

    # ── Overview KPIs ─────────────────────────────────────────────────────────
    users_with_rider = sum(
        1 for u in all_users
        if (u.rider_types and len(u.rider_types) > 0)
        or u.primary_rider_type or u.rider_type
    )
    users_with_equipment = db.session.query(
        EquipmentSetup.user_id.distinct()
    ).count()
    users_with_wishlist = sum(
        1 for u in all_users
        if u.wish_list_resorts and len(u.wish_list_resorts) > 0
    )
    users_with_visited = sum(
        1 for u in all_users
        if u.visited_resort_ids and len(u.visited_resort_ids) > 0
    )
    # "has a pass" = has at least one canonical pass that isn't no_pass/no_pass_yet
    _real_pass = frozenset({"epic", "ikon", "other"})
    users_with_pass = 0
    for u in all_users:
        raw = u.pass_type or ''
        parts = [p.strip() for p in raw.split(',') if p.strip()]
        if any(normalize_pass(p) in _real_pass for p in parts):
            users_with_pass += 1

    overview = {
        'total':          total,
        'with_rider':     (users_with_rider,     pct(users_with_rider)),
        'with_equipment': (users_with_equipment, pct(users_with_equipment)),
        'with_wishlist':  (users_with_wishlist,  pct(users_with_wishlist)),
        'with_visited':   (users_with_visited,   pct(users_with_visited)),
        'with_pass':      (users_with_pass,       pct(users_with_pass)),
    }

    # ── Rider mix ─────────────────────────────────────────────────────────────
    rt_ctr = Counter()
    for u in all_users:
        rts = u.rider_types or []
        if not rts:
            if u.primary_rider_type:
                rts = [u.primary_rider_type]
            elif u.rider_type:
                rts = [u.rider_type]
        for rt in rts:
            if rt:
                rt_ctr[rt] += 1
    rider_mix = sorted(
        [(rt, cnt, pct(cnt)) for rt, cnt in rt_ctr.items()],
        key=lambda x: -x[1]
    )

    # ── Pass insights ─────────────────────────────────────────────────────────
    pass_ctr = Counter()
    users_with_any_pass = 0
    for u in all_users:
        raw = u.pass_type or ''
        parts = [p.strip() for p in raw.split(',') if p.strip()]
        has_real = False
        for p in parts:
            norm = normalize_pass(p)
            if norm:
                pass_ctr[norm] += 1
                if norm in _real_pass:
                    has_real = True
        if has_real:
            users_with_any_pass += 1

    pass_rows = [
        (slug, PASS_DISPLAY_MAP.get(slug, slug), pass_ctr.get(slug, 0), pct(pass_ctr.get(slug, 0)))
        for slug in CANONICAL_PASS_ORDER
    ]

    # ── Equipment completion ───────────────────────────────────────────────────
    total_setups = EquipmentSetup.query.count()
    avg_setups = round(total_setups / users_with_equipment, 1) if users_with_equipment else 0

    ski_eq = db.session.query(EquipmentSetup.user_id.distinct()).filter(
        EquipmentSetup.discipline == EquipmentDiscipline.SKIER
    ).count()
    sb_eq = db.session.query(EquipmentSetup.user_id.distinct()).filter(
        EquipmentSetup.discipline == EquipmentDiscipline.SNOWBOARDER
    ).count()
    with_boots = db.session.query(EquipmentSetup.user_id.distinct()).filter(
        EquipmentSetup.boot_brand.isnot(None),
        EquipmentSetup.boot_brand != ''
    ).count()
    with_bindings = db.session.query(EquipmentSetup.user_id.distinct()).filter(
        EquipmentSetup.binding_brand.isnot(None),
        EquipmentSetup.binding_brand != ''
    ).count()

    eq_summary = {
        'users_with_eq': users_with_equipment,
        'avg_setups':    avg_setups,
        'ski':           (ski_eq,        pct(ski_eq)),
        'snowboard':     (sb_eq,         pct(sb_eq)),
        'boots':         (with_boots,    pct(with_boots)),
        'bindings':      (with_bindings, pct(with_bindings)),
    }

    # ── Equipment brand tables ─────────────────────────────────────────────────
    def _brand_table(col, disc_filter=None):
        q = db.session.query(col, func.count()).filter(
            col.isnot(None), col != ''
        )
        if disc_filter is not None:
            q = q.filter(disc_filter)
        rows = q.group_by(col).order_by(func.count().desc()).all()
        eq_base = users_with_equipment or 1
        return [(b, cnt, round(cnt / eq_base * 100)) for b, cnt in rows if b]

    ski_brands      = _brand_table(EquipmentSetup.brand,         EquipmentSetup.discipline == EquipmentDiscipline.SKIER)
    sb_brands       = _brand_table(EquipmentSetup.brand,         EquipmentSetup.discipline == EquipmentDiscipline.SNOWBOARDER)
    boot_brands     = _brand_table(EquipmentSetup.boot_brand)
    binding_brands  = _brand_table(EquipmentSetup.binding_brand)

    return render_template(
        "admin_user_insights.html",
        active_tab          = "user_insights",
        now                 = _fmt_admin_now(),
        total               = total,
        overview            = overview,
        rider_mix           = rider_mix,
        users_with_any_pass = users_with_any_pass,
        pass_rows           = pass_rows,
        eq_summary          = eq_summary,
        ski_brands          = ski_brands,
        sb_brands           = sb_brands,
        boot_brands         = boot_brands,
        binding_brands      = binding_brands,
    )


@app.route("/admin/growth")
@login_required
@admin_required
def admin_growth():
    """Admin Growth — acquisition funnel, activation, virality, retention proxy."""
    from collections import Counter as _GCtr
    from sqlalchemy import func
    from services.pass_utils import normalize_pass

    now_dt      = datetime.utcnow()
    thirty_ago  = now_dt - timedelta(days=30)
    _no_pass    = frozenset({"no_pass", "no_pass_yet"})

    # ── Base cohort: non-seeded users ─────────────────────────────────────────
    all_users = User.query.filter_by(is_seeded=False).all()
    total     = len(all_users) or 1

    def _pct(n, d=None):
        d = d if d is not None else total
        return round(n / d * 100) if d else 0

    # ── 1. Funnel ─────────────────────────────────────────────────────────────
    users_onboarded   = sum(1 for u in all_users if u.onboarding_completed_at)
    users_first_trip  = sum(1 for u in all_users if u.first_trip_created_at)

    # Users with at least one non-seeded Friend edge on either side
    _friend_uids = {
        r[0] for r in db.session.query(Friend.user_id).filter_by(is_seeded=False).all()
    } | {
        r[0] for r in db.session.query(Friend.friend_id).filter_by(is_seeded=False).all()
    }
    users_first_friend = len(_friend_uids & {u.id for u in all_users})

    # Avg days from signup → first trip
    trip_deltas = [
        (u.first_trip_created_at - u.created_at).days
        for u in all_users
        if u.first_trip_created_at and u.created_at
        and (u.first_trip_created_at - u.created_at).days >= 0
    ]
    avg_days_to_trip = round(sum(trip_deltas) / len(trip_deltas)) if trip_deltas else None

    # Avg days from signup → first friend (earliest Friend edge where user is owner)
    _first_friend_map = {
        uid: ts
        for uid, ts in db.session.query(
            Friend.user_id, func.min(Friend.created_at)
        ).filter_by(is_seeded=False).group_by(Friend.user_id).all()
    }
    friend_deltas = [
        (ts - u.created_at).days
        for u in all_users
        for ts in [_first_friend_map.get(u.id)]
        if ts and u.created_at and (ts - u.created_at).days >= 0
    ]
    avg_days_to_friend = round(sum(friend_deltas) / len(friend_deltas)) if friend_deltas else None

    funnel = dict(
        total              = total,
        onboarded          = users_onboarded,
        onboarding_rate    = _pct(users_onboarded),
        first_trip         = users_first_trip,
        first_trip_rate    = _pct(users_first_trip),
        first_friend       = users_first_friend,
        first_friend_rate  = _pct(users_first_friend),
        avg_days_to_trip   = avg_days_to_trip,
        avg_days_to_friend = avg_days_to_friend,
    )

    # ── 2. Activation ─────────────────────────────────────────────────────────
    _stage_order  = ["new", "onboarding", "active"]
    _stage_labels = {"new": "New", "onboarding": "Onboarding", "active": "Active"}
    stage_ctr     = _GCtr(u.lifecycle_stage or "new" for u in all_users)
    lifecycle_rows = [
        (_stage_labels.get(s, s.title()), stage_ctr.get(s, 0), _pct(stage_ctr.get(s, 0)))
        for s in _stage_order
    ]
    for s, cnt in stage_ctr.items():
        if s not in _stage_order:
            lifecycle_rows.append((s.title(), cnt, _pct(cnt)))

    pass_on_file = 0
    for u in all_users:
        if not u.pass_type:
            continue
        for raw in str(u.pass_type).split(","):
            raw = raw.strip()
            if not raw:
                continue
            norm = normalize_pass(raw)
            if norm and norm not in _no_pass:
                pass_on_file += 1
                break

    avail_uids    = {r[0] for r in db.session.query(UserAvailability.user_id).distinct().all()}
    users_avail   = len(avail_uids & {u.id for u in all_users})
    users_wishlist = sum(1 for u in all_users if u.wish_list_resorts and len(u.wish_list_resorts) > 0)
    users_visited  = sum(1 for u in all_users if u.visited_resort_ids and len(u.visited_resort_ids) > 0)

    activation = dict(
        lifecycle_rows  = lifecycle_rows,
        pass_on_file    = pass_on_file,
        pass_pct        = _pct(pass_on_file),
        wishlist        = users_wishlist,
        wishlist_pct    = _pct(users_wishlist),
        avail           = users_avail,
        avail_pct       = _pct(users_avail),
        mountains       = users_visited,
        mountains_pct   = _pct(users_visited),
    )

    # ── 3. Virality ───────────────────────────────────────────────────────────
    total_invites    = Invitation.query.count()
    accepted_invites = Invitation.query.filter_by(status="accepted").count()
    inv_accept_rate  = _pct(accepted_invites, total_invites)

    inv_30d          = Invitation.query.filter(Invitation.created_at >= thirty_ago).count()
    inv_30d_accepted = Invitation.query.filter(
        Invitation.created_at >= thirty_ago,
        Invitation.status == "accepted",
    ).count()
    inv_accept_rate_30d = _pct(inv_30d_accepted, inv_30d)

    senders          = db.session.query(Invitation.sender_id.distinct()).count()
    pct_sent_invite  = _pct(senders)

    referred         = sum(1 for u in all_users if u.invited_by_user_id)
    organic          = total - referred
    avg_inv_per_user = round(total_invites / total, 2)
    k_factor         = round(avg_inv_per_user * (inv_accept_rate / 100), 3)

    virality = dict(
        total_invites       = total_invites,
        accepted_invites    = accepted_invites,
        inv_accept_rate     = inv_accept_rate,
        inv_30d             = inv_30d,
        inv_30d_accepted    = inv_30d_accepted,
        inv_accept_rate_30d = inv_accept_rate_30d,
        senders             = senders,
        pct_sent_invite     = pct_sent_invite,
        organic             = organic,
        referred            = referred,
        organic_pct         = _pct(organic),
        referred_pct        = _pct(referred),
        avg_inv_per_user    = avg_inv_per_user,
        k_factor            = k_factor,
    )

    # ── 4. Retention proxy ────────────────────────────────────────────────────
    def _ret(n_days):
        cutoff  = now_dt - timedelta(days=n_days)
        cohort  = [u for u in all_users if u.created_at and u.created_at <= cutoff]
        if not cohort:
            return None, 0, 0
        retained = sum(
            1 for u in cohort
            if u.last_active_at
            and u.last_active_at >= (u.created_at + timedelta(days=n_days))
        )
        return round(retained / len(cohort) * 100), retained, len(cohort)

    d1_pct,  d1_ret,  d1_coh  = _ret(1)
    d7_pct,  d7_ret,  d7_coh  = _ret(7)
    d30_pct, d30_ret, d30_coh = _ret(30)

    retention = dict(
        d1  = dict(pct=d1_pct,  retained=d1_ret,  cohort=d1_coh),
        d7  = dict(pct=d7_pct,  retained=d7_ret,  cohort=d7_coh),
        d30 = dict(pct=d30_pct, retained=d30_ret, cohort=d30_coh),
    )

    return render_template(
        "admin_growth.html",
        active_tab = "growth",
        now        = _fmt_admin_now(),
        funnel     = funnel,
        activation = activation,
        virality   = virality,
        retention  = retention,
    )


@app.route("/admin/activation")
@login_required
@admin_required
def admin_activation():
    """Activation Dashboard v1 — 8-step funnel from signup to first group trip."""
    from services.pass_utils import normalize_pass

    # no_pass slugs: excludes "none" in addition to legacy no_pass variants
    _no_pass_slugs = frozenset({"no_pass", "no_pass_yet", "none"})
    now_dt = datetime.utcnow()

    # ── Base cohort: non-seeded users ─────────────────────────────────────────
    # Load all users so Python-computed properties (is_core_profile_complete,
    # is_active_user) can be evaluated without replicating complex SQL.
    all_users = User.query.filter(User.is_seeded == False).all()
    total = len(all_users)
    ns_ids = [u.id for u in all_users]

    def _pct(n, d):
        if not d:
            return 0
        return round(n / d * 100)

    # ── Step 1: Total Users ───────────────────────────────────────────────────
    s1 = total

    # ── Step 2: Onboarding Complete ───────────────────────────────────────────
    # Source of truth: is_core_profile_complete (the mandatory onboarding gate).
    # lifecycle_stage is NOT used — it is stale for users whose state has changed
    # since the last explicit update_lifecycle_stage() call.
    s2 = sum(1 for u in all_users if u.is_core_profile_complete)

    # ── Step 3: Has Real Pass ─────────────────────────────────────────────────
    # Subset of onboarded users who hold a committed pass (not no_pass/none).
    s3 = 0
    for u in all_users:
        pt = u.pass_type
        if not pt:
            continue
        for raw in str(pt).split(","):
            raw = raw.strip().lower()
            if not raw:
                continue
            norm = normalize_pass(raw)
            slug = norm if norm else raw
            if slug not in _no_pass_slugs:
                s3 += 1
                break

    # ── Step 4: Availability Added ────────────────────────────────────────────
    # Source of truth: open_dates JSON column (UserAvailability table has 0 rows)
    s4 = sum(
        1 for u in all_users
        if isinstance(u.open_dates, list) and len(u.open_dates) > 0
    )

    # ── Step 5: Wishlist Added ────────────────────────────────────────────────
    s5 = sum(
        1 for u in all_users
        if isinstance(u.wish_list_resorts, list) and len(u.wish_list_resorts) > 0
    )

    # ── Step 6: Activated User ────────────────────────────────────────────────
    # Source of truth: is_active_user = is_core_profile_complete AND has_started_planning.
    # lifecycle_stage is NOT used — it undercounts due to stale writes.
    s6 = sum(1 for u in all_users if u.is_active_user)

    # ── Social signals (shown separately, not in funnel conversion) ───────────
    friend_count = db.session.query(Friend.user_id.distinct()).filter(
        Friend.user_id.in_(ns_ids)
    ).count()

    # Accepted group-trip participants who are not the trip owner
    group_trip_participant_ids = set(
        r[0] for r in db.session.query(SkiTripParticipant.user_id).join(
            SkiTrip, SkiTrip.id == SkiTripParticipant.trip_id
        ).filter(
            SkiTrip.is_group_trip == True,
            SkiTripParticipant.status == GuestStatus.ACCEPTED,
            SkiTripParticipant.role != ParticipantRole.OWNER,
            SkiTripParticipant.user_id.in_(ns_ids),
        ).all()
    )
    group_trip_count = len(group_trip_participant_ids)

    social = {
        "friend_count":      friend_count,
        "friend_pct":        _pct(friend_count, total),
        "group_trip_count":  group_trip_count,
        "group_trip_pct":    _pct(group_trip_count, total),
    }

    # ── Build funnel steps ────────────────────────────────────────────────────
    raw_steps = [
        ("Total Users",         s1),
        ("Onboarding Complete", s2),
        ("Has Real Pass",       s3),
        ("Availability Added",  s4),
        ("Wishlist Added",      s5),
        ("Activated User",      s6),
    ]

    funnel_steps = []
    for i, (name, count) in enumerate(raw_steps):
        vs_total = _pct(count, total)
        if i == 0:
            vs_prev = 100
        else:
            prev_count = raw_steps[i - 1][1]
            vs_prev = 0 if not prev_count else min(100, round(count / prev_count * 100))
        funnel_steps.append({
            "name":     name,
            "count":    count,
            "vs_prev":  vs_prev,
            "vs_total": vs_total,
        })

    # ── Biggest Dropoff ───────────────────────────────────────────────────────
    dropoff = None
    worst_drop = 0
    for i in range(1, len(funnel_steps)):
        drop = funnel_steps[i]["vs_prev"] - 100
        if drop < worst_drop:
            worst_drop = drop
            dropoff = {
                "from_step": funnel_steps[i - 1]["name"],
                "to_step":   funnel_steps[i]["name"],
                "delta":     drop,
            }

    # ── KPI cards ─────────────────────────────────────────────────────────────
    kpis = {
        "activation_rate":   _pct(s6, total),
        "onboarding_pct":    _pct(s2, total),
        "availability_pct":  _pct(s4, total),
        "real_pass_pct":     _pct(s3, total),
        "wishlist_pct":      _pct(s5, total),
    }

    # ── Activation Insight ────────────────────────────────────────────────────
    activation_insight = {
        "onboarding_pct":  _pct(s2, total),
        "activation_pct":  _pct(s6, total),
    } if total > 0 else None

    return render_template(
        "admin_activation.html",
        active_tab         = "activation",
        now                = _fmt_admin_now(),
        funnel_steps       = funnel_steps,
        kpis               = kpis,
        dropoff            = dropoff,
        total              = total,
        social             = social,
        activation_insight = activation_insight,
    )


@app.route("/admin/resort-intelligence")
@login_required
@admin_required
def admin_resort_intelligence():
    """Resort Intelligence v1 — destination demand from wishlist, trips, and page views."""
    from collections import defaultdict
    from datetime import timedelta
    from sqlalchemy import func

    now_dt = datetime.utcnow()
    VIEW_TRACKING_LAUNCH = datetime(2026, 5, 27)
    tracking_age_days = (now_dt - VIEW_TRACKING_LAUNCH).days

    # ── Non-seeded user base ─────────────────────────────────────────────────
    all_users = User.query.filter(User.is_seeded == False).all()
    ns_ids    = [u.id for u in all_users]
    total_users = len(all_users)

    # ── Per-resort signals ────────────────────────────────────────────────────
    # Wishlist: User.wish_list_resorts (JSON list of resort IDs)
    wl_count  = defaultdict(int)   # resort_id → wishlist count
    wl_users  = defaultdict(set)   # resort_id → set of user_ids (for conversion)
    for u in all_users:
        for rid in (u.wish_list_resorts or []):
            wl_count[rid]  += 1
            wl_users[rid].add(u.id)

    # Trips: SkiTrip grouped by resort_id
    trip_rows = db.session.query(
        SkiTrip.resort_id,
        func.count(SkiTrip.id).label("trips"),
        func.sum(db.cast(SkiTrip.is_group_trip, db.Integer)).label("group_trips"),
    ).filter(
        SkiTrip.user_id.in_(ns_ids),
        SkiTrip.resort_id != None,
    ).group_by(SkiTrip.resort_id).all()

    trip_count  = {r.resort_id: r.trips       for r in trip_rows}
    gt_count    = {r.resort_id: r.group_trips  for r in trip_rows}

    # Trip user sets (for wishlist → trip conversion)
    trip_user_rows = db.session.query(SkiTrip.resort_id, SkiTrip.user_id).filter(
        SkiTrip.user_id.in_(ns_ids), SkiTrip.resort_id != None
    ).all()
    trip_users = defaultdict(set)
    for row in trip_user_rows:
        trip_users[row.resort_id].add(row.user_id)

    # Page views
    view_rows = db.session.query(
        MountainPageView.resort_id,
        func.count(MountainPageView.id).label("views"),
    ).group_by(MountainPageView.resort_id).all()
    view_count = {r.resort_id: r.views for r in view_rows}

    # ── All resort IDs with any signal ───────────────────────────────────────
    all_active_ids = set(wl_count) | set(trip_count) | set(view_count)

    # ── Fetch resort names in one query ──────────────────────────────────────
    resort_map = {}
    if all_active_ids:
        for r in Resort.query.filter(Resort.id.in_(all_active_ids)).all():
            resort_map[r.id] = r.name

    # ── Build unified per-resort rows ─────────────────────────────────────────
    rows = []
    for rid in all_active_ids:
        wl  = wl_count.get(rid, 0)
        tr  = trip_count.get(rid, 0)
        gt  = int(gt_count.get(rid, 0) or 0)
        vw  = view_count.get(rid, 0)
        # Wishlist → Trip conversion: % of wishlisters who also tripped here
        wl_uid = wl_users.get(rid, set())
        tr_uid = trip_users.get(rid, set())
        wl_trip_conv = round(len(wl_uid & tr_uid) / len(wl_uid) * 100) if wl_uid else None
        # Trip → Group %
        tr_gt_conv = round(gt / tr * 100) if tr else None
        rows.append({
            "id":           rid,
            "name":         resort_map.get(rid, f"Resort {rid}"),
            "wishlists":    wl,
            "trips":        tr,
            "group_trips":  gt,
            "views":        vw,
            "wl_trip_conv": wl_trip_conv,
            "tr_gt_conv":   tr_gt_conv,
        })

    # ── Section 0: Executive KPIs ─────────────────────────────────────────────
    total_wishlists   = sum(wl_count.values())
    total_trips_count = sum(trip_count.values())
    total_group_trips = sum(int(v or 0) for v in gt_count.values())
    total_views       = sum(view_count.values())
    trip_planners     = db.session.query(SkiTrip.user_id.distinct()).filter(
        SkiTrip.user_id.in_(ns_ids)
    ).count()
    resorts_with_activity = len(all_active_ids)

    exec_kpis = dict(
        resorts_with_activity = resorts_with_activity,
        trip_planners         = trip_planners,
        total_wishlists       = total_wishlists,
        total_group_trips     = total_group_trips,
        total_views           = total_views,
    )

    # ── Section 1: Demand table — top 25 by trips DESC ────────────────────────
    demand_table = sorted(rows, key=lambda r: (-r["trips"], -r["wishlists"]))[:25]

    # ── Section 2: Funnel — resorts with ≥1 trip, top 10 ─────────────────────
    funnel_rows = sorted(
        [r for r in rows if r["trips"] > 0],
        key=lambda r: -r["trips"]
    )[:10]

    # ── Section 3: Leaderboards ───────────────────────────────────────────────
    lb_most_wishlisted = sorted(
        [r for r in rows if r["wishlists"] > 0],
        key=lambda r: -r["wishlists"]
    )[:5]
    lb_most_planned = sorted(
        [r for r in rows if r["trips"] > 0],
        key=lambda r: -r["trips"]
    )[:5]
    lb_wl_trip_conv = sorted(
        [r for r in rows if r["wl_trip_conv"] is not None and r["wishlists"] >= 1],
        key=lambda r: -r["wl_trip_conv"]
    )[:5]

    leaderboards = dict(
        most_wishlisted = lb_most_wishlisted,
        most_planned    = lb_most_planned,
        wl_trip_conv    = lb_wl_trip_conv,
    )

    # ── Section 6: Market Insight ─────────────────────────────────────────────
    top_planned    = lb_most_planned[0]  if lb_most_planned    else None
    top_wishlisted = lb_most_wishlisted[0] if lb_most_wishlisted else None
    insight = None
    if top_planned and top_wishlisted:
        if top_planned["id"] == top_wishlisted["id"]:
            insight = f"{top_planned['name']} leads both planning and aspiration."
        else:
            insight = (
                f"{top_planned['name']} is the most planned destination. "
                f"{top_wishlisted['name']} is the most aspirational destination."
            )

    # ── Section 7: Destination Status ────────────────────────────────────────
    if tracking_age_days < 30:
        confidence = "Low"
    elif tracking_age_days < 180:
        confidence = "Growing"
    else:
        confidence = "High"

    total_resorts = Resort.query.count()

    status = dict(
        total_resorts          = total_resorts,
        resorts_with_activity  = resorts_with_activity,
        confidence             = confidence,
        tracking_age_days      = tracking_age_days,
    )

    return render_template(
        "admin_resort_intelligence.html",
        active_tab    = "resort_intelligence",
        now           = _fmt_admin_now(),
        exec_kpis     = exec_kpis,
        demand_table  = demand_table,
        funnel_rows   = funnel_rows,
        leaderboards  = leaderboards,
        insight       = insight,
        status        = status,
        total_views   = total_views,
    )


@app.route("/admin/product")
@login_required
@admin_required
def admin_product():
    """Product Dashboard v1 — feature adoption, activation drivers, power users."""
    from collections import Counter, defaultdict

    now_dt = datetime.utcnow()

    # ── Base population: non-seeded users ───────────────────────────────
    ns_users = User.query.filter(User.is_seeded == False).all()
    ns_ids   = [u.id for u in ns_users]
    total    = len(ns_users)
    if total == 0:
        return render_template("admin_product.html", active_tab="product",
                               now=_fmt_admin_now(),
                               total=0, kpis={}, adoption_table=[], drivers=[],
                               journey=[], underused=[], power={}, insight=None, status={})

    # ── Helper: safe pct ─────────────────────────────────────────────────
    def pct(n, d):
        return round(n / d * 100) if d else 0

    # ── Pre-compute feature memberships ─────────────────────────────────
    no_pass_slugs = {'no_pass', 'no_pass_yet', 'none', ''}

    def _has_real_pass(u):
        pt = u.pass_type or ''
        for raw in str(pt).split(','):
            if raw.strip().lower() not in no_pass_slugs:
                return True
        return False

    def _has_avail(u):
        return isinstance(u.open_dates, list) and len(u.open_dates) > 0

    def _has_wl(u):
        return isinstance(u.wish_list_resorts, list) and len(u.wish_list_resorts) > 0

    # Equipment: users with at least one EquipmentSetup row
    equip_user_ids = set(
        r[0] for r in db.session.query(EquipmentSetup.user_id.distinct())
        .filter(EquipmentSetup.user_id.in_(ns_ids)).all()
    )

    # Friends
    friend_user_ids = set(
        r[0] for r in db.session.query(Friend.user_id.distinct())
        .filter(Friend.user_id.in_(ns_ids)).all()
    )

    # Trips
    trip_rows = db.session.query(
        SkiTrip.user_id, db.func.count(SkiTrip.id).label('n')
    ).filter(SkiTrip.user_id.in_(ns_ids)).group_by(SkiTrip.user_id).all()
    trip_count_map = {r.user_id: r.n for r in trip_rows}
    trip_user_ids  = set(trip_count_map.keys())

    # Group trips
    gt_owner_ids = set(
        r[0] for r in db.session.query(SkiTrip.user_id.distinct())
        .filter(SkiTrip.user_id.in_(ns_ids), SkiTrip.is_group_trip == True).all()
    )

    # Invite tokens generated
    invite_user_ids = set(
        r[0] for r in db.session.query(InviteToken.inviter_id.distinct())
        .filter(InviteToken.inviter_id.in_(ns_ids)).all()
    )

    # ── Section 0: Product KPIs ──────────────────────────────────────────
    activated_n       = sum(1 for u in ns_users if u.is_active_user)
    profile_n         = sum(1 for u in ns_users if u.is_core_profile_complete)
    pass_n            = sum(1 for u in ns_users if _has_real_pass(u))
    avail_n           = sum(1 for u in ns_users if _has_avail(u))
    wl_n              = sum(1 for u in ns_users if _has_wl(u))
    friend_n          = len(friend_user_ids)

    kpis = dict(
        activated_pct = pct(activated_n, total),
        activated_n   = activated_n,
        profile_pct   = pct(profile_n, total),
        profile_n     = profile_n,
        pass_pct      = pct(pass_n, total),
        pass_n        = pass_n,
        avail_pct     = pct(avail_n, total),
        avail_n       = avail_n,
        wl_pct        = pct(wl_n, total),
        wl_n          = wl_n,
        friend_pct    = pct(friend_n, total),
        friend_n      = friend_n,
    )

    # ── Section 1: Feature Adoption Table ───────────────────────────────
    trip_n    = len(trip_user_ids)
    equip_n   = len(equip_user_ids)
    invite_n  = len(invite_user_ids)

    raw_adoption = [
        dict(feature='Invite',       users=invite_n,  pct=pct(invite_n, total)),
        dict(feature='Profile',      users=profile_n, pct=pct(profile_n, total)),
        dict(feature='Friends',      users=friend_n,  pct=pct(friend_n, total)),
        dict(feature='Pass',         users=pass_n,    pct=pct(pass_n, total)),
        dict(feature='Equipment',    users=equip_n,   pct=pct(equip_n, total)),
        dict(feature='Wishlist',     users=wl_n,      pct=pct(wl_n, total)),
        dict(feature='Trip',         users=trip_n,    pct=pct(trip_n, total)),
        dict(feature='Availability', users=avail_n,   pct=pct(avail_n, total)),
    ]
    adoption_table = sorted(raw_adoption, key=lambda r: r['pct'], reverse=True)

    # ── Section 2: Activation Drivers ───────────────────────────────────
    baseline_act = activated_n
    baseline_pct = pct(activated_n, total)

    def driver_row(label, with_ids):
        with_u  = [u for u in ns_users if u.id in with_ids]
        with_n  = len(with_u)
        act_n   = sum(1 for u in with_u if u.is_active_user)
        act_pct = pct(act_n, with_n)
        lift    = round(act_pct / baseline_pct, 1) if baseline_pct else 0
        return dict(feature=label, with_n=with_n, act_n=act_n,
                    act_pct=act_pct, lift=lift)

    drivers_raw = [
        driver_row('Equipment',    equip_user_ids),
        driver_row('Availability', set(u.id for u in ns_users if _has_avail(u))),
        driver_row('Wishlist',     set(u.id for u in ns_users if _has_wl(u))),
        driver_row('Friends',      friend_user_ids),
    ]
    drivers = sorted(drivers_raw, key=lambda r: r['lift'], reverse=True)

    # ── Section 3: Product Journey ───────────────────────────────────────
    onboarding_n = profile_n          # is_core_profile_complete
    real_pass_n  = pass_n
    journey = [
        dict(label='Total Users',        n=total,        pct=100),
        dict(label='Onboarding Complete', n=onboarding_n, pct=pct(onboarding_n, total)),
        dict(label='Has Real Pass',       n=real_pass_n,  pct=pct(real_pass_n, total)),
        dict(label='Availability Added',  n=avail_n,      pct=pct(avail_n, total)),
        dict(label='Wishlist Added',      n=wl_n,         pct=pct(wl_n, total)),
        dict(label='Activated User',      n=activated_n,  pct=pct(activated_n, total)),
    ]

    # ── Section 4: Underused Features (bottom 5) ─────────────────────────
    underused = sorted(adoption_table, key=lambda r: r['pct'])[:5]

    # ── Section 5: Power Users ───────────────────────────────────────────
    scores = []
    for u in ns_users:
        s = 0
        if _has_real_pass(u):  s += 1
        if _has_avail(u):      s += 1
        if _has_wl(u):         s += 1
        if u.id in equip_user_ids: s += 1
        if u.id in friend_user_ids: s += 1
        tc = trip_count_map.get(u.id, 0)
        s += tc * 2
        if u.id in gt_owner_ids: s += 3
        if tc >= 2:            s += 1
        scores.append(s)

    avg_score = round(sum(scores) / len(scores), 1) if scores else 0
    dist = Counter()
    for s in scores:
        if s <= 2:   dist['0–2'] += 1
        elif s <= 5: dist['3–5'] += 1
        elif s <= 8: dist['6–8'] += 1
        else:        dist['9+']  += 1
    dist_max = max(dist.values()) if dist else 1

    power = dict(
        avg_score = avg_score,
        dist      = [
            dict(label='0–2', n=dist.get('0–2', 0), pct=pct(dist.get('0–2',0), total)),
            dict(label='3–5', n=dist.get('3–5', 0), pct=pct(dist.get('3–5',0), total)),
            dict(label='6–8', n=dist.get('6–8', 0), pct=pct(dist.get('6–8',0), total)),
            dict(label='9+',  n=dist.get('9+',  0), pct=pct(dist.get('9+', 0), total)),
        ],
        dist_max  = dist_max,
        total     = total,
    )

    # ── Section 6: Product Insight ───────────────────────────────────────
    insight = None
    if baseline_pct > 0 and total >= 5:
        best = drivers[0] if drivers else None
        if best and best['with_n'] >= 3 and best['lift'] >= 1.5:
            insight = (
                f"{best['feature']} users activate at {best['lift']}× the baseline rate "
                f"({best['act_pct']}% vs {baseline_pct}% overall) — "
                f"yet only {pct(best['with_n'], total)}% of users have used this feature."
            )

    # ── Section 7: Product Status ─────────────────────────────────────────
    tracked_features  = 8
    meaningful_thresh = 20   # > 20% adoption
    meaningful_n      = sum(1 for r in adoption_table if r['pct'] > meaningful_thresh)
    # Confidence: Growing — view tracking < 60 days, user base < 100
    confidence = 'Growing'

    status = dict(
        tracked_features  = tracked_features,
        meaningful_n      = meaningful_n,
        confidence        = confidence,
        total_users       = total,
        baseline_pct      = baseline_pct,
    )

    return render_template(
        "admin_product.html",
        active_tab     = "product",
        now            = _fmt_admin_now(),
        total          = total,
        kpis           = kpis,
        adoption_table = adoption_table,
        drivers        = drivers,
        baseline_pct   = baseline_pct,
        journey        = journey,
        underused      = underused,
        power          = power,
        insight        = insight,
        status         = status,
    )


@app.route("/admin/messaging")
@login_required
@admin_required
def admin_messaging():
    """Messaging Dashboard v1 — invite links, trip invites, push reach, friend growth."""
    from collections import Counter

    now_dt = datetime.utcnow()

    # ── Base population: non-seeded users ───────────────────────────────
    ns_ids = [r[0] for r in db.session.query(User.id).filter(User.is_seeded == False).all()]
    total  = len(ns_ids)

    def pct(n, d):
        return round(n / d * 100) if d else 0

    # ── InviteToken (friend invite links) ───────────────────────────────
    it_generated = db.session.query(InviteToken.id).filter(
        InviteToken.inviter_id.in_(ns_ids)
    ).count()
    it_redeemed = db.session.query(InviteToken.id).filter(
        InviteToken.inviter_id.in_(ns_ids),
        InviteToken.used_at != None
    ).count()
    it_users = db.session.query(InviteToken.inviter_id.distinct()).filter(
        InviteToken.inviter_id.in_(ns_ids)
    ).count()

    # ── Trip Invites (Invitation where trip_id IS NOT NULL) + TripInviteToken ─
    ti_rows = db.session.query(
        Invitation.status, db.func.count(Invitation.id)
    ).filter(
        Invitation.sender_id.in_(ns_ids),
        Invitation.trip_id != None
    ).group_by(Invitation.status).all()
    ti_status = {s: n for s, n in ti_rows}
    ti_direct_sent     = sum(ti_status.values())
    ti_direct_accepted = ti_status.get('accepted', 0)
    ti_direct_pending  = ti_status.get('pending', 0)
    ti_direct_cancelled= ti_status.get('cancelled', 0)

    from models import TripInviteToken
    ti_link_sent = db.session.query(TripInviteToken.id).filter(
        TripInviteToken.inviter_user_id.in_(ns_ids)
    ).count()

    ti_total_sent     = ti_direct_sent + ti_link_sent
    ti_total_accepted = ti_direct_accepted   # links don't have per-recipient status

    # ── SkiTripParticipant corroboration ────────────────────────────────
    stp_rows = db.session.query(
        SkiTripParticipant.status, db.func.count(SkiTripParticipant.id)
    ).filter(
        SkiTripParticipant.user_id.in_(ns_ids),
        SkiTripParticipant.role != ParticipantRole.OWNER
    ).group_by(SkiTripParticipant.status).all()
    stp_status = {(s.value if hasattr(s, 'value') else s): n for s, n in stp_rows}

    # ── Friendships formed (Friend table, bidirectional) ─────────────────
    friend_rows = db.session.query(Friend.id).filter(
        Friend.user_id.in_(ns_ids)
    ).count()
    friendships_formed = friend_rows // 2 if friend_rows else 0

    # ── Push reach ───────────────────────────────────────────────────────
    push_users = db.session.query(PushDeviceToken.user_id.distinct()).filter(
        PushDeviceToken.user_id.in_(ns_ids)
    ).count()
    push_disabled = db.session.query(User.id).filter(
        User.id.in_(ns_ids),
        User.push_notifications_enabled == False
    ).count()

    # ── MEL queries (Sections A–E) ───────────────────────────────────────
    from datetime import timedelta
    from services.messaging_constants import EventName as _EN

    _now = now_dt
    _l7  = _now - timedelta(days=7)
    _l30 = _now - timedelta(days=30)
    _l90 = _now - timedelta(days=90)

    def _mel_window(cutoff=None):
        q = db.session.query(
            MessageEventLog.delivery_status,
            db.func.count(MessageEventLog.id)
        )
        if cutoff:
            q = q.filter(MessageEventLog.created_at >= cutoff)
        rows = q.group_by(MessageEventLog.delivery_status).all()
        d = {s: n for s, n in rows}
        total = sum(d.values())
        return dict(
            total   = total,
            sent    = d.get('sent', 0),
            skipped = d.get('skipped', 0),
            failed  = d.get('failed', 0),
            pending = d.get('pending', 0),
        )

    _mel_all = _mel_window()
    _mel_l7  = _mel_window(_l7)
    _mel_l30 = _mel_window(_l30)
    _mel_l90 = _mel_window(_l90)

    mel_volume = dict(
        all = _mel_all,
        l7  = _mel_l7,
        l30 = _mel_l30,
        l90 = _mel_l90,
    )

    # Section B: Event Type Activity leaderboard
    _evt_rows = db.session.query(
        MessageEventLog.event_name,
        MessageEventLog.delivery_status,
        db.func.count(MessageEventLog.id)
    ).group_by(
        MessageEventLog.event_name,
        MessageEventLog.delivery_status,
    ).all()

    _evt_map = {}
    for _en, _st, _cnt in _evt_rows:
        if _en not in _evt_map:
            _evt_map[_en] = dict(event_name=_en, total=0, sent=0, skipped=0, failed=0, pending=0)
        _evt_map[_en]['total'] += _cnt
        if _st in ('sent', 'skipped', 'failed', 'pending'):
            _evt_map[_en][_st] += _cnt
        else:
            _evt_map[_en]['pending'] += _cnt

    mel_event_leaderboard = sorted(_evt_map.values(), key=lambda r: r['total'], reverse=True)
    for _r in mel_event_leaderboard:
        _r['sent_pct'] = pct(_r['sent'], _r['total'])

    # Section C: Delivery Health (all-time from MEL)
    mel_health = dict(
        total       = _mel_all['total'],
        sent        = _mel_all['sent'],
        skipped     = _mel_all['skipped'],
        failed      = _mel_all['failed'],
        sent_pct    = pct(_mel_all['sent'],    _mel_all['total']),
        skipped_pct = pct(_mel_all['skipped'], _mel_all['total']),
        failed_pct  = pct(_mel_all['failed'],  _mel_all['total']),
    )

    # Section D: Suppression Breakdown
    _supp_rows = db.session.query(
        MessageEventLog.suppression_reason,
        db.func.count(MessageEventLog.id)
    ).filter(
        MessageEventLog.delivery_status == 'skipped',
        MessageEventLog.suppression_reason.isnot(None),
    ).group_by(
        MessageEventLog.suppression_reason,
    ).order_by(
        db.func.count(MessageEventLog.id).desc()
    ).all()

    _total_skipped = _mel_all['skipped'] or 1
    mel_suppression = [
        dict(reason=_r, count=_n, share_pct=pct(_n, _total_skipped))
        for _r, _n in _supp_rows
    ]

    # Section E: Network Effect Signals
    _net_event_names = [
        _EN.FRIEND_REQUEST_CREATED,
        _EN.FRIEND_REQUEST_ACCEPTED,
        _EN.TRIP_INVITE_CREATED,
        _EN.TRIP_INVITE_ACCEPTED,
    ]
    _net_label_map = {
        _EN.FRIEND_REQUEST_CREATED:  'Friend Request Created',
        _EN.FRIEND_REQUEST_ACCEPTED: 'Friend Request Accepted',
        _EN.TRIP_INVITE_CREATED:     'Trip Invite Created',
        _EN.TRIP_INVITE_ACCEPTED:    'Trip Invite Accepted',
    }
    _net_rows = db.session.query(
        MessageEventLog.event_name,
        MessageEventLog.delivery_status,
        db.func.count(MessageEventLog.id)
    ).filter(
        MessageEventLog.event_name.in_(_net_event_names)
    ).group_by(
        MessageEventLog.event_name,
        MessageEventLog.delivery_status,
    ).all()

    _net_map = {}
    for _en, _st, _cnt in _net_rows:
        if _en not in _net_map:
            _net_map[_en] = dict(event_name=_en, label=_net_label_map[_en],
                                 total=0, sent=0, skipped=0, failed=0)
        _net_map[_en]['total'] += _cnt
        if _st in ('sent', 'skipped', 'failed'):
            _net_map[_en][_st] += _cnt

    mel_network = [
        _net_map.get(_en, dict(event_name=_en, label=_net_label_map[_en],
                               total=0, sent=0, skipped=0, failed=0))
        for _en in _net_event_names
    ]
    for _r in mel_network:
        _r['sent_pct'] = pct(_r['sent'], _r['total'])

    # ── Section 0: Messaging KPIs ────────────────────────────────────────
    invitations_sent     = it_generated + ti_total_sent
    invitations_accepted = it_redeemed  + ti_total_accepted
    acceptance_rate      = pct(invitations_accepted, invitations_sent)
    push_reach_pct       = pct(push_users, total)
    push_optout_pct      = pct(push_disabled, total)

    kpis = dict(
        invitations_sent     = invitations_sent,
        invitations_accepted = invitations_accepted,
        acceptance_rate      = acceptance_rate,
        friendships_formed   = friendships_formed,
        push_reach_pct       = push_reach_pct,
        push_reach_n         = push_users,
        push_optout_pct      = push_optout_pct,
        push_optout_n        = push_disabled,
    )

    # ── Section 1: Message Volume Table ─────────────────────────────────
    volume_rows = [
        dict(action='Invite Links Generated', volume=it_generated,       accepted=None,          conv=None),
        dict(action='Invite Links Redeemed',  volume=it_redeemed,        accepted=None,          conv=pct(it_redeemed, it_generated)),
        dict(action='Friendships Formed',     volume=friendships_formed,  accepted=None,          conv=pct(friendships_formed, it_redeemed) if it_redeemed else None),
        dict(action='Push Reach',             volume=push_users,          accepted=None,          conv=pct(push_users, total)),
        dict(action='Trip Invites Sent',      volume=ti_total_sent,       accepted=ti_total_accepted, conv=pct(ti_total_accepted, ti_direct_sent) if ti_direct_sent else None),
        dict(action='Trip Invites Accepted',  volume=ti_total_accepted,   accepted=None,          conv=None),
    ]
    volume_rows = sorted(volume_rows, key=lambda r: r['volume'], reverse=True)

    # ── Section 2: Invite Link Funnel ────────────────────────────────────
    invite_funnel = [
        dict(label='Invite Links Generated', n=it_generated,       step_pct=100,
             step_conv=None, note=''),
        dict(label='Invite Links Redeemed',  n=it_redeemed,        step_pct=pct(it_redeemed, it_generated),
             step_conv=pct(it_redeemed, it_generated), note='redemption rate'),
        dict(label='Friendships Formed',     n=friendships_formed, step_pct=pct(friendships_formed, it_generated),
             step_conv=pct(friendships_formed, it_redeemed) if it_redeemed else None, note='est. from redeemed'),
    ]
    invite_overall_conv = pct(friendships_formed, it_generated)

    # ── Section 3: Friend Growth ──────────────────────────────────────────
    friend_growth = [
        dict(label='Users Who Generated Invite Links', n=it_users,         pct_of_total=pct(it_users, total)),
        dict(label='Invite Links Generated',           n=it_generated,     pct_of_total=None),
        dict(label='Invite Links Redeemed',            n=it_redeemed,      pct_of_total=pct(it_redeemed, it_generated)),
        dict(label='Friendships Formed',               n=friendships_formed, pct_of_total=pct(friendships_formed, it_redeemed) if it_redeemed else 0),
    ]

    # ── Section 4: Trip Invite Funnel ─────────────────────────────────────
    trip_funnel = dict(
        sent      = ti_direct_sent,
        link_sent = ti_link_sent,
        accepted  = ti_direct_accepted,
        pending   = ti_direct_pending,
        cancelled = ti_direct_cancelled,
        accept_pct  = pct(ti_direct_accepted, ti_direct_sent),
        pending_pct = pct(ti_direct_pending,  ti_direct_sent),
        cancel_pct  = pct(ti_direct_cancelled,ti_direct_sent),
    )

    # ── Section 5: Message Effectiveness ─────────────────────────────────
    effectiveness = []
    if it_generated >= 5:
        effectiveness.append(dict(action='Invite Link Redemption', conv=pct(it_redeemed, it_generated), n=it_generated))
    if ti_direct_sent >= 3:
        effectiveness.append(dict(action='Trip Invite Acceptance', conv=pct(ti_direct_accepted, ti_direct_sent), n=ti_direct_sent))
    if it_redeemed >= 3:
        effectiveness.append(dict(action='Redemption → Friendship', conv=pct(friendships_formed, it_redeemed), n=it_redeemed))
    effectiveness = sorted(effectiveness, key=lambda r: r['conv'], reverse=True)
    eff_max = max((r['conv'] for r in effectiveness), default=1) or 1

    # ── Section 6: Underused Channels ────────────────────────────────────
    underused_channels = [
        dict(channel='Trip Invites',   usage_n=ti_total_sent,   usage_pct=pct(ti_total_sent,   total), note='users who sent any trip invite'),
        dict(channel='Invite Links',   usage_n=it_users,        usage_pct=pct(it_users,         total), note='users who generated a link'),
        dict(channel='Push Reach',     usage_n=push_users,      usage_pct=push_reach_pct,        note='users with push token registered'),
        dict(channel='Friend Growth',  usage_n=friendships_formed, usage_pct=pct(friendships_formed, total), note='unique friendships vs user count'),
    ]
    underused_channels = sorted(underused_channels, key=lambda r: r['usage_pct'])

    # ── Section 7: Messaging Insight ─────────────────────────────────────
    insight = None
    candidates = []
    if it_generated >= 5:
        candidates.append((pct(it_redeemed, it_generated),
            f"{pct(it_redeemed, it_generated)}% of invite links are redeemed — "
            f"{it_redeemed} of {it_generated} generated links converted."))
    if push_reach_pct > 0:
        candidates.append((push_reach_pct,
            f"{push_reach_pct}% of users have granted push notification permissions ({push_users} of {total})."))
    if ti_direct_sent >= 3:
        candidates.append((pct(ti_direct_accepted, ti_direct_sent),
            f"Trip invites convert at {pct(ti_direct_accepted, ti_direct_sent)}% — "
            f"{ti_direct_accepted} of {ti_direct_sent} direct invites accepted."))
    # MEL-based insight candidates
    if _mel_all['total'] >= 5:
        _sent_rate = pct(_mel_all['sent'], _mel_all['total'])
        candidates.append((_sent_rate,
            f"{_sent_rate}% of push attempts were accepted by OneSignal "
            f"({_mel_all['sent']} of {_mel_all['total']} total events sent)."))
    if mel_event_leaderboard and _mel_all['total'] >= 5:
        _top = mel_event_leaderboard[0]
        _top_pct = pct(_top['total'], _mel_all['total'])
        candidates.append((_top_pct,
            f"{_top['event_name']} accounts for {_top_pct}% of all push activity "
            f"({_top['total']} of {_mel_all['total']} total events)."))
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        insight = candidates[0][1]

    # ── Section 8: Messaging Status ──────────────────────────────────────
    status = dict(
        tracked_signals  = 16,   # full audit: 9 GREEN + 4 YELLOW + 7 RED = 20 audited, 16 tracked
        reliable_signals = 9,    # GREEN signals
        confidence       = 'Growing',
        total_users      = total,
    )

    return render_template(
        "admin_messaging.html",
        active_tab              = "messaging",
        now                     = _fmt_admin_now(),
        total                   = total,
        kpis                    = kpis,
        volume_rows             = volume_rows,
        invite_funnel           = invite_funnel,
        invite_overall_conv     = invite_overall_conv,
        friend_growth           = friend_growth,
        trip_funnel             = trip_funnel,
        effectiveness           = effectiveness,
        eff_max                 = eff_max,
        underused_channels      = underused_channels,
        insight                 = insight,
        status                  = status,
        mel_volume              = mel_volume,
        mel_event_leaderboard   = mel_event_leaderboard,
        mel_health              = mel_health,
        mel_suppression         = mel_suppression,
        mel_network             = mel_network,
    )


@app.route("/admin/retention")
@login_required
@admin_required
def admin_retention():
    """Retention Dashboard v1 — proxy metrics from last_active_at and existing models."""
    from datetime import timedelta

    now_dt     = datetime.utcnow()
    one_ago    = now_dt - timedelta(days=1)
    seven_ago  = now_dt - timedelta(days=7)
    thirty_ago = now_dt - timedelta(days=30)
    fourteen_ago = now_dt - timedelta(days=14)

    # ── Base cohort ───────────────────────────────────────────────────────────
    non_seeded = User.query.filter(User.is_seeded == False)

    # ── Section 1: KPIs ──────────────────────────────────────────────────────
    dau = non_seeded.filter(User.last_active_at >= one_ago).count()
    wau = non_seeded.filter(User.last_active_at >= seven_ago).count()
    mau = non_seeded.filter(User.last_active_at >= thirty_ago).count()

    # Prior-period WAU (7-14 days ago)
    prior_wau = non_seeded.filter(
        User.last_active_at >= fourteen_ago,
        User.last_active_at < seven_ago,
    ).count()
    wau_delta = wau - prior_wau

    # Prior MAU: suppress if prior window likely has 0 (app < 60 days old)
    oldest_user = db.session.query(db.func.min(User.created_at)).filter(User.is_seeded == False).scalar()
    app_age_days = (now_dt - oldest_user).days if oldest_user else 0
    suppress_mau_delta = app_age_days < 60

    prior_mau = non_seeded.filter(
        User.last_active_at >= (now_dt - timedelta(days=60)),
        User.last_active_at < thirty_ago,
    ).count() if not suppress_mau_delta else None

    stickiness = round(dau / mau * 100) if mau else 0

    returning_l30 = non_seeded.filter(
        User.last_active_at >= thirty_ago,
        User.created_at < thirty_ago,
    ).count()

    # ── Sparkline: 14-day daily active counts ────────────────────────────────
    from collections import defaultdict
    sparkline_raw = defaultdict(int)
    rows = db.session.query(User.last_active_at).filter(
        User.is_seeded == False,
        User.last_active_at >= fourteen_ago,
    ).all()
    for (ts,) in rows:
        if ts:
            sparkline_raw[ts.strftime("%Y-%m-%d")] += 1
    sparkline = []
    for i in range(13, -1, -1):
        day = (now_dt - timedelta(days=i)).strftime("%Y-%m-%d")
        sparkline.append(sparkline_raw.get(day, 0))
    spark_max = max(sparkline) if any(sparkline) else 1

    kpis = dict(
        dau=dau, wau=wau, mau=mau,
        prior_wau=prior_wau, wau_delta=wau_delta,
        prior_mau=prior_mau, suppress_mau_delta=suppress_mau_delta,
        stickiness=stickiness,
        returning_l30=returning_l30,
        sparkline=sparkline, spark_max=spark_max,
    )

    # ── Section 2: D1/D7/D30 return rates ────────────────────────────────────
    def _ret_window(days):
        cutoff = now_dt - timedelta(days=days)
        cohort = non_seeded.filter(User.created_at <= cutoff).count()
        retained = non_seeded.filter(
            User.created_at <= cutoff,
            User.last_active_at >= (User.created_at + timedelta(days=days)),
        ).count()
        pct = round(retained / cohort * 100) if cohort else None
        return dict(cohort=cohort, retained=retained, pct=pct, small_sample=cohort < 5)

    return_rates = dict(d1=_ret_window(1), d7=_ret_window(7), d30=_ret_window(30))

    # ── Section 4: Retention by behavior (D7 proxy) ───────────────────────────
    def _behavior_ret(q):
        cohort = q.count()
        retained = q.filter(User.last_active_at >= User.created_at + timedelta(days=7)).count()
        pct = round(retained / cohort * 100) if cohort else 0
        return dict(users=cohort, retained=retained, pct=pct)

    non_seeded_ids = db.session.query(User.id).filter(User.is_seeded == False)

    has_friend_ids   = db.session.query(Friend.user_id.distinct()).filter(Friend.user_id.in_(non_seeded_ids))
    has_trip_ids     = db.session.query(SkiTrip.user_id.distinct()).filter(SkiTrip.user_id.in_(non_seeded_ids))

    # Availability: open_dates non-empty (canonical write target)
    avail_rows = db.session.query(User.id, User.open_dates, User.created_at, User.last_active_at).filter(User.is_seeded == False).all()
    wishlist_rows = db.session.query(User.id, User.wish_list_resorts, User.created_at, User.last_active_at).filter(User.is_seeded == False).all()
    friend_id_set = {r[0] for r in has_friend_ids}
    trip_id_set   = {r[0] for r in has_trip_ids}

    def _seg_from_rows(rows):
        """rows: list of (id, json_col, created_at, last_active_at)"""
        cohort = retained = 0
        for uid, jcol, cat, laa in rows:
            try:
                has_data = isinstance(jcol, list) and len(jcol) > 0
            except Exception:
                has_data = False
            if has_data:
                cohort += 1
                if laa and cat and laa >= cat + timedelta(days=7):
                    retained += 1
        pct = round(retained / cohort * 100) if cohort else 0
        return dict(users=cohort, retained=retained, pct=pct)

    avail_seg    = _seg_from_rows(avail_rows)
    wishlist_seg = _seg_from_rows(wishlist_rows)

    # friend/trip segments via query
    def _seg_from_id_set(id_set, all_rows):
        cohort = retained = 0
        for uid, _, cat, laa in all_rows:
            if uid in id_set:
                cohort += 1
                if laa and cat and laa >= cat + timedelta(days=7):
                    retained += 1
        pct = round(retained / cohort * 100) if cohort else 0
        return dict(users=cohort, retained=retained, pct=pct)

    # all non-seeded rows for cross-referencing
    all_ns_rows = db.session.query(User.id, User.open_dates, User.created_at, User.last_active_at).filter(User.is_seeded == False).all()

    friend_seg   = _seg_from_id_set(friend_id_set, all_ns_rows)
    trip_seg     = _seg_from_id_set(trip_id_set, all_ns_rows)

    # Browse-only: not in any of the above sets, and no availability, no wishlist
    avail_ids   = {uid for uid, jcol, _, _ in avail_rows if isinstance(jcol, list) and len(jcol) > 0}
    wishlist_ids = {uid for uid, jcol, _, _ in wishlist_rows if isinstance(jcol, list) and len(jcol) > 0}
    active_ids  = friend_id_set | trip_id_set | avail_ids | wishlist_ids
    browse_cohort = retained_browse = 0
    for uid, _, cat, laa in all_ns_rows:
        if uid not in active_ids:
            browse_cohort += 1
            if laa and cat and laa >= cat + timedelta(days=7):
                retained_browse += 1
    browse_pct = round(retained_browse / browse_cohort * 100) if browse_cohort else 0
    browse_seg = dict(users=browse_cohort, retained=retained_browse, pct=browse_pct)

    by_behavior = [
        dict(label="Has Friend",      **friend_seg),
        dict(label="Has Trip",        **trip_seg),
        dict(label="Has Availability",**avail_seg),
        dict(label="Has Wishlist",    **wishlist_seg),
        dict(label="Browse Only",     **browse_seg),
    ]
    beh_max_pct = max((r["pct"] for r in by_behavior), default=1) or 1

    # ── Section 6: Biggest retention driver ──────────────────────────────────
    action_segs = [r for r in by_behavior if r["label"] != "Browse Only"]
    best_seg    = max(action_segs, key=lambda r: r["pct"]) if action_segs else None
    driver = None
    if best_seg and best_seg["pct"] > 0:
        gap = best_seg["pct"] - browse_pct
        if gap >= 5:
            driver = dict(
                label=best_seg["label"],
                pct=best_seg["pct"],
                browse_pct=browse_pct,
                gap=gap,
            )

    # ── Section 7: Retention confidence context ───────────────────────────────
    if app_age_days < 61:
        confidence = "Low"
    elif app_age_days < 181:
        confidence = "Growing"
    else:
        confidence = "High"
    oldest_user_display = oldest_user.strftime("%b %d, %Y") if oldest_user else "—"

    status = dict(
        app_age_days=app_age_days,
        oldest_user=oldest_user_display,
        confidence=confidence,
    )

    return render_template(
        "admin_retention.html",
        active_tab   = "retention",
        now          = _fmt_admin_now(),
        kpis         = kpis,
        return_rates = return_rates,
        by_behavior  = by_behavior,
        beh_max_pct  = beh_max_pct,
        driver       = driver,
        status       = status,
    )


@app.route("/admin/user-intelligence")
@login_required
@admin_required
def admin_user_intelligence():
    """User Intelligence Dashboard v1 — demographics, pass, rider, destination, social, activation."""
    from collections import Counter, defaultdict
    from utils.countries import STATE_ABBR_MAP

    now_dt = datetime.utcnow()

    # ── Base population: non-seeded users ───────────────────────────────
    ns_users = User.query.filter(User.is_seeded == False).all()
    ns_ids   = [u.id for u in ns_users]
    total    = len(ns_users)

    def pct(n, d):
        return round(n / d * 100) if d else 0

    # ── Pass helpers ─────────────────────────────────────────────────────
    no_pass_slugs = {'no_pass', 'no_pass_yet', 'none', ''}

    def _has_real_pass(u):
        pt = u.pass_type or ''
        for raw in str(pt).split(','):
            if raw.strip().lower() not in no_pass_slugs:
                return True
        return False

    def _pass_slugs(u):
        pt = u.pass_type or ''
        return [r.strip().lower() for r in str(pt).split(',') if r.strip().lower() not in no_pass_slugs]

    def _has_avail(u):
        return isinstance(u.open_dates, list) and len(u.open_dates) > 0

    def _has_wl(u):
        return isinstance(u.wish_list_resorts, list) and len(u.wish_list_resorts) > 0

    # ── Pre-compute feature sets ─────────────────────────────────────────
    equip_user_ids = set(
        r[0] for r in db.session.query(EquipmentSetup.user_id.distinct())
        .filter(EquipmentSetup.user_id.in_(ns_ids)).all()
    )
    friend_user_ids = set(
        r[0] for r in db.session.query(Friend.user_id.distinct())
        .filter(Friend.user_id.in_(ns_ids)).all()
    )
    trip_rows = db.session.query(
        SkiTrip.user_id, db.func.count(SkiTrip.id).label('n')
    ).filter(SkiTrip.user_id.in_(ns_ids)).group_by(SkiTrip.user_id).all()
    trip_count_map = {r.user_id: r.n for r in trip_rows}

    gt_owner_ids = set(
        r[0] for r in db.session.query(SkiTrip.user_id.distinct())
        .filter(SkiTrip.user_id.in_(ns_ids), SkiTrip.is_group_trip == True).all()
    )

    # ── Product Score (identical to Product Dashboard) ──────────────────
    def product_score(u):
        s = 0
        if _has_real_pass(u):        s += 1
        if _has_avail(u):            s += 1
        if _has_wl(u):               s += 1
        if u.id in equip_user_ids:   s += 1
        if u.id in friend_user_ids:  s += 1
        tc = trip_count_map.get(u.id, 0)
        s += tc * 2
        if u.id in gt_owner_ids:     s += 3
        if tc >= 2:                  s += 1
        return s

    scores       = [product_score(u) for u in ns_users]
    avg_score    = round(sum(scores) / len(scores), 1) if scores else 0
    power_ids    = set(u.id for u, s in zip(ns_users, scores) if s >= 6)
    power_users  = [u for u in ns_users if u.id in power_ids]

    # ── Friend count map ─────────────────────────────────────────────────
    fr_count_map = defaultdict(int)
    for r in db.session.query(Friend.user_id, db.func.count(Friend.id).label('n'))\
            .filter(Friend.user_id.in_(ns_ids)).group_by(Friend.user_id).all():
        fr_count_map[r.user_id] = r.n

    # ── Section 0: User KPIs ─────────────────────────────────────────────
    friend_counts   = [fr_count_map.get(u.id, 0) for u in ns_users]
    trip_counts     = [trip_count_map.get(u.id, 0) for u in ns_users]
    wl_sizes        = [len(u.wish_list_resorts) if _has_wl(u) else 0 for u in ns_users]

    # Distinct resorts planned per user
    resort_rows = db.session.query(SkiTrip.user_id, SkiTrip.resort_id)\
        .filter(SkiTrip.user_id.in_(ns_ids), SkiTrip.resort_id != None).all()
    user_resort_sets = defaultdict(set)
    for uid, rid in resort_rows:
        user_resort_sets[uid].add(rid)
    planned_sizes = [len(user_resort_sets.get(u.id, set())) for u in ns_users]

    avg_friends         = round(sum(friend_counts) / total, 1) if total else 0
    avg_trips           = round(sum(trip_counts) / total, 1) if total else 0
    avg_wishlist        = round(sum(wl_sizes) / total, 1) if total else 0
    avg_planned_resorts = round(sum(planned_sizes) / total, 1) if total else 0

    kpis = dict(
        total               = total,
        avg_friends         = avg_friends,
        avg_trips           = avg_trips,
        avg_wishlist        = avg_wishlist,
        avg_planned_resorts = avg_planned_resorts,
        avg_score           = avg_score,
    )

    # ── Section 1: User Profile — Top States ────────────────────────────
    # Canadian province abbreviations (common ones)
    CA_PROVINCES = {'AB','BC','MB','NB','NL','NS','NT','NU','ON','PE','QC','SK','YT'}

    def state_to_country(abbr):
        if not abbr:
            return 'Unknown'
        if abbr.upper() in STATE_ABBR_MAP:
            return 'United States'
        if abbr.upper() in CA_PROVINCES:
            return 'Canada'
        return 'Other'

    state_counter   = Counter(u.home_state for u in ns_users if u.home_state)
    country_counter = Counter()
    for u in ns_users:
        if u.home_state:
            country_counter[state_to_country(u.home_state)] += 1

    top_states = [
        dict(state=abbr, name=STATE_ABBR_MAP.get(abbr, abbr), users=n, pct=pct(n, total))
        for abbr, n in state_counter.most_common(10)
    ]
    top_countries = [
        dict(country=country, users=n, pct=pct(n, total))
        for country, n in country_counter.most_common()
    ]
    state_coverage = pct(sum(1 for u in ns_users if u.home_state), total)

    # ── Section 2: Pass Intelligence ─────────────────────────────────────
    epic_n  = sum(1 for u in ns_users if 'epic' in _pass_slugs(u))
    ikon_n  = sum(1 for u in ns_users if 'ikon' in _pass_slugs(u))
    other_n = sum(
        1 for u in ns_users
        if any(s not in ('epic', 'ikon') for s in _pass_slugs(u)) and _has_real_pass(u)
        and not any(s in ('epic', 'ikon') for s in _pass_slugs(u))
    )
    none_n  = sum(1 for u in ns_users if not _has_real_pass(u))

    pass_table = sorted([
        dict(label='Epic',         users=epic_n,  pct=pct(epic_n,  total)),
        dict(label='Ikon',         users=ikon_n,  pct=pct(ikon_n,  total)),
        dict(label='Other Passes', users=other_n, pct=pct(other_n, total)),
        dict(label='No Pass',      users=none_n,  pct=pct(none_n,  total)),
    ], key=lambda r: r['users'], reverse=True)

    # ── Section 3: Rider Profile ─────────────────────────────────────────
    DISCIPLINES = {'Skier', 'Snowboarder', 'Telemark', 'Cross-Country'}

    def _disciplines(u):
        if not u.rider_types:
            return set()
        tags = u.rider_types if isinstance(u.rider_types, list) else [u.rider_types]
        return {str(t) for t in tags} & DISCIPLINES

    skiers     = sum(1 for u in ns_users if 'Skier' in _disciplines(u))
    boarders   = sum(1 for u in ns_users if 'Snowboarder' in _disciplines(u))
    telemark   = sum(1 for u in ns_users if 'Telemark' in _disciplines(u))
    xc         = sum(1 for u in ns_users if 'Cross-Country' in _disciplines(u))
    multi      = sum(1 for u in ns_users if len(_disciplines(u)) > 1)
    unknown_rt = sum(1 for u in ns_users if not _disciplines(u))

    rider_table = [
        dict(label='Skiers',          users=skiers,   pct=pct(skiers,   total)),
        dict(label='Snowboarders',     users=boarders, pct=pct(boarders, total)),
        dict(label='Multi-Discipline', users=multi,    pct=pct(multi,    total)),
        dict(label='Telemark',         users=telemark, pct=pct(telemark, total)),
        dict(label='Cross-Country',    users=xc,       pct=pct(xc,       total)),
    ]

    # ── Section 4: Destination Intelligence ──────────────────────────────
    # Wishlisted
    wl_resort_counter = Counter()
    wl_user_counter   = Counter()
    for u in ns_users:
        if _has_wl(u):
            seen = set()
            for rid in u.wish_list_resorts:
                wl_resort_counter[rid] += 1
                if rid not in seen:
                    wl_user_counter[rid] += 1
                    seen.add(rid)

    # Planned (trips per resort, distinct users per resort)
    planned_trip_counter = Counter(r[1] for r in resort_rows if r[1])
    planned_user_counter = Counter()
    for uid, rid in resort_rows:
        planned_user_counter[rid] += 1  # note: may double-count if same user has 2 trips at same resort
    # Distinct users per resort
    user_resort_resort = defaultdict(set)
    for uid, rid in resort_rows:
        if rid:
            user_resort_resort[rid].add(uid)
    planned_uniq_users = {rid: len(uids) for rid, uids in user_resort_resort.items()}

    # Resolve resort names
    top_wl_ids      = [rid for rid, _ in wl_user_counter.most_common(5)]
    top_plan_ids    = [rid for rid, _ in Counter(planned_trip_counter).most_common(5)]
    all_resort_ids  = list(set(top_wl_ids + top_plan_ids))
    resort_name_map = {r.id: r.name for r in Resort.query.filter(Resort.id.in_(all_resort_ids)).all()} if all_resort_ids else {}

    top_wishlisted = [
        dict(name=resort_name_map.get(rid, str(rid)),
             users=wl_user_counter[rid],
             trips=wl_resort_counter[rid])
        for rid in top_wl_ids
    ]
    top_planned = [
        dict(name=resort_name_map.get(rid, str(rid)),
             users=planned_uniq_users.get(rid, 0),
             trips=planned_trip_counter[rid])
        for rid in top_plan_ids
    ]

    # ── Section 5: Social Graph ───────────────────────────────────────────
    sorted_friends  = sorted(friend_counts)
    median_friends  = sorted_friends[len(sorted_friends) // 2] if sorted_friends else 0
    zero_friends_n  = sum(1 for c in friend_counts if c == 0)
    five_plus_n     = sum(1 for c in friend_counts if c >= 5)
    one_to_four_n   = sum(1 for c in friend_counts if 1 <= c <= 4)

    social = dict(
        avg_friends   = avg_friends,
        median        = median_friends,
        zero_n        = zero_friends_n,
        zero_pct      = pct(zero_friends_n, total),
        one_four_n    = one_to_four_n,
        one_four_pct  = pct(one_to_four_n, total),
        five_plus_n   = five_plus_n,
        five_plus_pct = pct(five_plus_n, total),
        dist_max      = max(zero_friends_n, one_to_four_n, five_plus_n, 1),
    )

    # ── Section 6: Activation Correlates ─────────────────────────────────
    baseline_act = sum(1 for u in ns_users if u.is_active_user)
    baseline_pct_val = pct(baseline_act, total)

    def correlate(label, with_ids_set):
        with_u  = [u for u in ns_users if u.id in with_ids_set]
        with_n  = len(with_u)
        act_n   = sum(1 for u in with_u if u.is_active_user)
        act_pct = pct(act_n, with_n)
        lift    = round(act_pct / baseline_pct_val, 2) if baseline_pct_val else 0
        return dict(
            segment = label,
            with_n  = with_n,
            act_n   = act_n,
            act_pct = act_pct,
            lift    = lift,
        )

    epic_ids   = set(u.id for u in ns_users if 'epic' in _pass_slugs(u))
    skier_ids  = set(u.id for u in ns_users if 'Skier' in _disciplines(u))
    avail_ids  = set(u.id for u in ns_users if _has_avail(u))
    wl_ids     = set(u.id for u in ns_users if _has_wl(u))

    correlates_raw = [
        correlate('Epic Pass',         epic_ids),
        correlate('Equipment Setup',   equip_user_ids),
        correlate('Friends',           friend_user_ids),
        correlate('Availability Set',  avail_ids),
        correlate('Wishlist',          wl_ids),
        correlate('Skier',             skier_ids),
    ]
    correlates = sorted(correlates_raw, key=lambda r: r['lift'], reverse=True)
    corr_max   = max((r['act_pct'] for r in correlates), default=1) or 1

    # ── Section 7: Power User Profile ────────────────────────────────────
    regular_users = [u for u in ns_users if u.id not in power_ids]

    def _avg(users, fn):
        vals = [fn(u) for u in users]
        return round(sum(vals) / len(vals), 1) if vals else 0

    power_profile = dict(
        power_n           = len(power_users),
        regular_n         = len(regular_users),
        total             = total,
        # All users
        all_avg_friends   = avg_friends,
        all_avg_trips     = avg_trips,
        all_avg_wl        = avg_wishlist,
        all_avail_pct     = pct(len(avail_ids), total),
        all_equip_pct     = pct(len(equip_user_ids), total),
        # Power users
        pow_avg_friends   = _avg(power_users, lambda u: fr_count_map.get(u.id, 0)),
        pow_avg_trips     = _avg(power_users, lambda u: trip_count_map.get(u.id, 0)),
        pow_avg_wl        = _avg(power_users, lambda u: len(u.wish_list_resorts) if _has_wl(u) else 0),
        pow_avail_pct     = pct(sum(1 for u in power_users if _has_avail(u)), max(len(power_users), 1)),
        pow_equip_pct     = pct(sum(1 for u in power_users if u.id in equip_user_ids), max(len(power_users), 1)),
    )

    # ── Section 8: User Insight ───────────────────────────────────────────
    insight = None
    # Candidate 1: Colorado concentration
    co_n = state_counter.get('CO', 0)
    co_pct_val = pct(co_n, sum(state_counter.values()))
    if co_pct_val >= 50 and co_n >= 5:
        insight = (
            f"{co_pct_val}% of users with a home state are based in Colorado "
            f"({co_n} of {sum(state_counter.values())} users with location data)."
        )
    # Candidate 2: Skier activation lift (override if stronger)
    skier_corr = next((r for r in correlates if r['segment'] == 'Skier'), None)
    if skier_corr and skier_corr['with_n'] >= 5 and skier_corr['lift'] >= 1.5:
        lift_disp = f"{skier_corr['lift']}×"
        skier_insight = (
            f"Skiers activate at {skier_corr['act_pct']}% vs {baseline_pct_val}% baseline "
            f"— {lift_disp} the rate of other rider types."
        )
        if insight is None:
            insight = skier_insight
    # Candidate 3: Power user density (fallback)
    if insight is None and len(power_users) >= 2:
        insight = (
            f"Power users maintain {power_profile['pow_avg_friends']}× more friend connections "
            f"on average ({power_profile['pow_avg_friends']} vs {avg_friends} overall)."
        )

    # ── Section 9: User Status ────────────────────────────────────────────
    status = dict(
        tracked_attributes  = 10,
        reliable_n          = 8,   # Green: state, pass, rider, wl, planned, avail, friends, trips
        confidence          = 'Growing',
        caveats             = [
            'Country is derived from state information — no canonical country field exists.',
            'Equipment ownership uses EquipmentSetup rows (behavioral signal). '
            'User.equipment_status excluded — defaults to "have own equipment".',
        ],
    )

    return render_template(
        "admin_user_intelligence.html",
        active_tab          = "user-intelligence",
        now                 = _fmt_admin_now(),
        kpis                = kpis,
        top_states          = top_states,
        top_countries       = top_countries,
        state_coverage      = state_coverage,
        pass_table          = pass_table,
        rider_table         = rider_table,
        unknown_rt          = unknown_rt,
        top_wishlisted      = top_wishlisted,
        top_planned         = top_planned,
        social              = social,
        correlates          = correlates,
        corr_max            = corr_max,
        baseline_pct        = baseline_pct_val,
        baseline_act        = baseline_act,
        power_profile       = power_profile,
        insight             = insight,
        status              = status,
    )


@app.route("/admin/posthog-funnels")
@login_required
@admin_required
def admin_posthog_funnels():
    """PostHog Funnels Dashboard v1 — feasibility audit + activation/social funnels."""
    import services.posthog_query as ph_q

    now_dt       = datetime.utcnow()
    no_pass_slugs = {"no_pass", "no_pass_yet", "none", ""}

    def _has_real_pass(u):
        return any(
            r.strip().lower() not in no_pass_slugs
            for r in (u.pass_type or "").split(",") if r.strip()
        )

    # ── DB proxy counts (always available, used for audit + estimates) ────
    ns_users = User.query.filter(User.is_seeded == False).all()
    ns_ids   = [u.id for u in ns_users]

    def _distinct_ids(query_obj):
        return set(r[0] for r in query_obj.all()) if ns_ids else set()

    trip_uids   = _distinct_ids(
        db.session.query(SkiTrip.user_id.distinct()).filter(SkiTrip.user_id.in_(ns_ids))
    )
    friend_uids = _distinct_ids(
        db.session.query(Friend.user_id.distinct()).filter(Friend.user_id.in_(ns_ids))
    )
    invite_uids = _distinct_ids(
        db.session.query(InviteToken.inviter_id.distinct())
        .filter(InviteToken.inviter_id.in_(ns_ids))
    )

    db_counts = {
        "signup_started":       len(ns_users),
        "signup_completed":     sum(
            1 for u in ns_users
            if (u.login_count or 0) > 0 or u.lifecycle_stage in ("onboarding", "active")
        ),
        "onboarding_completed": sum(1 for u in ns_users if u.is_core_profile_complete),
        "pass_added":           sum(1 for u in ns_users if _has_real_pass(u)),
        "availability_added":   sum(
            1 for u in ns_users
            if isinstance(u.open_dates, list) and len(u.open_dates) > 0
        ),
        "wishlist_added":       sum(
            1 for u in ns_users
            if isinstance(u.wish_list_resorts, list) and len(u.wish_list_resorts) > 0
        ),
        "trip_created":         len(trip_uids),
        "friend_connected":     len(friend_uids),
        "invite_generated":     len(invite_uids),
    }

    # ── PostHog audit (returns quickly; cached 30 min) ────────────────────
    ph_audit    = ph_q.fetch_event_audit()
    has_ph_creds = ph_audit["has_credentials"]
    ph_event_map = {e["name"]: e for e in ph_audit.get("events", [])}

    # ── Build audit rows ──────────────────────────────────────────────────
    INSTRUMENTATION_DATE = "2026-05-29"
    SERVER_EVENTS = {
        "pass_added", "availability_added", "wishlist_added",
        "trip_created", "friend_connected", "invite_generated",
    }
    AUTH_EVENTS = {"signup_started", "signup_completed", "onboarding_completed"}
    LABEL_MAP = {
        "signup_started":       "Signup Started",
        "signup_completed":     "Signup Completed",
        "onboarding_completed": "Onboarding Completed",
        "pass_added":           "Pass Added",
        "availability_added":   "Availability Added",
        "wishlist_added":       "Wishlist Added",
        "trip_created":         "Trip Created",
        "friend_connected":     "Friend Connected",
        "invite_generated":     "Invite Generated",
    }

    audit_rows = []
    for ev in [
        "signup_started", "signup_completed", "onboarding_completed",
        "pass_added", "availability_added", "wishlist_added",
        "trip_created", "friend_connected", "invite_generated",
    ]:
        ph_ev = ph_event_map.get(ev)
        if ph_ev and ph_ev.get("found"):
            audit_rows.append({
                "name":        ev,
                "label":       LABEL_MAP[ev],
                "user_count":  ph_ev["user_count"],
                "event_count": ph_ev["event_count"],
                "earliest":    ph_ev["earliest"],
                "latest":      ph_ev["latest"],
                "source":      "posthog",
                "deployed":    True,
            })
        else:
            audit_rows.append({
                "name":        ev,
                "label":       LABEL_MAP[ev],
                "user_count":  db_counts.get(ev, 0),
                "event_count": db_counts.get(ev, 0),
                "earliest":    INSTRUMENTATION_DATE if ev in SERVER_EVENTS else "auth flow",
                "latest":      INSTRUMENTATION_DATE,
                "source":      "db_proxy",
                "deployed":    ev in SERVER_EVENTS or ev in AUTH_EVENTS,
            })

    # ── Confidence + data gate ────────────────────────────────────────────
    if has_ph_creds and ph_audit["sufficient_data"]:
        confidence      = ph_audit["confidence"]
        sufficient_data = True
        data_source     = "posthog"
    else:
        confidence      = "low"
        sufficient_data = False
        data_source     = "db_proxy"

    # ── PostHog funnels (only when data is sufficient) ────────────────────
    funnel1 = ph_q.fetch_activation_funnel() if sufficient_data else None
    funnel2 = ph_q.fetch_social_funnel()     if sufficient_data else None
    ttv     = ph_q.fetch_time_to_value()     if sufficient_data else None

    # ── DB proxy funnels (shown as estimates when PostHog data is thin) ───
    def _build_proxy_funnel(steps):
        top = db_counts.get(steps[0][0], 1) or 1
        result = []
        prev = None
        for ev, label in steps:
            uc = db_counts.get(ev, 0)
            result.append({
                "label":      label,
                "user_count": uc,
                "pct_of_top": round(uc / top * 100) if top else 0,
                "vs_prev":    round(uc / prev * 100) if prev else None,
            })
            prev = uc or 1
        return result

    db_funnel1 = _build_proxy_funnel([
        ("signup_started",       "Signup Started"),
        ("signup_completed",     "Signup Completed"),
        ("onboarding_completed", "Onboarding Completed"),
        ("pass_added",           "Pass Added"),
        ("availability_added",   "Availability Added"),
        ("wishlist_added",       "Wishlist Added"),
        ("trip_created",         "Trip Created"),
    ])
    db_funnel2 = _build_proxy_funnel([
        ("signup_completed",  "Signup Completed"),
        ("friend_connected",  "Friend Connected"),
        ("invite_generated",  "Invite Generated"),
    ])

    # ── Biggest dropoff (from DB proxy funnel) ────────────────────────────
    biggest_drop = None
    biggest_delta = 0
    for i in range(1, len(db_funnel1)):
        prev_uc = db_funnel1[i - 1]["user_count"] or 1
        curr_uc = db_funnel1[i]["user_count"]
        delta = round((prev_uc - curr_uc) / prev_uc * 100) if prev_uc else 0
        if delta > biggest_delta:
            biggest_delta = delta
            biggest_drop  = (db_funnel1[i - 1]["label"], db_funnel1[i]["label"], delta)

    # ── Activation Insight ────────────────────────────────────────────────
    if biggest_drop:
        _f, _t, _d = biggest_drop
        insight = (
            f'"{_f} → {_t}" is the largest activation dropoff, '
            f"losing {_d}% of users at that step."
        )
    else:
        insight = "Insufficient data to generate an activation insight yet."

    # Days since instrumentation deployed
    from datetime import date as _date
    _deploy = _date(2026, 5, 29)
    days_live = (now_dt.date() - _deploy).days

    return render_template(
        "admin_posthog_funnels.html",
        active_tab           = "posthog_funnels",
        now                  = _fmt_admin_now(),
        audit_rows           = audit_rows,
        confidence           = confidence,
        sufficient_data      = sufficient_data,
        data_source          = data_source,
        has_ph_creds         = has_ph_creds,
        ph_error             = ph_audit.get("error"),
        funnel1              = funnel1,
        funnel2              = funnel2,
        ttv                  = ttv,
        db_funnel1           = db_funnel1,
        db_funnel2           = db_funnel2,
        insight              = insight,
        biggest_drop         = biggest_drop,
        ph_events_tracked    = 9,
        ph_events_active     = len(SERVER_EVENTS),
        instrumentation_date = INSTRUMENTATION_DATE,
        total_users          = len(ns_users),
        days_live            = days_live,
    )


@app.route("/admin/mountain-intelligence")
@login_required
@admin_required
def admin_mountain_intelligence():
    """Mountain Intelligence v1 — forward-looking demand: attention, conversion, latent demand, social potential, partnerships."""
    from collections import defaultdict
    from datetime import datetime as _dt

    now_dt = datetime.utcnow()
    VIEW_TRACKING_LAUNCH = datetime(2026, 5, 27)
    tracking_age_days = (now_dt - VIEW_TRACKING_LAUNCH).days

    # ── Non-seeded user base ──────────────────────────────────────────────
    all_users = User.query.filter(User.is_seeded == False).all()
    ns_ids    = [u.id for u in all_users]
    total_users = len(all_users)

    # ── Signal 2: Wishlist — per-resort count + distinct user sets ────────
    wl_count = defaultdict(int)
    wl_users = defaultdict(set)
    for u in all_users:
        for rid in (u.wish_list_resorts or []):
            wl_count[rid]  += 1
            wl_users[rid].add(u.id)

    # ── Signal 3/4: Trips + Group Trips per resort ────────────────────────
    trip_rows = db.session.query(
        SkiTrip.resort_id,
        func.count(SkiTrip.id).label("trips"),
        func.sum(db.cast(SkiTrip.is_group_trip, db.Integer)).label("group_trips"),
    ).filter(
        SkiTrip.user_id.in_(ns_ids),
        SkiTrip.resort_id != None,
    ).group_by(SkiTrip.resort_id).all()

    trip_count = {r[0]: r[1] for r in trip_rows}
    gt_count   = {r[0]: int(r[2] or 0) for r in trip_rows}

    # Wishlist→Trip conversion: % of wishlisters who also tripped here
    trip_user_rows = db.session.query(SkiTrip.resort_id, SkiTrip.user_id).filter(
        SkiTrip.user_id.in_(ns_ids), SkiTrip.resort_id != None
    ).all()
    trip_users = defaultdict(set)
    for row in trip_user_rows:
        trip_users[row[0]].add(row[1])

    # Concentration check: user_id accounting for >50% of a resort's trips
    trip_per_user_resort = db.session.query(
        SkiTrip.resort_id, SkiTrip.user_id, func.count(SkiTrip.id).label("cnt")
    ).filter(
        SkiTrip.user_id.in_(ns_ids), SkiTrip.resort_id != None
    ).group_by(SkiTrip.resort_id, SkiTrip.user_id).all()

    concentration_flags = set()  # resort_ids where one user has >50% of trips
    resort_user_trip_counts = defaultdict(list)
    for row in trip_per_user_resort:
        resort_user_trip_counts[row[0]].append(row[2])
    for rid, counts in resort_user_trip_counts.items():
        total = sum(counts)
        if total >= 3 and max(counts) / total > 0.5:
            concentration_flags.add(rid)

    # ── Signal 1: Page views per resort ──────────────────────────────────
    view_rows = db.session.query(
        MountainPageView.resort_id,
        func.count(MountainPageView.id).label("views"),
    ).group_by(MountainPageView.resort_id).all()
    view_count = {r[0]: r[1] for r in view_rows}

    # ── Signal 8: TripInviteToken per resort (via trip_id → resort_id) ───
    from models import TripInviteToken
    invite_rows = db.session.query(
        SkiTrip.resort_id, func.count(TripInviteToken.id).label("invites")
    ).join(TripInviteToken, TripInviteToken.trip_id == SkiTrip.id
    ).filter(SkiTrip.resort_id != None
    ).group_by(SkiTrip.resort_id).all()
    invite_count = {r[0]: r[1] for r in invite_rows}

    # ── Signal 7: Pass compatibility — users with a qualifying pass ───────
    from models import ResortPass
    pass_map_rows = db.session.query(ResortPass.resort_id, ResortPass.pass_name).all()
    resort_passes = defaultdict(set)
    for row in pass_map_rows:
        resort_passes[row[0]].add(row[1].lower() if row[1] else "")

    pass_slug_map = {
        "epic": "epic", "ikon": "ikon", "indy": "indy",
        "mountaincollective": "mountaincollective",
        "powderalliance": "powderalliance",
        "freedom": "freedom", "skicalifornia": "skicalifornia",
    }
    addressable_count = defaultdict(int)
    for u in all_users:
        user_pass = (u.pass_type or "").lower().replace("_", "").replace(" ", "")
        canonical = pass_slug_map.get(user_pass)
        if canonical:
            for rid, passes in resort_passes.items():
                if canonical in passes:
                    addressable_count[rid] += 1

    # ── All resort IDs with any signal ────────────────────────────────────
    all_active_ids = set(wl_count) | set(trip_count) | set(view_count)

    # ── Fetch resort names in one query ───────────────────────────────────
    resort_map = {}
    if all_active_ids:
        for r in Resort.query.filter(Resort.id.in_(all_active_ids)).all():
            resort_map[r.id] = r.name

    # ── Build unified per-resort rows ─────────────────────────────────────
    rows = []
    for rid in all_active_ids:
        wl  = wl_count.get(rid, 0)
        tr  = trip_count.get(rid, 0)
        gt  = gt_count.get(rid, 0)
        vw  = view_count.get(rid, 0)
        inv = invite_count.get(rid, 0)
        adr = addressable_count.get(rid, 0)
        conc = rid in concentration_flags

        wl_uid = wl_users.get(rid, set())
        tr_uid = trip_users.get(rid, set())
        wl_trip_conv = round(len(wl_uid & tr_uid) / len(wl_uid) * 100) if wl_uid else None
        tr_gt_conv   = round(gt / tr * 100) if tr else None

        # Conversion % with sample guards
        v2w_pct  = round(wl / vw * 100) if vw >= 5 and wl > 0 else None
        w2t_pct  = round(len(wl_uid & tr_uid) / len(wl_uid) * 100) if len(wl_uid) >= 3 else None
        overall_pct = round(tr / vw * 100) if vw >= 5 and tr > 0 else None

        # Partnership score
        score = vw * 1 + wl * 5 + tr * 10 + gt * 15

        rows.append({
            "id":           rid,
            "name":         resort_map.get(rid, f"Resort {rid}"),
            "views":        vw,
            "wishlists":    wl,
            "trips":        tr,
            "group_trips":  gt,
            "invites":      inv,
            "addressable":  adr,
            "concentration": conc,
            "wl_trip_conv": wl_trip_conv,
            "tr_gt_conv":   tr_gt_conv,
            "v2w_pct":      v2w_pct,
            "w2t_pct":      w2t_pct,
            "overall_pct":  overall_pct,
            "score":        score,
        })

    # ── Section 0: KPIs ───────────────────────────────────────────────────
    total_wishlists   = sum(wl_count.values())
    total_trips_count = sum(trip_count.values())
    total_group_trips = sum(gt_count.values())
    total_views       = sum(view_count.values())
    distinct_view_resorts = len(view_count)
    avg_views = round(total_views / distinct_view_resorts, 1) if distinct_view_resorts else 0

    mi_kpis = dict(
        resorts_with_activity = len(all_active_ids),
        total_views           = total_views,
        total_wishlists       = total_wishlists,
        total_planned_trips   = total_trips_count,
        total_group_trips     = total_group_trips,
        avg_views_per_resort  = avg_views,
    )

    # ── Section 1: Attention Leaderboard (views ≥1) ───────────────────────
    attention_rows = sorted(
        [r for r in rows if r["views"] >= 1],
        key=lambda r: -r["views"]
    )[:15]
    total_views_sum = total_views or 1  # avoid div/0 for share %

    # ── Section 2: Demand Leaderboard ────────────────────────────────────
    demand_wish = sorted(
        [r for r in rows if r["wishlists"] > 0],
        key=lambda r: (-r["wishlists"], -r["trips"])
    )[:10]
    demand_trips = sorted(
        [r for r in rows if r["trips"] > 0],
        key=lambda r: (-r["trips"], -r["wishlists"])
    )[:10]

    # ── Section 3: Conversion Funnel — resorts with any conversion data ───
    funnel_rows = sorted(
        [r for r in rows if r["trips"] > 0 or r["wishlists"] >= 3 or r["views"] >= 5],
        key=lambda r: -(r["trips"] * 10 + r["wishlists"] * 5 + r["views"])
    )[:10]

    # ── Section 4: Latent Demand (wish≥2, trips=0) ────────────────────────
    latent_rows = sorted(
        [r for r in rows if r["wishlists"] >= 2 and r["trips"] == 0],
        key=lambda r: (-r["wishlists"], -r["views"])
    )

    # ── Section 5: Overperformers (trips≥5, views < trips/2) ─────────────
    overperformer_rows = sorted(
        [r for r in rows if r["trips"] >= 5 and r["views"] < r["trips"] / 2],
        key=lambda r: -r["trips"]
    )

    # ── Section 6: Social Potential ───────────────────────────────────────
    # Signal = multiple distinct wishlist users + group trips + trip invites
    social_rows = sorted(
        [r for r in rows if r["group_trips"] > 0 or r["invites"] > 0 or r["wishlists"] >= 2],
        key=lambda r: -(r["group_trips"] * 15 + r["invites"] * 5 + r["wishlists"] * 3)
    )[:10]

    # ── Section 7: Partnership Watchlist ─────────────────────────────────
    partnership_rows = sorted(
        [r for r in rows],
        key=lambda r: -r["score"]
    )[:15]

    # ── Section 8: Mountain Insight ───────────────────────────────────────
    mi_insight = None

    # Rule 1: Aspiration gap — most-wishlisted with 0 trips
    aspiration_gap = sorted(
        [r for r in rows if r["wishlists"] >= 2 and r["trips"] == 0],
        key=lambda r: -r["wishlists"]
    )
    if aspiration_gap:
        top = aspiration_gap[0]
        mi_insight = (
            f"{top['name']} has {top['wishlists']} wishlists but 0 planned trips — "
            f"high aspiration with no commitment yet."
        )

    # Rule 2: Conversion leader — highest wish→trip with ≥2 wishlist users (only if no rule 1)
    if not mi_insight:
        conv_candidates = sorted(
            [r for r in rows if r["wl_trip_conv"] is not None and len(wl_users.get(r["id"], set())) >= 2],
            key=lambda r: -r["wl_trip_conv"]
        )
        if conv_candidates:
            top = conv_candidates[0]
            mi_insight = (
                f"{top['name']} is converting {top['wl_trip_conv']}% of wishlisters "
                f"into planned trips — highest conversion of any resort."
            )

    # Rule 3: Trip leader — skip if concentration flag
    if not mi_insight:
        trip_leaders = sorted(
            [r for r in rows if r["trips"] > 0 and not r["concentration"]],
            key=lambda r: -r["trips"]
        )
        if trip_leaders:
            top = trip_leaders[0]
            mi_insight = (
                f"{top['name']} leads all resorts with {top['trips']} planned trip{'s' if top['trips'] != 1 else ''}."
            )
        else:
            # Fall back to trip leader with concentration note
            trip_all = sorted([r for r in rows if r["trips"] > 0], key=lambda r: -r["trips"])
            if trip_all:
                top = trip_all[0]
                mi_insight = (
                    f"{top['name']} has the most planned trips ({top['trips']}) — "
                    f"note: high trip concentration from a small number of users."
                )

    # Rule 4: Latent demand with group activity
    if not mi_insight and latent_rows:
        for r in latent_rows:
            if r["group_trips"] > 0:
                mi_insight = (
                    f"{r['name']} has {r['wishlists']} wishlists and an active group trip — "
                    f"social demand is building."
                )
                break

    # Rule 5: Attention signal
    if not mi_insight and attention_rows:
        top = attention_rows[0]
        mi_insight = f"{top['name']} is attracting the most page traffic with {top['views']} view{'s' if top['views'] != 1 else ''}."

    # ── Section 9: Status ─────────────────────────────────────────────────
    if tracking_age_days < 30:
        mi_confidence = "Low"
    elif tracking_age_days < 180:
        mi_confidence = "Growing"
    else:
        mi_confidence = "High"

    total_resorts = Resort.query.filter_by(is_active=True).count()
    reliable_signals = 3  # Views (growing), Wishlist, Trips
    tracked_signals  = 7  # all 7 non-red signals

    mi_status = dict(
        total_resorts        = total_resorts,
        resorts_with_activity = len(all_active_ids),
        confidence           = mi_confidence,
        tracking_age_days    = tracking_age_days,
        tracked_signals      = tracked_signals,
        reliable_signals     = reliable_signals,
    )

    return render_template(
        "admin_mountain_intelligence.html",
        active_tab         = "mountain_intelligence",
        now                = _fmt_admin_now(),
        mi_kpis            = mi_kpis,
        attention_rows     = attention_rows,
        total_views_sum    = total_views_sum,
        demand_wish        = demand_wish,
        demand_trips       = demand_trips,
        funnel_rows        = funnel_rows,
        latent_rows        = latent_rows,
        overperformer_rows = overperformer_rows,
        social_rows        = social_rows,
        partnership_rows   = partnership_rows,
        mi_insight         = mi_insight,
        mi_status          = mi_status,
        rows               = rows,
    )


@app.route("/admin/activation-intel")
@login_required
@admin_required
def admin_activation_intel():
    """Activation Intelligence v1 — funnel, milestones, correlations."""
    import json as _json
    from collections import Counter

    now_str = _admin_now().strftime("%b %d, %Y at %H:%M %Z")

    def pct(n, d):
        return round(n / d * 100) if d else 0

    def _nonempty(val):
        if val is None:
            return False
        if isinstance(val, list):
            return len(val) > 0
        if isinstance(val, str):
            try:
                p = _json.loads(val)
                return isinstance(p, list) and len(p) > 0
            except Exception:
                return False
        return False

    _no_pass = {'no_pass', 'no_pass_yet', 'none', ''}

    def _real_pass(pt):
        for s in (pt or '').split(','):
            if s.strip().lower() not in _no_pass:
                return True
        return False

    # ── Q1: all users (activation fields only) ───────────────────────────
    user_rows = db.session.execute(db.text(
        'SELECT id, lifecycle_stage, pass_type, open_dates, wish_list_resorts FROM "user"'
    )).fetchall()
    total = len(user_rows)

    # ── Q2: trip owner IDs ───────────────────────────────────────────────
    trip_owner_ids = {r[0] for r in db.session.execute(
        db.text("SELECT DISTINCT user_id FROM ski_trip")
    )}

    # ── Q3: trip counts per user ─────────────────────────────────────────
    trip_counts_map = {r[0]: r[1] for r in db.session.execute(
        db.text("SELECT user_id, COUNT(*) FROM ski_trip GROUP BY user_id")
    )}

    # ── Q4: friend counts per user ───────────────────────────────────────
    friend_counts_raw = {r[0]: r[1] for r in db.session.execute(
        db.text("SELECT user_id, COUNT(*) FROM friend GROUP BY user_id")
    )}
    friend_ids = set(friend_counts_raw.keys())

    # ── Q5: invite generator IDs (generic + trip) ────────────────────────
    invite_ids = {r[0] for r in db.session.execute(
        db.text("SELECT inviter_id FROM invite_token "
                "UNION SELECT inviter_user_id FROM trip_invite_token")
    )}

    # ── Q6: accepted trip participants ───────────────────────────────────
    trip_guest_ids = {r[0] for r in db.session.execute(
        db.text("SELECT DISTINCT user_id FROM ski_trip_participant WHERE status='accepted'")
    )}

    has_trip_ids = trip_owner_ids | trip_guest_ids

    # ── Derive per-user signal sets ──────────────────────────────────────
    pass_ids     = set()
    avail_ids    = set()
    wishlist_ids = set()
    all_ids      = set()
    lifecycle_ct = Counter()

    for uid, ls, pt, od, wl in user_rows:
        all_ids.add(uid)
        lifecycle_ct[ls or 'new'] += 1
        if _real_pass(pt):    pass_ids.add(uid)
        if _nonempty(od):     avail_ids.add(uid)
        if _nonempty(wl):     wishlist_ids.add(uid)

    active_ids = {uid for uid, ls, *_ in user_rows if ls == 'active'}

    n_signup   = total
    n_onboard  = lifecycle_ct.get('active', 0)
    n_pass     = len(pass_ids)
    n_avail    = len(avail_ids)
    n_wishlist = len(wishlist_ids)
    n_trip     = len(has_trip_ids)
    n_friend   = len(friend_ids)
    n_invite   = len(invite_ids)
    n_trip_w_friends = len(has_trip_ids & friend_ids)
    n_active = len(active_ids)

    # ── Section 0: KPIs ──────────────────────────────────────────────────
    kpis = [
        {'label': 'Total Users',            'value': total,                        'sub': 'all time'},
        {'label': 'Activated Users',         'value': n_onboard,                    'sub': 'lifecycle = active'},
        {'label': 'Activation Rate',         'value': f'{pct(n_onboard, total)}%',  'sub': 'of all users'},
        {'label': 'Pass Adoption',           'value': f'{pct(n_pass, total)}%',     'sub': f'{n_pass} of {total}'},
        {'label': 'Availability Adoption',   'value': f'{pct(n_avail, total)}%',    'sub': f'{n_avail} of {total}'},
        {'label': 'Wishlist Adoption',       'value': f'{pct(n_wishlist, total)}%', 'sub': f'{n_wishlist} of {total}'},
        {'label': 'Trip Creation Rate',      'value': f'{pct(n_trip, total)}%',     'sub': f'{n_trip} of {total}'},
        {'label': 'Friend Connection Rate',  'value': f'{pct(n_friend, total)}%',   'sub': f'{n_friend} of {total}'},
        {'label': 'Invite Generation Rate',  'value': f'{pct(n_invite, total)}%',   'sub': f'{n_invite} of {total}'},
        {'label': 'Trip Users With Friends', 'value': n_trip_w_friends,             'sub': f'of {n_trip} trip users'},
    ]

    # ── Section 1: Funnel ────────────────────────────────────────────────
    funnel = [
        {'label': 'Signup Completed',     'n': n_signup,   'pct': pct(n_signup, total)},
        {'label': 'Onboarding Completed', 'n': n_onboard,  'pct': pct(n_onboard, total)},
        {'label': 'Pass Added',           'n': n_pass,     'pct': pct(n_pass, total)},
        {'label': 'Availability Added',   'n': n_avail,    'pct': pct(n_avail, total)},
        {'label': 'Wishlist Added',       'n': n_wishlist, 'pct': pct(n_wishlist, total)},
        {'label': 'Trip Created',         'n': n_trip,     'pct': pct(n_trip, total)},
        {'label': 'Friend Connected',     'n': n_friend,   'pct': pct(n_friend, total)},
    ]

    # ── Section 2: Milestones ────────────────────────────────────────────
    milestones = [
        {'label': 'Users with Pass',
         'all_n': n_pass,     'all_pct': pct(n_pass, total),
         'act_n': len(pass_ids & active_ids),     'act_pct': pct(len(pass_ids & active_ids), n_active)},
        {'label': 'Users with Trip',
         'all_n': n_trip,     'all_pct': pct(n_trip, total),
         'act_n': len(has_trip_ids & active_ids),  'act_pct': pct(len(has_trip_ids & active_ids), n_active)},
        {'label': 'Users with Friends',
         'all_n': n_friend,   'all_pct': pct(n_friend, total),
         'act_n': len(friend_ids & active_ids),    'act_pct': pct(len(friend_ids & active_ids), n_active)},
        {'label': 'Users with Wishlist',
         'all_n': n_wishlist, 'all_pct': pct(n_wishlist, total),
         'act_n': len(wishlist_ids & active_ids),  'act_pct': pct(len(wishlist_ids & active_ids), n_active)},
        {'label': 'Users with Availability',
         'all_n': n_avail,    'all_pct': pct(n_avail, total),
         'act_n': len(avail_ids & active_ids),     'act_pct': pct(len(avail_ids & active_ids), n_active)},
    ]

    # ── Section 3: Segments ──────────────────────────────────────────────
    seg_ct = {'new': 0, 'onboarding': 0, 'activated': 0, 'engaged': 0, 'power': 0}
    for uid, ls, *_ in user_rows:
        score = (
            (1 if uid in pass_ids     else 0) +
            (1 if uid in avail_ids    else 0) +
            (1 if uid in wishlist_ids else 0) +
            (1 if uid in has_trip_ids else 0) +
            (1 if uid in friend_ids   else 0)
        )
        if ls == 'active':
            if score >= 4:    seg_ct['power']     += 1
            elif score >= 2:  seg_ct['engaged']   += 1
            else:             seg_ct['activated'] += 1
        elif ls == 'onboarding':
            seg_ct['onboarding'] += 1
        else:
            seg_ct['new'] += 1

    segments = [
        {'label': 'New',        'n': seg_ct['new'],        'pct': pct(seg_ct['new'], total),
         'desc': 'Signed up, not yet in onboarding',  'color': '#9E958A'},
        {'label': 'Onboarding', 'n': seg_ct['onboarding'], 'pct': pct(seg_ct['onboarding'], total),
         'desc': 'Currently in onboarding flow',       'color': '#B07D2E'},
        {'label': 'Activated',  'n': seg_ct['activated'],  'pct': pct(seg_ct['activated'], total),
         'desc': 'Active, 0–1 milestones completed',   'color': '#5C7A9E'},
        {'label': 'Engaged',    'n': seg_ct['engaged'],    'pct': pct(seg_ct['engaged'], total),
         'desc': 'Active, 2–3 milestones completed',   'color': '#4A7C59'},
        {'label': 'Power',      'n': seg_ct['power'],      'pct': pct(seg_ct['power'], total),
         'desc': 'Active, 4+ milestones completed',    'color': '#5C1219'},
    ]

    # ── Section 4: Drop-offs ─────────────────────────────────────────────
    dropoff_pairs = [
        ('Signup', n_signup,   'Onboarding', n_onboard),
        ('Onboarding', n_onboard, 'Trip Created', n_trip),
        ('Pass Added', n_pass, 'Availability', n_avail),
        ('Trip Created', n_trip, 'Availability', n_avail),
    ]
    dropoffs = []
    for fl, fn, tl, tn in dropoff_pairs:
        if fn > 0:
            dp = pct(fn - tn, fn)
            dropoffs.append({
                'from': fl, 'from_n': fn,
                'to':   tl, 'to_n':   tn,
                'drop_pct': dp,
                'severity': 'high' if dp >= 60 else ('mid' if dp >= 30 else 'low'),
            })
    dropoffs.sort(key=lambda x: -x['drop_pct'])

    # ── Section 5: Correlations ──────────────────────────────────────────
    def _tpu(uid_set):
        if not uid_set:
            return 0.0
        return sum(trip_counts_map.get(u, 0) for u in uid_set) / len(uid_set)

    corr_inputs = [
        ('Has Friends',      friend_ids,   all_ids - friend_ids),
        ('Has Wishlist',     wishlist_ids, all_ids - wishlist_ids),
        ('Has Availability', avail_ids,    all_ids - avail_ids),
        ('Has Pass',         pass_ids,     all_ids - pass_ids),
        ('Generated Invite', invite_ids,   all_ids - invite_ids),
    ]
    correlations = []
    for label, has_set, no_set in corr_inputs:
        has_tpu = round(_tpu(has_set), 2)
        no_tpu  = round(_tpu(no_set),  2)
        ratio   = round(has_tpu / no_tpu, 1) if no_tpu > 0.05 else (has_tpu or 0.0)
        cls     = 'green' if ratio >= 2.5 else ('yellow' if ratio >= 1.2 else 'red')
        correlations.append({
            'label':   label,
            'has_n':   len(has_set),
            'no_n':    len(no_set),
            'has_tpu': has_tpu,
            'no_tpu':  no_tpu,
            'ratio':   ratio,
            'class':   cls,
        })
    correlations.sort(key=lambda x: -x['ratio'])

    # ── Section 6: Dynamic Insight ───────────────────────────────────────
    green_corrs = [c for c in correlations if c['class'] == 'green']
    if green_corrs:
        top = green_corrs[0]
        rstr = f"{int(top['ratio'])}×" if top['ratio'] == int(top['ratio']) else f"{top['ratio']}×"
        raw = f"Users {top['label'].lower()} create {rstr} more trips."
    elif n_friend > 0:
        raw = f"{pct(n_friend, total)}% of users have connected with friends."
    else:
        raw = "Collecting more activation data."
    insight = raw[0].upper() + raw[1:] if raw else raw

    # ── Section 7: Status ────────────────────────────────────────────────
    status = {
        'tracked':    9,
        'green':      7,
        'yellow':     2,
        'red':        1,
        'confidence': 'Growing',
    }

    return render_template(
        "admin_activation_intel.html",
        active_tab   = 'activation_intel',
        now          = now_str,
        total        = total,
        kpis         = kpis,
        funnel       = funnel,
        milestones   = milestones,
        segments     = segments,
        dropoffs     = dropoffs,
        correlations = correlations,
        insight      = insight,
        status       = status,
    )


@app.route("/admin/growth-intel")
@login_required
@admin_required
def admin_growth_intel():
    """Growth Intelligence v1 — network, invite, and growth signals."""
    from datetime import datetime as _dt, timedelta
    from collections import defaultdict, Counter

    now_str = _admin_now().strftime("%b %d, %Y at %H:%M %Z")
    now = _dt.utcnow()
    l7_cutoff  = now - timedelta(days=7)
    l30_cutoff = now - timedelta(days=30)
    l90_cutoff = now - timedelta(days=90)

    def pct(n, d):
        return round(n / d * 100) if d else 0

    def _month_sort_key(k):
        if k == 'Undated':
            return (9999, 99)
        try:
            d = _dt.strptime(k, "%b %Y")
            return (d.year, d.month)
        except Exception:
            return (9998, 0)

    # ── Q1: user growth fields ────────────────────────────────────────────
    user_rows = db.session.execute(db.text(
        'SELECT id, created_at, last_active_at, invited_by_user_id, '
        'first_connection_at, first_trip_created_at FROM "user"'
    )).fetchall()
    total = len(user_rows)

    new_l7         = sum(1 for r in user_rows if r[1] and r[1] >= l7_cutoff)
    new_l30        = sum(1 for r in user_rows if r[1] and r[1] >= l30_cutoff)
    new_l90        = sum(1 for r in user_rows if r[1] and r[1] >= l90_cutoff)
    null_created   = sum(1 for r in user_rows if r[1] is None)
    wau            = sum(1 for r in user_rows if r[2] and r[2] >= l7_cutoff)
    mau            = sum(1 for r in user_rows if r[2] and r[2] >= l30_cutoff)
    null_last_active = sum(1 for r in user_rows if r[2] is None)
    referred       = sum(1 for r in user_rows if r[3] is not None)
    fly_friend_users = sum(1 for r in user_rows if r[4] is not None)
    fly_trip_users   = sum(1 for r in user_rows if r[5] is not None)

    # Monthly user cohorts
    monthly_cohorts = defaultdict(int)
    for r in user_rows:
        key = r[1].strftime("%b %Y") if r[1] else 'Undated'
        monthly_cohorts[key] += 1
    cohort_max = max(monthly_cohorts.values()) if monthly_cohorts else 1
    cohort_list = [
        {'label': k, 'count': v, 'pct': pct(v, cohort_max), 'is_undated': k == 'Undated'}
        for k, v in sorted(monthly_cohorts.items(), key=lambda x: _month_sort_key(x[0]))
    ]

    # ── Q2: friend network ────────────────────────────────────────────────
    friend_rows = db.session.execute(db.text(
        "SELECT user_id, created_at FROM friend"
    )).fetchall()
    total_friend_rows  = len(friend_rows)
    unique_pairs       = total_friend_rows // 2
    friend_counts      = defaultdict(int)
    for r in friend_rows:
        friend_counts[r[0]] += 1
    connected_users        = len(friend_counts)
    avg_friends_all        = round(total_friend_rows / total, 2) if total else 0
    avg_friends_connected  = round(total_friend_rows / connected_users, 2) if connected_users else 0

    monthly_friends = defaultdict(int)
    for r in friend_rows:
        if r[1]:
            monthly_friends[r[1].strftime("%b %Y")] += 1
    f_monthly_max = max((v // 2 for v in monthly_friends.values()), default=1)
    friend_monthly_list = [
        {'label': k, 'pairs': v // 2, 'pct': pct(v // 2, max(f_monthly_max, 1))}
        for k, v in sorted(monthly_friends.items(), key=lambda x: _month_sort_key(x[0]))
    ]

    dist_raw = Counter(friend_counts.values())
    friend_dist = [{'friends': k, 'users': v} for k, v in sorted(dist_raw.items())]
    dist_max = max((d['users'] for d in friend_dist), default=1)
    for d in friend_dist:
        d['pct'] = pct(d['users'], dist_max)

    if friend_counts:
        mc_id  = max(friend_counts, key=lambda u: friend_counts[u])
        mc_ct  = friend_counts[mc_id]
        mc_row = db.session.execute(
            db.text('SELECT first_name, last_name FROM "user" WHERE id = :uid'),
            {'uid': mc_id}
        ).fetchone()
        most_connected_name = f"{mc_row[0]} {mc_row[1]}" if mc_row else f"User {mc_id}"
    else:
        most_connected_name, mc_ct = "—", 0

    # ── Q3: invite performance ────────────────────────────────────────────
    inv_total, inv_used = db.session.execute(db.text(
        "SELECT COUNT(*), COUNT(used_at) FROM invite_token"
    )).fetchone()
    inv_total = inv_total or 0
    inv_used  = inv_used  or 0

    trip_inv_total, trip_inv_used = db.session.execute(db.text(
        "SELECT COUNT(*), COUNT(used_at) FROM trip_invite_token"
    )).fetchone()
    trip_inv_total = trip_inv_total or 0
    trip_inv_used  = trip_inv_used  or 0

    mel_row = db.session.execute(db.text(
        "SELECT "
        "  COUNT(*) FILTER (WHERE event_name = 'trip.invite.created'), "
        "  COUNT(*) FILTER (WHERE event_name = 'trip.invite.accepted'), "
        "  COUNT(*) FILTER (WHERE event_name = 'friend.request.created'), "
        "  COUNT(*) FILTER (WHERE event_name = 'friend.request.accepted') "
        "FROM message_event_log"
    )).fetchone()
    mel_trip_created  = mel_row[0] or 0
    mel_trip_accepted = mel_row[1] or 0
    mel_fr_created    = mel_row[2] or 0
    mel_fr_accepted   = mel_row[3] or 0

    part_accepted, part_total = db.session.execute(db.text(
        "SELECT COUNT(*) FILTER (WHERE status = 'accepted'), COUNT(*) "
        "FROM ski_trip_participant"
    )).fetchone()
    part_accepted = part_accepted or 0
    part_total    = part_total    or 0

    # Flywheel invite count = total invite_token rows
    fly_invite = inv_total
    fly_signup = referred

    # ── Section 0: KPIs ──────────────────────────────────────────────────
    stickiness_val = pct(wau, mau)
    kpis = [
        {'label': 'Total Users',       'value': total,
         'sub': 'all time'},
        {'label': 'New Users L30',     'value': new_l30,
         'sub': 'last 30 days',        'conf': 'yellow'},
        {'label': 'WAU',               'value': wau,
         'sub': 'active last 7 days',  'conf': 'yellow'},
        {'label': 'MAU',               'value': mau,
         'sub': 'active last 30 days', 'conf': 'yellow'},
        {'label': 'MAU/WAU Stickiness','value': f'{stickiness_val}%',
         'sub': f'{wau} of {mau} MAU', 'conf': 'yellow'},
        {'label': 'Connected Users',   'value': f'{connected_users} ({pct(connected_users, total)}%)',
         'sub': 'have ≥1 friend',      'highlight': True},
        {'label': 'Friend Connections','value': unique_pairs,
         'sub': 'unique pairs'},
        {'label': 'Avg Friends / User','value': avg_friends_all,
         'sub': f'{avg_friends_connected} per connected user'},
        {'label': 'Invite Acceptance', 'value': f'{pct(inv_used, inv_total)}%',
         'sub': f'{inv_used} of {inv_total} used'},
        {'label': 'Referred Signups',  'value': f'{referred} ({pct(referred, total)}%)',
         'sub': 'via friend invite',   'conf': 'yellow'},
    ]

    # ── Section 6: Dynamic Insight ───────────────────────────────────────
    conn_pct = pct(connected_users, total)
    if conn_pct >= 50:
        insight = f"{conn_pct}% of users have at least one friend connection."
    elif avg_friends_all >= 1.0:
        insight = f"The average BaseLodge user now has {avg_friends_all} friends."
    elif inv_total > 0:
        insight = f"{pct(inv_used, inv_total)}% of friend invitations are accepted."
    else:
        insight = "Collecting more growth data."

    # ── Section 7: Status ────────────────────────────────────────────────
    status = {
        'tracked':    15,
        'green':       9,
        'yellow':      5,
        'red':         1,
        'confidence': 'Growing',
        'caveats': [
            f'{null_created} accounts have no signup date (Oct 2024 cohort — never backfilled)',
            f'{null_last_active} accounts have no session activity recorded (WAU/MAU undercount possible)',
            'MEL invite metrics cover May 2026 only — historical data not available',
            'Viral coefficient (K-factor) intentionally omitted — cannot be reliably computed from current schema',
        ],
    }

    return render_template(
        "admin_growth_intel.html",
        active_tab            = 'growth_intel',
        now                   = now_str,
        total                 = total,
        kpis                  = kpis,
        new_l7                = new_l7,
        new_l30               = new_l30,
        new_l90               = new_l90,
        null_created          = null_created,
        cohort_list           = cohort_list,
        unique_pairs          = unique_pairs,
        connected_users       = connected_users,
        avg_friends_all       = avg_friends_all,
        avg_friends_connected = avg_friends_connected,
        friend_monthly_list   = friend_monthly_list,
        friend_dist           = friend_dist,
        most_connected_name   = most_connected_name,
        most_connected_ct     = mc_ct,
        inv_total             = inv_total,
        inv_used              = inv_used,
        trip_inv_total        = trip_inv_total,
        trip_inv_used         = trip_inv_used,
        mel_trip_created      = mel_trip_created,
        mel_trip_accepted     = mel_trip_accepted,
        mel_fr_created        = mel_fr_created,
        mel_fr_accepted       = mel_fr_accepted,
        part_accepted         = part_accepted,
        part_total            = part_total,
        fly_invite            = fly_invite,
        fly_signup            = fly_signup,
        fly_friend            = fly_friend_users,
        fly_trip              = fly_trip_users,
        wau                   = wau,
        mau                   = mau,
        null_last_active      = null_last_active,
        referred              = referred,
        insight               = insight,
        status                = status,
    )


@app.route("/admin/crm-intel")
@login_required
@admin_required
def admin_crm_intel():
    """CRM & Lifecycle Intelligence v1 — reachability, segments, audiences, power users."""
    import json as _json
    from datetime import datetime as _dt, timedelta
    from collections import defaultdict

    now_str = _admin_now().strftime("%b %d, %Y at %H:%M %Z")
    now = _dt.utcnow()

    def pct(n, d):
        return round(n / d * 100) if d else 0

    _no_pass = {'no_pass', 'no_pass_yet', 'none', ''}

    def _real_pass(pt):
        for s in (pt or '').split(','):
            if s.strip().lower() not in _no_pass:
                return True
        return False

    def _nonempty(val):
        if val is None:
            return False
        if isinstance(val, list):
            return len(val) > 0
        if isinstance(val, str):
            try:
                p = _json.loads(val)
                return isinstance(p, list) and len(p) > 0
            except Exception:
                return False
        return False

    # ── Q1: all users (all CRM fields) ───────────────────────────────────
    user_rows = db.session.execute(db.text(
        'SELECT id, lifecycle_stage, pass_type, open_dates, wish_list_resorts, '
        'last_active_at, created_at, invited_by_user_id, push_notifications_enabled '
        'FROM "user"'
    )).fetchall()
    total = len(user_rows)

    # ── Q2: push device tokens (active) ──────────────────────────────────
    push_rows = db.session.execute(db.text(
        "SELECT user_id, active FROM push_device_token"
    )).fetchall()
    push_active_ids = {r[0] for r in push_rows if r[1]}

    # MEL push delivery stats
    mel_push = db.session.execute(db.text(
        "SELECT "
        "  COUNT(*) FILTER (WHERE delivery_status='sent'),  "
        "  COUNT(*) FILTER (WHERE delivery_status='failed') "
        "FROM message_event_log WHERE channel='push'"
    )).fetchone()
    mel_sent   = mel_push[0] or 0
    mel_failed = mel_push[1] or 0
    push_fail_rate = pct(mel_failed, mel_sent + mel_failed) if (mel_sent + mel_failed) else 0

    # ── Q3: friend counts per user ────────────────────────────────────────
    friend_counts = {r[0]: r[1] for r in db.session.execute(db.text(
        "SELECT user_id, COUNT(*) FROM friend GROUP BY user_id"
    ))}
    friend_ids = set(friend_counts.keys())

    # ── Q4: trip ownership + participation ───────────────────────────────
    trip_owner_ids = {r[0] for r in db.session.execute(
        db.text("SELECT DISTINCT user_id FROM ski_trip")
    )}
    trip_part_ids = {r[0] for r in db.session.execute(
        db.text("SELECT DISTINCT user_id FROM ski_trip_participant WHERE status='accepted'")
    )}
    has_trip_ids = trip_owner_ids | trip_part_ids

    trip_counts_map = {r[0]: r[1] for r in db.session.execute(
        db.text("SELECT user_id, COUNT(*) FROM ski_trip GROUP BY user_id")
    )}

    # ── Q5: invite generators + counts ───────────────────────────────────
    invite_counts_map = {r[0]: r[1] for r in db.session.execute(
        db.text("SELECT inviter_id, COUNT(*) FROM invite_token GROUP BY inviter_id")
    )}
    trip_invite_counts = {r[0]: r[1] for r in db.session.execute(
        db.text("SELECT inviter_user_id, COUNT(*) FROM trip_invite_token GROUP BY inviter_user_id")
    )}
    for uid, ct in trip_invite_counts.items():
        invite_counts_map[uid] = invite_counts_map.get(uid, 0) + ct
    invite_ids = set(invite_counts_map.keys())

    # ── Derive per-user signal sets ───────────────────────────────────────
    pass_ids     = set()
    avail_ids    = set()
    wishlist_ids = set()
    all_ids      = set()
    lc_map       = {}

    for r in user_rows:
        uid, ls = r[0], r[1]
        all_ids.add(uid)
        lc_map[uid] = ls or 'new'
        if _real_pass(r[2]):   pass_ids.add(uid)
        if _nonempty(r[3]):    avail_ids.add(uid)
        if _nonempty(r[4]):    wishlist_ids.add(uid)

    # Engagement score per user (max 5)
    def _score(uid):
        return sum([uid in pass_ids, uid in avail_ids, uid in wishlist_ids,
                    uid in has_trip_ids, uid in friend_ids])

    # Segment counts
    seg = {'new': 0, 'onboarding': 0, 'activated': 0, 'engaged': 0, 'power': 0, 'untracked': 0}
    untracked_ids = set()
    for r in user_rows:
        uid, ls, la = r[0], r[1] or 'new', r[5]
        if la is None:
            untracked_ids.add(uid)
        sc = _score(uid)
        if ls == 'active':
            if sc >= 4:   seg['power']     += 1
            elif sc >= 2: seg['engaged']   += 1
            else:         seg['activated'] += 1
        elif ls == 'onboarding':
            seg['onboarding'] += 1
        else:
            seg['new'] += 1
    seg['untracked'] = len(untracked_ids)

    n_activated = seg['activated']
    n_engaged   = seg['engaged']
    n_power     = seg['power']
    n_push_active = len(push_active_ids)
    n_no_push     = total - n_push_active

    # Lifecycle × reachability
    active_ids_lc     = {r[0] for r in user_rows if (r[1] or 'new') == 'active'}
    new_ids_lc        = {r[0] for r in user_rows if (r[1] or 'new') == 'new'}
    onboard_ids_lc    = {r[0] for r in user_rows if (r[1] or 'new') == 'onboarding'}

    active_push  = len(active_ids_lc & push_active_ids)
    new_push     = len(new_ids_lc & push_active_ids)
    onboard_push = len(onboard_ids_lc & push_active_ids)

    # ── Section 0: CRM KPIs ──────────────────────────────────────────────
    kpis = [
        {'label': 'Reachable Users',    'value': n_push_active,
         'sub': f'{pct(n_push_active, total)}% have active push'},
        {'label': 'Unreachable Users',  'value': n_no_push,
         'sub': 'no active device token'},
        {'label': 'Activated Users',    'value': seg['activated'] + seg['engaged'] + seg['power'],
         'sub': 'lifecycle = active',   'highlight': True},
        {'label': 'Engaged Users',      'value': n_engaged,
         'sub': '2–3 milestones',       'conf': 'yellow'},
        {'label': 'Power Users',        'value': n_power,
         'sub': '4–5 milestones',       'conf': 'yellow'},
        {'label': 'Untracked',          'value': seg['untracked'],
         'sub': 'no session history'},
        {'label': 'Pass Users',         'value': len(pass_ids),
         'sub': f'{pct(len(pass_ids), total)}% of all users'},
        {'label': 'Connected Users',    'value': len(friend_ids),
         'sub': f'{pct(len(friend_ids), total)}% of all users'},
        {'label': 'Push Failure Rate',  'value': f'{push_fail_rate}%',
         'sub': f'{mel_failed} failed / {mel_sent+mel_failed} (May 2026)', 'conf': 'yellow',
         'alert': push_fail_rate >= 20},
    ]

    # ── Section 2: Lifecycle segments ────────────────────────────────────
    lifecycle_segments = [
        {'label': 'New',        'n': seg['new'],        'pct': pct(seg['new'], total),
         'criteria': 'lifecycle_stage = new', 'conf': 'green'},
        {'label': 'Onboarding', 'n': seg['onboarding'], 'pct': pct(seg['onboarding'], total),
         'criteria': 'lifecycle_stage = onboarding', 'conf': 'green'},
        {'label': 'Activated',  'n': seg['activated'],  'pct': pct(seg['activated'], total),
         'criteria': 'active, 0–1 milestones', 'conf': 'yellow'},
        {'label': 'Engaged',    'n': seg['engaged'],    'pct': pct(seg['engaged'], total),
         'criteria': 'active, 2–3 milestones', 'conf': 'yellow'},
        {'label': 'Power',      'n': seg['power'],      'pct': pct(seg['power'], total),
         'criteria': 'active, 4–5 milestones', 'conf': 'yellow'},
        {'label': 'Untracked',  'n': seg['untracked'],  'pct': pct(seg['untracked'], total),
         'criteria': 'last_active_at IS NULL', 'conf': 'yellow'},
    ]

    # ── Section 3: CRM Audiences ──────────────────────────────────────────
    aud_pass_no_avail       = len(pass_ids - avail_ids)
    aud_connected_no_trip   = len(friend_ids - has_trip_ids)
    aud_reachable_no_trip   = len(push_active_ids - has_trip_ids)
    aud_reachable_no_pass   = len(push_active_ids - pass_ids)
    aud_wishlist_no_trip    = len(wishlist_ids - has_trip_ids)
    aud_invite_no_friend    = len(invite_ids - friend_ids)
    aud_activated_no_friend = len(active_ids_lc - friend_ids)
    aud_trip_no_friend      = len(has_trip_ids - friend_ids)

    crm_audiences = [
        # High priority
        {'label': 'Pass, No Availability',      'size': aud_pass_no_avail,
         'priority': 'high',
         'opp': 'Pass holders who haven\'t set open dates — convert planners'},
        {'label': 'Connected, No Trip',          'size': aud_connected_no_trip,
         'priority': 'high',
         'opp': 'Have friends but no trip — social nudge to plan together'},
        {'label': 'Reachable, No Trip',          'size': aud_reachable_no_trip,
         'priority': 'high',
         'opp': 'Push-reachable users who have never planned'},
        # Medium priority
        {'label': 'Reachable, No Pass',          'size': aud_reachable_no_pass,
         'priority': 'medium',
         'opp': 'Push-reachable users missing a pass — pass selection nudge'},
        {'label': 'Wishlist, No Trip',           'size': aud_wishlist_no_trip,
         'priority': 'medium',
         'opp': 'Have intent (wishlist) but no planned trip'},
        {'label': 'Invite Generated, No Friend', 'size': aud_invite_no_friend,
         'priority': 'medium',
         'opp': 'Sent invites but no friend connection made yet'},
        # Low priority
        {'label': 'Activated, No Friends',       'size': aud_activated_no_friend,
         'priority': 'low',
         'opp': 'Active users who are isolated — friend discovery nudge'},
        {'label': 'Trip, No Friends',            'size': aud_trip_no_friend,
         'priority': 'low',
         'opp': 'Planning solo — potential for social conversion'},
    ]
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    crm_audiences.sort(key=lambda x: (priority_order[x['priority']], -x['size']))

    # ── Section 4: Dormancy ───────────────────────────────────────────────
    l30_cutoff = now - timedelta(days=30)
    l60_cutoff = now - timedelta(days=60)
    recently_inactive_30 = sum(
        1 for r in user_rows
        if r[5] is not None and r[5] < l30_cutoff
    )
    recently_inactive_60 = sum(
        1 for r in user_rows
        if r[5] is not None and r[5] < l60_cutoff
    )
    dormancy = {
        'untracked':          seg['untracked'],
        'recently_inactive_30': recently_inactive_30,
        'recently_inactive_60': recently_inactive_60,
    }

    # ── Section 5: Power users ────────────────────────────────────────────
    scored_users = []
    for uid in all_ids:
        fc = friend_counts.get(uid, 0)
        tc = trip_counts_map.get(uid, 0)
        ic = invite_counts_map.get(uid, 0)
        # Composite: normalise each to max, sum ranks
        scored_users.append({'uid': uid, 'friends': fc, 'trips': tc, 'invites': ic,
                              'composite': fc + tc + ic})
    scored_users.sort(key=lambda x: -x['composite'])

    # Fetch names for top 5
    top5_ids = [u['uid'] for u in scored_users[:5]]
    name_map = {}
    if top5_ids:
        name_rows = db.session.execute(
            db.text('SELECT id, first_name, last_name FROM "user" WHERE id = ANY(:ids)'),
            {'ids': top5_ids}
        ).fetchall()
        name_map = {r[0]: f"{r[1]} {r[2]}" for r in name_rows}

    power_users = []
    for rank, u in enumerate(scored_users[:5], 1):
        power_users.append({
            'rank':     rank,
            'name':     name_map.get(u['uid'], f"User {u['uid']}"),
            'friends':  u['friends'],
            'trips':    u['trips'],
            'invites':  u['invites'],
            'score':    u['composite'],
        })

    # ── Section 6: Dynamic Insight ────────────────────────────────────────
    # Ordered preference: largest actionable gap
    if aud_pass_no_avail >= 10:
        insight = (f"{aud_pass_no_avail} users have a pass but no availability set "
                   f"— the biggest planning gap.")
    elif aud_connected_no_trip >= 10:
        insight = (f"{aud_connected_no_trip} connected users haven't planned a trip together yet.")
    elif aud_reachable_no_trip >= 5:
        insight = (f"{aud_reachable_no_trip} push-reachable users have never planned a trip.")
    else:
        insight = f"{pct(len(friend_ids), total)}% of users are connected with friends."

    # ── Section 7: Status ────────────────────────────────────────────────
    status = {
        'crm_audited': 18, 'crm_green': 10, 'crm_yellow': 6, 'crm_red': 2,
        'ret_audited': 10, 'ret_green':  1, 'ret_yellow': 7, 'ret_red': 2,
        'confidence': 'Growing',
        'caveats': [
            'push_notifications_enabled defaults to true for all users — not a reliable opt-in signal',
            'login_count has not been backfilled — 30 of 36 users show 0 logins',
            'Dormancy signals limited by NULL last_active_at (9 untracked accounts)',
            'Retention Intelligence dashboard intentionally deferred — signals too noisy at current scale',
        ],
    }

    return render_template(
        "admin_crm_intel.html",
        active_tab         = 'crm_intel',
        now                = now_str,
        total              = total,
        kpis               = kpis,
        n_push_active      = n_push_active,
        n_no_push          = n_no_push,
        active_push        = active_push,
        new_push           = new_push,
        onboard_push       = onboard_push,
        n_active_lc        = len(active_ids_lc),
        n_new_lc           = len(new_ids_lc),
        n_onboard_lc       = len(onboard_ids_lc),
        mel_sent           = mel_sent,
        mel_failed         = mel_failed,
        push_fail_rate     = push_fail_rate,
        lifecycle_segments = lifecycle_segments,
        crm_audiences      = crm_audiences,
        dormancy           = dormancy,
        power_users        = power_users,
        insight            = insight,
        status             = status,
    )


# ══════════════════════════════════════════════════════════════════════════
# Admin — Active Today API + User Detail
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/api/active-today-users")
@login_required
@admin_required
def admin_api_active_today_users():
    """JSON list of users active today — exact same predicate as Founder Pulse Active count."""
    today_start = _admin_today_start_utc()
    users = User.query.filter(User.last_active_at >= today_start)\
                      .order_by(User.last_active_at.desc()).all()
    result = []
    for u in users:
        first = (u.first_name or "").strip()
        last  = (u.last_name  or "").strip()
        name  = f"{first} {last}".strip() or "Unknown User"
        result.append({
            "id":                u.id,
            "name":              name,
            "state":             u.home_state or "",
            "last_active_at_iso": u.last_active_at.isoformat() if u.last_active_at else None,
        })
    return jsonify(result)


@app.route("/admin/api/new-users-today")
@login_required
@admin_required
def admin_api_new_users_today():
    """JSON list of users who signed up today — same Denver midnight as Founder Pulse."""
    today_start = _admin_today_start_utc()
    users = User.query.filter(User.created_at >= today_start)\
                      .order_by(User.created_at.desc()).all()
    result = []
    for u in users:
        result.append({
            "user_id":    u.id,
            "first_name": (u.first_name or "").strip(),
            "last_name":  (u.last_name  or "").strip(),
            "email":      u.email or "",
            "state":      u.home_state or "",
            "created_at": u.created_at.isoformat() if u.created_at else None,
        })
    return jsonify(result)


@app.route("/admin/test-founder-app-open-push", methods=["POST"])
@login_required
@admin_required
def admin_test_founder_app_open_push():
    """Simulate the app-open founder push for a given user_id.

    POST /admin/test-founder-app-open-push?user_id=42

    Returns JSON describing every gate check so you can see exactly why
    a push would or would not send — without touching session throttle state.
    """
    user_id = request.args.get("user_id", type=int)
    if not user_id:
        return jsonify({"error": "user_id query param required (e.g. ?user_id=42)"}), 400

    # ── Gate 1: feature flag ──────────────────────────────────────────────────
    enabled = os.environ.get("FOUNDER_APP_OPEN_PUSH_ENABLED", "").lower() == "true"
    if not enabled:
        return jsonify({
            "sent": False,
            "reason": "feature_disabled",
            "fix": "Set FOUNDER_APP_OPEN_PUSH_ENABLED=true in Secrets / environment variables",
        })

    # ── Gate 2: user exists ───────────────────────────────────────────────────
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({"sent": False, "reason": "user_not_found", "user_id": user_id}), 404

    # ── Gate 3: not seeded ────────────────────────────────────────────────────
    if getattr(user, "is_seeded", False):
        return jsonify({"sent": False, "reason": "seeded_user", "user_id": user_id})

    # ── Gate 4: not a founder/admin ───────────────────────────────────────────
    admin_emails = {
        e.strip().lower()
        for e in os.environ.get("ALLOWED_ADMIN_EMAILS", "").split(",")
        if e.strip()
    }
    if user.email.lower() in admin_emails:
        return jsonify({
            "sent": False,
            "reason": "founder_user",
            "email": user.email,
            "note": "This is the founder account — it is intentionally excluded",
        })

    # ── Gate 5: founder (richard) has a push token ────────────────────────────
    richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
    if not richard:
        return jsonify({"sent": False, "reason": "no_founder_account"})

    # ── All gates passed — fire synchronously so we get the result inline ─────
    try:
        from services.push_providers import send_onesignal_push as _os_push

        first = (user.first_name or "").strip()
        last  = (user.last_name  or "").strip()
        name  = (first + " " + last).strip()
        state = (user.home_state or "").strip()

        if name and state:
            body = f"{name} opened the app · {state}"
        elif name:
            body = f"{name} opened the app"
        else:
            body = "Someone opened BaseLodge"

        result = _os_push([richard.id], "BaseLodge Opened", body)
        app.logger.warning(
            "[founder_app_open_push] TEST user_id=%d sent=%s skipped=%s error=%s body=%r",
            user_id, result.get("success"), result.get("skipped"), result.get("error"), body,
        )
        return jsonify({
            "sent":              result.get("success", False),
            "skipped":           result.get("skipped"),
            "skipped_reason":    result.get("skipped_reason"),
            "push_error":        result.get("error"),
            "push_body":         body,
            "richard_id":        richard.id,
            "user_id":           user_id,
            "note":              "Session throttle NOT updated — safe to call multiple times for QA",
        })
    except Exception as exc:
        app.logger.exception("[founder_app_open_push] TEST error: %s", exc)
        return jsonify({"sent": False, "reason": "exception", "error": str(exc)}), 500


@app.route("/admin/test-founder-signup-push", methods=["POST"])
@login_required
@admin_required
def admin_test_founder_signup_push():
    """TEST-ONLY — sends a hardcoded founder signup alert to richardbattlebaxter@gmail.com.
    Remove or disable after QA passes.
    """
    try:
        from services.push_providers import send_onesignal_push as _os_push
        richard = User.query.filter_by(email="richardbattlebaxter@gmail.com").first()
        if not richard:
            return jsonify({"success": False, "error": "richard account not found"}), 404

        title = "New BaseLodge User 🎿"
        body  = "Alex Smith just signed up (NJ)\nConnected to James Morgan"
        result = _os_push([richard.id], title, body)
        if result.get("success"):
            return jsonify({"success": True,
                            "skipped": result.get("skipped"),
                            "skipped_reason": result.get("skipped_reason"),
                            "provider_message_id": result.get("provider_message_id")})
        return jsonify({"success": False, "error": result.get("error")}), 500
    except Exception as exc:
        app.logger.exception("[TestFounderPush] unexpected error: %s", exc)
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/admin/users/<int:user_id>")
@login_required
@admin_required
def admin_user_detail(user_id):
    """Admin-only drilldown page for a single user."""
    from datetime import datetime as _dt
    from sqlalchemy import or_ as _or

    user = db.session.get(User, user_id)
    if not user:
        abort(404)

    today_start = _admin_today_start_utc()
    now_utc     = _dt.utcnow()

    is_active_today = bool(user.last_active_at and user.last_active_at >= today_start)

    trips_created = SkiTrip.query.filter_by(user_id=user_id)\
                                 .order_by(SkiTrip.start_date.desc().nullslast())\
                                 .limit(5).all()
    trips_created_count = SkiTrip.query.filter_by(user_id=user_id).count()

    trips_joined_count = SkiTripParticipant.query.filter_by(
        user_id=user_id, status=GuestStatus.ACCEPTED
    ).count()

    friend_count = Friend.query.filter(
        _or(Friend.user_id == user_id, Friend.friend_id == user_id)
    ).count()

    activity_days = db.session.execute(db.text(
        'SELECT COUNT(DISTINCT DATE(created_at)) FROM activity WHERE actor_user_id = :uid'
    ), {"uid": user_id}).scalar() or 0

    activity_label_map = {
        "trip_created":                        "Created Trip",
        "trip_updated":                        "Updated Trip",
        "friend_joined_trip":                  "Friend Joined Trip",
        "trip_invite_received":                "Received Trip Invite",
        "trip_invite_accepted":                "Accepted Trip Invite",
        "trip_invite_declined":                "Declined Trip Invite",
        "connection_accepted":                 "Made a New Friend",
        "trip_overlap":                        "Trip Overlap Found",
        "friend_trip_overlaps_availability":   "Availability Match",
        "carpool_offered":                     "Carpool Offered",
        "join_request_received":               "Received Join Request",
        "join_request_accepted":               "Join Request Accepted",
        "join_request_declined":               "Join Request Declined",
        "trip_location_changed":               "Changed Trip Location",
        "trip_pass_changed":                   "Updated Trip Pass",
    }

    act_rows = Activity.query.filter_by(actor_user_id=user_id)\
                             .order_by(Activity.created_at.desc()).limit(40).all()
    mpv_rows = MountainPageView.query.filter_by(user_id=user_id)\
                                     .order_by(MountainPageView.viewed_at.desc()).limit(40).all()

    combined = []
    for a in act_rows:
        raw = a.type.value if hasattr(a.type, "value") else str(a.type)
        label = activity_label_map.get(raw, raw.replace("_", " ").title())
        combined.append({"label": label, "ts": a.created_at})
    for v in mpv_rows:
        resort_name = (v.resort.name if v.resort else None) or "a mountain"
        combined.append({"label": f"Viewed {resort_name}", "ts": v.viewed_at})

    combined.sort(key=lambda x: x["ts"] or _dt.min, reverse=True)
    recent_activity = combined[:20]

    pass_raw = (user.pass_type or "").strip().lower()
    if not pass_raw or pass_raw in ("no_pass", "no_pass_yet", "none"):
        pass_display = "No pass on file"
    else:
        pass_display = user.pass_type

    state_full = STATE_NAMES.get(user.home_state or "", user.home_state or "")

    def _count_json_list(field):
        if not field:
            return 0
        try:
            parsed = json.loads(field) if isinstance(field, str) else field
            return len(parsed) if isinstance(parsed, list) else 0
        except Exception:
            return 0

    wish_count    = _count_json_list(user.wish_list_resorts)
    visited_count = _count_json_list(user.visited_resort_ids)

    def _fmt_trip_dates(trip):
        try:
            if trip.start_date and trip.end_date:
                sm = trip.start_date.strftime("%b %-d")
                em = trip.end_date.strftime("%-d") if trip.start_date.month == trip.end_date.month \
                     else trip.end_date.strftime("%b %-d")
                return f"{sm}–{em}"
            elif trip.start_date:
                return trip.start_date.strftime("%b %-d")
        except Exception:
            pass
        return ""

    trip_display = []
    for t in trips_created:
        resort_name = t.mountain or (t.resort.name if t.resort else None) or "Unknown"
        trip_display.append({
            "name":  resort_name,
            "dates": _fmt_trip_dates(t),
        })

    now_str = _admin_now().strftime("%b %d, %Y at %H:%M %Z")

    return render_template(
        "admin_user_detail.html",
        active_tab          = "dashboard",
        user                = user,
        is_active_today     = is_active_today,
        state_full          = state_full,
        pass_display        = pass_display,
        trips_created_count = trips_created_count,
        trips_joined_count  = trips_joined_count,
        friend_count        = friend_count,
        activity_days       = activity_days,
        recent_activity     = recent_activity,
        wish_count          = wish_count,
        visited_count       = visited_count,
        trip_display        = trip_display,
        now_str             = now_str,
        now_utc             = now_utc,
    )


# ============================================================================
# ADMIN — USER LOOKUP (auth diagnostic by email)
# ============================================================================

@app.route("/admin/user-lookup")
@login_required
@admin_required
def admin_user_lookup():
    """
    Safe auth diagnostic: look up a user by email and return key account fields.
    Never exposes password hashes or secrets.

    GET /admin/user-lookup?email=someone@example.com
    """
    email = request.args.get("email", "").lower().strip()
    if not email:
        return jsonify({"error": "Provide ?email= query parameter"}), 400

    user = User.query.filter(sa.func.lower(User.email) == email).first()
    if not user:
        return jsonify({"found": False, "email_queried": email}), 404

    return jsonify({
        "found": True,
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "auth_provider": user.auth_provider,
        "has_password": bool(user.password_hash),
        "is_verified": user.is_verified,
        "is_seeded": user.is_seeded,
        "lifecycle_stage": user.lifecycle_stage,
        "login_count": user.login_count,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "last_active_at": user.last_active_at.isoformat() if user.last_active_at else None,
        "password_changed_at": user.password_changed_at.isoformat() if user.password_changed_at else None,
        "admin_detail_url": url_for("admin_user_detail", user_id=user.id, _external=False),
    })


# ============================================================================
# ADMIN — APP STORE PERFORMANCE
# ============================================================================

@app.route("/admin/app-store")
@login_required
@admin_required
def admin_app_store():
    """App Store Performance — reads from AppStoreMetric table only.
    No live calls to Apple or Google at page-render time."""
    from models import AppStoreMetric
    from datetime import date, timedelta
    from collections import defaultdict

    today     = date.today()
    yesterday = today - timedelta(days=1)
    l7_start  = today - timedelta(days=7)
    l30_start = today - timedelta(days=30)

    rows = (
        AppStoreMetric.query
        .filter(AppStoreMetric.report_date >= l30_start)
        .order_by(AppStoreMetric.report_date.desc())
        .all()
    )

    def _rows_for(platform):
        return [r for r in rows if r.platform == platform]

    def _dl_sum(platform_rows):
        vals = [r.downloads for r in platform_rows if r.downloads is not None]
        return sum(vals) if vals else None

    def _latest_rating(platform_rows):
        for r in platform_rows:
            if r.rating is not None:
                return r.rating
        return None

    def _latest_reviews(platform_rows):
        for r in platform_rows:
            if r.review_count is not None:
                return r.review_count
        return None

    def _dl_yesterday(platform_rows):
        for r in platform_rows:
            if r.report_date == yesterday and r.downloads is not None:
                return r.downloads
        return None

    def _dl_l7(platform_rows):
        vals = [r.downloads for r in platform_rows
                if r.report_date >= l7_start and r.downloads is not None]
        return sum(vals) if vals else None

    def _sparkline_series(platform_rows):
        """Return 30-element daily download list for sparkline (oldest → newest)."""
        by_date = {r.report_date: (r.downloads or 0) for r in platform_rows
                   if r.downloads is not None}
        series = []
        for delta in range(29, -1, -1):
            d = today - timedelta(days=delta)
            series.append(by_date.get(d, 0))
        return series

    def _crash_rate_avg(platform_rows):
        vals = [r.crashes for r in platform_rows if r.crashes is not None]
        return round(sum(vals) / len(vals), 3) if vals else None

    ios_rows     = _rows_for("ios")
    android_rows = _rows_for("android")

    ios_kpis = dict(
        dl_yesterday = _dl_yesterday(ios_rows),
        dl_l7        = _dl_l7(ios_rows),
        dl_l30       = _dl_sum(ios_rows),
        rating       = _latest_rating(ios_rows),
        review_count = _latest_reviews(ios_rows),
        crash_rate   = _crash_rate_avg(ios_rows),
        spark        = _sparkline_series(ios_rows),
    )
    android_kpis = dict(
        dl_yesterday = _dl_yesterday(android_rows),
        dl_l7        = _dl_l7(android_rows),
        dl_l30       = _dl_sum(android_rows),
        rating       = _latest_rating(android_rows),
        review_count = _latest_reviews(android_rows),
        crash_rate   = _crash_rate_avg(android_rows),
        spark        = _sparkline_series(android_rows),
    )

    total_dl_l30 = (
        (ios_kpis["dl_l30"] or 0) + (android_kpis["dl_l30"] or 0)
        if ios_kpis["dl_l30"] is not None or android_kpis["dl_l30"] is not None
        else None
    )

    ios_configured     = all(os.environ.get(k) for k in (
        "ASC_KEY_P8", "ASC_KEY_ID", "ASC_ISSUER_ID", "ASC_VENDOR_NO"))
    android_configured = all(os.environ.get(k) for k in (
        "GOOGLE_PLAY_SERVICE_ACCOUNT_JSON", "GOOGLE_PLAY_PACKAGE_NAME"))

    has_data      = bool(rows)
    last_refreshed = (
        max((r.fetched_at for r in rows), default=None)
        if rows else None
    )
    last_refreshed_str = (
        last_refreshed.strftime("%b %d, %Y at %H:%M UTC") if last_refreshed else "Never"
    )

    return render_template(
        "admin_app_store.html",
        active_tab          = "app_store",
        now                 = _fmt_admin_now(),
        ios_kpis            = ios_kpis,
        android_kpis        = android_kpis,
        total_dl_l30        = total_dl_l30,
        has_data            = has_data,
        ios_configured      = ios_configured,
        android_configured  = android_configured,
        last_refreshed_str  = last_refreshed_str,
    )


@app.route("/admin/app-store/refresh", methods=["POST"])
@login_required
@admin_required
def admin_app_store_refresh():
    """Pull the latest metrics from Apple / Google and upsert into AppStoreMetric.

    - Checks for credentials before attempting each platform.
    - Fails gracefully per-platform; one failure does not abort the other.
    - Returns a flash message + redirect back to /admin/app-store.
    - Idempotent: re-running overwrites (upserts) the same (platform, date) rows.
    """
    from models import AppStoreMetric
    from datetime import datetime as _dt

    messages = []
    errors   = []

    # ── iOS / App Store Connect ───────────────────────────────────────────────
    ios_configured = all(os.environ.get(k) for k in (
        "ASC_KEY_P8", "ASC_KEY_ID", "ASC_ISSUER_ID", "ASC_VENDOR_NO"))

    if ios_configured:
        try:
            from services.app_store_client import fetch_daily_downloads, fetch_app_rating

            dl_rows = fetch_daily_downloads(days_back=30)
            rating  = fetch_app_rating()

            upserted = 0
            for row in dl_rows:
                existing = AppStoreMetric.query.filter_by(
                    platform="ios", report_date=row["report_date"]
                ).first()
                if existing:
                    existing.downloads  = row["downloads"]
                    existing.fetched_at = _dt.utcnow()
                else:
                    db.session.add(AppStoreMetric(
                        platform    = "ios",
                        report_date = row["report_date"],
                        downloads   = row["downloads"],
                        fetched_at  = _dt.utcnow(),
                    ))
                upserted += 1

            if rating:
                latest = AppStoreMetric.query.filter_by(platform="ios").order_by(
                    AppStoreMetric.report_date.desc()
                ).first()
                if latest:
                    latest.rating       = rating["rating"]
                    latest.review_count = rating["review_count"]
                    latest.fetched_at   = _dt.utcnow()

            db.session.commit()
            messages.append(f"iOS: {upserted} day(s) upserted.")
        except Exception as exc:
            db.session.rollback()
            errors.append(f"iOS fetch failed: {exc}")
    else:
        messages.append("iOS: skipped (ASC credentials not configured).")

    # ── Android / Google Play ─────────────────────────────────────────────────
    android_configured = all(os.environ.get(k) for k in (
        "GOOGLE_PLAY_SERVICE_ACCOUNT_JSON", "GOOGLE_PLAY_PACKAGE_NAME"))

    if android_configured:
        try:
            from services.play_store_client import fetch_daily_installs, fetch_daily_crashes

            install_rows = fetch_daily_installs(days_back=30)
            crash_rows   = fetch_daily_crashes(days_back=30)

            crash_by_date = {r["report_date"]: r["crashes"] for r in crash_rows}

            upserted = 0
            for row in install_rows:
                existing = AppStoreMetric.query.filter_by(
                    platform="android", report_date=row["report_date"]
                ).first()
                crashes_val = crash_by_date.get(row["report_date"])
                if existing:
                    existing.downloads  = row["downloads"]
                    existing.crashes    = crashes_val
                    existing.fetched_at = _dt.utcnow()
                else:
                    db.session.add(AppStoreMetric(
                        platform    = "android",
                        report_date = row["report_date"],
                        downloads   = row["downloads"],
                        crashes     = crashes_val,
                        fetched_at  = _dt.utcnow(),
                    ))
                upserted += 1

            db.session.commit()
            messages.append(f"Android: {upserted} day(s) upserted.")
        except Exception as exc:
            db.session.rollback()
            errors.append(f"Android fetch failed: {exc}")
    else:
        messages.append("Android: skipped (Play credentials not configured).")

    # ── Flash result ─────────────────────────────────────────────────────────
    if errors:
        flash("Refresh completed with errors — " + " | ".join(messages + errors), "error")
    else:
        flash("Refresh complete — " + " | ".join(messages), "success")

    return redirect(url_for("admin_app_store"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
